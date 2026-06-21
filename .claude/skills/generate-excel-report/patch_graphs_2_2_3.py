"""Re-chart existing reports: rebuild the ``Graphs`` worksheet in place.

Originally written for the 2.2.3 reports (whose ``Graphs`` charts pre-dated the
unified styling and auto-scaled their axes / plotted every row), this is a
version-neutral RE-CHART tool: it rebuilds the ``Graphs`` worksheet of any
existing report from the current shared ``CHART_SPECS`` so the embedded charts
pick up the latest FIXED axes (e.g. EP 0–3 kWh/km, mass 0–45000 kg, temperature
−5..30 °C) WITHOUT re-running the pipeline — the cheap way to update
already-generated reports after an axis / styling change. It rebuilds the
``Graphs`` worksheet so that:

  - axes are FIXED constants shared across all vehicles / reports, and
  - scatter + linear trendline are computed on filtered (driving-leg, in-range)
    data only,

using the exact same ``CHART_SPECS`` config that the generator
(``jolt_toolkit.report_generator.report_builder``) now uses — so the two
rendering paths (xlsxwriter generation / openpyxl patch) stay in lockstep.

This is an in-place maintenance pass: it does NOT re-run the pipeline and does
NOT bump the package version. It only rewrites the ``Graphs`` (and a hidden
``GraphsData``) sheet; the ``Report`` and ``Definitions`` sheets — including
``=NA()`` formulas and SRF hyperlinks — are preserved by the openpyxl round-trip.

Safety:
  - every file is backed up to a sibling ``.bak/`` directory before editing
    (existing ``.bak`` copies are never overwritten and never patched);
  - ``_finetuned.xlsx`` reports are patched too;
  - rows are read with ``data_only=True`` so ``=NA()`` cells resolve to ``None``
    and are filtered out, matching the generator's NaN handling.

Usage (from repo root, with ``PYTHONPATH=src`` or an editable install):

    # re-chart a whole report-database version subdir (optionally one REG):
    python .claude/skills/generate-excel-report/patch_graphs_2_2_3.py --version 2.2.6 [--glob REG]

    # re-chart an explicit directory (recursively) or a single xlsx file:
    python .claude/skills/generate-excel-report/patch_graphs_2_2_3.py excel_report_database/2.2.6/YK73WFN
    python .claude/skills/generate-excel-report/patch_graphs_2_2_3.py path/to/jolt_report_REG_start_end.xlsx

Options:
  PATH          optional explicit xlsx file or directory to re-chart. When given,
                it overrides --version (the version DB-root is not used); a
                directory is searched recursively for jolt_report_*.xlsx.
  --dry-run     report what would change, write nothing.
  --version     report-database version subdir to re-chart (default 2.2.3); used
                only when PATH is omitted.
  --glob        only re-chart files whose path matches this substring (e.g. a REG).
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.xml.constants import DRAWING_NS
from openpyxl.xml.functions import Element, SubElement

# Import the single shared chart-spec config + style + filtering helpers from the
# toolkit so the patch and the generator can never drift apart.
from jolt_toolkit.report_generator.report_builder import (
    CHART_STYLE,
    chart_row_step,
    chart_specs_for,
    empty_chart_note,
    empty_note_extent,
    _chart_subtitle,
    _filtered_chart_points,
)

REPO_ROOT = Path(__file__).resolve().parents[3]
DB_ROOT = REPO_ROOT / "excel_report_database"


class _AlphaColorChoice(ColorChoice):
    """A ``solidFill`` colour that carries an ``<a:alpha>`` transparency child.

    openpyxl's high-level ``ColorChoice.srgbClr`` is a plain string and cannot
    emit the ``<a:alpha val="..."/>`` element needed for a soft, semi-transparent
    marker fill. This subclass overrides ``to_tree`` to write
    ``<a:solidFill><a:srgbClr val=..><a:alpha val=../></a:srgbClr></a:solidFill>``
    directly, matching the opacity used by the xlsxwriter generation path
    (``CHART_STYLE['scatter_opacity']``). ``alpha`` is stored in DrawingML
    thousandths-of-a-percent (e.g. 72 % opacity → 72000).
    """

    def __init__(self, rgb: str, opacity_pct: int):
        super().__init__(srgbClr=rgb)
        self._alpha_thousandths = int(round(opacity_pct * 1000))

    def to_tree(self, tagname=None, idx=None, namespace=None):  # noqa: ARG002
        fill = Element(f"{{{DRAWING_NS}}}solidFill")
        clr = SubElement(fill, f"{{{DRAWING_NS}}}srgbClr", {"val": self.srgbClr})
        SubElement(clr, f"{{{DRAWING_NS}}}alpha", {"val": str(self._alpha_thousandths)})
        return fill


def _styled_marker() -> Marker:
    """Build the shared single-colour, semi-transparent, borderless marker."""
    marker = Marker(symbol="circle", size=CHART_STYLE["marker_size"])
    gp = GraphicalProperties()
    gp.solidFill = _AlphaColorChoice(
        CHART_STYLE["scatter_rgb"], CHART_STYLE["scatter_opacity"]
    )
    gp.line = LineProperties(noFill=True)  # no marker border
    marker.graphicalProperties = gp
    return marker


def _light_gridlines() -> ChartLines:
    """Build light-grey, high-transparency major gridlines.

    Mirrors the xlsxwriter path's ``major_gridlines`` (colour
    ``CHART_STYLE['grid_rgb']``, opacity ``grid_opacity``, width ``grid_width_emu``),
    so both rendering paths show the same soft grid. Minor gridlines are left
    unset (off).
    """
    gl = ChartLines()
    gp = GraphicalProperties()
    lp = LineProperties(w=CHART_STYLE["grid_width_emu"])
    # ``_AlphaColorChoice`` emits a full ``<a:solidFill><a:srgbClr><a:alpha/>``
    # block — the form a line fill needs (a bare ``<a:srgbClr>`` under ``<a:ln>``
    # is invalid and Excel silently falls back to black gridlines).
    lp.solidFill = _AlphaColorChoice(
        CHART_STYLE["grid_rgb"], CHART_STYLE["grid_opacity"]
    )
    gp.line = lp
    gl.spPr = gp
    return gl


def _write_empty_note(ws, gi: int, text: str) -> None:
    """Replace chart ``gi`` with a centred grey "no data" note panel.

    Mirrors the xlsxwriter generator path (``_write_excel_report``): when a chart
    has zero plottable points (e.g. a vehicle with no mass data → EP-vs-Mass can
    plot nothing) we draw a merged-cell note over the chart's footprint instead of
    leaving an empty scatter frame. A merged cell is used (not a textbox) because
    it is the one primitive both render paths write identically. The panel extent
    and message come from the shared toolkit helpers, and the style fields from the
    shared ``CHART_STYLE``, so the two paths stay in lockstep.

    Styling is applied to every cell of the range BEFORE merging: openpyxl forbids
    writing a value to a merged-away cell, and styling all cells gives the clean
    outer border / full fill that Excel renders for a merged range (interior
    borders are hidden by the merge, leaving a single bordered panel).
    """
    fr, fc, lr, lc = empty_note_extent(gi)  # 0-based
    r1, c1, r2, c2 = fr + 1, fc + 1, lr + 1, lc + 1  # openpyxl is 1-based

    fill = PatternFill(fill_type="solid", fgColor=CHART_STYLE["empty_note_fill_rgb"])
    side = Side(style="thin", color=CHART_STYLE["empty_note_border_rgb"])
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border

    top_left = ws.cell(row=r1, column=c1, value=text)
    top_left.font = Font(
        size=CHART_STYLE["empty_note_font_size"],
        italic=True,
        color=CHART_STYLE["empty_note_rgb"],
    )
    top_left.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _tick_text_props(size_pt: int) -> RichText:
    """A ``txPr`` rich-text block that sets the axis tick-label font size."""
    cp = CharacterProperties(sz=size_pt * 100)  # DrawingML wants size × 100
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def _axis_title(text: str, size_pt: int) -> Title:
    """An axis title rendered at ``size_pt`` (openpyxl ignores font on a bare str)."""
    cp = CharacterProperties(sz=size_pt * 100)
    para = Paragraph(
        pPr=ParagraphProperties(defRPr=cp), r=[RegularTextRun(rPr=cp, t=text)]
    )
    return Title(tx=Text(rich=RichText(p=[para])))


def _chart_title(title: str, subtitle: str, title_pt: int, subtitle_pt: int) -> Title:
    """A two-line chart title: bold ``title`` then a smaller ``subtitle`` line.

    Each line carries its own font size, so the fit-equation / R² subtitle reads
    smaller than the title band even though they share one title object.
    """
    cp1 = CharacterProperties(sz=title_pt * 100, b=True)
    paras = [
        Paragraph(
            pPr=ParagraphProperties(defRPr=cp1), r=[RegularTextRun(rPr=cp1, t=title)]
        )
    ]
    if subtitle:
        cp2 = CharacterProperties(sz=subtitle_pt * 100, b=False)
        paras.append(
            Paragraph(
                pPr=ParagraphProperties(defRPr=cp2),
                r=[RegularTextRun(rPr=cp2, t=subtitle)],
            )
        )
    return Title(tx=Text(rich=RichText(p=paras)))


def _read_report_rows(xlsx_path: Path):
    """Read the ``Report`` sheet into row tuples matching the generator layout.

    Returns ``(headers, rows)`` where ``headers`` is the full header tuple
    (including the leading 'Leg Number' column) and ``rows`` is a list of value
    tuples WITHOUT the Leg-Number column (i.e. ``row[0]`` is 'Leg Type'),
    exactly the shape ``_filtered_chart_points`` expects. ``=NA()`` cells read
    back as ``None`` under ``data_only=True`` and are left as-is (the filter
    drops non-numeric values).
    """
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    try:
        ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return None, []
        headers = tuple(h for h in header_row if h is not None)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            # Drop the leading Leg-Number column to match the generator's `rows`.
            rows.append(tuple(row[1:]))
        return headers, rows
    finally:
        wb.close()


def _rebuild_graphs(xlsx_path: Path, headers: tuple, rows: list) -> tuple[int, int]:
    """Rewrite the ``Graphs`` + hidden ``GraphsData`` sheets in place.

    Returns ``(n_charts, n_notes)`` — the number of scatter charts drawn and the
    number of charts replaced by a "no data" note panel. The ``Report`` /
    ``Definitions`` sheets are untouched.
    """
    specs = chart_specs_for(headers)
    # Row-relative index map: rows omit the leading 'Leg Number' column.
    row_col_idx = {h: i - 1 for i, h in enumerate(headers) if i >= 1}

    wb = load_workbook(str(xlsx_path))  # formulas preserved (data_only=False)
    # Drop any pre-existing chart sheets so the rebuild is idempotent.
    for name in ("Graphs", "GraphsData"):
        if name in wb.sheetnames:
            del wb[name]
    graphs_ws = wb.create_sheet("Graphs")
    data_ws = wb.create_sheet("GraphsData")
    data_ws.sheet_state = "hidden"

    data_col = 1  # 1-based column on GraphsData
    n_charts = 0
    n_notes = 0
    for gi, spec in enumerate(specs):
        x_hdr, y_hdr = spec["x_hdr"], spec["y_hdr"]
        if x_hdr not in row_col_idx or y_hdr not in row_col_idx:
            continue
        pts = _filtered_chart_points(rows, row_col_idx, spec)

        xcol, ycol = data_col, data_col + 1
        data_ws.cell(row=1, column=xcol, value=x_hdr)
        data_ws.cell(row=1, column=ycol, value=y_hdr)
        for di, (xv, yv) in enumerate(pts, start=2):
            data_ws.cell(row=di, column=xcol, value=xv)
            data_ws.cell(row=di, column=ycol, value=yv)
        n_pts = len(pts)
        # Advance the GraphsData column pair now so it stays in lockstep whether
        # this chart is drawn or replaced by a note.
        data_col += 2

        # Graceful degradation: with no plottable points (e.g. a vehicle whose
        # mass column is null for the whole period) replace the empty scatter
        # frame with a centred text note over the chart's footprint, matching the
        # xlsxwriter generator path and leaving the other charts' positions intact.
        if n_pts == 0:
            _write_empty_note(graphs_ws, gi, empty_chart_note(spec))
            n_notes += 1
            continue

        # Short title on its own band; fit equation + R² as a 2nd title line so
        # they no longer overlap each other or sit on top of the data points.
        title = spec.get("title", f"{y_hdr} vs {x_hdr}")
        subtitle = _chart_subtitle(pts) if CHART_STYLE["show_fit_in_subtitle"] else ""

        chart = ScatterChart()
        chart.scatterStyle = "marker"
        # Two-line title (bold title + smaller fit/R² subtitle), each line at its
        # own font size — openpyxl ignores fonts on a plain ``str`` title.
        chart.title = _chart_title(
            title,
            subtitle,
            CHART_STYLE["title_font_size"],
            CHART_STYLE["subtitle_font_size"],
        )
        # Axis titles as Title objects so the font size actually takes effect.
        # The DISPLAY title is decoupled from the data header: the series still
        # references ``x_hdr`` / ``y_hdr`` (and so does the GraphsData block above);
        # only the printed label may differ via the optional ``x_title`` /
        # ``y_title`` spec fields (e.g. "Vehicle gross weight (kg)" on the mass
        # chart while the column stays "Vehicle Mass (kg)").
        chart.x_axis.title = _axis_title(
            spec.get("x_title", x_hdr), CHART_STYLE["axis_title_font_size"]
        )
        chart.y_axis.title = _axis_title(
            spec.get("y_title", y_hdr), CHART_STYLE["axis_title_font_size"]
        )
        # Tick-label font size on both axes.
        chart.x_axis.txPr = _tick_text_props(CHART_STYLE["tick_font_size"])
        chart.y_axis.txPr = _tick_text_props(CHART_STYLE["tick_font_size"])
        chart.x_axis.scaling.min = spec["x_min"]
        chart.x_axis.scaling.max = spec["x_max"]
        chart.x_axis.majorUnit = spec["x_major"]
        chart.y_axis.scaling.min = spec["y_min"]
        chart.y_axis.scaling.max = spec["y_max"]
        chart.y_axis.majorUnit = spec["y_major"]
        # openpyxl hides axis lines/ticks unless delete is explicitly False.
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        # Light-grey, high-transparency major gridlines on BOTH axes; minor
        # gridlines stay off (matches the xlsxwriter path and the paper grid).
        chart.x_axis.majorGridlines = _light_gridlines()
        chart.y_axis.majorGridlines = _light_gridlines()
        chart.x_axis.minorGridlines = None
        chart.y_axis.minorGridlines = None
        # Legend in the top-right corner (matches the xlsxwriter path), with the
        # shared legend font size.
        chart.legend = Legend()
        chart.legend.position = CHART_STYLE["legend_pos_openpyxl"]
        chart.legend.overlay = False
        chart.legend.txPr = _tick_text_props(CHART_STYLE["legend_font_size"])

        if n_pts > 0:
            x_ref = Reference(data_ws, min_col=xcol, min_row=2, max_row=1 + n_pts)
            y_ref = Reference(data_ws, min_col=ycol, min_row=2, max_row=1 + n_pts)
            series = Series(y_ref, x_ref, title=spec.get("series_name", y_hdr))
            series.marker = _styled_marker()  # single steelblue, alpha, no border
            series.graphicalProperties.line.noFill = True  # scatter, no connecting line
            # Fit equation + R² live in the subtitle now, so the trendline only
            # draws the line itself (no overlapping floating labels).
            series.trendline = Trendline(
                trendlineType="linear", dispEq=False, dispRSqr=False
            )
            chart.series.append(series)

        chart.height = CHART_STYLE["chart_height_cm"]
        chart.width = CHART_STYLE["chart_width_cm"]
        # Pin the inner plot area with the shared CHART_STYLE margin fractions
        # (edge-anchored, same numbers as the xlsxwriter ``set_plotarea`` path)
        # so the big-font axis titles sit clear of the tick labels instead of
        # overlapping them. NB: set ``chart.layout`` (not ``chart.plot_area.layout``)
        # — openpyxl's ScatterChart._write copies ``chart.layout`` onto the plot
        # area and would otherwise overwrite a layout set directly on plot_area.
        # ``layoutTarget='inner'`` makes x/y/w/h address the plot box ITSELF
        # (excluding axis labels), matching xlsxwriter's set_plotarea so both
        # render paths place the box identically.
        chart.layout = Layout(
            manualLayout=ManualLayout(
                layoutTarget="inner",
                xMode="edge",
                yMode="edge",
                x=CHART_STYLE["plot_x"],
                y=CHART_STYLE["plot_y"],
                w=CHART_STYLE["plot_w"],
                h=CHART_STYLE["plot_h"],
            )
        )
        # Anchor each chart a fixed number of rows below the previous one. The
        # step (chart height in rows + gap) comes from the SHARED
        # ``chart_row_step`` helper in the toolkit — the same step the xlsxwriter
        # generator uses — so adjacent charts never overlap and both render paths
        # stack the charts identically.
        graphs_ws.add_chart(chart, f"A{gi * chart_row_step() + 1}")
        n_charts += 1

    wb.save(str(xlsx_path))
    wb.close()
    return n_charts, n_notes


def _iter_target_files(version: str, glob_filter: str | None):
    """Yield report xlsx paths under a version DB-root, excluding .bak copies."""
    root = DB_ROOT / version
    if not root.exists():
        print(f"[error] no such version dir: {root}", file=sys.stderr)
        return
    for p in sorted(root.rglob("jolt_report_*.xlsx")):
        parts = {seg.lower() for seg in p.parts}
        if ".bak" in parts:
            continue
        if glob_filter and glob_filter not in str(p):
            continue
        yield p


def _iter_path_files(path: Path, glob_filter: str | None):
    """Yield report xlsx paths under an explicit file or directory ``path``.

    A single ``.xlsx`` file is yielded directly; a directory is searched
    recursively for ``jolt_report_*.xlsx`` (excluding any ``.bak`` copies). This
    lets the re-chart tool target an arbitrary location (e.g. a tmp copy for
    verification, or one report-database REG dir) independent of the version
    DB-root used when ``PATH`` is omitted.
    """
    if not path.exists():
        print(f"[error] no such path: {path}", file=sys.stderr)
        return
    if path.is_file():
        if path.suffix.lower() == ".xlsx" and (
            not glob_filter or glob_filter in str(path)
        ):
            yield path
        return
    for p in sorted(path.rglob("jolt_report_*.xlsx")):
        parts = {seg.lower() for seg in p.parts}
        if ".bak" in parts:
            continue
        if glob_filter and glob_filter not in str(p):
            continue
        yield p


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "path",
        nargs="?",
        default=None,
        help=(
            "optional explicit xlsx file or directory to re-chart; overrides "
            "--version (a directory is searched recursively)."
        ),
    )
    ap.add_argument(
        "--dry-run", action="store_true", help="report what would change, write nothing"
    )
    ap.add_argument(
        "--version",
        default="2.2.3",
        help="report-database version subdir (default 2.2.3); used only when PATH is omitted",
    )
    ap.add_argument(
        "--glob",
        default=None,
        help="only re-chart files whose path contains this substring",
    )
    args = ap.parse_args()

    if args.path:
        files = list(_iter_path_files(Path(args.path), args.glob))
        source = str(args.path)
    else:
        files = list(_iter_target_files(args.version, args.glob))
        source = args.version
    if not files:
        print("no matching files found.")
        return 1

    print(
        f"found {len(files)} report(s) under {source}"
        f"{' matching ' + args.glob if args.glob else ''}."
    )
    patched = 0
    for p in files:
        try:
            headers, rows = _read_report_rows(p)
        except Exception as exc:  # noqa: BLE001 - report and continue
            print(f"  [skip] {p.name}: cannot read Report sheet ({exc})")
            continue
        if not headers:
            print(f"  [skip] {p.name}: empty / unreadable Report sheet")
            continue

        if args.dry_run:
            specs = chart_specs_for(headers)
            row_col_idx = {h: i - 1 for i, h in enumerate(headers) if i >= 1}
            counts = [
                len(_filtered_chart_points(rows, row_col_idx, s))
                for s in specs
                if s["x_hdr"] in row_col_idx and s["y_hdr"] in row_col_idx
            ]
            fuel = "diesel" if "Fuel Consumption (L/100km)" in headers else "EV"
            print(
                f"  [dry] {p.name}: {fuel}, {len(counts)} charts, "
                f"filtered points per chart = {counts}"
            )
            continue

        # Back up to a sibling .bak/ dir; never overwrite an existing backup.
        bak_dir = p.parent / ".bak"
        bak_dir.mkdir(exist_ok=True)
        bak_path = bak_dir / p.name
        if not bak_path.exists():
            shutil.copy2(p, bak_path)

        try:
            n_charts, n_notes = _rebuild_graphs(p, headers, rows)
            note_suffix = f", {n_notes} no-data note(s)" if n_notes else ""
            print(f"  [ok]  {p.name}: rebuilt {n_charts} charts{note_suffix}")
            patched += 1
        except Exception as exc:  # noqa: BLE001
            print(f"  [FAIL] {p.name}: {exc} — restoring from backup")
            if bak_path.exists():
                shutil.copy2(bak_path, p)

    if not args.dry_run:
        print(f"patched {patched}/{len(files)} report(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
