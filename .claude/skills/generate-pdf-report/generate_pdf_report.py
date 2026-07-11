"""从 JOLT 报告管线产物（excel_report_database/<ver>/<REG>/*.xlsx）生成工业界 PDF 简报。

流程：xlsx → 计算真实 KPI + 生成 JOLT 风格 matplotlib 图 → 渲染 Jinja2 模板
→ headless Chrome 出 PDF，并打印「字段适用性审计」（哪些是真实数据 / 哪些 N/A）。

用法（从仓库根运行）：
    python .claude/skills/generate-pdf-report/generate_pdf_report.py --reg YK73WFN --period 20250301_20250601
    python .claude/skills/generate-pdf-report/generate_pdf_report.py --reg YK73WFN --period 20250301_20250601 --base  # 用非 finetuned

产物写到 pdf_report_workspace/output_by_TBD/<REG>_<period>/（工作区；定稿后整体改名为 output_by_<YYYYMMDD> 快照，gitignored）。

图表风格刻意对齐 analysis/ 下的图：白底 + 浅网格 + 散点 + 线性拟合 + ±1σ 阴影带 + 图例。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
import sys
import time
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from jinja2 import Template
from scipy.optimize import curve_fit
from scipy.signal import find_peaks
from scipy.stats import gaussian_kde

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
from build_pdf import html_to_pdf  # noqa: E402

PROJECT = Path(__file__).resolve().parents[3]  # repo root (script lives at .claude/skills/generate-pdf-report/)
WORKSPACE = PROJECT / "pdf_report_workspace"  # artefact store: output_by_*/ + HERE-tile cache/
TEMPLATE = ROOT / "templates" / "report_template.html.j2"

DRIVE = {"Outbound", "Return", "In Transit", "Round Trip", "In House"}
# Away legs are real charging sessions (public/en-route chargers) and must count: Home-only
# under-counted fleet-wide (EV73SAL ~83, LN25NKE ~58 Away legs) and broke down entirely once the
# 2.2.8 recompute reclassified CMZ6260's depot charges to Away (Home 68->8: sessions showed 0
# against a 10.6 MWh charged total on the same page).
CHARGE = {"DC Home", "AC Home", "Charge Home", "DC Away", "AC Away", "Charge Away"}

# JOLT 配置（plot_config.json 子集 + vehicles.json 取电池容量）
PLOT_CFG = json.loads((PROJECT / "src/jolt_toolkit/configs/plot_config.json").read_text(encoding="utf-8"))
VEHICLES = json.loads((PROJECT / "src/jolt_toolkit/configs/vehicles.json").read_text(encoding="utf-8"))
# Skill-local per-vehicle briefing specs (unladen mass overrides + provenance). Kept inside the
# skill (not vehicles.json) to preserve the skill's self-containment; absent REG -> auto fallback.
_SPECS_PATH = ROOT / "briefing_vehicle_specs.json"
BRIEFING_SPECS = json.loads(_SPECS_PATH.read_text(encoding="utf-8")) if _SPECS_PATH.exists() else {}
OEM_BLUE = "#1f77b4"  # Volvo / matplotlib 默认蓝，贴合 analysis 图风格

# EP 合理范围：超出即视为不合理数据，直接剔除
EP_MIN, EP_MAX = 0.0, 3.0  # EP 图 y 轴范围（恒 0–3，与清洗界限分开）
# EP 清洗界限：EP ∉ [EP_CLEAN_MIN, EP_CLEAN_MAX] 视为不合理数据并剔除。下限 0.3 排除
# 极低 EP（极短/残段，如续航投影被 600/0.08 拉到上千 km 的伪值）；上限 3 排除异常高值。
EP_CLEAN_MIN, EP_CLEAN_MAX = 0.3, 3.0

# (Removed 2026-07-02) The SOC-quantisation guard `SOC_QUANT_GUARD_PCT` — a |SOC change| ≤ 1%
# filter that dropped short-trip EP — is gone. It was a downstream workaround for a jolt_toolkit
# ±1σ effective-capacity correction that re-derived short-trip energy from the coarse integer SOC
# (spurious low-EP band). Fixed at source in jolt_toolkit ≥ 2.2.7, so the guard is no longer needed
# (and would now discard valid short-trip points). Only the MIN_TRIP_KM distance filter remains.

# 有效 trip 的最小里程（km）：里程 < MIN_TRIP_KM（或里程缺失）的行驶段视为无效（极短/残段，
# EP 不可靠），从一切分析中剔除——OPERATING PERIOD、活跃天数、KPI、散点/拟合/点评统计均只基于
# 有效 trip（故运营周期 = 有效 event 决定的真实跨度）。
MIN_TRIP_KM = 3.0

# 轻/重载分组方式按车辆区分：此集合内的车保留**固定 1/3 三分位**的旧表述
# （"Lightest/Heaviest 1/3 of trips"，合作方风格基准，如 YK73WFN）；其余车用
# GVM 密度谷聚类（_gvm_cluster_split）。见 SKILL.md §5。
LEGACY_TERTILE_REGS = {"YK73WFN"}

# 单组分析（不分轻/重簇）：这些车主要满载运行、GVM 分簇无意义——改为只给
# **GVM > 阈值(t) 的离群剔除（IQR 1.5×）平均**（EP 与续航各一条 bullet）。值 = GVM 阈值(t)。
# 优先级高于 LEGACY_TERTILE_REGS / KDE 聚类。见 SKILL.md §5。
SINGLE_GROUP_REGS = {}  # none currently — YN25RSY moved to an AI-judged artic band (a band, when set in briefing_vehicle_specs.json, overrides single-group anyway via the guard below)

# 统一坐标轴范围与刻度（参数化，所有报告跨车辆/周期可比）
GVM_XLIM = (0, 45)               # gross vehicle mass 轴，t
GVM_XTICKS = range(0, 41, 10)    # 0/10/20/30/40
TEMP_XLIM = (-5, 30)           # ambient temperature 轴，°C
TEMP_XTICKS = range(0, 31, 10)  # 0/10/20/30（-5 起点不设刻度，避免拥挤）

# Reference full-laden gross mass (t) for the EP/Range conclusion + dashed fit extension.
# UK artic max ≈ 44 t; the briefing projects to 42 t. The dashed tail flags it as an
# extrapolation when a vehicle's observed GVM never reaches it (e.g. a rigid).
FULL_LADEN_T = 42.0

# 路线图底图请求像素尺寸 = 第 1 页地图卡显示窗口（约 370×530，竖版）的 2x。
# 必须与卡片纵横比一致：卡片用 object-fit:cover，比例不符会裁掉两侧（连带左上角图例）。
MAP_PX_W, MAP_PX_H = 740, 1060

# 三张散点图（EP-vs-GVM / Range-vs-GVM / EP-vs-temp）统一坐标区位置（figure 分数）。
# 钉死后绘图框像素跨图一致；边距按缩小后的轴字号 + y 刻度自适应（见 _finish_scatter）取定，
# 既容纳 Range 的 "Range (km)" 轴标题 + 三位 y 刻度、又让标签完整不被图边裁切。
# Square figures (1:1) for the 2x2 page-2 grid. No legend strip any more (legends removed),
# so the plot box reaches near the top; left/bottom leave room for the larger axis labels.
SCATTER_AXBOX = dict(left=0.160, right=0.945, top=0.945, bottom=0.135)

# ---- matplotlib 全局风格：对齐 analysis/ 图；字号约为原版 2x（标题/图例受图宽限制取 ~1.5x）----
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    # Axis labels deliberately smaller than the HTML chart titles (.chart-cell__title).
    "font.size": 16, "axes.titlesize": 18, "axes.labelsize": 16,
    "xtick.labelsize": 15, "ytick.labelsize": 15, "legend.fontsize": 13,
    "axes.grid": True, "grid.color": "#cfd3e0", "grid.linewidth": 0.7,
    "axes.edgecolor": "#9aa0b4", "axes.linewidth": 0.8, "axes.axisbelow": True,
    "figure.dpi": 200, "savefig.dpi": 200,
})

# 时间线 SVG 图标
ICON_TRUCK = '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M1 3h13v10H1z"/><path d="M14 6h4l3 3v4h-7z"/><circle cx="6" cy="17" r="2"/><circle cx="17" cy="17" r="2"/></svg>'
ICON_CLOCK = '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/></svg>'


def num(s):
    return pd.to_numeric(s, errors="coerce")


def _sum_or_na(s):
    """求和，但当整列没有任何有效值（数据不可用）时返回 NaN —— 渲染处的 f() 会显示
    '—' 而非误导性的 0（如本车未上报充电 AC/DC kWh 或再生能量）。"""
    return s.sum() if s.notna().any() else float("nan")


def _counter_recup(tr, reg, version):
    """周期内再生能量总量（kWh），取自 raw_telematics 的**累积再生计数器**（全覆盖、口径正确）。

    替代 xlsx 稀疏的 `Recuperation Energy` 列（覆盖常 <50%、低估）。对每个有效行驶段在
    [Start, End] 端点插值差分（与 data_analysis_workspace/energy_breakdown 同口径，复用版本化的
    `jolt_toolkit.analysis` 计数器机器）。返回 (total_kwh, n_covered, n_total)；无 raw_telematics、
    无该计数器（如 Scania/Mercedes/柴油）、或导入失败时返回 None → 调用方回退 xlsx 列。
    """
    raw_dir = PROJECT / "excel_report_database" / version / reg / "raw_telematics"
    if not raw_dir.is_dir():
        return None
    try:
        from jolt_toolkit.analysis import COL_RECUP, build_interp, delta, to_utc
    except Exception:
        return None
    st = tr["st"].dropna()
    et = tr["et"].dropna()
    if st.empty or et.empty:
        return None
    # 只读周期跨度内（±1 天缓冲）的 raw 文件，避免读全历史（某些车有数百个日文件）
    lo = (st.min() - pd.Timedelta(days=1)).normalize()
    hi = (et.max() + pd.Timedelta(days=1)).normalize()
    raws = []
    for p in sorted(raw_dir.glob("raw_*.csv")):
        m = re.search(r"raw_(\d{4}-\d{2}-\d{2})", p.name)
        if m and lo <= pd.Timestamp(m.group(1)) <= hi:
            raws.append(p)
    if not raws:
        return None
    need = {"eventDatetime", COL_RECUP}
    try:
        tel = pd.concat([pd.read_csv(p, dtype=str, usecols=lambda c: c in need) for p in raws],
                        ignore_index=True)
        interp = build_interp(tel, COL_RECUP)
    except Exception:
        return None
    if interp is None:  # 该车无再生计数器
        return None
    tot, ncov, ntot = 0.0, 0, 0
    for s, e in zip(tr["st"], tr["et"]):
        if pd.isna(s) or pd.isna(e):
            continue
        ntot += 1
        d = delta(interp, to_utc(s), to_utc(e))
        if pd.notna(d) and d >= 0:  # 跳过越界(NaN)与计数器复位(负)
            tot += d
            ncov += 1
    return (tot, ncov, ntot) if ntot else None


def _raw_kpi_totals(tr, reg, version):
    """Page-1 energy / distance / recuperation totals from the raw_telematics cumulative
    counters — robust to counter RESETS (drops back to a lower value) and spurious SPIKES
    (physically impossible single-step jumps).

    Unlike the xlsx driving-leg sums (which cover only valid driving legs ≥ MIN_TRIP_KM),
    these whole-period counter totals INCLUDE non-driving consumption (parked HVAC / battery
    thermal management / aux), so ``Total Energy Used`` reconciles with ``Total Energy
    Charged``; ``Total Distance`` is the odometer's true travelled km (incl. < 3 km / excluded
    legs). The page-2 per-leg analysis keeps the xlsx segment basis (this is page-1 only).

    Robust estimator = Σ of per-sample increments keeping only ``0 ≤ Δ ≤ max_rate × Δt``
    (negative Δ = reset → 0; Δ above the physical rate ceiling = spike → dropped). Energy /
    recuperation use a 1000 kW ceiling (far above real draw, only kills garbage); distance a
    130 km/h ceiling. Returns ``dict(energy_kwh, distance_km, recup_kwh, daily_km: Series)``
    with NaN / None for any channel the vehicle does not report, or ``None`` when there is no
    raw_telematics at all (caller then falls back to the driving-leg segment basis).
    """
    raw_dir = PROJECT / "excel_report_database" / version / reg / "raw_telematics"
    if not raw_dir.is_dir():
        return None
    st = tr["st"].dropna()
    et = tr["et"].dropna()
    if st.empty or et.empty:
        return None
    # Only read raw files spanning the briefing window (±1 day) — some vehicles have 100s of files.
    lo = (st.min() - pd.Timedelta(days=1)).normalize()
    hi = (et.max() + pd.Timedelta(days=1)).normalize()
    want = {"eventDatetime", "total_electric_energy_used", "odometer",
            "electric_energy_recuperation_watthours",
            "battery_pack_dc_watthours", "battery_pack_ac_watthours"}
    raws = []
    for p in sorted(raw_dir.glob("raw_*.csv")):
        m = re.search(r"raw_(\d{4}-\d{2}-\d{2})", p.name)
        if m and lo <= pd.Timestamp(m.group(1)) <= hi:
            raws.append(p)
    if not raws:
        return None
    try:
        tel = pd.concat([pd.read_csv(p, dtype=str, usecols=lambda c: c in want) for p in raws],
                        ignore_index=True)
    except Exception:
        return None
    tel["t"] = pd.to_datetime(tel.get("eventDatetime"), errors="coerce", utc=True)
    tel = tel.dropna(subset=["t"]).sort_values("t")
    if len(tel) < 2:
        return None

    GAP_HOURS = 6.0  # an increment spanning > this Δt is treated as a data-gap accumulation

    def _robust(col, max_rate, scale):
        """Σ valid increments of one cumulative counter + per-day Series (both ×scale).

        IMPORTANT: drop NaN for THIS column first — energy / recuperation counters are sparse
        (logged only on TIMER rows, NaN on the dense GPS rows), so diffing the raw column would
        pair a value against a NaN neighbour and lose almost everything; we diff between the
        actual consecutive readings and measure Δt between THOSE readings for the rate cap.

        The TOTAL keeps every valid increment (incl. one that spans a data gap — the counter still
        accumulated real distance/energy during the outage). The PER-DAY series additionally drops
        gap-spanning increments (Δt > GAP_HOURS): a multi-hour/day telematics outage accumulates a
        large jump that, dumped into the single day of the later reading, would give an absurd daily
        max (e.g. EX74JXY's 34-day gap → a 5,117 km "day"). Daily max must stay a real busiest day.
        """
        if col not in tel.columns:
            return float("nan"), None
        sub = pd.DataFrame({"t": tel["t"], "v": pd.to_numeric(tel[col], errors="coerce")}).dropna(subset=["v"])
        if len(sub) < 2:
            return float("nan"), None
        dv = sub["v"].diff()
        dt_h = sub["t"].diff().dt.total_seconds() / 3600.0
        ok = (dv >= 0) & (dv <= max_rate * dt_h)   # physical per-step ceiling; drops resets + spikes
        total = float(dv.where(ok, other=0.0).sum() * scale)
        inc_day = dv.where(ok & (dt_h <= GAP_HOURS), other=0.0) * scale  # exclude gap jumps from days
        per_day = inc_day.groupby(sub["t"].dt.date).sum()
        return total, per_day

    energy_kwh, _ = _robust("total_electric_energy_used", 1_000_000.0, 1 / 1000.0)
    recup_kwh, _ = _robust("electric_energy_recuperation_watthours", 1_000_000.0, 1 / 1000.0)
    distance_km, daily_km = _robust("odometer", 130.0, 1.0)
    # Charged from the raw battery-pack counters (whole-period) — captures ALL charging, unlike the
    # xlsx charge-leg sum which only counts SEGMENTED sessions and under-captures (5–36%, see SKILL).
    charged_dc_kwh, _ = _robust("battery_pack_dc_watthours", 1_000_000.0, 1 / 1000.0)
    charged_ac_kwh, _ = _robust("battery_pack_ac_watthours", 1_000_000.0, 1 / 1000.0)
    if not (pd.notna(energy_kwh) or pd.notna(distance_km)):
        return None
    return dict(energy_kwh=energy_kwh, distance_km=distance_km,
                recup_kwh=recup_kwh, daily_km=daily_km,
                charged_dc_kwh=charged_dc_kwh, charged_ac_kwh=charged_ac_kwh)


def _outlier_trimmed_mean(s):
    """IQR(1.5×) 剔除离群后的均值；返回 (mean, n_kept, n_total)。有效值 < 4 个则不剔除。"""
    x = pd.to_numeric(s, errors="coerce").dropna()
    n = len(x)
    if n < 4:
        return (float(x.mean()) if n else float("nan")), n, n
    q1, q3 = x.quantile(0.25), x.quantile(0.75)
    iqr = q3 - q1
    kept = x[(x >= q1 - 1.5 * iqr) & (x <= q3 + 1.5 * iqr)]
    return float(kept.mean()), len(kept), n


def _gvm_cluster_split(mass_t):
    """轻/重载分界阈值：取逐段 GVM(t) 概率密度（高斯 KDE）在**两个最高密度峰之间的谷**。

    重卡毛重常呈空载/满载多峰；密度谷比 2-means 的方差均衡中点更贴合「视觉上的簇边界」
    —— 当一簇样本远多于另一簇时（如多数行程满载、少数空载），2-means 会把界线拉向大簇、
    把大簇边缘点错划进小簇，故弃用。样本过少 / 无展度 / 单峰时退回中位数二分。
    返回单一阈值：GVM ≤ 阈值为轻簇、> 为重簇。
    """
    x = np.sort(np.asarray(mass_t, dtype=float))
    x = x[np.isfinite(x)]
    if len(x) < 8 or np.ptp(x) == 0:
        return float(np.median(x)) if len(x) else float("nan")
    try:
        grid = np.linspace(x.min(), x.max(), 400)
        y = gaussian_kde(x)(grid)
        peaks, _ = find_peaks(y)
        if len(peaks) >= 2:
            # 取密度最高的两个峰，谷 = 两峰之间密度最低点
            p_lo, p_hi = sorted(sorted(peaks, key=lambda i: y[i])[-2:])
            valley = p_lo + int(np.argmin(y[p_lo:p_hi + 1]))
            return float(grid[valley])
    except Exception:
        pass
    return float(np.median(x))  # 单峰 / KDE 失败：中位数二分


def _narrow_temp_window(gvw_t, widths=(2, 3, 4, 5), min_n=15):
    """Within a laden GVM series (tonnes), return (lo, hi) of the narrowest sliding window of width in
    ``widths`` (try 2 t first, widen only as needed) holding >= ``min_n`` points, centred on the
    densest part of the distribution. If even the widest width can't reach ``min_n`` (a sparse
    vehicle), return the densest window of the widest width — capped, never wider. Used to hold mass
    roughly constant for the EP-vs-temperature fit (the full laden cluster can span 20+ t)."""
    m = pd.to_numeric(pd.Series(gvw_t), errors="coerce").dropna().values
    if len(m) == 0:
        return None
    widest = None
    for w in widths:
        cnt_best, lo_best = -1, float(np.min(m))
        for lo in np.unique(m):                      # candidate window = [lo, lo+w]
            c = int(((m >= lo) & (m <= lo + w)).sum())
            if c > cnt_best:
                cnt_best, lo_best = c, float(lo)
        widest = (lo_best, lo_best + w)
        if cnt_best >= min_n:
            return widest                            # narrowest width meeting the minimum
    return widest                                    # cap at the widest width (sparse vehicle)


def _adaptive_temp_window(gvw_t, ep, temp, widths=(2, 3, 4, 5), min_n=15):
    """Choose the laden mass window for the EP-vs-temperature fit, WIDENING until the sign is right.

    Start narrow (mass held ~constant so the temperature effect is not confounded by mass), but a
    too-narrow window on a weakly-temperature-dependent / single-season period can yield a SPURIOUS
    POSITIVE slope (warmer → higher EP) that is just noise (R²≈0). So widen the densest window
    (2→3→4→5 t, then the whole laden cluster) until the EP-vs-temperature slope is ≤ 0 — the
    physically-expected colder→higher-EP direction. Returns ``((lo, hi) | None, status)``:
    ``status="ok"`` (a window with ≥ ``min_n`` valid points and slope ≤ 0 was found) or
    ``status="inconclusive"`` — when even the full laden cluster either (a) has **fewer than
    ``min_n`` valid points** (too sparse to fit a reliable slope, e.g. a sparse operator like
    EX74JXW/WELCH with ~6 laden trips → would otherwise show an absurd ~1 kWh/km-per-10 °C slope),
    or (b) still has a positive slope. The caller then states the period cannot characterise the
    temperature effect instead of asserting a noisy / over-fitted sign.
    """
    m = pd.to_numeric(pd.Series(list(gvw_t)), errors="coerce").reset_index(drop=True)
    e = pd.to_numeric(pd.Series(list(ep)), errors="coerce").reset_index(drop=True) if ep is not None else None
    t = pd.to_numeric(pd.Series(list(temp)), errors="coerce").reset_index(drop=True) if temp is not None else None
    mv = m.dropna()
    if mv.empty:
        return None, "ok"

    def _densest(w):
        cnt_best, lo_best = -1, float(mv.min())
        for lo in np.unique(mv.values):
            c = int(((mv >= lo) & (mv <= lo + w)).sum())
            if c > cnt_best:
                cnt_best, lo_best = c, float(lo)
        return (lo_best, lo_best + w), cnt_best

    def _fit(lo, hi):
        """(slope, n_valid) of EP-vs-temp over the window's valid (ep & temp present) points."""
        if e is None or t is None:
            return float("nan"), 0
        sel = (m >= lo) & (m <= hi) & e.notna() & t.notna()
        n = int(sel.sum())
        if n < 3 or float(t[sel].std() or 0.0) == 0:
            return float("nan"), n
        return float(np.polyfit(t[sel].values, e[sel].values, 1)[0]), n

    for w in widths:
        win = _densest(w)[0]
        s, n = _fit(*win)
        if n < min_n:
            continue                                 # too few valid points at this width → widen
        if pd.isna(s) or s <= 0:
            return win, "ok"                         # ≥min_n points, expected colder→higher-EP direction
    # no narrow window reached min_n with a non-positive slope → fall back to the full laden cluster
    full = (float(mv.min()), float(mv.max()))
    s_full, n_full = _fit(*full)
    if n_full < min_n:
        return full, "inconclusive"                  # too few valid points overall → unreliable temp fit
    if pd.isna(s_full) or s_full <= 0:
        return full, "ok"
    return full, "inconclusive"                      # enough points but slope still positive → caveat


def _compute_load_points(reg, tr):
    """Resolve the unladen / laden / full-laden reference masses + the unladen and laden ("dense")
    operating masks used by the page-2 conclusion and the temperature plot.

    This combines **data + judgement** (not pure clustering). Modes, by priority:

    - **Band mode** (``briefing_vehicle_specs.json`` ``unladen_band_t`` = [lo, hi], optional
      ``laden_min_t``): for an artic whose lightest GVM cluster is the *bobtail* tractor (e.g.
      ~12 t), not the real unladen mass. Set by AI judgement of the vehicle configuration and
      grounded in the observed distribution: unladen = the data points in [lo, hi) (tractor +
      empty trailer); laden = GVM ≥ laden_min (default hi); bobtail (< lo) is excluded.
    - **Single-group** (``SINGLE_GROUP_REGS``): laden = GVM > threshold; no unladen point.
    - **Legacy tertile** (``LEGACY_TERTILE_REGS``): unladen = lightest tertile, laden = heaviest.
    - **Default**: KDE density-valley split; laden = denser cluster, unladen = lighter cluster.

    Masses are the **median GVM of the data points** in each mask (so they track the data);
    ``unladen_mass_t`` may still pin the unladen marker explicitly. Returns the masks (tr-indexed)
    and the band thresholds (or None) for the verification workbook.
    """
    gvw_t = tr["mass"] / 1000.0
    spec = BRIEFING_SPECS.get(reg, {})
    band = spec.get("unladen_band_t")
    sg_thr = SINGLE_GROUP_REGS.get(reg)
    band_lo = band_hi = laden_min = None
    if band:                                     # judgement band (artic: bobtail vs unladen)
        band_lo, band_hi = float(band[0]), float(band[1])
        laden_min = float(spec.get("laden_min_t", band_hi))
        unladen_mask = (gvw_t >= band_lo) & (gvw_t < band_hi)
        dense_mask = gvw_t >= laden_min
    elif sg_thr is not None:                     # single-group: laden only
        dense_mask = gvw_t > sg_thr
        unladen_mask = pd.Series(False, index=gvw_t.index)
    elif reg in LEGACY_TERTILE_REGS:             # fixed tertiles
        loq, hiq = gvw_t.quantile(0.33), gvw_t.quantile(0.67)
        dense_mask, unladen_mask = gvw_t > hiq, gvw_t < loq
    else:                                        # default: KDE density-valley split — laden = the
        # HEAVIER cluster, unladen = the lighter (split by MASS, not by which cluster is denser: a
        # vehicle that runs mostly light, e.g. CMZ6260 at ~19 t, must not have its denser light
        # cluster mislabelled as "laden" — laden must always be the heavier observed mass).
        split = _gvm_cluster_split(gvw_t.values)
        dense_mask, unladen_mask = gvw_t > split, gvw_t <= split
    m_la = float(gvw_t[dense_mask].median()) if dense_mask.any() else float("nan")
    ov = spec.get("unladen_mass_t")
    if ov is not None:
        m_un = float(ov)
    else:
        m_un = float(gvw_t[unladen_mask].median()) if unladen_mask.any() else float("nan")
    # Temperature analysis: a tight mass window (2-5 t, adaptive to data volume) within the laden
    # cluster, so mass is held roughly constant for the EP-vs-temperature fit. Load points keep the
    # full dense_mask; only the temperature subset narrows.
    temp_lo = temp_hi = None
    temp_mask = dense_mask
    temp_status = "ok"
    if dense_mask.any():
        ep_d = tr["ep"][dense_mask] if "ep" in tr.columns else None
        tp_d = tr["temp"][dense_mask] if "temp" in tr.columns else None
        tw, temp_status = _adaptive_temp_window(gvw_t[dense_mask], ep_d, tp_d)
        if tw is not None:
            temp_lo, temp_hi = tw
            temp_mask = dense_mask & (gvw_t >= temp_lo) & (gvw_t <= temp_hi)
    return dict(unladen=m_un, laden=m_la, full=FULL_LADEN_T,
                unladen_mask=unladen_mask, dense_mask=dense_mask, temp_mask=temp_mask,
                temp_lo=temp_lo, temp_hi=temp_hi, temp_status=temp_status,
                band_lo=band_lo, band_hi=band_hi, laden_min=laden_min)


def _add_load_markers(ax, lp, predict, x_data_max):
    """Add unladen / laden / full-laden vertical dashed lines + labels and a dashed extension
    of the fitted curve from the observed data max out to the full-laden mass.

    ``predict(mass) -> y`` evaluates the fitted model (linear for EP, reciprocal for Range).
    """
    lo, hi = ax.get_ylim()
    rng = hi - lo
    # Stagger the labels vertically so adjacent markers never overlap regardless of spacing:
    # sort by mass and alternate a high/low level. Lowest mass (Unladen) lands on the high level
    # ("above"); the labels sit in the lower, emptier band (EP/Range rise with mass).
    marks = [(mt, lab, col) for mt, lab, col in
             [(lp["unladen"], "Unladen", "#3a3f55"),
              (lp["laden"], "Laden", "#3a3f55"),
              (lp["full"], "Full", "#c0392b")]
             if mt is not None and np.isfinite(mt)]
    marks.sort(key=lambda m: m[0])
    levels = [0.17, 0.045]
    for i, (mt, lab, col) in enumerate(marks):
        ax.axvline(mt, ls=(0, (4, 3)), color=col, lw=1.4, alpha=0.85, zorder=3)
        ax.text(mt, lo + rng * levels[i % 2], f"{lab}\n{mt:.0f} t", ha="center", va="bottom",
                fontsize=14, color=col, linespacing=0.95, zorder=6,
                bbox=dict(boxstyle="round,pad=0.12", fc="white", ec="none", alpha=0.78))
    if np.isfinite(x_data_max) and lp["full"] > x_data_max:
        xe = np.linspace(x_data_max, lp["full"], 40)
        ax.plot(xe, predict(xe), ls="--", color=OEM_BLUE, lw=1.6, alpha=0.9, zorder=4)


def _resolve_operator(reg, period):
    """运营方：先查 company_assignment.simple；轮换车（round_robin）则按报告期匹配日期段
    —— 优先取**报告期末日**所在段，其次取与报告期**有重叠**的段，皆无则回退车牌。"""
    ca = PLOT_CFG.get("company_assignment", {})
    if ca.get("simple", {}).get(reg):
        return ca["simple"][reg]
    segs = ca.get("round_robin", {}).get(reg, [])
    d0, d1 = period.split("_")
    # Companies whose segment OVERLAPS the operating period [d0, d1].
    overlap = {seg["company"] for seg in segs
               if seg.get("date_start", "0") <= d1 and d0 <= seg.get("date_end", "99999999")}
    if len(overlap) == 1:
        return next(iter(overlap))           # whole period sits with one operator
    if len(overlap) > 1:
        return "JOLT Partners"               # span crosses several operators (e.g. an --all-data briefing)
    # round-robin car but the period covers no segment → generic partner label, not the bare
    # registration (avoids "LN25NKE · LN25NKE"); a non-round-robin unconfigured car still falls back to reg.
    return "JOLT Partners" if segs else reg


def _coords(series):
    """从 '(lat, lon)' 字符串解析坐标，返回 (lat[], lon[])。"""
    lats, lons = [], []
    for v in series.dropna():
        m = re.findall(r"-?\d+\.?\d*", str(v))
        if len(m) >= 2:
            lats.append(float(m[0])); lons.append(float(m[1]))
    return np.array(lats), np.array(lons)


# ---- HERE 真实底图（Web Mercator 像素配准）----

def _load_here_key():
    """从仓库根 .env 读取 HERE_API_KEY（其次取环境变量）；没有则返回 None。"""
    env = PROJECT / ".env"
    if env.exists():
        for line in env.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("HERE_API_KEY="):
                return line.split("=", 1)[1].strip().strip("'\"") or None
    return os.environ.get("HERE_API_KEY")


def _merc_x(lon):
    """Longitude -> Web Mercator world fraction [0,1]."""
    return (np.asarray(lon, dtype=float) + 180.0) / 360.0


def _merc_y(lat):
    """Latitude -> Web Mercator world fraction [0,1] (y grows southwards)."""
    s = np.sin(np.radians(np.asarray(lat, dtype=float)))
    return 0.5 - np.log((1.0 + s) / (1.0 - s)) / (4.0 * np.pi)


def _ll_to_px(lat, lon, clat, clon, zoom, w, h):
    """(lat, lon) -> pixel coords on a w*h HERE map centred at (clat, clon)/zoom."""
    world = 256.0 * (2.0 ** zoom)
    x = (_merc_x(lon) - _merc_x(clon)) * world + w / 2.0
    y = (_merc_y(lat) - _merc_y(clat)) * world + h / 2.0
    return x, y


def _fit_center_zoom(lats, lons, w, h, pad=0.10):
    """Centre + Web-Mercator zoom so all points (plus *pad* margin) fit in w*h px.

    The map card displays the PNG with object-fit:contain (no cropping), so this fit alone
    decides what is visible — every origin/destination point is guaranteed inside the frame.
    """
    xf, yf = _merc_x(lons), _merc_y(lats)
    x0, x1 = float(xf.min()), float(xf.max())
    y0, y1 = float(yf.min()), float(yf.max())
    dx = max(x1 - x0, 1e-7); dy = max(y1 - y0, 1e-7)
    x0 -= dx * pad; x1 += dx * pad; y0 -= dy * pad; y1 += dy * pad
    zoom = min(math.log2(w / (256.0 * (x1 - x0))), math.log2(h / (256.0 * (y1 - y0))))
    zoom = max(3.0, min(zoom, 17.0))
    cx, cy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
    clon = cx * 360.0 - 180.0
    clat = math.degrees(math.atan(math.sinh(math.pi * (1.0 - 2.0 * cy))))
    return clat, clon, zoom, cx, cy


def _fetch_here_basemap(lats, lons, w, h, key):
    """拉取一张盖住所有点（含 15% 边距）的 HERE 静态底图（lite.day，带地名）。

    Returns (png_path, center_lat, center_lon, zoom); raises on network/API failure.
    Images are cached under cache/ keyed by (center, zoom, size) so re-runs are offline.
    """
    clat, clon, zoom, _, _ = _fit_center_zoom(lats, lons, w, h)
    zoom = round(zoom, 2)
    cache = WORKSPACE / "cache"
    cache.mkdir(parents=True, exist_ok=True)
    tag = hashlib.md5(f"{clat:.5f},{clon:.5f},{zoom},{w}x{h}".encode()).hexdigest()[:10]
    out = cache / f"here_{tag}.png"
    if not out.exists():
        url = (f"https://image.maps.hereapi.com/mia/v3/base/mc/"
               f"center:{clat:.5f},{clon:.5f};zoom={zoom}/{w}x{h}/png8"
               f"?apiKey={key}&style=lite.day")
        with urllib.request.urlopen(url, timeout=25) as resp:
            data = resp.read()
        if data[:8] != b"\x89PNG\r\n\x1a\n":
            raise RuntimeError("HERE Map Image API returned a non-PNG response")
        out.write_bytes(data)
    return out, clat, clon, zoom


CARTO_TILE = "https://basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}@2x.png"


def _fetch_carto_basemap(lats, lons, w, h):
    """匿名版底图：CARTO light_nolabels（OSM 数据，**无地名标注**），瓦片拼接后裁切。

    与 HERE 版同一坐标约定（centre/zoom 的 Web Mercator 像素配准，_ll_to_px 通用）。
    @2x 瓦片（512 px/瓦片）在 tile-zoom z_t 下等效 zoom = z_t + 1。
    Returns (png_path, center_lat, center_lon, zoom_eff); cached under cache/.
    """
    import io

    from PIL import Image

    clat, clon, z_star, cx, cy = _fit_center_zoom(lats, lons, w, h)
    z_t = max(2, int(math.floor(z_star)) - 1)  # 保证 zoom_eff = z_t+1 <= z_star，bbox 必能装下
    zoom_eff = z_t + 1
    world = 256.0 * (2 ** zoom_eff)
    x0, y0 = cx * world - w / 2.0, cy * world - h / 2.0

    cache = WORKSPACE / "cache"
    cache.mkdir(parents=True, exist_ok=True)
    tag = hashlib.md5(f"carto,{clat:.5f},{clon:.5f},{zoom_eff},{w}x{h}".encode()).hexdigest()[:10]
    out = cache / f"carto_{tag}.png"
    if not out.exists():
        ts = 512  # @2x tile size in zoom_eff pixel space
        xt0, xt1 = int(x0 // ts), int((x0 + w - 1) // ts)
        yt0, yt1 = int(y0 // ts), int((y0 + h - 1) // ts)
        mosaic = Image.new("RGB", ((xt1 - xt0 + 1) * ts, (yt1 - yt0 + 1) * ts), "#eef2f6")
        n_side = 2 ** z_t
        for xt in range(xt0, xt1 + 1):
            for yt in range(yt0, yt1 + 1):
                if not 0 <= yt < n_side:
                    continue  # beyond the poles: keep background fill
                url = CARTO_TILE.format(z=z_t, x=xt % n_side, y=yt)
                req = urllib.request.Request(url, headers={"User-Agent": "JOLT-briefing/1.0"})
                with urllib.request.urlopen(req, timeout=20) as resp:
                    tile = Image.open(io.BytesIO(resp.read())).convert("RGB")
                mosaic.paste(tile, ((xt - xt0) * ts, (yt - yt0) * ts))
        ox, oy = int(round(x0 - xt0 * ts)), int(round(y0 - yt0 * ts))
        mosaic.crop((ox, oy, ox + w, oy + h)).save(out)
    return out, clat, clon, zoom_eff


def _scatter_fit(ax, x, y, color=OEM_BLUE, unit="", fit=True):
    """JOLT 风格：散点 + 线性趋势线（无图例 / 无 ±1σ 阴影带）。

    拟合的斜率 / R² / σ 仍返回给调用方用于结论文字，但**不画进图里**（合作方要求图面只
    保留散点与趋势线）。``unit`` 参数保留以兼容调用方签名（现已不在图例中使用）。
    ``fit=False``：只画散点、不拟合、不画趋势线（温度分析判定为 inconclusive 时用——
    避免画出一条只是噪声的趋势线）。
    """
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    ax.scatter(x, y, s=9, color=color, alpha=0.35, edgecolors="none")
    # 需 ≥3 点且 x/y 各有方差才能拟合：否则（如某列数据缺失/全为占位 0）polyfit 的
    # Vandermonde 矩阵退化会抛 LinAlgError —— 此时只画散点、跳过拟合，不让整张图崩。
    if fit and len(x) >= 3 and np.ptp(x) > 0 and np.ptp(y) > 0:
        m, b = np.polyfit(x, y, 1)
        xs = np.linspace(x.min(), x.max(), 100)
        ys = m * xs + b
        resid = y - (m * x + b)
        sigma = float(resid.std())
        ss_res = float((resid ** 2).sum()); ss_tot = float(((y - y.mean()) ** 2).sum())
        r2 = 1 - ss_res / ss_tot if ss_tot else 0.0
        ax.plot(xs, ys, color=color, lw=1.7)
        return m, b, sigma, r2, len(x)
    return None


def recip_model(x, k, a, c):
    """Reciprocal model for range vs GVM: range = c / (k*x + a).

    Physically: EP rises ~linearly with mass (k*x + a), and range = capacity / EP,
    so range follows the reciprocal of a linear function of GVM.
    """
    return c / (k * x + a)


def _save(fig, path, top=None, box=None):
    # 不用 bbox_inches="tight"：保持 PNG 纵横比 == figsize，便于精确贴合槽位、消除留白。
    # box：固定坐标区位置（三张散点图共用 SCATTER_AXBOX）→ 不论 y 轴刻度位数多少，
    #      绘图框像素始终一致，避免 Range 图（"250/500" 三位刻度）被 tight_layout 挤小、
    #      固定磅值字号相对显大而与 EP-vs-GVM / EP-vs-temp 视觉不一致。
    # top：tight_layout 后再压低坐标区顶边，给锚定在坐标区上方的图例留出专用空间
    # （图表标题不画在 PNG 里，由 HTML 模板的 .chart-panel__title 渲染，避免大字号被图宽截断）。
    if box is not None:
        fig.subplots_adjust(**box)
    else:
        fig.tight_layout(pad=0.5)
        if top is not None:
            fig.subplots_adjust(top=top)
    fig.savefig(path, facecolor="white")
    plt.close(fig)


def _finish_scatter(ax):
    """散点图收尾：y 刻度按**有效位数自适应**——位数越多字号越小（≥3 位缩一档，
    如 Range 的 250/500），极宽刻度（≥4 位）再竖排旋转——使各散点图都能落进同一
    左边距、绘图框对齐又不裁切。EP（一位刻度）保持基准字号。"""
    lo, hi = ax.get_ylim()
    yt = [t for t in ax.get_yticks() if lo - 1e-9 <= t <= hi + 1e-9]
    digits = max((len(f"{t:.0f}") for t in yt), default=1)
    if digits >= 3:
        ax.tick_params(axis="y", labelsize=plt.rcParams["ytick.labelsize"] - 3)
    if digits >= 4:
        ax.tick_params(axis="y", labelrotation=90)


def build_charts(tr, ch, outdir, reg, cb, here_key=None, cap_kwh=None, anon=False,
                 rel="figures", load_pts=None, no_mass=False):
    # 文件名带 run token 即可破坏 Chrome 缓存；旧图 best-effort 清理
    # （不用 rmtree —— OneDrive 目录可能被占用导致 WinError 5）
    outdir.mkdir(parents=True, exist_ok=True)
    for old in outdir.glob("*.png"):
        try:
            old.unlink()
        except OSError:
            pass
    charts, stats = {}, {}
    SQ = 4.6  # square figures (1:1) for the 2x2 page-2 grid — easier to read trends
    lp = load_pts or {}
    def name(stem):
        return f"{stem}_{cb}.png"

    # 1) EP vs GVM (EP axis fixed 0–3, GVM axis GVM_XLIM). Adds unladen/laden vertical markers
    #    and a dashed extension of the linear fit out to FULL_LADEN_T.
    gvw_t = tr["mass"].values / 1000.0
    fig, ax = plt.subplots(figsize=(SQ, SQ))
    r = _scatter_fit(ax, gvw_t, tr["ep"].values, unit=" /t")
    ax.set_xlabel("Gross Vehicle Mass (t)"); ax.set_ylabel("EP (kWh/km)")
    ax.set_ylim(EP_MIN, EP_MAX); ax.set_yticks([0, 1, 2, 3])
    ax.set_xlim(*GVM_XLIM); ax.set_xticks(GVM_XTICKS)
    if r:
        m, b, sigma, r2, _ = r
        stats["gvm_slope"], stats["gvm_intercept"], stats["gvm_sigma"], stats["gvm_r2"] = m, b, sigma, r2
        if lp:
            _add_load_markers(ax, lp, lambda xx: m * xx + b, float(np.nanmax(gvw_t)))
    _finish_scatter(ax)
    _save(fig, outdir / name("ep_gvm"), box=SCATTER_AXBOX); charts["gvm"] = f"{rel}/{name('ep_gvm')}"

    # 1b) Range vs GVM: projected full-charge range = effective capacity / per-trip EP
    #     (needs vehicles.json capacity). Reciprocal-model fit; same unladen/laden markers +
    #     dashed extension of the reciprocal curve to FULL_LADEN_T.
    if cap_kwh:
        rng = cap_kwh / tr["ep"].values
        fig, ax = plt.subplots(figsize=(SQ, SQ))
        mask = np.isfinite(gvw_t) & np.isfinite(rng)
        x, y = gvw_t[mask], rng[mask]
        ax.scatter(x, y, s=9, color=OEM_BLUE, alpha=0.35, edgecolors="none")
        fitted = False
        if len(x) >= 3 and np.ptp(x) > 0 and np.ptp(y) > 0:
            try:
                (rk, ra, rc), _ = curve_fit(recip_model, x, y,
                                            p0=[0.03, 0.3, float(cap_kwh)], maxfev=20000)
                fitted = True
            except (RuntimeError, ValueError):
                fitted = False
        if fitted:
            # 模型对 (k,a,c) 同乘缩放不变（尺度自由度）：归一化钉 c = 电池容量，
            # 此时 k·x+a 即 EP(kWh/km) 的线性拟合，方程对读者有物理含义
            scale = float(cap_kwh) / rc
            rk, ra, rc = rk * scale, ra * scale, rc * scale
            xs = np.linspace(x.min(), x.max(), 200)
            ys = recip_model(xs, rk, ra, rc)
            resid = y - recip_model(x, rk, ra, rc)
            sigma = float(resid.std())
            ss_res = float((resid ** 2).sum()); ss_tot = float(((y - y.mean()) ** 2).sum())
            r2 = 1 - ss_res / ss_tot if ss_tot else 0.0
            ax.plot(xs, ys, color=OEM_BLUE, lw=1.7)
            stats["range_fit"] = (rk, ra, rc, r2); stats["range_sigma"] = sigma
        ax.set_xlabel("Gross Vehicle Mass (t)"); ax.set_ylabel("Range (km)")
        ax.set_ylim(0, np.nanmax(rng) * 1.05)
        ax.set_xlim(*GVM_XLIM); ax.set_xticks(GVM_XTICKS)
        if fitted and lp:
            _add_load_markers(ax, lp, lambda xx: recip_model(xx, rk, ra, rc), float(x.max()))
        _finish_scatter(ax)
        _save(fig, outdir / name("range_gvm"), box=SCATTER_AXBOX)
        charts["range"] = f"{rel}/{name('range_gvm')}"

    # 1c/1d) No-mass distribution variant: EP & projected-range histograms with a KDE overlay
    #        and mean (dashed red) / median (dotted green) markers. Replace the GVM scatters.
    if no_mass:
        epv = pd.to_numeric(tr["ep"], errors="coerce").dropna().values
        fig, ax = plt.subplots(figsize=(SQ, SQ))
        if len(epv):
            ax.hist(epv, bins=np.linspace(EP_MIN, EP_MAX, 25), color=OEM_BLUE,
                    alpha=0.82, edgecolor="white")
            if len(epv) >= 3 and np.ptp(epv) > 0:
                try:
                    xs = np.linspace(EP_MIN, EP_MAX, 200)
                    ax2 = ax.twinx(); ax2.plot(xs, gaussian_kde(epv)(xs), color="#e67e22", lw=1.7)
                    ax2.set_yticks([]); ax2.set_ylim(bottom=0)
                except Exception:
                    pass
            ax.axvline(float(np.mean(epv)), color="#c0392b", lw=1.4, ls="--")
            ax.axvline(float(np.median(epv)), color="#27ae60", lw=1.4, ls=":")
        ax.set_xlabel("EP (kWh/km)"); ax.set_ylabel("Trips")
        ax.set_xlim(EP_MIN, EP_MAX); ax.set_xticks([0, 1, 2, 3]); ax.set_ylim(bottom=0)
        _finish_scatter(ax)  # digit-adaptive y ticks — 4-digit counts (T88RNW ~1000) push the ylabel off-canvas otherwise
        _save(fig, outdir / name("ep_dist"), box=SCATTER_AXBOX)
        charts["ep_dist"] = f"{rel}/{name('ep_dist')}"
        if cap_kwh and len(epv):
            rngv = cap_kwh / epv
            rmax = float(np.nanmax(rngv)) * 1.05 if len(rngv) else 1.0
            fig, ax = plt.subplots(figsize=(SQ, SQ))
            ax.hist(rngv, bins=np.linspace(0, rmax, 25), color=OEM_BLUE, alpha=0.82, edgecolor="white")
            if len(rngv) >= 3 and np.ptp(rngv) > 0:
                try:
                    xs = np.linspace(0, rmax, 200)
                    ax2 = ax.twinx(); ax2.plot(xs, gaussian_kde(rngv)(xs), color="#e67e22", lw=1.7)
                    ax2.set_yticks([]); ax2.set_ylim(bottom=0)
                except Exception:
                    pass
            ax.axvline(float(np.mean(rngv)), color="#c0392b", lw=1.4, ls="--")
            ax.axvline(float(np.median(rngv)), color="#27ae60", lw=1.4, ls=":")
            ax.set_xlabel("Projected Range (km)"); ax.set_ylabel("Trips")
            ax.set_xlim(0, rmax); ax.set_ylim(bottom=0)
            _finish_scatter(ax)
            _save(fig, outdir / name("range_dist"), box=SCATTER_AXBOX)
            charts["range_dist"] = f"{rel}/{name('range_dist')}"

    # 2) EP vs ambient temperature — restricted to the dense (laden) cluster so mass is held
    #    roughly constant: a control-variable view of the temperature effect.
    # Temperature held at roughly constant mass: use the narrow temp_mask (a tight laden window) so
    # the temperature effect is not confounded by mass. No-mass variant → fit over ALL trips.
    sub = lp.get("temp_mask"); sub = sub if sub is not None else lp.get("dense_mask")
    td = tr if no_mass else (tr[sub] if sub is not None else tr)
    fig, ax = plt.subplots(figsize=(SQ, SQ))
    # inconclusive（laden 窗口样本过少 / 即便放宽到整簇仍正相关）→ 只画散点、不画趋势线（结论改文字
    # 说明）。仅对 LADEN(有质量)分析生效;no-mass 变体按全行程拟合,与 laden 低样本判定无关,始终拟合。
    _temp_fit = no_mass or lp.get("temp_status", "ok") != "inconclusive"
    r = _scatter_fit(ax, td["temp"].values, td["ep"].values, unit=" /°C", fit=_temp_fit)
    ax.set_xlabel("Ambient Temperature (°C)"); ax.set_ylabel("EP (kWh/km)")
    ax.set_ylim(EP_MIN, EP_MAX); ax.set_yticks([0, 1, 2, 3])
    ax.set_xlim(*TEMP_XLIM); ax.set_xticks(TEMP_XTICKS)
    _finish_scatter(ax)
    _save(fig, outdir / name("ep_temp"), box=SCATTER_AXBOX); charts["temp"] = f"{rel}/{name('ep_temp')}"
    if r:
        (stats["temp_slope"], stats["temp_intercept"], stats["temp_sigma"],
         stats["temp_r2"], stats["temp_n"]) = r

    # 3) Charging start SoC histogram (square)
    fig, ax = plt.subplots(figsize=(SQ, SQ))
    ssoc = ch["ssoc"].dropna().values
    ax.hist(ssoc, bins=range(0, 101, 10), color=OEM_BLUE, alpha=0.85, edgecolor="white")
    ax.set_xlabel("Charging Start SoC (%)"); ax.set_ylabel("Sessions")
    ax.set_xlim(0, 100); ax.set_ylim(bottom=0); ax.set_xticks(range(0, 101, 20))
    _finish_scatter(ax)
    _save(fig, outdir / name("soc_hist"), box=SCATTER_AXBOX); charts["soc"] = f"{rel}/{name('soc_hist')}"

    # 路线图：命名版用 HERE 真实底图（lite.day，带地名）；匿名版用 CARTO light_nolabels
    # （保留线和点、无地名——SteerCo 匿名化要求）；失败/无 key 退回无底图示意。
    olat, olon = _coords(tr["Origin (Lat, Lon)"]); dlat, dlon = _coords(tr["Destination (Lat, Lon)"])
    n = min(len(olat), len(dlat))
    basemap, attrib = None, None
    if n:
        try:
            all_lat = np.concatenate([olat[:n], dlat[:n]])
            all_lon = np.concatenate([olon[:n], dlon[:n]])
            if anon:
                basemap = _fetch_carto_basemap(all_lat, all_lon, MAP_PX_W, MAP_PX_H)
                attrib = "© OpenStreetMap © CARTO"
            elif here_key:
                basemap = _fetch_here_basemap(all_lat, all_lon, MAP_PX_W, MAP_PX_H, here_key)
                attrib = "map © HERE"
        except Exception as exc:
            print(f"  [map] 底图获取失败（{exc}），退回无底图示意")
            basemap, attrib = None, None
    if basemap:
        map_png, clat, clon, zoom = basemap
        fig = plt.figure(figsize=(MAP_PX_W / 200.0, MAP_PX_H / 200.0))
        ax = fig.add_axes([0, 0, 1, 1])  # full-bleed: 底图铺满整张 PNG
        ax.imshow(plt.imread(map_png), extent=[0, MAP_PX_W, MAP_PX_H, 0])
        # 起讫点不作区分：统一样式、无图例
        ox, oy = _ll_to_px(olat[:n], olon[:n], clat, clon, zoom, MAP_PX_W, MAP_PX_H)
        dx, dy = _ll_to_px(dlat[:n], dlon[:n], clat, clon, zoom, MAP_PX_W, MAP_PX_H)
        for i in range(n):
            ax.plot([ox[i], dx[i]], [oy[i], dy[i]], color=OEM_BLUE, lw=1.0, alpha=0.3)
        ax.scatter(np.concatenate([ox, dx]), np.concatenate([oy, dy]), s=24,
                   color="#e74c3c", alpha=0.9, edgecolors="white", linewidths=0.5, zorder=5)
        ax.set_xlim(0, MAP_PX_W); ax.set_ylim(MAP_PX_H, 0)
        ax.set_xticks([]); ax.set_yticks([]); ax.grid(False)
        for sp in ax.spines.values():
            sp.set_edgecolor("#c7ccda")
        fig.savefig(outdir / name("route_map"), dpi=200)
        plt.close(fig)
    else:
        fig, ax = plt.subplots(figsize=(3.7, 2.66))
        fig.patch.set_facecolor("#eef2f6"); ax.set_facecolor("#eef2f6")
        for i in range(n):
            ax.plot([olon[i], dlon[i]], [olat[i], dlat[i]], color=OEM_BLUE, lw=0.7, alpha=0.18)
        # 起讫点不作区分：统一样式、无图例
        ax.scatter(np.concatenate([olon, dlon]), np.concatenate([olat, dlat]), s=14,
                   color="#e74c3c", alpha=0.55, edgecolors="none")
        ax.set_xticks([]); ax.set_yticks([]); ax.margins(0.18)
        for sp in ax.spines.values():
            sp.set_edgecolor("#c7ccda")
        _save(fig, outdir / name("route_map"))
    charts["map"] = f"{rel}/{name('route_map')}"
    charts["map_attrib"] = attrib  # None => 无底图示意（caption 显示 "indicative"）

    return charts, stats


def _parse_period(period):
    """'YYYYMMDD_YYYYMMDD' -> (start_ts, end_ts) at midnight (end = last *included* day)."""
    a, b = period.split("_")
    return pd.Timestamp(a), pd.Timestamp(b)


def _resolve_xlsx(rdir, reg, period, finetuned):
    """Locate the source xlsx for a briefing period (precondition relaxed).

    Either an xlsx for the *exact* period exists, OR an xlsx whose ``[start, end]``
    range *covers* the requested window exists. Exact match wins (no subsetting);
    otherwise the **tightest** covering report is used and its legs are later subset
    to the window in :func:`compute`.

    Returns ``(path, subset_to_window)``. Raises ``FileNotFoundError`` if neither
    an exact-period report nor a covering one is present.
    """
    # 1) exact period — finetuned preferred (when requested), then base
    suffix = "_finetuned" if finetuned else ""
    for cand in (rdir / f"jolt_report_{reg}_{period}{suffix}.xlsx",
                 rdir / f"jolt_report_{reg}_{period}.xlsx"):
        if cand.exists():
            return cand, False
    # 2) covering fallback — tightest [start,end] that contains the window
    tgt_s, tgt_e = _parse_period(period)
    pat = re.compile(rf"^jolt_report_{re.escape(reg)}_(\d{{8}})_(\d{{8}})(_finetuned)?\.xlsx$")
    covering = []  # (span_days, kind_rank, path)
    for p in rdir.glob(f"jolt_report_{reg}_*.xlsx"):
        m = pat.match(p.name)
        if not m:
            continue
        cs, ce = _parse_period(f"{m.group(1)}_{m.group(2)}")
        if cs <= tgt_s and ce >= tgt_e:
            is_ft = bool(m.group(3))
            kind_rank = 0 if is_ft == finetuned else 1  # honour finetuned preference
            covering.append(((ce - cs).days, kind_rank, p))
    if covering:
        covering.sort(key=lambda t: (t[0], t[1]))  # tightest span, then preferred kind
        return covering[0][2], True
    raise FileNotFoundError(
        f"No xlsx for {reg} {period} under {rdir} — neither an exact-period report "
        f"nor one whose [start,end] covers the window. Run /generate-excel-report first.")


def _read_all_xlsx(rdir, reg, finetuned):
    """All-data mode: read & concatenate EVERY period report for a vehicle, deduping
    quarterly-boundary overlaps by (Start Time, Leg Type). Returns (df, last_path, False)."""
    pat = re.compile(rf"^jolt_report_{re.escape(reg)}_(\d{{8}})_(\d{{8}})(_finetuned)?\.xlsx$")
    by_period = {}
    for p in rdir.glob(f"jolt_report_{reg}_*.xlsx"):
        m = pat.match(p.name)
        if not m:
            continue
        key = f"{m.group(1)}_{m.group(2)}"; is_ft = bool(m.group(3))
        cur = by_period.get(key)  # one file per period; honour the finetuned preference
        if cur is None or (is_ft == finetuned and cur[1] != finetuned):
            by_period[key] = (p, is_ft)
    if not by_period:
        raise FileNotFoundError(
            f"No xlsx for {reg} under {rdir} — run /generate-excel-report first.")
    paths = [by_period[k][0] for k in sorted(by_period)]
    df = pd.concat([pd.read_excel(p) for p in paths], ignore_index=True)
    n0 = len(df)
    df = df.drop_duplicates(subset=["Start Time (UTC)", "Leg Type"], keep="first").reset_index(drop=True)
    print(f"  [all-data] {len(paths)} period report(s) for {reg} concatenated: {n0}->{len(df)} legs (deduped)")
    return df, paths[-1], False


def compute(reg, period, version, finetuned, all_data=False):
    rdir = PROJECT / "excel_report_database" / version / reg
    if all_data:
        df, xlsx, subset = _read_all_xlsx(rdir, reg, finetuned)
    else:
        xlsx, subset = _resolve_xlsx(rdir, reg, period, finetuned)
        df = pd.read_excel(xlsx)
    tr_all = df[df["Leg Type"].isin(DRIVE)].copy()
    ch = df[df["Leg Type"].isin(CHARGE)].copy()
    for c, col in [("d", "Distance (km)"), ("ep", "Energy Performance (kWh/km)"),
                   ("ec", "Energy Change (kWh)"), ("temp", "Average Temperature (C)"),
                   ("mass", "Vehicle Mass (kg)"), ("recup", "Recuperation Energy (kWh)")]:
        tr_all[c] = num(tr_all[col])
    # A Vehicle Mass of 0 means "no GVM signal" — some vehicles report 0 rather than blank when the
    # GCW channel is absent (e.g. T88RNW / YN75NMA on the SRF-free recompute path). Treat <= 0 as
    # MISSING so the no-mass distribution-variant detection, the GVM scatter and Median GVM all
    # correctly exclude it (otherwise a 0 counts as notna and the vehicle is mis-classified as mass-bearing).
    tr_all.loc[tr_all["mass"] <= 0, "mass"] = np.nan
    # Operator is DATA-DRIVEN: read the report's per-leg `Operator` column (the generator derives it
    # from SRF leg.trip.trial.description / vehicle.organisation.name). This is the source of truth for
    # the page header and the per-operator split — NOT the manual plot_config company_assignment.
    tr_all["operator"] = tr_all["Operator"].astype(str).str.strip() if "Operator" in tr_all.columns else ""
    tr_all["st"] = pd.to_datetime(tr_all["Start Time (UTC)"], errors="coerce")
    tr_all["et"] = pd.to_datetime(tr_all["End Time (UTC)"], errors="coerce")
    tr_all["date"] = tr_all["st"].dt.date
    # Two trip sets, by page:
    #   tr_all = EVERY driving leg — the PAGE-1 operations dashboard is UNFILTERED (Active Days,
    #            Driving Legs, Median GVM, timeline and the totals count every leg the vehicle drove,
    #            incl. < MIN_TRIP_KM residual legs; no EP nulling).
    #   tr     = the cleaned analysis set, used ONLY for the PAGE-2 figures / fits / conclusions:
    #            valid-trip distance filter + EP cleaning (the old SOC-quantisation guard was
    #            removed once jolt_toolkit ≥ 2.2.7 fixed the short-trip EP at source — see below).
    tr = tr_all.copy()
    # 有效 trip 过滤（仅第 2 页分析）：里程 < MIN_TRIP_KM（或缺失）的行驶段极短/残段、EP 不可靠。
    n_before = len(tr)
    tr = tr[tr["d"] >= MIN_TRIP_KM].copy()
    if len(tr) < n_before:
        print(f"  [Trip 过滤·第2页] 剔除 {n_before - len(tr)} 个里程 < {MIN_TRIP_KM} km（或缺失）的行驶段")
    # EP 超出 [EP_CLEAN_MIN, EP_CLEAN_MAX] 视为不合理数据，剔除（散点/拟合/点评统计均不计）
    bad = (tr["ep"] < EP_CLEAN_MIN) | (tr["ep"] > EP_CLEAN_MAX)
    tr.loc[bad, "ep"] = np.nan
    if bad.sum():
        print(f"  [EP 清洗] 剔除 {int(bad.sum())} 个超出 [{EP_CLEAN_MIN}, {EP_CLEAN_MAX}] 的 EP 值")
    # NOTE: the earlier |SOC change| ≤ 1% "SOC-quantisation guard" was REMOVED (2026-07-02).
    # It was a downstream workaround for a jolt_toolkit bug where the ±1σ effective-capacity
    # correction back-calculated short-trip energy from the coarse integer SOC, producing a
    # spurious low-EP band. That ROOT CAUSE is now fixed in jolt_toolkit ≥ 2.2.7 (MODE A: EP
    # keeps the counter energy; capacity uses ΔSOC-weighted aggregation), so short-trip EP is
    # already correct at source. Keeping the guard here would now wrongly discard those
    # now-valid short-trip points, so it is gone; only the MIN_TRIP_KM distance filter remains.
    ch["ssoc"] = num(ch["Start SOC (%)"]); ch["esoc"] = num(ch["End SOC (%)"])
    ch["ac"] = num(ch["Energy Charged AC (kWh)"]); ch["dc"] = num(ch["Energy Charged DC (kWh)"])
    ch["st"] = pd.to_datetime(ch["Start Time (UTC)"], errors="coerce")
    if subset:
        # 选用覆盖周期的更宽报告时，按 Start Time 把行程/充电裁剪到目标窗口（含止日整天）。
        tgt_s, tgt_e = _parse_period(period)
        win_lo = tgt_s
        win_hi = tgt_e + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        n_tr0, n_ch0 = len(tr), len(ch)
        tr = tr[(tr["st"] >= win_lo) & (tr["st"] <= win_hi)].copy()
        tr_all = tr_all[(tr_all["st"] >= win_lo) & (tr_all["st"] <= win_hi)].copy()
        ch = ch[(ch["st"] >= win_lo) & (ch["st"] <= win_hi)].copy()
        print(f"  [covering] 选用覆盖周期报告 {xlsx.name}，裁剪到窗口 {period}："
              f"trips {n_tr0}->{len(tr)}, charges {n_ch0}->{len(ch)}")
    return df, tr_all, tr, ch, xlsx, subset


def median_time(times):
    mins = times.dropna().dt.hour * 60 + times.dropna().dt.minute
    if mins.empty:
        return "—"
    m = int(mins.median())
    return f"{m // 60:02d}:{m % 60:02d}"


# ---- 人工核实（human-in-the-loop verification）----

def _cell(x):
    """pandas value -> openpyxl-writable cell value (NaN/NaT -> None, tz stripped)."""
    if x is None or (isinstance(x, float) and math.isnan(x)) or pd.isna(x):
        return None
    if isinstance(x, pd.Timestamp):
        return x.tz_localize(None).to_pydatetime() if x.tzinfo else x.to_pydatetime()
    return x


def build_verification_workbook(path, tr_all, tr, ch, v):
    """Write the human-verification workbook (verification_<REG>_<period>.xlsx).

    Two INDEPENDENT computation paths: this script (pandas) produced the briefing
    values; the workbook recomputes each of them with native Excel formulas over
    the raw legs copied into the leg sheets. Any discrepancy shows up as FAIL. The
    reviewer checks FAILs, spot-checks PASSes and the MANUAL rows, and records the
    outcome in the Verified by / Date / Notes columns.

    Two leg sheets, matching the two-page split: `Trips` = the CLEANED page-2 analysis
    set (distance ≥ MIN_TRIP_KM, EP cleaned; the SOC-quantisation guard was removed
    2026-07-02) — page-2 EP/GVM formulas run
    over it; `AllTrips` (+ the all-legs `Daily`) = EVERY driving leg — the UNFILTERED
    page-1 operational counts/totals (Active Days, Driving Legs, Median GVM, timeline,
    segment totals) recompute over it.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", start_color="E8EAF6")

    # ---- Trips sheet (valid driving legs: distance ≥ MIN_TRIP_KM, EP cleaned to [EP_CLEAN_MIN, EP_CLEAN_MAX]) ----
    ws = wb.active
    ws.title = "Trips"
    ws.append(["Date", "Start", "End", "Distance_km", "EnergyChange_kWh",
               "EP_kWhkm_cleaned", "GVM_t", "Temp_C", "Recup_kWh", "Range_km",
               "", "EffCap_kWh"])
    cap = v.get("cap_kwh")
    for _, r in tr.iterrows():
        ws.append([_cell(r["date"]), _cell(r["st"]), _cell(r["et"]), _cell(r["d"]),
                   _cell(r["ec"]), _cell(r["ep"]),
                   _cell(r["mass"] / 1000.0 if pd.notna(r["mass"]) else None),
                   _cell(r["temp"]), _cell(r["recup"]), None])
    tn = len(tr) + 1  # last data row
    ws["L2"] = cap if cap else None
    if cap:
        for i in range(2, tn + 1):
            ws.cell(row=i, column=10).value = f'=IF($F{i}="","",$L$2/$F{i})'
    for c in ws[1]:
        c.font = bold; c.fill = head_fill
    ws.freeze_panes = "A2"

    # ---- Charges sheet ----
    wc = wb.create_sheet("Charges")
    wc.append(["Start", "StartSOC_pct", "EndSOC_pct", "AC_kWh", "DC_kWh"])
    ch_st = pd.to_datetime(ch["Start Time (UTC)"], errors="coerce")
    for (_, r), st_val in zip(ch.iterrows(), ch_st):
        wc.append([_cell(st_val), _cell(r["ssoc"]), _cell(r["esoc"]),
                   _cell(r["ac"]), _cell(r["dc"])])
    cn = len(ch) + 1
    for c in wc[1]:
        c.font = bold; c.fill = head_fill
    wc.freeze_panes = "A2"

    # ---- AllTrips sheet (EVERY driving leg — page-1 unfiltered basis) ----
    # Page-1 operational counts/totals are computed over ALL legs (tr_all), which the cleaned Trips
    # sheet cannot reproduce; this sheet is the independent all-legs recompute for those rows.
    wat = wb.create_sheet("AllTrips")
    wat.append(["Date", "Distance_km", "EnergyChange_kWh", "GVM_t", "Recup_kWh"])
    for _, r in tr_all.iterrows():
        wat.append([_cell(r["date"]), _cell(r["d"]), _cell(r["ec"]),
                    _cell(r["mass"] / 1000.0 if pd.notna(r["mass"]) else None), _cell(r["recup"])])
    an = len(tr_all) + 1  # last AllTrips data row
    for c in wat[1]:
        c.font = bold; c.fill = head_fill
    wat.freeze_panes = "A2"

    # ---- Daily sheet (per-day sums over ALL legs; page-1 is unfiltered) ----
    wd = wb.create_sheet("Daily")
    wd.append(["Date", "Distance_km", "TractionEnergy_kWh"])
    dates = sorted(d for d in tr_all["date"].dropna().unique())
    for i, d in enumerate(dates, start=2):
        wd.cell(row=i, column=1).value = d
        wd.cell(row=i, column=2).value = f"=SUMIF(AllTrips!$A$2:$A${an},A{i},AllTrips!$B$2:$B${an})"
        wd.cell(row=i, column=3).value = f"=SUMPRODUCT((AllTrips!$A$2:$A${an}=A{i})*ABS(AllTrips!$C$2:$C${an}))"
    dn = len(dates) + 1
    for c in wd[1]:
        c.font = bold; c.fill = head_fill

    # ---- Audit sheet ----
    wa = wb.create_sheet("Audit", 0)
    wa["A1"] = f"Verification — {v['reg']} {v['period']} (pipeline {v['version']}, source: {v['fname']})"
    wa["A1"].font = Font(bold=True, size=13)
    wa["A2"] = ("Each briefing number below is recomputed with native Excel formulas over the raw legs "
                "(Trips/Charges/Daily sheets) — an independent path from the generator. Review every FAIL, "
                "spot-check PASS and MANUAL rows, fill in the sign-off columns, then regenerate with "
                '--final --verified-by "<name>".')
    wa["A2"].alignment = Alignment(wrap_text=True)
    wa.merge_cells("A2:K2")

    tr_f, tr_g, tr_h = f"Trips!$F$2:$F${tn}", f"Trips!$G$2:$G${tn}", f"Trips!$H$2:$H${tn}"
    # AllTrips ranges (page-1 unfiltered, all legs): A=Date B=Distance C=EnergyChange D=GVM E=Recup
    at_a, at_b, at_c, at_d, at_e = (f"AllTrips!$A$2:$A${an}", f"AllTrips!$B$2:$B${an}",
                                    f"AllTrips!$C$2:$C${an}", f"AllTrips!$D$2:$D${an}",
                                    f"AllTrips!$E$2:$E${an}")
    # Page-1 distance/energy/recup totals may use the RAW-TELEMATICS counter basis (see
    # _raw_kpi_totals), decided PER FIELD: distance from the raw odometer, energy from the raw
    # used-energy counter — independently. A raw-basis total is a whole-period counter sum the leg
    # sheets cannot reproduce, so that row is flagged CHECK MANUALLY (not FAIL) with the leg-sum
    # formula kept in the note. Operational counts / charging / SoC stay leg-recomputed.
    raw_dist = v.get("dist_basis") == "raw"
    raw_energy = v.get("energy_basis") == "raw"
    raw_charge = v.get("charge_basis") == "raw"
    _raw_note = "PAGE-1 RAW-TELEMATICS counter total (robust to resets/spikes, incl. non-driving). "
    rows = [
        ("Page 1 · Summary", "Active days", v["ndays"],
         f"=COUNTA(Daily!$A$2:$A${dn})", 0.5,
         "Distinct dates with ≥1 driving leg (ALL legs — page-1 unfiltered)", "Start Time (UTC) → date"),
        ("Page 1 · Summary", "Driving legs", v["n_tr"],
         f"=COUNTA({at_a})", 0.5,
         "ALL driving legs (page-1 unfiltered — every leg the vehicle drove, incl. < "
         f"{MIN_TRIP_KM:.0f} km legs)", "Leg Type ∈ DRIVE"),
        ("Page 1 · Summary", "Charging sessions", v["n_ch"],
         f"=ROWS(Charges!$A$2:$A${cn})", 0.5, "Count of charging legs (Leg Type ∈ CHARGE)", "Leg Type"),
        ("Page 1 · Summary", "Median GVM (t)", v["gvw_med"],
         f"=MEDIAN({at_d})", 0.05,
         "Median of per-leg GVM over ALL legs (page-1 'Median GVM' tile)", "Vehicle Mass (kg)/1000"),
        ("Page 1 · Timeline", "Trips per day", v["trips_per_day"],
         f"=COUNTA({at_a})/COUNTA(Daily!$A$2:$A${dn})", 0.05,
         "All driving legs ÷ active days (page-1 unfiltered)", "—"),
        ("Page 1 · Timeline", "Daily average distance (km)", v["daily_avg_km"],
         ("manual" if raw_dist else f"=AVERAGE(Daily!$B$2:$B${dn})"), (None if raw_dist else 0.5),
         (_raw_note + "Total raw odometer km ÷ active days. Leg-sum mean: =AVERAGE(Daily!$B$2:$B$" + str(dn) + ")"
          if raw_dist else "Mean of per-day distance sums"), "odometer (raw)" if raw_dist else "Distance (km)"),
        ("Page 1 · Timeline", "Median first departure (hh:mm)", None, "manual", None,
         f"Briefing shows ≈ {v['med_dep']} — spot-check per-day minima of Trips!B", "Start Time (UTC)"),
        ("Page 1 · Timeline", "Median last arrival (hh:mm)", None, "manual", None,
         f"Briefing shows ≈ {v['med_arr']} — spot-check per-day maxima of Trips!C", "End Time (UTC)"),
        ("Page 1 · Performance", "Total distance (km)", v["tot_km"],
         ("manual" if raw_dist else f"=SUM({at_b})"), (None if raw_dist else 0.5),
         (_raw_note + "Odometer travelled km. All-legs sum: =SUM(" + at_b + ")"
          if raw_dist else "Sum over ALL driving legs"), "odometer (raw)" if raw_dist else "Distance (km)"),
        ("Page 1 · Performance", "Max daily distance (km)", v["daily_max_km"],
         ("manual" if raw_dist else f"=MAX(Daily!$B$2:$B${dn})"), (None if raw_dist else 0.5),
         (_raw_note + "Max per-day odometer km. Leg-sum max: =MAX(Daily!$B$2:$B$" + str(dn) + ")"
          if raw_dist else "Max of per-day distance sums"), "odometer (raw)" if raw_dist else "Distance (km)"),
        ("Page 1 · Performance", "Total energy used (kWh)", v["tot_e"],
         ("manual" if raw_energy else f"=SUMPRODUCT(ABS({at_c}))"), (None if raw_energy else 0.5),
         (_raw_note + "Σ used-energy counter incl. non-driving (reconciles with Total Energy "
          "Charged). All-legs sum: =SUMPRODUCT(ABS(" + at_c + "))"
          if raw_energy else "Σ|energy change| over ALL driving legs"),
         "total_electric_energy_used (raw)" if raw_energy else "Energy Change (kWh)"),
        ("Page 1 · Performance", "Mean energy performance (kWh/km)", v["mean_ep"],
         ("manual" if raw_energy else f"=SUMPRODUCT(ABS({at_c}))/SUM({at_b})"),
         (None if raw_energy else 0.005),
         (_raw_note + "Raw total energy ÷ raw odometer distance (same raw basis)"
          if raw_energy else "Total energy ÷ total distance (NOT mean of per-leg EP); all-legs basis"), "—"),
        ("Page 1 · Performance", "Energy recuperated (kWh)", v["recup"],
         ("manual" if v["recup_src"] in ("counter", "raw_total") else f"=SUM({at_e})"),
         (None if v["recup_src"] in ("counter", "raw_total") else 0.5),
         (_raw_note + "Σ raw_telematics recuperation counter over the whole period (robust to "
          "resets/spikes; NOT the sparse xlsx column). Sparse xlsx-column SUM for comparison: "
          f"=SUM({at_e})"
          if v["recup_src"] == "raw_total"
          else f"Counter-based: Σ raw_telematics recuperation counter over {v['recup_cov']}/{v['recup_tot']} "
          f"legs (full coverage; NOT the sparse xlsx column). Sparse xlsx-column SUM for comparison: "
          f"=SUM({at_e})"
          if v["recup_src"] == "counter"
          else f"Σ recuperation; only {v['recup_cov']}/{v['recup_tot']} legs have a value (known undercount)"),
         "raw_telematics recuperation counter" if v["recup_src"] in ("counter", "raw_total")
         else "Recuperation Energy (kWh)"),
        ("Page 1 · Performance", "Regen recovery (% of energy used)", v.get("recup_pct"),
         ("manual" if v["recup_src"] in ("counter", "raw_total")
          else f"=SUM({at_e})/SUMPRODUCT(ABS({at_c}))*100"),
         (None if v["recup_src"] in ("counter", "raw_total") else 0.5),
         ("Energy recuperated ÷ total energy used × 100 (both raw-counter based, audited above)"
          if v["recup_src"] in ("counter", "raw_total")
          else "Σ recuperation ÷ Σ|energy change| × 100"),
         "—"),
        ("Page 1 · Charging", "Total energy charged (kWh)", v["tot_ch"],
         ("manual" if raw_charge else f"=SUM(Charges!$D$2:$D${cn})+SUM(Charges!$E$2:$E${cn})"),
         (None if raw_charge else 0.5),
         (_raw_note + "Σ raw battery_pack DC+AC counter (ALL charging). Event charge-leg sum (under-"
          f"captures): =SUM(Charges!$D$2:$D${cn})+SUM(Charges!$E$2:$E${cn})"
          if raw_charge else "AC + DC"),
         "battery_pack DC+AC (raw)" if raw_charge else "Energy Charged AC/DC (kWh)"),
        ("Page 1 · Charging", "AC charged (kWh)", v["ac"],
         ("manual" if raw_charge else f"=SUM(Charges!$D$2:$D${cn})"), (None if raw_charge else 0.5),
         (_raw_note + "raw battery_pack_ac counter" if raw_charge else ""),
         "battery_pack_ac (raw)" if raw_charge else "Energy Charged AC (kWh)"),
        ("Page 1 · Charging", "DC charged (kWh)", v["dc"],
         ("manual" if raw_charge else f"=SUM(Charges!$E$2:$E${cn})"), (None if raw_charge else 0.5),
         (_raw_note + "raw battery_pack_dc counter" if raw_charge else ""),
         "battery_pack_dc (raw)" if raw_charge else "Energy Charged DC (kWh)"),
        ("Page 1 · Charging", "Mean start SoC (%)", v["mean_ssoc"],
         f"=AVERAGE(Charges!$B$2:$B${cn})", 0.5, "", "Start SOC (%)"),
        ("Page 1 · Charging", "Mean end SoC (%)", v["mean_esoc"],
         f"=AVERAGE(Charges!$C$2:$C${cn})", 0.5, "", "End SOC (%)"),
        ("Page 2 · EP vs GVM", "GVM min (t)", v["gvm_min"],
         f"=MIN(Trips!$G$2:$G${tn})", 0.05, "", "Vehicle Mass (kg)/1000"),
        ("Page 2 · EP vs GVM", "GVM max (t)", v["gvm_max"],
         f"=MAX(Trips!$G$2:$G${tn})", 0.05, "", ""),
        ("Page 2 · EP vs GVM", "n trips (EP & GVM present)", v["n_pairs"],
         f'=COUNTIFS({tr_f},"<>",{tr_g},"<>")', 0.5, "Legs with both cleaned EP and GVM", ""),
        ("Page 2 · EP vs GVM", "Fit slope (kWh/km per t)", v["gvm_slope"],
         f"=SLOPE({tr_f},{tr_g})", 0.0005, "Least-squares linear fit (blank EP rows excluded)", ""),
        ("Page 2 · EP vs GVM", "Fit R²", v["gvm_r2"],
         f"=RSQ({tr_f},{tr_g})", 0.005, "", ""),
    ]
    # 轻/重载分组（按车辆区分，三选一）：SINGLE_GROUP_REGS 只给 GVM>阈值的离群剔除均值（单组）；
    # LEGACY_TERTILE_REGS 用 33%/67% 三分位（Excel PERCENTILE 独立复算，AVERAGEIFS 引用 D 列）；
    # 其余用 GVM 密度谷聚类阈值（KDE 无 Excel 单公式 → manual 行，AVERAGEIFS 引用 C 列）。
    # 各分支同时备好 range_rows（续航对照行），供下方 if cap 段拼接。
    if v.get("band_lo") is not None:
        lo, hi, lm, mu = v["band_lo"], v["band_hi"], v["laden_min"], v["mass_unladen"]
        rows += [
            ("Page 2 · Conclusion", f"Unladen EP @ {mu:g} t (kWh/km)", v["ep_unladen"],
             f"=SLOPE({tr_f},{tr_g})*{mu:g}+INTERCEPT({tr_f},{tr_g})", 0.02,
             f"EP-vs-GVM trend at the unladen mass — the unladen band ({lo:g}-{hi:g} t = tractor + "
             f"empty trailer) is sparse, so the trend is used; bobtail (< {lo:g} t) excluded", ""),
            ("Page 2 · Conclusion", f"Laden EP, GVM >= {lm:g} t (kWh/km)", v["ep_laden_pt"],
             f'=AVERAGEIFS({tr_f},{tr_g},">={lm:g}")', 0.01, "Mean EP of laden operation", ""),
        ]
        range_rows = ([("Page 2 · Conclusion", f"Unladen range @ {mu:g} t (km)", v["rng_unladen"],
                        f"={cap:g}/(SLOPE({tr_f},{tr_g})*{mu:g}+INTERCEPT({tr_f},{tr_g}))", 6.0,
                        "capacity / unladen trend EP", ""),
                       ("Page 2 · Conclusion", f"Laden range, GVM >= {lm:g} t (km)", v["rng_laden_pt"],
                        f'={cap:g}/AVERAGEIFS({tr_f},{tr_g},">={lm:g}")', 3.0,
                        "capacity / laden mean EP", "")] if cap else [])
    elif v["reg"] in SINGLE_GROUP_REGS:
        thr = SINGLE_GROUP_REGS[v["reg"]]
        rows += [
            ("Page 2 · EP vs GVM", f"n trips GVM > {thr:g} t (EP present)", v["n_laden_tot"],
             f'=COUNTIFS({tr_g},">{thr:g}",{tr_f},"<>")', 0.5, "Laden trips with cleaned EP (before IQR trim)", ""),
            ("Page 2 · EP vs GVM", f"Avg EP, laden GVM>{thr:g}t (IQR-trimmed)", v["ep_laden"],
             "manual", None,
             f"IQR(1.5x)-trimmed mean of EP for GVM > {thr:g} t (n={v['n_laden']} of {v['n_laden_tot']}); "
             f'untrimmed cross-check =AVERAGEIFS({tr_f},{tr_g},">{thr:g}")', ""),
        ]
        ep_laden_row = 6 + len(rows)
        range_rows = ([("Page 2 · Range vs GVM", f"Avg range, laden GVM>{thr:g}t (km)", v["rng_laden"],
                        f'={cap:g}/$C${ep_laden_row}', 1.0, "capacity / trimmed laden mean EP", "")]
                      if cap else [])
    elif v["reg"] in LEGACY_TERTILE_REGS:
        rows += [
            ("Page 2 · EP vs GVM", "GVM 33% quantile (t)", v["gvw_lo"],
             f"=PERCENTILE({tr_g},0.33)", 0.05, "Same interpolation as pandas quantile()", ""),
            ("Page 2 · EP vs GVM", "GVM 67% quantile (t)", v["gvw_hi"],
             f"=PERCENTILE({tr_g},0.67)", 0.05, "", ""),
        ]
        q67_row = 6 + len(rows); q33_row = q67_row - 1
        lo_crit, hi_crit = f'"<"&$D${q33_row}', f'">"&$D${q67_row}'
        rows += [
            ("Page 2 · EP vs GVM", "Average EP, lightest 1/3 (kWh/km)", v["ep_light"],
             f'=AVERAGEIFS({tr_f},{tr_g},{lo_crit})', 0.005, "Mean EP where GVM < 33% quantile", ""),
            ("Page 2 · EP vs GVM", "Average EP, heaviest 1/3 (kWh/km)", v["ep_heavy"],
             f'=AVERAGEIFS({tr_f},{tr_g},{hi_crit})', 0.005, "Mean EP where GVM > 67% quantile", ""),
        ]
        range_rows = ([("Page 2 · Range vs GVM", "Average range, lighter group (km)", v.get("rng_light", float("nan")),
                        f'={cap:g}/AVERAGEIFS({tr_f},{tr_g},{lo_crit})', 1.0,
                        "capacity / cluster mean EP (monotonic with EP, never inverts)", ""),
                       ("Page 2 · Range vs GVM", "Average range, heavier group (km)", v.get("rng_heavy", float("nan")),
                        f'={cap:g}/AVERAGEIFS({tr_f},{tr_g},{hi_crit})', 1.0, "", "")] if cap else [])
    else:
        rows += [
            ("Page 2 · EP vs GVM", "GVM light/heavy cluster split (t)", v["gvw_split"],
             "manual", None,
             "Data-driven KDE density-valley split of per-trip GVM (not a fixed quantile); "
             "compare with the visible gap / bimodality in the EP-vs-GVM scatter", ""),
        ]
        split_row = 6 + len(rows)
        lo_crit, hi_crit = f'"<="&$C${split_row}', f'">"&$C${split_row}'
        rows += [
            ("Page 2 · EP vs GVM", "Average EP, lighter cluster (kWh/km)", v["ep_light"],
             f'=AVERAGEIFS({tr_f},{tr_g},{lo_crit})', 0.005, "Mean EP where GVM <= cluster split", ""),
            ("Page 2 · EP vs GVM", "Average EP, heavier cluster (kWh/km)", v["ep_heavy"],
             f'=AVERAGEIFS({tr_f},{tr_g},{hi_crit})', 0.005, "Mean EP where GVM > cluster split", ""),
        ]
        range_rows = ([("Page 2 · Range vs GVM", "Average range, lighter group (km)", v.get("rng_light", float("nan")),
                        f'={cap:g}/AVERAGEIFS({tr_f},{tr_g},{lo_crit})', 1.0,
                        "capacity / cluster mean EP (monotonic with EP, never inverts)", ""),
                       ("Page 2 · Range vs GVM", "Average range, heavier group (km)", v.get("rng_heavy", float("nan")),
                        f'={cap:g}/AVERAGEIFS({tr_f},{tr_g},{hi_crit})', 1.0, "", "")] if cap else [])
    if cap:
        rows += [
            ("Page 2 · Range vs GVM", "Effective battery capacity (kWh)", None, "manual", None,
             f"Briefing uses {cap:g} kWh — confirm against src/jolt_toolkit/configs/vehicles.json "
             "(effective_capacity_kwh) / SRF", "vehicles.json"),
            ("Page 2 · Range vs GVM", "Reciprocal fit k (≈, kWh/km per t)", v["rng_k"],
             f"=SLOPE({tr_f},{tr_g})", max(abs(v["rng_k"]) * 0.15, 0.003),
             "APPROX check: reciprocal fit minimises range-space error; linearised equivalent "
             "is the EP slope (expect agreement within ~15%)", ""),
            ("Page 2 · Range vs GVM", "Reciprocal fit a (≈, kWh/km)", v["rng_a"],
             f"=INTERCEPT({tr_f},{tr_g})", max(abs(v["rng_a"]) * 0.15, 0.05), "APPROX (see above)", ""),
            ("Page 2 · Range vs GVM", "Reciprocal fit R²", None, "manual", None,
             f"Briefing shows {v['rng_r2']:.2f} — nonlinear fit, no native Excel equivalent; "
             "compare with the chart legend", ""),
        ] + range_rows
    rows += [
        ("Page 2 · Conclusion", "Fully-laden EP @42 t (kWh/km)", v["ep_full"],
         f"=SLOPE({tr_f},{tr_g})*{FULL_LADEN_T:g}+INTERCEPT({tr_f},{tr_g})", 0.02,
         "EP-vs-GVM linear fit extrapolated to the full-laden reference mass", "—"),
    ] + ([("Page 2 · Conclusion", "Fully-laden range @42 t (km)", v["rng_full"],
           f"={cap:g}/(SLOPE({tr_f},{tr_g})*{FULL_LADEN_T:g}+INTERCEPT({tr_f},{tr_g}))", 5.0,
           "capacity / (EP fit @42 t)", "")] if cap else []) + [
        ("Page 2 · Conclusion", "Unladen reference mass (t)", v["mass_unladen"],
         "manual", None,
         "Lighter-cluster median GVM (or briefing_vehicle_specs.json override) — cross-check "
         "against the model kerb mass", "Vehicle Mass (kg)/1000"),
        ("Page 2 · Conclusion", "Laden reference mass (t)", v["mass_laden"],
         "manual", None, "Median GVM of the denser (laden) cluster", "Vehicle Mass (kg)/1000"),
        ("Page 2 · EP vs temperature (laden)", "Fit slope (kWh/km per °C)", v["temp_slope"],
         "manual", None,
         f"Laden-cluster subset (mass controlled), n={v['temp_n']}, "
         f"{v['temp_min']:.0f}–{v['temp_max']:.0f} °C — compare with the chart legend "
         "(KDE subset, no native Excel filter)", "Average Temperature (C)"),
        ("Page 2 · EP vs temperature (laden)", "Fit R²", v["temp_r2"],
         "manual", None, "Laden-cluster subset — compare with the chart legend", ""),
        ("Page 2 · Charging SoC", "Sessions starting <40% SoC (%)", v["pct_low"],
         f'=COUNTIF(Charges!$B$2:$B${cn},"<40")/ROWS(Charges!$B$2:$B${cn})*100', 0.5,
         "Denominator = all sessions (matches generator: blank SoC counts as not-<40)", "Start SOC (%)"),
        ("Page 2 · Charging SoC", "Median start SoC (%)", v["med_start"],
         f"=MEDIAN(Charges!$B$2:$B${cn})", 0.5, "", ""),
        ("Page 2 · Charging SoC", "DC share of charged energy (%)", v["dc_share"],
         f"=SUM(Charges!$E$2:$E${cn})/(SUM(Charges!$D$2:$D${cn})+SUM(Charges!$E$2:$E${cn}))*100",
         0.5, "", ""),
    ]

    header = ["Section", "Item", "Briefing value", "Recomputed (Excel)", "Tolerance",
              "Status", "Definition / filter", "Source column(s)", "Verified by", "Date", "Notes"]
    wa.append([])  # row 3 spacer
    wa.append([])  # row 4
    wa.append([])  # row 5
    wa.append(header)  # row 6
    hr = 6
    for c in wa[hr]:
        c.font = bold; c.fill = head_fill
    for i, (sec, item, val, formula, tol, defn, src) in enumerate(rows, start=hr + 1):
        wa.cell(row=i, column=1).value = sec
        wa.cell(row=i, column=2).value = item
        wa.cell(row=i, column=3).value = None if val is None or pd.isna(val) else float(val)
        wa.cell(row=i, column=4).value = formula
        wa.cell(row=i, column=5).value = tol
        if formula == "manual":
            wa.cell(row=i, column=6).value = "CHECK MANUALLY"
        else:
            wa.cell(row=i, column=6).value = (
                f'=IF(ABS($C{i}-$D{i})<=$E{i},"PASS","FAIL")')
        wa.cell(row=i, column=7).value = defn
        wa.cell(row=i, column=8).value = src
    last = hr + len(rows)
    wa.conditional_formatting.add(
        f"F{hr + 1}:F{last}",
        CellIsRule(operator="equal", formula=['"FAIL"'],
                   fill=PatternFill("solid", start_color="FFC7CE"), font=Font(color="9C0006", bold=True)))
    wa.conditional_formatting.add(
        f"F{hr + 1}:F{last}",
        CellIsRule(operator="equal", formula=['"PASS"'], font=Font(color="1E7B33", bold=True)))
    wa.conditional_formatting.add(
        f"F{hr + 1}:F{last}",
        CellIsRule(operator="equal", formula=['"CHECK MANUALLY"'],
                   fill=PatternFill("solid", start_color="FFF2CC")))
    widths = {"A": 24, "B": 36, "C": 15, "D": 22, "E": 10, "F": 17, "G": 58, "H": 26,
              "I": 14, "J": 12, "K": 28}
    for col, w in widths.items():
        wa.column_dimensions[col].width = w
    wa.freeze_panes = f"A{hr + 1}"

    wb.save(path)
    return len(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--reg", default="YK73WFN")
    ap.add_argument("--period", default="20250301_20250601")
    ap.add_argument("--version", default="2.2.7")
    ap.add_argument("--base", action="store_true", help="用非 finetuned 的 base xlsx")
    ap.add_argument("--anon", action="store_true",
                    help="匿名化展示版：隐去运营方与车牌、底图换无地名 CARTO，"
                         "产物为 report_anon_<period>.*（与命名版共存）")
    ap.add_argument("--all-data", action="store_true",
                    help="读取该车所有周期报告并合并（全部已采集数据，而非单一周期；忽略 --period）")
    ap.add_argument("--min-operator-trips", type=int, default=20,
                    help="round-robin 拆分时,某运营方有效行程 < 此值则跳过(默认 20)。调低可为稀疏"
                         "运营方(如 EX74JXW/WELCH_TRANSPORT ~11 行程)强制出报告;注意样本过少时"
                         "载重点/温度趋势等统计不可靠。")
    ap.add_argument("--page1-basis", choices=["raw", "segment"], default="raw",
                    help="第 1 页能量/距离总量口径：raw=raw_telematics 计数器（含非行驶、真里程，默认）；"
                         "segment=excel report 分段（行驶腿）口径。第 2 页分析始终为分段口径。"
                         "segment 版产物加 _xlsxkpi 后缀，与 raw 版在同一目录共存。")
    args = ap.parse_args()
    df, tr_all, tr, ch, xlsx_path, _cov = compute(args.reg, args.period, args.version,
                                                   finetuned=not args.base, all_data=args.all_data)
    fname = xlsx_path.name
    # no_mass 按**全车**数据判定（在 operator 拆分之前），使同一车各 operator 子集变体一致。
    veh_no_mass = int(tr["mass"].notna().sum()) < max(1, int(0.05 * len(tr)))
    # Operators are DATA-DRIVEN from the report's `Operator` column. --all-data + >1 distinct operator
    # (each with ≥20 valid trips) → one briefing per operator; sparser operators are skipped. Otherwise a
    # single briefing labelled by the dominant operator. (plot_config company_assignment no longer used.)
    def _valid_n(o):
        return int(((tr["operator"] == o) & tr["ep"].notna()).sum())
    ops_all = [o for o in tr["operator"].dropna().unique()
               if str(o).strip() and str(o).strip().lower() != "nan"]
    _min_op = args.min_operator_trips
    eligible = [o for o in ops_all if _valid_n(o) >= _min_op]
    if args.all_data and len(eligible) > 1:
        print(f"[operator-split] {args.reg}: {len(eligible)} operator(s) → " + ", ".join(eligible))
        for o in ops_all:
            if o not in eligible:
                print(f"[operator-split] {args.reg} / {o}: only {_valid_n(o)} valid trips (<{_min_op}) — skipped")
        for o in eligible:
            _emit_briefing(args, tr_all, tr, ch, fname, veh_no_mass, op_filter=o)
    else:
        _emit_briefing(args, tr_all, tr, ch, fname, veh_no_mass, op_filter=None)


def _emit_briefing(args, tr_all_full, tr_full, ch_full, fname, veh_no_mass, op_filter=None):
    # op_filter = an operator name → trips filtered by the data-driven `operator` column, charges by that
    # operator's trip date span. None → single briefing, labelled by the dominant operator in the data.
    # tr_all = every driving leg (PAGE-1 unfiltered counts); tr = the cleaned PAGE-2 analysis set. Both
    # are operator-scoped identically; the charge span + operating period stay keyed to the VALID trips
    # (tr) so the OPERATING PERIOD label and the output directory naming are unchanged.
    if op_filter is not None:
        tr = tr_full[tr_full["operator"] == op_filter].copy()
        tr_all = tr_all_full[tr_all_full["operator"] == op_filter].copy()
        d0f, d1f = tr["st"].min(), tr["et"].max()
        ch = ch_full[(ch_full["st"] >= d0f) & (ch_full["st"] <= d1f)].copy()
        operator = op_filter
    else:
        tr, tr_all, ch = tr_full.copy(), tr_all_full.copy(), ch_full.copy()
        _op = tr["operator"][tr["operator"].astype(str).str.strip().ne("")
                             & tr["operator"].astype(str).str.lower().ne("nan")]
        operator = _op.mode().iloc[0] if not _op.mode().empty else args.reg
    spec = PLOT_CFG["vehicle_specs"].get(args.reg, {})
    make = spec.get("make", "")
    model = VEHICLES.get(args.reg, {}).get("model", "")
    vehicle_model = (f"{make} {model}".strip() or make)  # 真实型号取自 vehicles.json

    # 运营周期 = 有效 trip 的真实跨度（首发车 → 末到达）；输出（文件夹/PDF/标签）一律按此命名。
    d0, d1 = tr["st"].min(), tr["et"].max()
    op_period = (f"{d0:%Y%m%d}_{d1:%Y%m%d}" if pd.notna(d0) and pd.notna(d1) else args.period)
    _op_tag = re.sub(r"[^A-Za-z0-9]+", "_", str(operator)).strip("_") or "OP"
    # ALWAYS operator-scoped naming (<REG>_<OPERATOR>_<period>), for both the per-operator split AND
    # the single-operator case — so every briefing's dir/filename carries its operator consistently.
    tag = f"{_op_tag}_{op_period}"

    # Working set = output_by_TBD; on finalisation the whole dir is renamed to
    # output_by_<YYYYMMDD> (frozen snapshot) and a fresh output_by_TBD is created
    # (naming scheme of 2026-07-10; see pdf_report_workspace/README.md).
    outdir = WORKSPACE / "output_by_TBD" / f"{args.reg}_{tag}"
    fig_rel = "figures_anon" if args.anon else "figures"  # 匿名版独立图目录，与命名版产物共存
    fig_dir = outdir / fig_rel
    cb = int(time.time())  # run token：图片文件名缓存破坏
    _veh_cfg = VEHICLES.get(args.reg, {})
    cap_kwh = _veh_cfg.get("effective_capacity_kwh")
    # Rated-capacity fallback: some vehicles have no telematics-derived effective capacity
    # (LN25NKE: every leg's energy is soc_estimate = ΔSOC × the SRF rated capacity, so no
    # independent estimate exists). Fall back to the rated capacity so the Range figure and
    # bullets still exist — for soc_estimate legs the capacity cancels in range = cap/EP
    # (≡ km-per-%SOC × 100), so no new assumption enters; the wording must then say
    # "rated", never "effective".
    cap_is_rated = False
    if not cap_kwh:
        cap_kwh = _veh_cfg.get("srf_capacity_kwh") or _veh_cfg.get("nominal_kwh")
        cap_is_rated = bool(cap_kwh)
    # No usable mass channel (e.g. YN75NMA, T88RNW) → distribution variant instead of the GVM scatters.
    # Use the vehicle-level flag (computed before any operator filter) so per-operator splits are consistent.
    no_mass = veh_no_mass
    load_pts = _compute_load_points(args.reg, tr)
    charts, st = build_charts(tr, ch, fig_dir, args.reg, cb,
                              here_key=_load_here_key(), cap_kwh=cap_kwh,
                              anon=args.anon, rel=fig_rel, load_pts=load_pts, no_mass=no_mass)

    # ---- 真实 KPI ----
    # PAGE 1 is UNFILTERED: every operational count / total below is computed over tr_all (all driving
    # legs, incl. < MIN_TRIP_KM legs). Only PAGE 2 (charts / fits / conclusions) uses
    # the cleaned `tr`. Active Days, Driving Legs, timeline and Median GVM therefore reflect every trip.
    daily_km = tr_all.dropna(subset=["date"]).groupby("date")["d"].sum()
    tot_km = tr_all["d"].sum(); ndays = tr_all["date"].nunique()
    # Driving Legs = ALL driving legs (unfiltered) — not the analysed-leg (chart n) count any more.
    n_legs = int(len(tr_all))
    # ── Page-1 totals: RAW-TELEMATICS basis (whole-period cumulative counters) ──────────────
    # Total Distance / Total Energy Used / Energy Recuperated on PAGE 1 come from the
    # raw_telematics counters (robust to resets + spikes, see _raw_kpi_totals), NOT the filtered
    # driving-leg sums — so they INCLUDE non-driving consumption (parked HVAC/thermal/aux) and
    # reconcile with Total Energy Charged, and distance is the odometer's true travelled km.
    # PAGE 2 (per-leg EP/GVM scatters, conclusions) keeps the xlsx segment basis unchanged.
    # Vehicles with no raw_telematics / counters fall back to the driving-leg segment basis.
    # Decided PER FIELD (not all-or-nothing): Total Distance uses the raw odometer whenever it is
    # populated; Total Energy Used uses the raw used-energy counter whenever populated, else the
    # driving-leg / SOC-derived segment sum. They are INDEPENDENT — e.g. Scania/DAF/Mercedes
    # populate the odometer but NOT the energy counter, so they get a raw (complete) Total Distance
    # yet keep the segment Total Energy Used. Mean EP always divides Total Energy Used by distance ON
    # THE SAME BASIS as that energy (raw odometer for raw energy; driving-leg km for segment energy)
    # so it stays a valid per-driving-km efficiency, not (driving energy ÷ all-travel distance).
    tot_e_seg = tr_all["ec"].abs().sum()  # driving-leg energy over ALL legs (SOC/counter-derived)
    tot_km_seg = tot_km                     # driving-leg distance (tr_all["d"].sum() above)
    tot_e, tot_km = tot_e_seg, tot_km_seg
    # --page1-basis segment → 完全复刻 excel report 分段口径（不查 raw 计数器），产物为对照版。
    raw_kpi = _raw_kpi_totals(tr_all, args.reg, args.version) if args.page1_basis == "raw" else None
    dist_basis = energy_basis = "segment"
    if raw_kpi:
        if pd.notna(raw_kpi.get("energy_kwh")):
            tot_e, energy_basis = raw_kpi["energy_kwh"], "raw"
        if pd.notna(raw_kpi.get("distance_km")):
            tot_km, dist_basis = raw_kpi["distance_km"], "raw"
            if raw_kpi.get("daily_km") is not None and len(raw_kpi["daily_km"]):
                daily_km = raw_kpi["daily_km"]  # per-day odometer km (robust)
    ep_dist = tot_km if energy_basis == "raw" else tot_km_seg  # Mean-EP distance matches energy basis
    mean_ep = tot_e / ep_dist if ep_dist else float("nan")
    # Daily-average distance = Total Distance ÷ ACTIVE days (ties to the Active Days tile & headline);
    # for the segment basis this equals the per-active-day mean. Max = the true busiest single day.
    daily_avg_km = tot_km / ndays if ndays else float("nan")
    daily_max_km = daily_km.max() if len(daily_km) else float("nan")
    print(f"  [page-1 KPI basis] distance={dist_basis}, energy={energy_basis} "
          f"(raw = raw_telematics counter, incl. non-driving / true odometer; segment = driving-leg sum)")
    # 再生能量：再生计数器有值就用整段稳健计数器总量（独立于能量/距离口径；不漏复位/越界段，
    # ≈项目"~20%"基准）；否则回退 _counter_recup（逐段插值差分）/ 稀疏 xlsx 列
    # （DAF/Mercedes 无该列、Scania 整列空 → 显示 '—'）。
    if raw_kpi and pd.notna(raw_kpi.get("recup_kwh")):
        recup, recup_cov, recup_tot, recup_src = raw_kpi["recup_kwh"], 0, 0, "raw_total"
    else:
        _cr = _counter_recup(tr_all, args.reg, args.version)
        if _cr is not None:
            recup, recup_cov, recup_tot, recup_src = _cr[0], _cr[1], _cr[2], "counter"
        else:
            recup = _sum_or_na(tr_all["recup"]); recup_cov = int(tr_all["recup"].notna().sum())
            recup_tot = len(tr_all); recup_src = "xlsx"
    # 制动能量回收比例 = 再生回收 ÷ page-1 总用电量 ×100（raw 口径分母为整段总用电）。
    recup_pct = recup / tot_e * 100 if (pd.notna(recup) and tot_e) else float("nan")
    # Charged: raw battery-pack DC+AC counter (whole-period — captures ALL charging) in raw mode when
    # available; else the xlsx charge-leg sum (event — only segmented sessions, under-captures 5–36%).
    # SoC stats (median start / mean end) always stay event-based (no raw equivalent).
    ac, dc = _sum_or_na(ch["ac"]), _sum_or_na(ch["dc"])
    charge_basis = "segment"
    if raw_kpi and (pd.notna(raw_kpi.get("charged_dc_kwh")) or pd.notna(raw_kpi.get("charged_ac_kwh"))):
        charge_basis = "raw"
        dc, ac = raw_kpi.get("charged_dc_kwh"), raw_kpi.get("charged_ac_kwh")
    tot_ch = np.nansum([ac, dc]) if (pd.notna(ac) or pd.notna(dc)) else float("nan")
    # A vehicle WITH charging sessions but a charged-energy total of 0 is NOT "0 kWh charged" — its
    # AC/DC channels are unpopulated (some legs carry a placeholder 0.0, not NaN, so _sum_or_na sums to
    # 0). Show "—" rather than a self-contradictory 0 (e.g. EX74JXY: 146 charge legs, SoC 52→96 %, AC/DC
    # counters present-but-all-NaN). Event basis only — the raw battery_pack basis has a true value.
    if charge_basis == "segment" and len(ch) > 0 and not (pd.notna(tot_ch) and tot_ch > 0):
        ac = dc = tot_ch = float("nan")
    med_start = ch["ssoc"].median(); mean_end = ch["esoc"].mean()
    pct_low = (ch["ssoc"] < 40).mean() * 100
    dc_share = dc / tot_ch * 100 if (pd.notna(tot_ch) and tot_ch) else float("nan")
    print(f"  [page-1 charge basis] {charge_basis} "
          f"({'raw battery_pack DC+AC counter' if charge_basis == 'raw' else 'event charge-leg AC+DC'})")
    gvw_med = float("nan") if no_mass else tr_all["mass"].median() / 1000.0  # PAGE 1: all legs
    # Operating-period label: show the start year too when the span crosses a calendar year
    # (e.g. an --all-data briefing 11 Jun 2024 → 1 Dec 2025), otherwise keep the compact form.
    period_label = (
        (f"{d0.day} {d0:%b %Y} – {d1.day} {d1:%b %Y}" if d0.year != d1.year
         else f"{d0.day} {d0:%b} – {d1.day} {d1:%b %Y}") if pd.notna(d0) else args.period)
    # 轻/重载对照：默认 GVM 密度谷聚类自适应分界；LEGACY_TERTILE_REGS 内的车保留固定 1/3 三分位
    # 旧表述（合作方风格基准）。冷/暖对照；满电续航投影（容量/EP）。
    gvw_t = tr["mass"] / 1000.0
    legacy_tertile = args.reg in LEGACY_TERTILE_REGS
    if legacy_tertile:
        gvw_lo, gvw_hi = gvw_t.quantile(0.33), gvw_t.quantile(0.67)
        gvw_split = float("nan")
        light_m = gvw_t < gvw_lo; heavy_m = gvw_t > gvw_hi  # 旧法剔除中间 1/3
        light_label, heavy_label = "Lightest 1/3 of trips", "Heaviest 1/3 of trips"
        light_cond = f"For GVM < {gvw_lo:.0f} t"; heavy_cond = f"For GVM > {gvw_hi:.0f} t"
    else:
        gvw_lo = gvw_hi = float("nan")
        gvw_split = _gvm_cluster_split(gvw_t.values)
        light_m = gvw_t <= gvw_split; heavy_m = gvw_t > gvw_split
        light_label, heavy_label = "Lighter GVM cluster", "Heavier GVM cluster"
        light_cond = f"For GVM < {gvw_split:.0f} t"; heavy_cond = f"For GVM ≥ {gvw_split:.0f} t"
    ep_light = tr.loc[light_m, "ep"].mean(); ep_heavy = tr.loc[heavy_m, "ep"].mean()
    ep_cold = tr.loc[tr["temp"] < 5, "ep"].mean(); ep_warm = tr.loc[tr["temp"] > 15, "ep"].mean()
    if cap_kwh:
        # 簇续航 = 容量 ÷ **簇平均 EP**（不是「逐段续航 cap/EP 的均值」）。后者被低 EP 段（高续航）
        # 主导、且与上方 EP 点评不自洽 → 会出现「重簇续航 > 轻簇」的反直觉反转；前者随簇 EP 单调，
        # 必与 EP 点评一致（EP 低 → 续航高）。
        rng_light = cap_kwh / ep_light if pd.notna(ep_light) else float("nan")
        rng_heavy = cap_kwh / ep_heavy if pd.notna(ep_heavy) else float("nan")

    # 单组模式（SINGLE_GROUP_REGS，如 YN25RSY）：不分轻/重簇，只给 GVM > 阈值的离群剔除均值
    sg_thr = SINGLE_GROUP_REGS.get(args.reg)
    # A configured AI-judged band (briefing_vehicle_specs.json) always takes priority over the
    # single-group treatment (matching _compute_load_points), so the unladen line is not suppressed.
    single_group = (sg_thr is not None) and (load_pts.get("band_lo") is None)
    ep_laden = rng_laden = float("nan"); n_laden = n_laden_tot = 0
    if single_group:
        ep_laden, n_laden, n_laden_tot = _outlier_trimmed_mean(tr.loc[gvw_t > sg_thr, "ep"])
        rng_laden = (cap_kwh / ep_laden) if (cap_kwh and pd.notna(ep_laden)) else float("nan")

    # ---- Load-point conclusion values: unladen / laden / fully-laden EP & range (±1σ) ----
    # Unladen / laden EP = mean of the ACTUAL data points in each operating mask (chosen by
    # judgement band or clustering — see _compute_load_points); ± = within-mask std. Fully-laden
    # (42 t) has no data → EP-vs-GVM linear fit extrapolated; ± = fit residual σ.
    dense_mask = load_pts["dense_mask"]; unladen_mask = load_pts["unladen_mask"]
    m_un, m_la, m_full = load_pts["unladen"], load_pts["laden"], load_pts["full"]
    gvm_slope = st.get("gvm_slope", float("nan")); gvm_icpt = st.get("gvm_intercept", float("nan"))
    gvm_sig = st.get("gvm_sigma", float("nan"))
    is_band = load_pts.get("band_lo") is not None

    def _ep_fit(mt):
        return (gvm_slope * mt + gvm_icpt) if (pd.notna(gvm_slope) and pd.notna(mt)) else float("nan")

    # Laden EP = mean EP of the actual laden data points (ample data); ± = within-cluster std.
    if single_group:
        ep_la = ep_laden; ep_la_sd = float(tr.loc[gvw_t > sg_thr, "ep"].std()); n_la = int(n_laden)
    else:
        ep_la = float(tr.loc[dense_mask, "ep"].mean()); ep_la_sd = float(tr.loc[dense_mask, "ep"].std())
        n_la = int((dense_mask & tr["ep"].notna()).sum())
    # Unladen EP: in band mode the tractor+empty-trailer band is sparse (a handful of trips,
    # easily skewed) → read the EP-vs-GVM trend at the unladen mass instead; otherwise use the
    # lighter-cluster mean. Fully-laden (42 t) is always the trend extrapolated.
    # ± = fit residual σ for trend values, within-cluster std for means.
    has_un = (not single_group) and bool(unladen_mask.any())
    n_un = int((unladen_mask & tr["ep"].notna()).sum()) if has_un else 0
    if not has_un:
        ep_un = float("nan"); ep_un_sd = float("nan")
    elif is_band:
        ep_un = _ep_fit(m_un); ep_un_sd = gvm_sig
    else:
        ep_un = float(tr.loc[unladen_mask, "ep"].mean()); ep_un_sd = float(tr.loc[unladen_mask, "ep"].std())
    ep_fl = _ep_fit(m_full); ep_fl_sd = gvm_sig

    def _rng_pair(ep, ep_sd):
        if not cap_kwh or pd.isna(ep) or ep <= 0:
            return float("nan"), float("nan")
        r = cap_kwh / ep
        return r, (r * ep_sd / ep if pd.notna(ep_sd) else float("nan"))
    rng_un, rng_un_sd = _rng_pair(ep_un, ep_un_sd)
    rng_la, rng_la_sd = _rng_pair(ep_la, ep_la_sd)
    rng_fl, rng_fl_sd = _rng_pair(ep_fl, ep_fl_sd)

    dt = tr[load_pts.get("temp_mask", dense_mask)]   # narrow laden window (mass ~constant for temp)
    td_min, td_max = dt["temp"].min(), dt["temp"].max()
    laden_gvm = dt["mass"] / 1000.0
    laden_gmin, laden_gmax = laden_gvm.min(), laden_gvm.max()
    t_slope = st.get("temp_slope", float("nan")); t_r2 = st.get("temp_r2", float("nan"))
    t_n = int(st.get("temp_n", 0)); t_per10 = t_slope * 10 if pd.notna(t_slope) else float("nan")

    # plain-English number formatters (NaN -> "—"; ± omitted when unknown)
    _pep = lambda v: "—" if pd.isna(v) else f"{v:.2f}"
    _psd = lambda v: "" if pd.isna(v) else f" (±{v:.2f})"
    _prn = lambda v: "—" if pd.isna(v) else f"{v:,.0f}"
    _prsd = lambda v: "" if pd.isna(v) else f" (±{v:,.0f})"
    conclusion_points = []
    if has_un and pd.notna(m_un):
        conclusion_points.append(
            f"Unladen (~{m_un:.0f} t): energy performance {_pep(ep_un)}{_psd(ep_un_sd)} kWh/km, "
            f"range {_prn(rng_un)}{_prsd(rng_un_sd)} km.")
    conclusion_points.append(
        f"Laden (~{m_la:.0f} t): energy performance {_pep(ep_la)}{_psd(ep_la_sd)} kWh/km, "
        f"range {_prn(rng_la)}{_prsd(rng_la_sd)} km.")
    conclusion_points.append(
        f"Fully laden ({m_full:.0f} t, projected to rated GVW): energy performance "
        f"{_pep(ep_fl)}{_psd(ep_fl_sd)} kWh/km, range {_prn(rng_fl)}{_prsd(rng_fl_sd)} km.")
    if pd.notna(gvm_slope):
        conclusion_points.append(
            f"Each extra tonne of load adds ~{gvm_slope:.2f} kWh/km.")
    if load_pts.get("temp_status") == "inconclusive" or pd.isna(t_per10):
        # Even widening the mass window to the whole laden cluster could not show the expected
        # colder→higher-EP trend (a narrow / single-season temperature range with no real signal):
        # state that honestly rather than asserting a noise-driven direction.
        conclusion_points.append(
            "Temperature: the temperature range observed in this period is too limited to "
            "characterise a reliable temperature effect on energy performance.")
    else:
        conclusion_points.append(
            f"Temperature (laden trips, {laden_gmin:.0f}–{laden_gmax:.0f} t): energy performance changes "
            f"~{abs(t_per10):.2f} kWh/km per 10 °C {'colder' if t_per10 < 0 else 'warmer'}.")

    # No-mass variant: replace the load-point conclusions with EP & projected-range distribution stats.
    ep_mean = ep_med = ep_sd_d = ep_q1 = ep_q3 = float("nan"); ep_n_d = 0
    rng_mean = rng_med = rng_q1 = rng_q3 = float("nan")
    if no_mass:
        epv = tr["ep"].dropna()
        ep_n_d = int(len(epv))
        ep_mean = float(epv.mean()); ep_med = float(epv.median()); ep_sd_d = float(epv.std())
        ep_q1 = float(epv.quantile(0.25)); ep_q3 = float(epv.quantile(0.75))
        conclusion_points = [
            f"Energy performance averaged {_pep(ep_mean)} kWh/km over {ep_n_d} trips."]
        if cap_kwh and ep_n_d:
            rngv = cap_kwh / epv
            rng_mean = float(rngv.mean()); rng_med = float(rngv.median())
            rng_q1 = float(rngv.quantile(0.25)); rng_q3 = float(rngv.quantile(0.75))
            conclusion_points.append(
                f"Projected full-charge range averaged {_prn(rng_mean)} km, "
                f"{'rated' if cap_is_rated else 'effective'} capacity {cap_kwh:.0f} kWh ÷ per-trip EP.")
        if pd.notna(t_per10):
            conclusion_points.append(
                f"Temperature (all trips, {tr['temp'].min():.0f}–{tr['temp'].max():.0f} °C): energy "
                f"performance changes ~{abs(t_per10):.2f} kWh/km per 10 °C {'colder' if t_per10 < 0 else 'warmer'}.")
        conclusion_points.append(
            "This vehicle does not report gross vehicle mass, so the load dependence of energy "
            "performance cannot be assessed; the figures above characterise the observed spread.")

    # page-1 summary: concise plain-English takeaways for partners. Kept to a high-level overview
    # + charging behaviour + a data-availability note; the load/range/temperature detail lives on
    # page 2 (not duplicated here).
    summary_points = [
        f"Over {ndays} active days the vehicle covered {tot_km:,.0f} km "
        f"(~{daily_avg_km:.0f} km/day) using {tot_e:,.0f} kWh, averaging {mean_ep:.2f} kWh/km.",
        f"Charging: {len(ch)} sessions over the period, typically plugged in from a median "
        f"{med_start:.0f}% start SoC and charged to a mean {mean_end:.0f}%.",
    ]
    # Regenerative-braking recovery ratio (only when the vehicle reports recuperation, e.g. the
    # Volvo/Renault telematics counter); skipped where there is no regen channel (shown as "—" below).
    if pd.notna(recup_pct):
        summary_points.append(
            f"Regenerative braking recovered ~{recup:,.0f} kWh over the period, "
            f"about {recup_pct:.0f}% of the energy used.")
    _missing = []
    if pd.isna(tot_ch):
        _missing.append("charged energy (AC/DC)")
    if pd.isna(recup):
        _missing.append("energy recuperated")
    if no_mass:
        _missing.append("gross vehicle mass")
    if _missing:
        # Oxford-style list join so 3 items read "A, B, and C" (not "A and B and C").
        if len(_missing) == 1:
            joined = _missing[0]
        elif len(_missing) == 2:
            joined = " and ".join(_missing)
        else:
            joined = ", ".join(_missing[:-1]) + ", and " + _missing[-1]
        summary_points.append(
            f"Data channels: {joined} are not reported by this vehicle telematics "
            f"(shown as “—”).")

    # ---- 人工核实：把简报中出现的每个数字汇成清单，写核实工作簿（Excel 公式独立复算）----
    vals = dict(
        reg=args.reg, period=op_period, version=args.version, fname=fname,
        cap_kwh=cap_kwh, ndays=ndays, n_tr=n_legs, n_ch=len(ch),
        gvw_med=gvw_med, trips_per_day=n_legs / ndays if ndays else float("nan"),
        daily_avg_km=daily_avg_km, daily_max_km=daily_max_km,
        med_dep=median_time(tr_all.groupby("date")["st"].min()),
        med_arr=median_time(tr_all.groupby("date")["et"].max()),
        tot_km=tot_km, tot_e=tot_e, mean_ep=mean_ep, recup=recup, recup_cov=int(recup_cov),
        recup_tot=int(recup_tot), recup_src=recup_src, recup_pct=recup_pct,
        dist_basis=dist_basis, energy_basis=energy_basis, charge_basis=charge_basis,
        tot_ch=tot_ch, ac=ac, dc=dc, mean_ssoc=ch["ssoc"].mean(), mean_esoc=mean_end,
        med_start=med_start, pct_low=pct_low, dc_share=dc_share,
        gvm_min=gvw_t.min(), gvm_max=gvw_t.max(),
        n_pairs=int((tr["ep"].notna() & tr["mass"].notna()).sum()),
        n_ep=int(tr["ep"].notna().sum()),
        gvm_slope=st.get("gvm_slope", float("nan")), gvm_r2=st.get("gvm_r2", float("nan")),
        gvw_split=gvw_split, gvw_lo=gvw_lo, gvw_hi=gvw_hi, ep_light=ep_light, ep_heavy=ep_heavy,
        sg_thr=sg_thr, ep_laden=ep_laden, n_laden=n_laden, n_laden_tot=n_laden_tot, rng_laden=rng_laden,
        temp_min=td_min, temp_max=td_max, temp_n=t_n,
        temp_slope=st.get("temp_slope", float("nan")), temp_r2=st.get("temp_r2", float("nan")),
        gvm_intercept=gvm_icpt, gvm_sigma=gvm_sig,
        mass_unladen=m_un, mass_laden=m_la, mass_full=m_full,
        ep_unladen=ep_un, ep_laden_pt=ep_la, ep_full=ep_fl,
        rng_unladen=rng_un, rng_laden_pt=rng_la, rng_full=rng_fl,
        band_lo=load_pts.get("band_lo"), band_hi=load_pts.get("band_hi"),
        laden_min=load_pts.get("laden_min"), n_un=n_un, n_la=n_la,
    )
    if cap_kwh and "range_fit" in st:
        vals.update(rng_k=st["range_fit"][0], rng_a=st["range_fit"][1],
                    rng_r2=st["range_fit"][3], rng_light=rng_light, rng_heavy=rng_heavy)

    verif_path, n_audit = None, 0
    if not args.anon and not no_mass:  # 命名版才出核实工作簿；no_mass 分布变体暂不出（mass-based 审计不适用，留作后续）
        verif_path = outdir / f"verification_{args.reg}_{tag}{'_xlsxkpi' if args.page1_basis == 'segment' else ''}.xlsx"
        try:
            n_audit = build_verification_workbook(verif_path, tr_all, tr, ch, vals)
        except PermissionError:
            verif_path = verif_path.with_name(f"{verif_path.stem}_{int(time.time())}.xlsx")
            n_audit = build_verification_workbook(verif_path, tr_all, tr, ch, vals)
            print(f"[verify] 原核实工作簿被占用，改写到：{verif_path.name}")

    # PAGE-2 footnote: the trip-filtering rule that shapes the PAGE-2 figures (only the MIN_TRIP_KM
    # valid-trip distance filter now — the SOC-quantisation guard was removed once jolt_toolkit ≥ 2.2.7
    # fixed short-trip EP at source). It lives on PAGE 2 because PAGE 1 is unfiltered (all trips) — the
    # filtering only affects the analysis charts. Shown on BOTH the named and the anon version
    # (methodology, not an identifying data source); the page-1 footer stays empty.
    filter_note = (
        f"Figures exclude driving legs shorter than {MIN_TRIP_KM:.0f} km."
    )
    # ---- 匿名化展示版（SteerCo）：隐去运营方与车牌；型号/数据/图保持不变 ----
    if args.anon:
        operator_disp, reg_disp = "JOLT MEMBER", ""
        analysis_sub = f"JOLT Member · {vehicle_model}"
        source_ops, source_analysis = "", filter_note  # 匿名版：第2页显示筛选脚注，不显示数据来源
    else:
        # Display name: SRF operator tokens are underscore-joined (e.g. "WELCH_TRANSPORT",
        # "PORT_EXPRESS_DAIMLER") — show them with spaces. The raw `operator` value is kept for the
        # per-leg filter, and the dir/filename tag (`_op_tag`) keeps its own filesafe underscores.
        operator_clean = str(operator).replace("_", " ")
        operator_disp, reg_disp = operator_clean, args.reg
        analysis_sub = f"{operator_clean} · {args.reg} · {vehicle_model}"
        source_ops, source_analysis = "", filter_note  # 脚注在第2页(图过滤说明)；第1页页脚空；来源页脚已移除

    def f(v, d=0, suf=""):
        return "—" if pd.isna(v) else f"{v:,.{d}f}{suf}"

    # Page-2 2x2 chart grid (titles rendered by HTML; the daily-energy figure was removed).
    # Laden mass threshold (lower bound of the laden mask) for the temperature-chart title.
    laden_thr = (load_pts.get("laden_min") if load_pts.get("laden_min") is not None
                 else (sg_thr if single_group
                       else (gvw_hi if legacy_tertile else gvw_split)))
    laden_lbl = (f"laden, {laden_gmin:.0f}–{laden_gmax:.0f} t"
                 if pd.notna(laden_gmin) and pd.notna(laden_gmax) else "laden")  # narrow temp window
    if no_mass:
        charts_grid = [dict(title="Energy Performance Distribution", img=charts["ep_dist"])]
        if charts.get("range_dist"):
            charts_grid.append(dict(title="Projected Range Distribution", img=charts["range_dist"]))
        charts_grid += [
            dict(title="Energy Performance vs Ambient Temperature (all trips)", img=charts["temp"]),
            dict(title="Charging Start State of Charge", img=charts["soc"]),
        ]
    else:
        charts_grid = [dict(title="Energy Performance vs Gross Vehicle Mass", img=charts["gvm"])]
        if charts.get("range"):
            charts_grid.append(dict(title="Projected Range vs Gross Vehicle Mass", img=charts["range"]))
        charts_grid += [
            dict(title=f"Energy Performance vs Ambient Temperature ({laden_lbl})", img=charts["temp"]),
            dict(title="Charging Start State of Charge", img=charts["soc"]),
        ]

    ctx = dict(
        reg=reg_disp, operator=operator_disp, vehicle_model=vehicle_model.upper(),
        period_label=period_label,
        timeline=[
            dict(tag="DAY START", icon=ICON_TRUCK, time=f"≈ {median_time(tr_all.groupby('date')['st'].min())}",
                 label="Median first departure"),
            dict(tag="", icon=ICON_CLOCK, time=f"≈ {n_legs/ndays:.1f} trips/day",
                 label=f"{daily_avg_km:.0f} km / day average"),
            dict(tag="DAY END", icon=ICON_TRUCK, time=f"≈ {median_time(tr_all.groupby('date')['et'].max())}",
                 label="Median last arrival"),
        ],
        operating_days=f"OPERATING PERIOD · {period_label.upper()} · {ndays} ACTIVE DAYS",
        summary=[
            dict(k="Active Days", v=f(ndays)),
            dict(k="Driving Legs", v=f(n_legs)),
            dict(k="Charging Sessions", v=f(len(ch))),
            dict(k="Median GVM", v=f(gvw_med, 1, " t")),
        ],
        perf_title="VEHICLE PERFORMANCE",
        perf_rows=[
            dict(k="Total Distance", v=f(tot_km, 0, " km"), cls="num"),
            dict(k="Daily Avg Distance", v=f(daily_avg_km, 0, " km"), cls="num"),
            dict(k="Max Daily Distance", v=f(daily_max_km, 0, " km"), cls="num"),
            dict(k="Total Energy Used", v=f(tot_e, 0, " kWh"), cls="num"),
            dict(k="Mean Energy Performance", v=f(mean_ep, 2, " kWh/km"), cls="num"),
            dict(k="Energy Recuperated", v=f(recup, 0, " kWh"), cls="num pos"),
        ],
        charge_title="CHARGING SESSIONS",
        charge_rows=[
            dict(k="No. of Charging Sessions", v=f(len(ch)), cls="num"),
            dict(k="Total Energy Charged", v=f(tot_ch, 0, " kWh"), cls="num"),
            dict(k="— AC Charged", v=f(ac, 0, " kWh"), cls="num"),
            dict(k="— DC Charged", v=f(dc, 0, " kWh"), cls="num"),
            dict(k="Mean Start SoC", v=f(ch["ssoc"].mean(), 0, " %"), cls="num"),
            dict(k="Mean End SoC", v=f(mean_end, 0, " %"), cls="num"),
        ],
        map_img=charts["map"],
        map_caption=charts.get("map_attrib") or "indicative",
        source_ops=source_ops,
        summary_title="Summary",
        summary_points=summary_points,
        analysis_title="Performance Analysis",
        analysis_sub=analysis_sub,
        charts_grid=charts_grid,
        conclusion_title="Conclusions",
        conclusion_points=conclusion_points,
        source_analysis=source_analysis,
    )

    # Card VALUES render at full, consistent font — the flex layout (label wraps to ≤2 lines, value
    # `flex:0 0 auto`) already fits every realistic value (today's longest is "193,759 kWh" = 11 ch).
    # Only a pathologically long string (≥13 ch — never seen) is shrunk, a last-resort overflow guard.
    def _num_size_cls(s):
        return " v-xs" if len(str(s)) >= 13 else ""
    for _rows in (ctx.get("perf_rows") or [], ctx.get("charge_rows") or []):
        for _r in _rows:
            _r["cls"] = (str(_r.get("cls", "")) + _num_size_cls(_r.get("v", ""))).strip()

    html = Template(TEMPLATE.read_text(encoding="utf-8")).render(**ctx)
    # 匿名版文件名不含车牌（便于直接转发），与命名版产物在同一目录共存
    _bs = "_xlsxkpi" if args.page1_basis == "segment" else ""  # segment 对照版后缀，与 raw 版共存
    stem = f"report_anon_{tag}{_bs}" if args.anon else f"report_{args.reg}_{tag}{_bs}"
    out_html = outdir / f"{stem}.html"
    out_html.write_text(html, encoding="utf-8")
    out_pdf = outdir / f"{stem}.pdf"
    out_pdf = html_to_pdf(out_html, out_pdf)  # 若原名被占用会返回带时间戳的新名

    # 保持输出目录清洁：自动清理生成器自己产生的历史时间戳副本（`_<unix时间戳>` 后缀，
    # 来自规范名被占用时的回退写入），使 PDF 只剩命名版与匿名版各一份最新。
    # 只匹配 10 位时间戳后缀 —— 不碰用户手动命名的其它文件。
    keep = {out_pdf.name} | ({verif_path.name} if verif_path is not None else set())
    stale = re.compile(r"_(?:\d{10})\.(?:pdf|xlsx)$")
    for p in list(outdir.glob("report*.pdf")) + list(outdir.glob("verification_*.xlsx")):
        if p.name not in keep and stale.search(p.name):
            # A timestamped copy NEWER than its canonical is a locked-canonical fallback from a
            # previous run (e.g. the PDF was open in a viewer) — the canonical is the STALE one.
            # Promote the fallback instead of deleting fresh output (deleting it left T88RNW's
            # canonical two days stale on 2026-07-10).
            canon = p.with_name(re.sub(r"_\d{10}(?=\.(?:pdf|xlsx)$)", "", p.name))
            if canon.exists() and canon.name not in keep and p.stat().st_mtime > canon.stat().st_mtime:
                try:
                    os.replace(p, canon)
                    print(f"  [clean] 规范名已过期，用较新的时间戳副本顶替: {canon.name}")
                except OSError:
                    print(f"  [clean] 规范名仍被占用，保留较新副本: {p.name}")
                continue
            try:
                p.unlink()
                print(f"  [clean] 删除旧时间戳副本: {p.name}")
            except OSError:
                print(f"  [clean] 旧副本被占用、暂留: {p.name}")

    # ---- 适用性审计 ----
    print("\n" + "=" * 64)
    print(f"  字段适用性审计 — {args.reg} {op_period}")
    print("=" * 64)
    print("  [真实管线数据]")
    _na = lambda v: "—" if pd.isna(v) else v  # 数据不可用打印 '—'，与简报一致
    for k, v in [("Active days", ndays), ("Driving legs", n_legs), ("Charging sessions", len(ch)),
                 ("Total distance km", round(tot_km)), ("Mean EP kWh/km", round(mean_ep, 3)),
                 ("Total energy kWh", round(tot_e)), ("Charged kWh", _na(round(tot_ch, 0) if pd.notna(tot_ch) else tot_ch)),
                 ("Recup kWh (cov %d/%d)" % (recup_cov, len(tr)), _na(round(recup, 0) if pd.notna(recup) else recup)),
                 ("Median GVM t", round(gvw_med, 1))]:
        print(f"    ✓ {k:32s} {_na(v)}")
    print("  [N/A — 需运营方 / 商业数据，JOLT 管线不产出]")
    for k in ["Customers served", "Cargo tonnage (仅有 GVW)", "Types of goods",
              "Charging cost / rebates (£)", "CO₂ & fuel saved vs diesel (需柴油基线)",
              "Fixed daily schedule (Session 1/break/2)"]:
        print(f"    ✗ {k}")
    print(f"\n  HTML: {out_html}\n  PDF : {out_pdf}")
    if args.anon:
        print("  匿名版：运营方/车牌已隐去，底图为 CARTO 无地名版；数字与命名版一致，"
              "核实请用命名版的核实工作簿。")
    else:
        print(f"  核实工作簿: {verif_path}（{n_audit} 项审计）")
        print("  人工核实：打开核实工作簿，复核 FAIL 项、抽查 PASS 项、完成 CHECK MANUALLY 项，"
              "并填写 Verified by / Date / Notes 列存档。")


if __name__ == "__main__":
    main()
