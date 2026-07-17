# Canonical home since v3.1.0 (P1 copy, 2026-07-17): moved here from
# src/jolt_toolkit/report_generator/finetune.py — the report-finetuner skill now
# owns the finetune library (the package original is removed in P2).
# Invocation pattern (repo root, jolt_toolkit importable via the `jolt` env or
# PYTHONPATH=src):
#   import sys; sys.path.insert(0, r".claude/skills/report-finetuner/code")
#   from finetune import MergeOp, SplitOp, DeleteOp, apply_operations, ...
#
# TODO(P2b): the RENDERING imports below still reach into the package and will be
# rewired to the report-visuals skill CLI in P2b:
#   - jolt_toolkit.report_generator.segment_algorithms.plot_leg_validation
#     (imported inside regenerate_figures) -> repaint via the report-visuals CLI
#     (--finetuned overlay path) instead of importing the painter;
#   - jolt_toolkit.report_generator.validation_generator.ValidationGenerator
#     (imported inside _try_fetch_logger for Logger speed/mass) -> same CLI.
#   (html_viewer is NOT imported — regenerate_inspect_html is self-contained; the
#   docstring mention of report_builder._write_html_viewer is descriptive only.)
# The NON-rendering imports (report_builder HEADERS/DIESEL_HEADERS,
# segment_algorithms constants / VEHICLE_CONFIG / _agg_mass / resolve_mass_agg /
# cluster_mass_data) stay on the package.
"""
finetune.py
===========
对已生成的 xlsx report 做 **人工后处理** 式的分段修正。

使用场景
--------
分段算法 (`segment_algorithms.py`) 即使精调参数后仍会有视觉可见的分段错误
（多切 / 漏切 / 误分类）。我们在主对话里用 LLM vision 审查
`validation_figures/*.png`，得到一份 "操作清单"（merge / split / delete），
然后调用本模块把操作应用到 xlsx 上，并重新生成 validation figures 和
inspect HTML。

关键约定
--------
- 原 xlsx **不被改动**；所有输出都带 ``_finetuned`` 后缀
- openpyxl 保留原有的绿 / 红 / 白 leg-type 底色；被修改过的行额外覆盖
  浅黄色 ``#FFFFCC`` 作为 finetune 标记
- 操作清单按行号降序应用（避免行号偏移）
- 每个操作后重算 Distance / Energy / EP / SOC / AvgSpeed 等派生字段
- 一些需要 pipeline-level 上下文的列（Elevation/Mass-corrected EP、Effective
  Capacity、Wire Energy Efficiency 等）会被置为 None 并在 ``Finetune Log``
  里注明 "需在下次全量 rerun 后同步"
- 应用完所有操作后，Leg Number / Cumulative Distance / Cumulative CO2 会被
  从头重编号

公开 API
--------
- :class:`MergeOp` / :class:`SplitOp` / :class:`DeleteOp` — 操作 dataclass
- :func:`apply_operations` — 对 xlsx 应用操作清单并写出 ``*_finetuned.xlsx``
- :func:`reconstruct_segs_from_xlsx` — 从 xlsx 反推出 (charge_segs,
  discharge_segs) 列表，供 :func:`segment_algorithms.plot_leg_validation` 使用
- :func:`regenerate_figures` — 基于 ``_finetuned.xlsx`` 重新生成 validation PNG
- :func:`regenerate_inspect_html` — 基于 ``_finetuned.xlsx`` 重写 inspect HTML

smoke test
----------
在模块末尾的 ``if __name__ == "__main__":`` 区块里对
``excel_report_database/2.2.2/YK73WFN/jolt_report_YK73WFN_20240601_20240901.xlsx`` 做一个
空操作列表（``[]``）的 :func:`apply_operations`，验证能原样输出 _finetuned
xlsx，然后对 ``2024-06-11_0000`` 这天做 :func:`reconstruct_segs_from_xlsx`
+ :func:`regenerate_figures`。
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from jolt_toolkit.report_generator.report_builder import DIESEL_HEADERS, HEADERS
from jolt_toolkit.report_generator.segment_algorithms import (
    MASS_COL,
    MOVING_SPEED_THRESHOLD_KMH,
    ODO_COL,
    RECUP_COL,
    SOC_COL,
    TIME_COL,
    VEHICLE_CONFIG,
    _agg_mass,
    resolve_mass_agg,
)

logger = logging.getLogger(__name__)

# ── 列号常量（xlsx 是 1-based；openpyxl 的 cell(row=, column=) 也 1-based）──
# HEADERS[0] == 'Leg Number' 占 column=1，所以 HEADERS[i] 位于 column=i+1。

# ── 浅黄色 finetune 标记 ────────────────────────────────────────────────
_FINETUNE_FILL = PatternFill(
    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
)

# ── 电车 / 柴油 leg-type 颜色（用于新 row 的底色）─────────────────────────
_COLOR_TRIP = "C6EFCE"  # 绿
_COLOR_CHARGE = "FFC7CE"  # 红
_COLOR_STOP = "FFFFFF"  # 白

_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# 被 apply_operations 清空（因为无法独立重算）的列。
# 这些列需要 pipeline-level context（effective capacity、Logger speed array
# 等）才能重建；下次完整 rerun 后会自动回填。
_RECOMPUTE_SKIP_COLS = (
    "Energy Performance Corrected by Elevation Difference (kWh/km)",
    "Energy Performance Kinetics Corrected (kWh/km)",
    "Battery Capacity (kWh)",
    "Wire Energy Efficiency (kWh/kWh)",
    "Battery Power (kW)",
    "Peak Charging (kW)",
    "Average Charging (kW)",
    "Energy based on motor power (kWh)",
    "Energy Output from Charger (kWh)",
    "CO2 level (g/kWh)",
    "CO2 for event (g)",
    "Cumulative CO2 (g)",
    "Recuperation Energy (kWh)",
    "Histogram of Accelerator Pedal Position",
    "Histogram of Decelerator Pedal Position",
)


# =============================================================================
# Operation dataclasses
# =============================================================================


@dataclass
class MergeOp:
    """把连续的若干行合并为一行。

    通常用途：把中间的 Stop 吸收掉，或把两段 In Transit 拼接。

    合并后的 leg type = 第一行的 leg type（或由 ``new_type`` 覆盖）。
    合并后的时间窗口 = 第一行 Start Time ~ 最后一行 End Time。
    """

    rows: list[int] = field(default_factory=list)
    new_type: str | None = None
    reason: str = ""
    type: Literal["merge"] = "merge"

    def __post_init__(self):
        if len(self.rows) < 2:
            raise ValueError(f"MergeOp requires >= 2 rows, got {self.rows!r}")
        # 必须严格连续
        sorted_rows = sorted(self.rows)
        if sorted_rows != list(range(sorted_rows[0], sorted_rows[-1] + 1)):
            raise ValueError(
                f"MergeOp rows must be consecutive 1-based xlsx row numbers, "
                f"got {self.rows!r}"
            )
        self.rows = sorted_rows


@dataclass
class SplitOp:
    """把一行在指定时间点 ``at_time`` 拆成两行。

    ``at_time`` 必须落在原行 ``[Start Time, End Time]`` 区间内（UTC 字符串
    ``'YYYY-MM-DD HH:MM:SS'``）。拆出的两段 leg type 都继承原行；如果
    ``new_types=(t1, t2)`` 则分别覆盖。
    """

    row: int = 0
    at_time: str = ""
    new_types: tuple[str, str] | None = None
    reason: str = ""
    type: Literal["split"] = "split"

    def __post_init__(self):
        if self.row < 2:
            raise ValueError(
                f"SplitOp.row must be >= 2 (row 1 is header), got {self.row}"
            )
        if not self.at_time:
            raise ValueError("SplitOp.at_time is required")


@dataclass
class DeleteOp:
    """删除一行（假 trip / 噪声段）。"""

    row: int = 0
    reason: str = ""
    type: Literal["delete"] = "delete"

    def __post_init__(self):
        if self.row < 2:
            raise ValueError(
                f"DeleteOp.row must be >= 2 (row 1 is header), got {self.row}"
            )


Operation = MergeOp | SplitOp | DeleteOp


# =============================================================================
# 辅助：xlsx header / row access
# =============================================================================


def _detect_headers(ws) -> tuple[str, ...]:
    """从 worksheet 第一行读出 HEADERS 布局（EV 或 DIESEL）。"""
    header_row = [c.value for c in ws[1]]
    headers = tuple(h for h in header_row if h is not None)
    if headers == HEADERS:
        return HEADERS
    if headers == DIESEL_HEADERS:
        return DIESEL_HEADERS
    # Fallback: use the detected tuple directly (unknown but still usable for
    # column lookups) — log a warning so the caller knows this xlsx might have
    # been edited manually.
    logger.warning(
        "xlsx header row does not exactly match HEADERS or DIESEL_HEADERS "
        "(n=%d); continuing with raw header tuple. This likely means the "
        "source code layout has changed since the report was generated.",
        len(headers),
    )
    return tuple(headers)


def _col_of(headers: tuple, name: str) -> int:
    """Return 1-based column index of ``name`` (None if absent)."""
    try:
        return headers.index(name) + 1
    except ValueError:
        return 0


def _row_values(ws, row_idx: int, n_cols: int) -> list:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, n_cols + 1)]


def _to_utc(v) -> pd.Timestamp | None:
    """Robust datetime → UTC Timestamp. Returns None if not parseable."""
    if v is None:
        return None
    try:
        ts = pd.Timestamp(v)
    except Exception:
        return None
    if pd.isna(ts):
        return None
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    else:
        ts = ts.tz_convert("UTC")
    return ts


def _excel_duration_days(ts_start, ts_end) -> float:
    s = _to_utc(ts_start)
    e = _to_utc(ts_end)
    if s is None or e is None:
        return float("nan")
    return max((e - s).total_seconds(), 0.0) / 86400.0


# =============================================================================
# Raw CSV lookup
# =============================================================================


def _find_raw_csvs(raw_dir: Path) -> dict[str, Path]:
    """Map leg suffix (e.g. '2024-06-11_0000') → raw CSV path."""
    out: dict[str, Path] = {}
    if not raw_dir.exists():
        return out
    for p in sorted(raw_dir.glob("raw_*.csv")):
        m = re.match(r"raw_(.+)\.csv$", p.name)
        if m:
            out[m.group(1)] = p
    return out


def _load_raw_df(csv_path: Path) -> pd.DataFrame:
    """Load a raw telematics CSV into a dataframe with parsed UTC time."""
    df = pd.read_csv(csv_path, dtype=str, low_memory=False)
    if TIME_COL in df.columns:
        df[TIME_COL] = pd.to_datetime(df[TIME_COL], errors="coerce", utc=True)
    return df


def _raw_df_for_time_window(
    raw_dir: Path, t_start, t_end, _cache: dict | None = None
) -> pd.DataFrame | None:
    """Concatenate all raw CSVs whose time window overlaps [t_start, t_end]."""
    ts_s = _to_utc(t_start)
    ts_e = _to_utc(t_end)
    if ts_s is None or ts_e is None:
        return None
    if _cache is None:
        _cache = {}
    frames = []
    for suffix, csv_path in _find_raw_csvs(raw_dir).items():
        # Fast pre-filter by filename date
        m = re.match(r"(\d{4}-\d{2}-\d{2})_", suffix)
        if m:
            try:
                day = pd.Timestamp(m.group(1)).tz_localize("UTC")
                if day + pd.Timedelta(days=1) < ts_s or day > ts_e:
                    continue
            except Exception:
                pass
        if csv_path in _cache:
            df = _cache[csv_path]
        else:
            df = _load_raw_df(csv_path)
            _cache[csv_path] = df
        if TIME_COL not in df.columns or df.empty:
            continue
        mask = (df[TIME_COL] >= ts_s) & (df[TIME_COL] <= ts_e)
        if mask.any():
            frames.append(df.loc[mask])
    if not frames:
        return None
    out = pd.concat(frames, ignore_index=True).sort_values(TIME_COL)
    return out.reset_index(drop=True)


# =============================================================================
# 派生字段重算
# =============================================================================


def _veh_cfg(reg: str) -> dict:
    return VEHICLE_CONFIG.get(reg, {})


def _soc_endpoints(df: pd.DataFrame) -> tuple[float | None, float | None]:
    if SOC_COL not in df.columns:
        return None, None
    soc = pd.to_numeric(df[SOC_COL], errors="coerce")
    soc = soc[soc.notna() & (soc > 0)]
    if soc.empty:
        return None, None
    return float(soc.iloc[0]), float(soc.iloc[-1])


def _energy_delta_kwh(df: pd.DataFrame, col: str) -> float | None:
    """Cumulative Wh counter → kWh delta over the window (end − start)."""
    if col not in df.columns:
        return None
    vals = pd.to_numeric(df[col], errors="coerce").dropna()
    if len(vals) < 2:
        return None
    return float(vals.iloc[-1] - vals.iloc[0]) / 1000.0


def _odometer_distance_km(df: pd.DataFrame) -> float | None:
    if ODO_COL not in df.columns:
        return None
    vals = pd.to_numeric(df[ODO_COL], errors="coerce").dropna()
    if len(vals) < 2:
        return None
    d = float(vals.iloc[-1] - vals.iloc[0])
    return d if d > 0 else None


def _mass_mean(
    df: pd.DataFrame, mass_col: str, method: str = "mean", speed_col: str | None = None
) -> tuple[float | None, float | None]:
    """Aggregate the window's vehicle mass into ``(mass_kg, cv)``.

    v2.2.6: mirrors ``report_builder._get_vehicle_mass`` — filter to valid (> 0)
    and, when a ``speed_col`` is present, prefer moving-only samples (>= 2), then
    aggregate by ``method`` via the shared ``_agg_mass``. Returns ``(None, None)``
    (not nan) to keep the finetune row-rewrite contract.
    """
    if mass_col not in df.columns:
        return None, None
    vals = pd.to_numeric(df[mass_col], errors="coerce")
    valid = vals.notna() & (vals > 0)
    if speed_col and speed_col in df.columns:
        spd = pd.to_numeric(df[speed_col], errors="coerce")
        moving = valid & spd.notna() & (spd > MOVING_SPEED_THRESHOLD_KMH)
        if moving.sum() >= 2:
            valid = moving
    sel = vals[valid]
    # ``mad_tw_mean`` needs a per-sample time axis aligned to ``sel``; build it
    # from TIME_COL only for that method (other methods are byte-identical).
    timestamps = None
    if (method or "").lower() == "mad_tw_mean" and TIME_COL in df.columns:
        timestamps = pd.to_datetime(df[TIME_COL], errors="coerce", utc=True).loc[
            sel.index
        ]
    mass_kg, cv = _agg_mass(sel, method, timestamps=timestamps)
    if not np.isfinite(mass_kg):
        return None, None
    return mass_kg, (cv if np.isfinite(cv) else None)


def _elev_diff(df: pd.DataFrame, alt_col: str | None) -> float | None:
    if alt_col is None or alt_col not in df.columns:
        return None
    vals = pd.to_numeric(df[alt_col], errors="coerce").dropna()
    if len(vals) < 2:
        return None
    return round(float(vals.iloc[-1] - vals.iloc[0]), 1)


def _avg_speed_kmh(distance_km: float | None, t_start, t_end) -> float | None:
    if distance_km is None or distance_km <= 0:
        return None
    s = _to_utc(t_start)
    e = _to_utc(t_end)
    if s is None or e is None:
        return None
    hours = max((e - s).total_seconds(), 0) / 3600.0
    if hours <= 0:
        return None
    return round(distance_km / hours, 2)


def _recompute_leg_derived(
    row: list, headers: tuple, raw_dir: Path, reg: str, raw_cache: dict | None = None
) -> dict:
    """Recompute distance / energy / EP / SOC / avg speed / mass for a row.

    Returns a dict describing which fields were rewritten vs cleared.
    Updates ``row`` in-place. ``row`` is a list of length ``len(headers)``
    (1-based aligned to xlsx columns: row[0] == Leg Number slot).
    """
    cfg = _veh_cfg(reg)
    mass_col = cfg.get("mass_col", MASS_COL)
    altitude_col = cfg.get("altitude_col")
    total_energy_col = cfg.get("total_energy_col")
    moving_energy_col = cfg.get("moving_energy_col")
    ac_col = cfg.get("ac_col")
    dc_col = cfg.get("dc_col")
    # v2.2.6: 与报告同口径的稳健质量聚合（vehicle > pipeline > 默认 'mean'）
    mass_agg = resolve_mass_agg(reg)
    speed_col = cfg.get("speed_col", "wheel_based_speed")

    def _set(name: str, value):
        c = _col_of(headers, name)
        if c:
            row[c - 1] = value

    def _get(name: str):
        c = _col_of(headers, name)
        if not c:
            return None
        return row[c - 1]

    status: dict = {"rewritten": [], "cleared": [], "skipped": []}

    leg_type = _get("Leg Type")
    t_start = _get("Start Time (UTC)")
    t_end = _get("End Time (UTC)")
    is_stop = isinstance(leg_type, str) and leg_type.strip().lower() == "stop"
    is_charge = (
        isinstance(leg_type, str)
        and re.match(r"^(AC|DC|Charge|Mix|estimated)", leg_type, re.IGNORECASE)
        is not None
    )
    is_trip = (not is_stop) and (not is_charge)

    # ── Duration（对所有 leg type 都有效）─────────────────────────────────
    dur_days = _excel_duration_days(t_start, t_end)
    if not np.isnan(dur_days):
        _set("Duration (HH:MM:SS)", dur_days)
        status["rewritten"].append("Duration (HH:MM:SS)")

    # ── Stop 行：distance / speed / elevation diff 强制为 0，其他列不动 ──
    if is_stop:
        _set("Distance (km)", 0.0)
        _set("Average Speed (km/h)", 0.0)
        if "Elevation Difference (m)" in headers:
            _set("Elevation Difference (m)", 0.0)
        status["rewritten"] += [
            "Distance (km)",
            "Average Speed (km/h)",
            "Elevation Difference (m)",
        ]
        # Clear EP columns explicitly for Stop rows
        for c in (
            "Energy Performance (kWh/km)",
            "Energy Performance Corrected by Elevation Difference (kWh/km)",
            "Energy Performance Kinetics Corrected (kWh/km)",
            "Energy Change (kWh)",
        ):
            _set(c, None)
            status["cleared"].append(c)
        return status

    # ── Trip / Charge 行：从 raw CSV 切窗口 ─────────────────────────────
    df_win = _raw_df_for_time_window(raw_dir, t_start, t_end, _cache=raw_cache)
    if df_win is None or df_win.empty:
        status["skipped"].append("raw_csv_missing")
        return status

    # Distance
    distance_km = _odometer_distance_km(df_win)
    if distance_km is not None:
        _set("Distance (km)", round(distance_km, 3))
        status["rewritten"].append("Distance (km)")
    else:
        # Fallback: keep existing distance if set, else clear
        _set("Distance (km)", None)
        status["cleared"].append("Distance (km)")

    # Avg speed
    avg_sp = _avg_speed_kmh(distance_km, t_start, t_end)
    if avg_sp is not None:
        _set("Average Speed (km/h)", avg_sp)
        status["rewritten"].append("Average Speed (km/h)")

    # Mass
    mass_mean, mass_cv = _mass_mean(
        df_win, mass_col, method=mass_agg, speed_col=speed_col
    )
    if mass_mean is not None:
        _set("Vehicle Mass (kg)", mass_mean)
        status["rewritten"].append("Vehicle Mass (kg)")
    if mass_cv is not None:
        _set("Vehicle Mass CV (reliability)", mass_cv)
        status["rewritten"].append("Vehicle Mass CV (reliability)")

    # Elevation difference
    elev = _elev_diff(df_win, altitude_col)
    if elev is not None:
        _set("Elevation Difference (m)", elev)
        status["rewritten"].append("Elevation Difference (m)")

    # SOC endpoints
    soc_s, soc_e = _soc_endpoints(df_win)
    if soc_s is not None:
        _set("Start SOC (%)", round(soc_s, 1))
        status["rewritten"].append("Start SOC (%)")
    if soc_e is not None:
        _set("End SOC (%)", round(soc_e, 1))
        status["rewritten"].append("End SOC (%)")
    if soc_s is not None and soc_e is not None:
        _set("SOC Change (%)", round(soc_e - soc_s, 1))
        status["rewritten"].append("SOC Change (%)")

    # Energy change — priority: total_energy → moving_energy → AC/DC (charge)
    energy_change_kwh: float | None = None
    energy_source = _get("Energy Source")
    if is_trip:
        if total_energy_col:
            dE = _energy_delta_kwh(df_win, total_energy_col)
            if dE is not None:
                # trip is net consumption → sign negative
                energy_change_kwh = -abs(dE) if dE > 0 else dE
                energy_source = "total_energy"
        if energy_change_kwh is None and moving_energy_col:
            dE = _energy_delta_kwh(df_win, moving_energy_col)
            if dE is not None:
                energy_change_kwh = -abs(dE) if dE > 0 else dE
                energy_source = "moving_energy"
        if energy_change_kwh is None and soc_s is not None and soc_e is not None:
            # SOC estimate fallback — but we don't have effective_capacity_kwh
            # here, so mark as skipped and clear the cell
            nominal = (
                cfg.get("effective_capacity_kwh")
                or cfg.get("srf_capacity_kwh")
                or cfg.get("nominal_kwh")
            )
            if nominal:
                energy_change_kwh = (soc_e - soc_s) / 100.0 * nominal
                energy_source = "soc_estimate"
    elif is_charge:
        # Charge: AC+DC or total_energy delta. Keep sign positive.
        if ac_col and dc_col:
            dE_ac = _energy_delta_kwh(df_win, ac_col) or 0.0
            dE_dc = _energy_delta_kwh(df_win, dc_col) or 0.0
            total = dE_ac + dE_dc
            if total > 0:
                energy_change_kwh = total
                energy_source = "ac_dc"
                _set("Energy Charged AC (kWh)", round(dE_ac, 3))
                _set("Energy Charged DC (kWh)", round(dE_dc, 3))
                status["rewritten"] += [
                    "Energy Charged AC (kWh)",
                    "Energy Charged DC (kWh)",
                ]

    if energy_change_kwh is not None:
        _set("Energy Change (kWh)", round(energy_change_kwh, 3))
        _set("Energy Source", energy_source)
        status["rewritten"] += ["Energy Change (kWh)", "Energy Source"]

    # Energy Performance (only for discharge trips with distance > 0)
    if is_trip and energy_change_kwh is not None and distance_km and distance_km > 0:
        ep = abs(energy_change_kwh) / distance_km
        _set("Energy Performance (kWh/km)", round(ep, 4))
        status["rewritten"].append("Energy Performance (kWh/km)")
    else:
        # Clear EP columns for charge / Stop / missing-distance rows
        _set("Energy Performance (kWh/km)", None)
        status["cleared"].append("Energy Performance (kWh/km)")

    # Clear pipeline-level columns (needs downstream rerun)
    for c in _RECOMPUTE_SKIP_COLS:
        if _col_of(headers, c):
            _set(c, None)
            status["cleared"].append(c)

    return status


# =============================================================================
# xlsx I/O and operations
# =============================================================================


def _wb_sort_and_renumber(ws, n_data_rows: int, headers: tuple) -> int:
    """Sort data rows by Start Time, renumber Leg Number and Cumulative Distance.

    Returns the new ``n_data_rows`` (unchanged here; rows were already
    sort-stable before this call). Does NOT touch the header row.
    """
    c_leg = _col_of(headers, "Leg Number")
    c_start = _col_of(headers, "Start Time (UTC)")
    c_type = _col_of(headers, "Leg Type")
    c_dist = _col_of(headers, "Distance (km)")
    c_cum = _col_of(headers, "Cumulative Distance (km)")
    n_cols = len(headers)

    # Collect existing rows, sort by start time
    rows_data = []
    for r in range(2, 2 + n_data_rows):
        vals = _row_values(ws, r, n_cols)
        start = _to_utc(vals[c_start - 1]) if c_start else None
        rows_data.append((start, vals))
    rows_data.sort(key=lambda t: (t[0] is None, t[0]))

    # Rewrite rows in sorted order
    cum = 0.0
    for i, (_, vals) in enumerate(rows_data):
        r = 2 + i
        vals[c_leg - 1] = i + 1  # Leg Number 1-based
        leg_type = vals[c_type - 1] if c_type else None
        is_trip = (
            isinstance(leg_type, str)
            and leg_type.strip().lower() != "stop"
            and not re.match(r"^(AC|DC|Charge|Mix|estimated)", leg_type, re.IGNORECASE)
        )
        # Cumulative Distance: running total over trips only; Stop rows carry
        # over, charge rows leave empty
        dist = vals[c_dist - 1] if c_dist else None
        if c_cum:
            if is_trip and isinstance(dist, (int, float)) and not _is_nan(dist):
                cum += float(dist)
                vals[c_cum - 1] = round(cum, 3)
            elif isinstance(leg_type, str) and leg_type.strip().lower() == "stop":
                vals[c_cum - 1] = round(cum, 3) if cum > 0 else None
            else:  # charge
                vals[c_cum - 1] = None
        # Write back
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c).value = v

    return len(rows_data)


def _is_nan(v) -> bool:
    try:
        return bool(np.isnan(v))
    except (TypeError, ValueError):
        return False


def _mark_row_finetuned(ws, row_idx: int, n_cols: int):
    """Overlay a light-yellow fill on every cell of the row."""
    for c in range(1, n_cols + 1):
        ws.cell(row=row_idx, column=c).fill = _FINETUNE_FILL


def _apply_merge(
    ws, op: MergeOp, headers: tuple, raw_dir: Path, reg: str, raw_cache: dict
) -> dict:
    """Collapse ``op.rows`` into a single row at ``op.rows[0]``; delete the rest."""
    n_cols = len(headers)
    first_r = op.rows[0]
    last_r = op.rows[-1]

    # Capture before-summary
    before_leg_types = [
        ws.cell(row=r, column=_col_of(headers, "Leg Type")).value for r in op.rows
    ]

    # Build merged row from first_r but taking End Time / Destination from last_r
    first_vals = _row_values(ws, first_r, n_cols)
    last_vals = _row_values(ws, last_r, n_cols)

    carry_end_cols = (
        "End Time (UTC)",
        "Destination (Lat, Lon)",
        "Destination Place",
        "End SOC (%)",
    )
    for cname in carry_end_cols:
        c = _col_of(headers, cname)
        if c:
            first_vals[c - 1] = last_vals[c - 1]

    if op.new_type is not None:
        c = _col_of(headers, "Leg Type")
        first_vals[c - 1] = op.new_type

    # Write back to first_r
    for c, v in enumerate(first_vals, start=1):
        ws.cell(row=first_r, column=c).value = v

    # Delete rows from last_r → first_r + 1 (descending to avoid shifts)
    for r in range(last_r, first_r, -1):
        ws.delete_rows(r, 1)

    # Recompute derived fields on the merged row
    merged_vals = _row_values(ws, first_r, n_cols)
    recompute_status = _recompute_leg_derived(
        merged_vals, headers, raw_dir, reg, raw_cache=raw_cache
    )
    for c, v in enumerate(merged_vals, start=1):
        ws.cell(row=first_r, column=c).value = v

    _mark_row_finetuned(ws, first_r, n_cols)

    return {
        "op": "merge",
        "rows": op.rows,
        "new_type": op.new_type,
        "reason": op.reason,
        "before_leg_types": before_leg_types,
        "after_leg_type": ws.cell(
            row=first_r, column=_col_of(headers, "Leg Type")
        ).value,
        "recompute": recompute_status,
    }


def _apply_split(
    ws, op: SplitOp, headers: tuple, raw_dir: Path, reg: str, raw_cache: dict
) -> dict:
    """Split ``op.row`` at ``op.at_time`` into two rows; insert a new row below."""
    n_cols = len(headers)
    r = op.row
    vals = _row_values(ws, r, n_cols)

    c_start = _col_of(headers, "Start Time (UTC)")
    c_end = _col_of(headers, "End Time (UTC)")
    c_type = _col_of(headers, "Leg Type")

    orig_start = _to_utc(vals[c_start - 1])
    orig_end = _to_utc(vals[c_end - 1])
    split_ts = _to_utc(op.at_time)
    if split_ts is None:
        raise ValueError(f"Could not parse SplitOp.at_time={op.at_time!r}")
    if orig_start is None or orig_end is None:
        raise ValueError(f"Row {r} has unparsable Start/End Time")
    if not (orig_start < split_ts < orig_end):
        raise ValueError(
            f"SplitOp.at_time={op.at_time} outside row {r} window "
            f"[{orig_start}, {orig_end}]"
        )

    before_leg_type = vals[c_type - 1]

    # Build two new rows by copy
    row_a = list(vals)
    row_b = list(vals)

    # Row A ends at split_ts
    row_a[c_end - 1] = split_ts.tz_convert(None).to_pydatetime()
    row_b[c_start - 1] = split_ts.tz_convert(None).to_pydatetime()

    if op.new_types is not None:
        t1, t2 = op.new_types
        row_a[c_type - 1] = t1
        row_b[c_type - 1] = t2

    # Origin/destination: row A keeps original origin, row B inherits original
    # destination. We cannot know the true geographical mid-point without
    # slicing the raw CSV, so for row A's destination and row B's origin we use
    # the best-effort mid GPS reading from the raw data.
    df_mid = _raw_df_for_time_window(raw_dir, orig_start, split_ts, _cache=raw_cache)
    df_after = _raw_df_for_time_window(raw_dir, split_ts, orig_end, _cache=raw_cache)

    def _last_gps(df):
        if df is None or df.empty:
            return None, None
        for c_lat, c_lon in (
            ("latitude", "longitude"),
            ("gnss_latitude", "gnss_longitude"),
        ):
            if c_lat in df.columns and c_lon in df.columns:
                la = pd.to_numeric(df[c_lat], errors="coerce").dropna()
                lo = pd.to_numeric(df[c_lon], errors="coerce").dropna()
                if len(la) and len(lo):
                    idx = min(la.index.max(), lo.index.max())
                    return float(la.loc[idx]), float(lo.loc[idx])
        return None, None

    mid_lat, mid_lon = _last_gps(df_mid)
    if mid_lat is not None and mid_lon is not None:
        mid_str = f"Point({mid_lat:.6f} {mid_lon:.6f})"
        c_dest = _col_of(headers, "Destination (Lat, Lon)")
        c_org = _col_of(headers, "Origin (Lat, Lon)")
        if c_dest:
            row_a[c_dest - 1] = mid_str
        if c_org:
            row_b[c_org - 1] = mid_str
        # Invalidate place fields; user can re-geocode by running LoggerPatcher
        c_dest_p = _col_of(headers, "Destination Place")
        c_org_p = _col_of(headers, "Origin Place")
        if c_dest_p:
            row_a[c_dest_p - 1] = None
        if c_org_p:
            row_b[c_org_p - 1] = None

    # Write row A in place of the original, then insert row B below
    for c, v in enumerate(row_a, start=1):
        ws.cell(row=r, column=c).value = v
    ws.insert_rows(r + 1, 1)
    for c, v in enumerate(row_b, start=1):
        ws.cell(row=r + 1, column=c).value = v

    # Recompute both rows
    for rr in (r, r + 1):
        rv = _row_values(ws, rr, n_cols)
        status = _recompute_leg_derived(rv, headers, raw_dir, reg, raw_cache=raw_cache)
        for c, v in enumerate(rv, start=1):
            ws.cell(row=rr, column=c).value = v
        _mark_row_finetuned(ws, rr, n_cols)

    return {
        "op": "split",
        "row": r,
        "at_time": op.at_time,
        "new_types": op.new_types,
        "reason": op.reason,
        "before_leg_type": before_leg_type,
        "after_leg_types": (
            ws.cell(row=r, column=c_type).value,
            ws.cell(row=r + 1, column=c_type).value,
        ),
    }


def _apply_delete(ws, op: DeleteOp, headers: tuple) -> dict:
    """Simply delete ``op.row``."""
    c_type = _col_of(headers, "Leg Type")
    before_leg_type = ws.cell(row=op.row, column=c_type).value if c_type else None
    ws.delete_rows(op.row, 1)
    return {
        "op": "delete",
        "row": op.row,
        "reason": op.reason,
        "before_leg_type": before_leg_type,
    }


def _write_finetune_log(wb, entries: list[dict]):
    """Append a 'Finetune Log' sheet with all applied operations."""
    if "Finetune Log" in wb.sheetnames:
        ws = wb["Finetune Log"]
    else:
        ws = wb.create_sheet(title="Finetune Log")

    # Header row
    header_font = Font(bold=True)
    headers = (
        "Timestamp",
        "Operation",
        "Rows Affected",
        "Reason",
        "Before Summary",
        "After Summary",
    )
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    for i, entry in enumerate(entries, start=2):
        if entry["op"] == "merge":
            rows_str = ",".join(str(r) for r in entry["rows"])
            before = " | ".join(str(x) for x in entry["before_leg_types"])
            after = str(entry["after_leg_type"])
        elif entry["op"] == "split":
            rows_str = f"{entry['row']} @ {entry['at_time']}"
            before = str(entry["before_leg_type"])
            after = " | ".join(str(x) for x in entry["after_leg_types"])
        elif entry["op"] == "delete":
            rows_str = str(entry["row"])
            before = str(entry["before_leg_type"])
            after = "(deleted)"
        else:
            rows_str = before = after = ""
        ws.cell(row=i, column=1, value=now)
        ws.cell(row=i, column=2, value=entry["op"])
        ws.cell(row=i, column=3, value=rows_str)
        ws.cell(row=i, column=4, value=entry.get("reason", ""))
        ws.cell(row=i, column=5, value=before)
        ws.cell(row=i, column=6, value=after)

    # Column widths
    widths = (24, 12, 24, 50, 40, 40)
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


# =============================================================================
# Public: apply_operations
# =============================================================================


def apply_operations(
    xlsx_path: str | Path,
    operations: list,
    raw_telematics_dir: str | Path,
    out_path: str | Path | None = None,
) -> Path:
    """Apply the operation list to ``xlsx_path`` and write ``*_finetuned.xlsx``.

    - Operations are applied in **descending row order** to avoid index drift
    - Each operation triggers a best-effort recompute of derived fields using
      the raw telematics CSVs in ``raw_telematics_dir``
    - After all ops, rows are re-sorted by Start Time and ``Leg Number`` +
      ``Cumulative Distance`` are re-numbered from 1
    - Modified rows are overlaid with a light-yellow fill (``#FFFFCC``)
    - A ``Finetune Log`` worksheet is appended recording each operation

    Returns the path of the generated xlsx.
    """
    xlsx_path = Path(xlsx_path)
    raw_dir = Path(raw_telematics_dir)

    if out_path is None:
        out_path = xlsx_path.with_name(xlsx_path.stem + "_finetuned" + xlsx_path.suffix)
    else:
        out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(xlsx_path))
    if "Report" not in wb.sheetnames:
        raise ValueError(f'{xlsx_path} has no "Report" sheet')
    ws = wb["Report"]
    headers = _detect_headers(ws)
    n_cols = len(headers)

    # Detect number of data rows (scan until Leg Number column goes empty)
    c_leg = _col_of(headers, "Leg Number")
    n_data_rows = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=c_leg).value in (None, ""):
            break
        n_data_rows += 1

    raw_cache: dict = {}
    log_entries: list[dict] = []

    # Apply ops in descending row order (rows param is xlsx 1-based)
    # — for MergeOp, key = max row; for SplitOp, key = row; for DeleteOp, key = row
    def _op_key(op) -> int:
        if isinstance(op, MergeOp):
            return max(op.rows)
        if isinstance(op, (SplitOp, DeleteOp)):
            return op.row
        raise TypeError(f"Unknown operation: {op!r}")

    for op in sorted(operations, key=_op_key, reverse=True):
        if isinstance(op, MergeOp):
            entry = _apply_merge(
                ws, op, headers, raw_dir, _reg_from_name(xlsx_path.name), raw_cache
            )
        elif isinstance(op, SplitOp):
            entry = _apply_split(
                ws, op, headers, raw_dir, _reg_from_name(xlsx_path.name), raw_cache
            )
        elif isinstance(op, DeleteOp):
            entry = _apply_delete(ws, op, headers)
        else:
            raise TypeError(f"Unknown operation type: {type(op).__name__}")
        log_entries.append(entry)

    # Recount rows after edits
    n_data_rows = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=c_leg).value in (None, ""):
            break
        n_data_rows += 1

    # Sort + renumber — only when at least one op modified rows. A no-op
    # apply_operations([]) must produce a bit-for-bit output (modulo metadata
    # that openpyxl necessarily rewrites: shared-strings, style indices).
    if n_data_rows > 0 and operations:
        _wb_sort_and_renumber(ws, n_data_rows, headers)

    # Write Finetune Log
    if log_entries:
        _write_finetune_log(wb, log_entries)

    wb.save(str(out_path))
    logger.info(
        "finetune: wrote %s (%d operations, %d rows)",
        out_path.name,
        len(operations),
        n_data_rows,
    )
    return out_path


def _reg_from_name(filename: str) -> str:
    """Parse 'jolt_report_REG_DATE_DATE[.*].xlsx' → 'REG'."""
    m = re.match(r"jolt_report_([A-Z0-9]+)_\d{8}_\d{8}", filename)
    if not m:
        raise ValueError(f"Cannot infer vehicle REG from filename: {filename}")
    return m.group(1)


# =============================================================================
# Public: reconstruct_segs_from_xlsx
# =============================================================================

# 严格白名单：反推 segs 时只接受下列 leg type，其它（包括 Stop / 空值 /
# 未识别字符串）一律跳过，防止 Stop 行被误判为 discharge（SOC drop 来源于
# auxiliary drain 或 DC rebalance，并非真正的行驶放电）。
_DISCHARGE_LEG_TYPES = frozenset(
    {
        "In Transit",
        "Outbound",
        "Return",
        "Round Trip",
        "In House",
    }
)
# Charge 行由前缀正则判定，覆盖 "AC Home/Away", "DC Home/Away",
# "AC/DC Home/Away", "Charge Home/Away", "Mix <anything>", "estimated …"。
_CHARGE_LEG_PREFIX_RE = re.compile(r"^(AC|DC|Charge|Mix|estimated)", re.IGNORECASE)


def reconstruct_segs_from_xlsx(
    xlsx_path: str | Path,
    date: str,
) -> tuple[list[dict], list[dict]]:
    """Read xlsx and build (charge_segs, discharge_segs) for ``date``.

    The returned seg dicts follow the schema produced by
    :mod:`segment_algorithms` (start_time, end_time, start_soc, end_soc,
    delta_soc_pct, delta_energy_kwh, effective_capacity_kwh, energy_source,
    odo_start_km, odo_end_km, lat_start/lon_start/lat_end/lon_end for
    discharge; latitude/longitude for charge) — at least enough for
    :func:`plot_leg_validation` to render.

    Leg-type filtering uses a strict whitelist:

    * Discharge: :data:`_DISCHARGE_LEG_TYPES` (``In Transit`` / ``Outbound`` /
      ``Return`` / ``Round Trip`` / ``In House``).
    * Charge: any leg type whose first token matches ``AC`` / ``DC`` /
      ``Charge`` / ``Mix`` / ``estimated`` (:data:`_CHARGE_LEG_PREFIX_RE`).
    * ``Stop`` rows and any unrecognised leg types are silently skipped
      (Stop) or skipped with a ``warning`` log (unknown). This prevents
      Stop rows — whose SOC drop reflects auxiliary drain / DC rebalance
      rather than actual discharge — from being counted as discharge segs.

    The ``date`` string is ``YYYY-MM-DD`` (UTC); any leg whose
    ``[Start, End]`` overlaps that calendar day is included.
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    ws = wb["Report"]
    headers = _detect_headers(ws)
    n_cols = len(headers)

    day_start = pd.Timestamp(date).tz_localize("UTC")
    day_end = day_start + pd.Timedelta(days=1)

    def _col(name):
        return _col_of(headers, name)

    charge_segs: list[dict] = []
    discharge_segs: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or _col("Leg Number") == 0:
            continue
        leg_no = row[_col("Leg Number") - 1]
        if leg_no in (None, ""):
            break
        leg_type = row[_col("Leg Type") - 1] if _col("Leg Type") else None
        if not isinstance(leg_type, str):
            continue
        lt_low = leg_type.strip().lower()

        t_s = _to_utc(row[_col("Start Time (UTC)") - 1])
        t_e = _to_utc(row[_col("End Time (UTC)") - 1])
        if t_s is None or t_e is None:
            continue
        # Day overlap
        if t_e < day_start or t_s >= day_end:
            continue

        # Strict whitelist: Stop 静默跳过；未识别 leg type 打 warning 并跳过。
        if lt_low == "stop":
            continue  # skipped — parked gap, never a trip or charge
        stripped_lt = leg_type.strip()
        is_charge = _CHARGE_LEG_PREFIX_RE.match(stripped_lt) is not None
        is_discharge = stripped_lt in _DISCHARGE_LEG_TYPES
        if not is_charge and not is_discharge:
            logger.warning(
                "reconstruct_segs_from_xlsx: unrecognised Leg Type %r on "
                "row starting %s — skipped (not discharge, not charge)",
                leg_type,
                t_s,
            )
            continue

        def _get(name):
            c = _col(name)
            if not c:
                return None
            v = row[c - 1]
            return v

        start_soc = _num_or_nan(_get("Start SOC (%)"))
        end_soc = _num_or_nan(_get("End SOC (%)"))
        d_soc = (
            end_soc - start_soc
            if (not np.isnan(start_soc) and not np.isnan(end_soc))
            else float("nan")
        )
        d_energy = _num_or_nan(_get("Energy Change (kWh)"))
        eff_cap = _num_or_nan(_get("Battery Capacity (kWh)"))
        dist = _num_or_nan(_get("Distance (km)"))
        energy_source = _get("Energy Source") or ""
        # Parse location points
        lat_s, lon_s = _parse_point(_get("Origin (Lat, Lon)"))
        lat_e, lon_e = _parse_point(_get("Destination (Lat, Lon)"))

        seg = {
            "start_time": t_s,
            "end_time": t_e,
            "start_soc": start_soc if not np.isnan(start_soc) else None,
            "end_soc": end_soc if not np.isnan(end_soc) else None,
            "delta_soc_pct": d_soc if not np.isnan(d_soc) else None,
            "delta_energy_kwh": d_energy if not np.isnan(d_energy) else None,
            "effective_capacity_kwh": eff_cap if not np.isnan(eff_cap) else None,
            "energy_source": energy_source,
        }
        if is_charge:
            # Charge segs need latitude / longitude (single point)
            seg["latitude"] = lat_s if lat_s is not None else lat_e
            seg["longitude"] = lon_s if lon_s is not None else lon_e
            seg["charge_type"] = leg_type
            charge_segs.append(seg)
        else:
            # Discharge: endpoints + odometer deltas
            seg["lat_start"] = lat_s
            seg["lon_start"] = lon_s
            seg["lat_end"] = lat_e
            seg["lon_end"] = lon_e
            if not np.isnan(dist) and dist > 0:
                # Fake odo endpoints so plot_leg_validation's EP annotation
                # works; real odo is not stored in xlsx.
                seg["odo_start_km"] = 0.0
                seg["odo_end_km"] = dist
            discharge_segs.append(seg)

    return charge_segs, discharge_segs


def _num_or_nan(v) -> float:
    if v is None or v == "":
        return float("nan")
    try:
        return float(v)
    except (TypeError, ValueError):
        return float("nan")


def _parse_point(s) -> tuple[float | None, float | None]:
    if not isinstance(s, str):
        return None, None
    m = re.match(r"Point\(([+-]?\d+\.?\d*)\s+([+-]?\d+\.?\d*)\)", s)
    if not m:
        return None, None
    try:
        return float(m.group(1)), float(m.group(2))
    except ValueError:
        return None, None


def _cumulative_relative_kwh(
    df_leg: pd.DataFrame,
    *cols: str,
) -> tuple[pd.Series, pd.DatetimeIndex] | tuple[None, None]:
    """Build a relative-kWh cumulative series from one or more Wh columns.

    Returns ``(values_kwh, times_utc)`` where ``values_kwh[i]`` is the sum of
    ``cols`` at time ``times_utc[i]`` minus the first sample (so the series
    starts at 0). Mirrors :func:`segment_algorithms._build_energy_series`
    exactly so ``_anchor_*_rel_kwh`` values land on the same curve the
    original algorithm's triangles sit on.
    """
    if not cols or TIME_COL not in df_leg.columns:
        return None, None
    missing = [c for c in cols if c not in df_leg.columns]
    if missing:
        return None, None
    sub = df_leg[[TIME_COL, *cols]].copy()
    for c in cols:
        sub[c] = pd.to_numeric(sub[c], errors="coerce")
    sub[TIME_COL] = pd.to_datetime(sub[TIME_COL], errors="coerce", utc=True)
    sub = sub.dropna(subset=[TIME_COL, *cols]).sort_values(TIME_COL)
    if len(sub) < 2:
        return None, None
    combined = sub[list(cols)].sum(axis=1)
    base = combined.iloc[0]
    values_kwh = (combined - base) / 1000.0
    times_utc = pd.DatetimeIndex(sub[TIME_COL].values, tz="UTC")
    return values_kwh.reset_index(drop=True), times_utc


def _interp_at(
    values: pd.Series,
    times: pd.DatetimeIndex,
    t: pd.Timestamp,
) -> float:
    """Linear interpolation of ``values[times]`` at timestamp ``t``.

    Clamps to first / last sample when ``t`` falls outside the range so the
    nearest-endpoint fallback still produces a sane anchor rather than NaN.
    """
    if values is None or times is None or len(values) == 0:
        return float("nan")
    if t <= times[0]:
        return float(values.iloc[0])
    if t >= times[-1]:
        return float(values.iloc[-1])
    idx_right = int(times.searchsorted(t, side="left"))
    if idx_right <= 0:
        return float(values.iloc[0])
    if idx_right >= len(values):
        return float(values.iloc[-1])
    t_l = times[idx_right - 1]
    t_r = times[idx_right]
    v_l = float(values.iloc[idx_right - 1])
    v_r = float(values.iloc[idx_right])
    span = (t_r - t_l).total_seconds()
    if span <= 0:
        return v_r
    frac = (t - t_l).total_seconds() / span
    return v_l + (v_r - v_l) * frac


def attach_anchors_from_df(
    segs: list[dict],
    df_leg: pd.DataFrame,
    *,
    kind: str,
    ac_col: str,
    dc_col: str,
    panel3_col: str,
) -> None:
    """Populate ``_anchor_*`` fields on reconstructed segs in place.

    Charge segs → anchors sit on the AC+DC cumulative kWh curve (Panel 2).
    Discharge segs → anchors sit on the ``panel3_col`` cumulative kWh curve
    (Panel 3). The time axis is taken from each seg's own ``start_time`` /
    ``end_time`` (the xlsx rows are the authoritative boundaries after
    finetune); values are linearly interpolated from the raw CSV.

    If required columns are missing or the leg CSV is empty the seg is left
    untouched (``_mark_anchors_stored`` then silently skips it).
    """
    if not segs:
        return
    if kind == "charge":
        values, times = _cumulative_relative_kwh(df_leg, ac_col, dc_col)
    elif kind == "discharge":
        values, times = _cumulative_relative_kwh(df_leg, panel3_col)
    else:
        raise ValueError(
            f"attach_anchors_from_df: kind must be "
            f'"charge" or "discharge", got {kind!r}'
        )
    if values is None or times is None:
        return
    for seg in segs:
        t_s = _to_utc(seg.get("start_time"))
        t_e = _to_utc(seg.get("end_time"))
        if t_s is None or t_e is None:
            continue
        v_s = _interp_at(values, times, t_s)
        v_e = _interp_at(values, times, t_e)
        seg["_anchor_start_time"] = t_s
        seg["_anchor_end_time"] = t_e
        seg["_anchor_start_rel_kwh"] = (
            round(v_s, 4) if not np.isnan(v_s) else float("nan")
        )
        seg["_anchor_end_rel_kwh"] = (
            round(v_e, 4) if not np.isnan(v_e) else float("nan")
        )


# =============================================================================
# Public: regenerate_figures
# =============================================================================


def regenerate_figures(
    xlsx_path: str | Path,
    raw_telematics_dir: str | Path,
    out_dir: str | Path,
    suffix: str = "_finetuned",
    *,
    original_xlsx_path: str | Path | None = None,
) -> int:
    """Regenerate validation figures from ``xlsx_path``.

    For every ``raw_<date>_<idx>.csv`` inside the xlsx's ``[start, end]``
    date range we:
      1. Reconstruct ``(charge_segs, discharge_segs)`` for that date from
         ``xlsx_path`` (= the finetuned xlsx) via
         :func:`reconstruct_segs_from_xlsx`
      2. If ``original_xlsx_path`` is provided (or auto-detected by stripping
         the ``_finetuned`` suffix from ``xlsx_path``), also reconstruct the
         **original** segs and compare per date:
           - Segs identical (same start/end + leg_type tuple)
             → **skip** entirely — no ``_finetuned.png`` is produced. The
             inspect HTML generator falls back to the original
             ``validation_<...>.png`` for that day.
           - Segs differ
             → draw with base = original (red/green), overlay = finetuned
             (orange/cyan) so visual comparison is obvious
      3. If no original xlsx is available (legacy behaviour), draw with only
         the finetuned segs and no overlay.

    Returns the number of ``_finetuned.png`` files produced (overlay + plain
    draws only; skipped-unchanged days contribute zero).
    """
    from jolt_toolkit.report_generator.segment_algorithms import (
        AC_COL,
        DC_COL,
        MOVING_COL,
        TOTAL_ENERGY_COL,
        cluster_mass_data,
        plot_leg_validation,
    )

    xlsx_path = Path(xlsx_path)
    raw_dir = Path(raw_telematics_dir)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    reg = _reg_from_name(xlsx_path.name)
    cfg = _veh_cfg(reg)
    ac_col = cfg.get("ac_col", AC_COL)
    dc_col = cfg.get("dc_col", DC_COL)
    total_col = cfg.get("total_energy_col", TOTAL_ENERGY_COL)
    moving_col = cfg.get("moving_energy_col", MOVING_COL)
    mass_col = cfg.get("mass_col", MASS_COL)
    speed_col = cfg.get("speed_col", "wheel_based_speed")
    mass_agg = resolve_mass_agg(reg)  # v2.2.6: 图上质量与报告同口径稳健聚合

    # Derive [period_start, period_end] from xlsx name so we skip raw CSVs
    # outside the report's date range (a vehicle's raw_telematics/ folder is
    # shared across all periods).
    m = re.search(r"(\d{8})_(\d{8})", xlsx_path.name)
    if m:
        period_start = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}"
        period_end = f"{m.group(2)[:4]}-{m.group(2)[4:6]}-{m.group(2)[6:]}"
    else:
        period_start = "0000-00-00"
        period_end = "9999-99-99"

    # Resolve original xlsx path. If caller didn't pass one, try stripping the
    # _finetuned suffix from xlsx_path.stem to find the source xlsx.
    if original_xlsx_path is None:
        stem = xlsx_path.stem
        if stem.endswith("_finetuned"):
            candidate = xlsx_path.with_name(
                stem[: -len("_finetuned")] + xlsx_path.suffix
            )
            if candidate.exists():
                original_xlsx_path = candidate
    else:
        original_xlsx_path = Path(original_xlsx_path)
        if not original_xlsx_path.exists():
            logger.warning(
                "regenerate_figures: original_xlsx_path=%s not found; "
                "falling back to no-overlay rendering",
                original_xlsx_path,
            )
            original_xlsx_path = None

    # Best-effort Logger fetch (optional)
    logger_speed_all, logger_mass_all = _try_fetch_logger(reg, xlsx_path.name)

    # Pre-scan segs for the entire period once (xlsx reload is expensive).
    # - segs_by_date_ft  : from the finetuned xlsx
    # - segs_by_date_orig: from the original xlsx (optional)
    segs_by_date_ft: dict[str, tuple[list[dict], list[dict]]] = {}
    segs_by_date_orig: dict[str, tuple[list[dict], list[dict]]] = {}

    raw_csvs = sorted(raw_dir.glob("raw_*.csv"))
    n_out = 0
    n_skipped = 0
    n_overlay = 0
    n_plain = 0
    for csv_path in raw_csvs:
        m2 = re.match(r"raw_(\d{4}-\d{2}-\d{2})_(\d+)\.csv$", csv_path.name)
        if not m2:
            continue
        date_str, idx_str = m2.group(1), m2.group(2)
        if not (period_start <= date_str <= period_end):
            continue
        leg_suffix = f"{date_str}_{idx_str}"

        df_leg = pd.read_csv(csv_path, dtype=str, low_memory=False)
        if df_leg.empty or SOC_COL not in df_leg.columns:
            continue
        # 在 plot_leg_validation 之前跑 mass 聚类，生成 ``mass_cluster`` 列。
        # Panel 4 的 seg-mean mass 标注（含 overlay ``[FT] XX.X t``）依赖该列
        # 做去噪聚合；若缺失则 Panel 4 会退化为只画原始散点（见 v2.2.2
        # 限制），finetune 图与 production 图因此不一致。
        if mass_col in df_leg.columns:
            try:
                # v2.2.4: 传 speed_col 让 finetune 图的 mass 聚类与 production 一致
                # （聚类均值只用行驶中读数；speed 列缺失时自动回退全部有效读数）。
                df_leg = cluster_mass_data(
                    df_leg, mass_col=mass_col, speed_col=speed_col
                )
            except Exception as exc:
                logger.debug(
                    "regenerate_figures: cluster_mass_data failed "
                    "on %s (%s) — Panel 4 mass labels skipped",
                    csv_path.name,
                    exc,
                )

        if date_str not in segs_by_date_ft:
            segs_by_date_ft[date_str] = reconstruct_segs_from_xlsx(xlsx_path, date_str)
        ft_c_segs, ft_d_segs = segs_by_date_ft[date_str]

        orig_c_segs: list[dict] = []
        orig_d_segs: list[dict] = []
        if original_xlsx_path is not None:
            if date_str not in segs_by_date_orig:
                segs_by_date_orig[date_str] = reconstruct_segs_from_xlsx(
                    original_xlsx_path, date_str
                )
            orig_c_segs, orig_d_segs = segs_by_date_orig[date_str]

        # Keep only segs that overlap this leg's time window
        t_s, t_e = _leg_time_bounds(df_leg)
        ft_c_segs = [s for s in ft_c_segs if _overlap(s, t_s, t_e)]
        ft_d_segs = [s for s in ft_d_segs if _overlap(s, t_s, t_e)]
        orig_c_segs = [s for s in orig_c_segs if _overlap(s, t_s, t_e)]
        orig_d_segs = [s for s in orig_d_segs if _overlap(s, t_s, t_e)]

        out_path = out_dir / f"validation_{reg}_{leg_suffix}{suffix}.png"

        # ── Path 1: no original → fall back to plain finetuned draw ────────
        if original_xlsx_path is None:
            panel3_col = moving_col
            if ft_d_segs and ft_d_segs[0].get("energy_source") == "total_energy":
                panel3_col = total_col
            leg_logger_spd = _slice_logger(logger_speed_all, t_s, t_e)
            leg_logger_mass = _slice_logger(logger_mass_all, t_s, t_e)
            # 从 df_leg 的累积能量列线性插值出 ▼▲ 锚点位置，使 plain 图的
            # Panel 2 / 3 三角形与原 production 图风格一致。
            attach_anchors_from_df(
                ft_c_segs,
                df_leg,
                kind="charge",
                ac_col=ac_col,
                dc_col=dc_col,
                panel3_col=panel3_col,
            )
            attach_anchors_from_df(
                ft_d_segs,
                df_leg,
                kind="discharge",
                ac_col=ac_col,
                dc_col=dc_col,
                panel3_col=panel3_col,
            )
            try:
                plot_leg_validation(
                    df_leg,
                    ft_c_segs,
                    ft_d_segs,
                    reg,
                    leg_suffix,
                    out_path,
                    ac_col=ac_col,
                    dc_col=dc_col,
                    panel3_col=panel3_col,
                    mass_col=mass_col,
                    speed_col=speed_col,
                    logger_speed_df=leg_logger_spd,
                    logger_mass_df=leg_logger_mass,
                    mass_agg=mass_agg,
                )
                n_out += 1
                n_plain += 1
            except Exception as exc:
                logger.warning("regenerate_figures: skip %s (%s)", leg_suffix, exc)
            continue

        # ── Path 2: segs unchanged → skip entirely ───────────────────────
        # No ``_finetuned.png`` is produced for unchanged days; the inspect
        # HTML falls back to the original ``validation_<...>.png`` so the
        # viewer still shows every day. Stale ``_finetuned.png`` from a
        # previous run is removed so the HTML falls back correctly.
        if _segs_equal(ft_c_segs, orig_c_segs) and _segs_equal(ft_d_segs, orig_d_segs):
            if out_path.exists():
                try:
                    out_path.unlink()
                except Exception as exc:
                    logger.warning(
                        "regenerate_figures: could not remove stale %s " "(%s)",
                        out_path.name,
                        exc,
                    )
            n_skipped += 1
            continue

        # ── Path 3: segs changed → draw base=original + overlay=finetuned ─
        panel3_col = moving_col
        # Pick panel3 col based on the original (base) segs for consistency
        # with how the original PNG was drawn.
        if orig_d_segs and orig_d_segs[0].get("energy_source") == "total_energy":
            panel3_col = total_col
        elif ft_d_segs and ft_d_segs[0].get("energy_source") == "total_energy":
            panel3_col = total_col
        leg_logger_spd = _slice_logger(logger_speed_all, t_s, t_e)
        leg_logger_mass = _slice_logger(logger_mass_all, t_s, t_e)
        # 为 base（原 segs）和 overlay（finetuned segs）都补齐 ▼▲ 锚点，
        # 让 Panel 2 / 3 能同时画出红绿 ▼▲（original）+ 橙青 ▼▲（[FT]）。
        attach_anchors_from_df(
            orig_c_segs,
            df_leg,
            kind="charge",
            ac_col=ac_col,
            dc_col=dc_col,
            panel3_col=panel3_col,
        )
        attach_anchors_from_df(
            orig_d_segs,
            df_leg,
            kind="discharge",
            ac_col=ac_col,
            dc_col=dc_col,
            panel3_col=panel3_col,
        )
        attach_anchors_from_df(
            ft_c_segs,
            df_leg,
            kind="charge",
            ac_col=ac_col,
            dc_col=dc_col,
            panel3_col=panel3_col,
        )
        attach_anchors_from_df(
            ft_d_segs,
            df_leg,
            kind="discharge",
            ac_col=ac_col,
            dc_col=dc_col,
            panel3_col=panel3_col,
        )
        try:
            plot_leg_validation(
                df_leg,
                orig_c_segs,
                orig_d_segs,
                reg,
                leg_suffix,
                out_path,
                ac_col=ac_col,
                dc_col=dc_col,
                panel3_col=panel3_col,
                mass_col=mass_col,
                speed_col=speed_col,
                logger_speed_df=leg_logger_spd,
                logger_mass_df=leg_logger_mass,
                overlay_charge_segs=ft_c_segs,
                overlay_discharge_segs=ft_d_segs,
                mass_agg=mass_agg,
            )
            n_out += 1
            n_overlay += 1
        except Exception as exc:
            logger.warning("regenerate_figures: skip %s (%s)", leg_suffix, exc)

    logger.info(
        "regenerate_figures: skipped %d, overlaid %d (plain=%d) — "
        "wrote %d PNG(s) to %s",
        n_skipped,
        n_overlay,
        n_plain,
        n_out,
        out_dir,
    )
    return n_out


def _segs_equal(a: list[dict], b: list[dict]) -> bool:
    """Compare two seg lists by (start_time, end_time) tuples.

    Used by :func:`regenerate_figures` to decide whether a date's finetuned
    segs match the original segs byte-for-byte (so we can skip redrawing and
    just copy the original PNG). We intentionally ignore numeric fields
    (kwh, soc_estimate floating-point noise) and compare only the temporal
    boundary tuples, which fully determine segment identity.
    """
    if len(a) != len(b):
        return False
    ka = sorted((_to_utc(s["start_time"]), _to_utc(s["end_time"])) for s in a)
    kb = sorted((_to_utc(s["start_time"]), _to_utc(s["end_time"])) for s in b)
    return ka == kb


def _leg_time_bounds(df: pd.DataFrame):
    times = pd.to_datetime(df[TIME_COL], errors="coerce", utc=True).dropna()
    if times.empty:
        return None, None
    return times.min(), times.max()


def _overlap(seg: dict, t_s, t_e) -> bool:
    if t_s is None or t_e is None:
        return True
    try:
        s = _to_utc(seg["start_time"])
        e = _to_utc(seg["end_time"])
    except KeyError:
        return True
    return s <= t_e and e >= t_s


def _slice_logger(df: pd.DataFrame | None, t_s, t_e):
    if df is None or df.empty or t_s is None or t_e is None:
        return None
    try:
        sliced = df.loc[t_s:t_e]
        return sliced if not sliced.empty else None
    except Exception:
        return None


def _try_fetch_logger(reg: str, xlsx_name: str):
    """Best-effort SRF fetch for Logger speed / mass.

    Returns ``(speed_df, mass_df)`` or ``(None, None)`` on any failure. Do not
    block finetune on missing Logger data. Skips silently if
    ``SRF_API_KEY`` is not set in the environment.
    """
    import os as _os

    if not _os.environ.get("SRF_API_KEY"):
        return None, None
    try:
        from jolt_toolkit.report_generator.validation_generator import (
            ValidationGenerator,
        )

        cfg = _veh_cfg(reg)
        reg_srf = cfg.get("srf_reg")
        if not reg_srf:
            return None, None
        m = re.search(r"(\d{8})_(\d{8})", xlsx_name)
        if not m:
            return None, None
        ds = datetime.strptime(m.group(1), "%Y%m%d")
        de = datetime.strptime(m.group(2), "%Y%m%d")
        gen = ValidationGenerator()
        return gen._fetch_logger_data(reg_srf, ds, de)
    except Exception as exc:
        logger.info("regenerate_figures: Logger fetch skipped (%s)", exc)
        return None, None


# =============================================================================
# Public: regenerate_inspect_html
# =============================================================================


def regenerate_inspect_html(
    xlsx_path: str | Path,
    out_path: str | Path | None = None,
    fig_suffix: str = "_finetuned",
) -> Path:
    """Generate an inspect HTML pointing at ``validation_*{fig_suffix}.png``.

    If ``out_path`` is None, writes
    ``inspect_<basename>_finetuned.html`` next to the xlsx.

    This is a lightweight variant of
    :func:`report_builder._write_html_viewer` — we can't reuse the original
    directly because it globs ``validation_{REG}_*.png`` unconditionally and
    would pick up the non-finetuned figures too.
    """
    xlsx_path = Path(xlsx_path)
    reg = _reg_from_name(xlsx_path.name)
    m = re.search(r"(\d{8})_(\d{8})", xlsx_path.name)
    if not m:
        raise ValueError(f"Cannot parse date range from {xlsx_path.name}")
    ds_str, de_str = m.group(1), m.group(2)
    period_start = f"{ds_str[:4]}-{ds_str[4:6]}-{ds_str[6:]}"
    period_end = f"{de_str[:4]}-{de_str[4:6]}-{de_str[6:]}"

    out_dir = xlsx_path.parent
    fig_dir = out_dir / "validation_figures"
    if not fig_dir.exists():
        raise FileNotFoundError(f"{fig_dir} does not exist")

    # Enumerate every original ``validation_<reg>_<date>_<idx>.png`` in the
    # period. For each one, if a ``*_finetuned.png`` sibling exists → the
    # day was modified by apply_operations / regenerate_figures; else →
    # unchanged, fall back to the original. This lets the HTML cover every
    # day of the period even when most days have no finetuned output.
    all_originals = sorted(
        p
        for p in fig_dir.glob(f"validation_{reg}_*.png")
        if not p.stem.endswith(fig_suffix)
    )
    if not all_originals:
        raise FileNotFoundError(f"No validation_{reg}_*.png figures in {fig_dir}")

    date_re = re.compile(rf"validation_{re.escape(reg)}_(\d{{4}}-\d{{2}}-\d{{2}})_")
    figs: list[Path] = []
    is_modified: list[bool] = []
    for p in all_originals:
        dm = date_re.match(p.name)
        if not (dm and period_start <= dm.group(1) <= period_end):
            continue
        ft_png = p.with_name(p.stem + fig_suffix + p.suffix)
        if ft_png.exists():
            figs.append(ft_png)
            is_modified.append(True)
        else:
            figs.append(p)
            is_modified.append(False)
    if not figs:
        raise FileNotFoundError(
            f"No figures in date range {period_start}..{period_end}"
        )

    rel = [f"validation_figures/{p.name}" for p in figs]
    # Append a lightweight tag so users can see at a glance which days were
    # touched by the finetune pass and which fell back to the original PNG.
    labels = [
        p.stem + (" [modified]" if mod else " (unchanged — original)")
        for p, mod in zip(figs, is_modified)
    ]

    if out_path is None:
        # Avoid doubling the suffix when xlsx stem already ends with it
        # (e.g. jolt_report_X_..._finetuned.xlsx → inspect_..._finetuned.html,
        # NOT ..._finetuned_finetuned.html).
        stem = xlsx_path.stem
        if not stem.endswith(fig_suffix):
            stem += fig_suffix
        out_path = out_dir / ("inspect_" + stem + ".html")
    else:
        out_path = Path(out_path)

    imgs_js = "[\n" + ",\n".join(f'        "{r}"' for r in rel) + "\n    ]"
    labels_js = "[\n" + ",\n".join(f'        "{l}"' for l in labels) + "\n    ]"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>JOLT Inspection (finetuned) — {reg} {period_start} – {period_end}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ display: flex; height: 100vh; font-family: Arial, sans-serif; }}
  #sidebar {{ width: 260px; min-width: 160px; background: #1e1e2e; color: #cdd6f4;
              overflow-y: auto; flex-shrink: 0; padding: 8px 0; }}
  #sidebar h2 {{ font-size: 13px; padding: 8px 12px 4px; color: #89b4fa;
                 border-bottom: 1px solid #313244; margin-bottom: 4px; }}
  #sidebar ul {{ list-style: none; }}
  #sidebar li {{ padding: 5px 12px; cursor: pointer; font-size: 12px;
                  border-left: 3px solid transparent; }}
  #sidebar li:hover {{ background: #313244; }}
  #sidebar li.active {{ background: #313244; border-left-color: #f9e2af;
                         color: #f9e2af; }}
  #main {{ flex: 1; display: flex; flex-direction: column;
            background: #11111b; overflow: hidden; }}
  #toolbar {{ display: flex; align-items: center; gap: 10px;
               padding: 8px 16px; background: #181825; border-bottom: 1px solid #313244; }}
  #toolbar button {{ padding: 4px 14px; background: #313244; color: #cdd6f4;
                      border: 1px solid #45475a; border-radius: 4px;
                      cursor: pointer; font-size: 13px; }}
  #toolbar button:hover {{ background: #45475a; }}
  #counter {{ color: #a6adc8; font-size: 13px; }}
  #label {{ color: #f9e2af; font-size: 12px; font-family: monospace;
             flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  #img-wrap {{ flex: 1; display: flex; align-items: center;
               justify-content: center; overflow: auto; padding: 12px; }}
  #fig {{ max-width: 100%; max-height: 100%; object-fit: contain; }}
  #badge {{ background: #f9e2af; color: #1e1e2e; padding: 2px 8px;
            border-radius: 3px; font-size: 11px; font-weight: bold; }}
  /* Row tags: subtle grey 'unchanged' vs amber 'modified', small size */
  .tag {{ font-size: 10px; margin-left: 6px; padding: 1px 4px;
          border-radius: 2px; font-family: sans-serif; }}
  .tag-mod {{ background: #f9e2af; color: #1e1e2e; font-weight: bold; }}
  .tag-unchanged {{ background: transparent; color: #7f849c;
                    font-style: italic; }}
</style>
</head>
<body>
<div id="sidebar">
  <h2>{reg} &nbsp;{period_start} – {period_end}<br><span id="badge">FINETUNED</span></h2>
  <ul id="list"></ul>
</div>
<div id="main">
  <div id="toolbar">
    <button onclick="go(-1)">&#9664; Prev</button>
    <button onclick="go(1)">Next &#9654;</button>
    <span id="counter"></span>
    <span id="label"></span>
  </div>
  <div id="img-wrap"><img id="fig" src="" alt="validation figure"></div>
</div>
<script>
const imgs   = {imgs_js};
const labels = {labels_js};
let cur = 0;
const listEl    = document.getElementById("list");
const figEl     = document.getElementById("fig");
const counterEl = document.getElementById("counter");
const labelEl   = document.getElementById("label");
labels.forEach((lb, i) => {{
  const li = document.createElement("li");
  // Strip the ``validation_<REG>_`` prefix so the sidebar shows
  // ``<date>_<idx>[_finetuned] [modified]`` (or the unchanged variant).
  // Split off the bracketed tag and render it as a coloured pill.
  const short = lb.replace(/^validation_[^_]+_/, "");
  const m = short.match(/^(.*?)\s+(\[modified\]|\(unchanged — original\))$/);
  if (m) {{
    const base = document.createTextNode(m[1]);
    li.appendChild(base);
    const tag = document.createElement("span");
    tag.textContent = m[2];
    tag.className = "tag " + (m[2] === "[modified]" ? "tag-mod" : "tag-unchanged");
    li.appendChild(tag);
  }} else {{
    li.textContent = short;
  }}
  li.onclick = () => show(i);
  listEl.appendChild(li);
}});
function show(i) {{
  cur = i;
  figEl.src = imgs[i];
  counterEl.textContent = (i + 1) + " / " + imgs.length;
  labelEl.textContent   = labels[i];
  document.querySelectorAll("#list li").forEach((el, j) =>
    el.classList.toggle("active", j === i));
  document.querySelectorAll("#list li")[i]
    .scrollIntoView({{block: "nearest"}});
}}
function go(d) {{ show((cur + d + imgs.length) % imgs.length); }}
document.addEventListener("keydown", e => {{
  if (e.key === "ArrowLeft"  || e.key === "ArrowUp")   go(-1);
  if (e.key === "ArrowRight" || e.key === "ArrowDown")  go(1);
}});
show(0);
</script>
</body>
</html>
"""
    out_path.write_text(html, encoding="utf-8")
    logger.info(
        "finetune: wrote inspect HTML %s (%d figures)", out_path.name, len(figs)
    )
    return out_path


# =============================================================================
# Smoke test
# =============================================================================

if __name__ == "__main__":
    import sys
    import tempfile

    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s"
    )

    project_root = Path(__file__).resolve().parents[3]
    xlsx = (
        project_root
        / "excel_report_database"
        / "2.2.2"
        / "YK73WFN"
        / "jolt_report_YK73WFN_20240601_20240901.xlsx"
    )
    raw_dir = xlsx.parent / "raw_telematics"
    fig_dir = xlsx.parent / "validation_figures"

    if not xlsx.exists():
        print(f"[smoke] xlsx not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    print("[smoke] 1. apply_operations with [] (no-op)", flush=True)
    out = apply_operations(xlsx, [], raw_dir)
    print(f"[smoke]    -> {out}", flush=True)
    assert out.exists(), "finetuned xlsx not created"
    assert out.name.endswith("_finetuned.xlsx")

    print("[smoke] 2. reconstruct_segs_from_xlsx for 2024-06-11", flush=True)
    c_segs, d_segs = reconstruct_segs_from_xlsx(out, "2024-06-11")
    print(
        f"[smoke]    -> {len(c_segs)} charge segs, " f"{len(d_segs)} discharge segs",
        flush=True,
    )

    print("[smoke] 3. regenerate_figures — single-day test for 2024-06-11", flush=True)
    # Use a temp dir so we don't pollute the real validation_figures/ during
    # smoke tests — then move a sample into validation_figures/ for visual
    # inspection.
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        # Copy only the 2024-06-11 csv and let regenerate_figures scan it
        single_csv_dir = tmp_path / "raw_one"
        single_csv_dir.mkdir()
        import shutil as _sh

        src_csv = raw_dir / "raw_2024-06-11_0000.csv"
        if src_csv.exists():
            _sh.copy2(src_csv, single_csv_dir / src_csv.name)
        n = regenerate_figures(out, single_csv_dir, fig_dir, suffix="_finetuned")
        print(f"[smoke]    -> {n} PNG(s) produced", flush=True)

    print("[smoke] 4. regenerate_inspect_html", flush=True)
    try:
        html_out = regenerate_inspect_html(out)
        print(f"[smoke]    -> {html_out}", flush=True)
    except FileNotFoundError as exc:
        print(f"[smoke]    skipped: {exc}", flush=True)

    print("[smoke] DONE", flush=True)
