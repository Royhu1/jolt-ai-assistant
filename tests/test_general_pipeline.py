"""Contract tests for the v3.1.0 general fallback pipeline.

The general fallback pipeline (``report_generator.general_pipeline``) lets any
registration NOT in vehicles.json still produce a structurally valid report.
These tests cover the pure, offline surface (registration-spacing resolution,
EV column auto-detection, runtime-config assembly), the capacity post-processing
tolerating an all-fallback report, and the two guarantees wired into the
generator (no vehicles.json write-back for a runtime config; a structurally
complete empty report is always writable). No network / no SRF key.
"""

from __future__ import annotations

import datetime
from time import perf_counter
from unittest.mock import Mock

import pandas as pd
import pytest

from jolt_toolkit.report_generator import general_pipeline as gp
from jolt_toolkit.report_generator._generator import JOLTReportGenerator
from jolt_toolkit.report_generator.capacity import (
    _IDX_BPOWER,
    _IDX_CAP,
    _IDX_DISTANCE,
    _IDX_DURATION,
    _IDX_ELEV,
    _IDX_ENERGY,
    _IDX_EPERF,
    _IDX_EPERF_CORR,
    _IDX_EPERF_KIN,
    _IDX_ESOURCE,
    _IDX_MASS,
    _IDX_SOC_CHANGE,
    _IDX_START,
    _correct_effective_capacity,
)
from jolt_toolkit.report_generator.columns import HEADERS
from jolt_toolkit.report_generator.report_builder import (
    DIESEL_HEADERS,
    _write_excel_report,
)

# ── Registration-spacing resolution (pure) ───────────────────────────────────


@pytest.mark.parametrize(
    "reg_input, must_contain",
    [
        ("YK73WFN", "YK73 WFN"),  # current UK format LLNN LLL
        ("AB12CDE", "AB12 CDE"),
        ("CMZ6260", "CMZ 6260"),  # NI-style 3-alpha + 4-digit
        ("WU70GLV", "WU70 GLV"),  # diesel, spaced form
        ("T88RNW", "T88 RNW"),  # 6-char form
    ],
)
def test_reg_spacing_variants_include_srf_spelling(reg_input, must_contain):
    variants = gp.reg_spacing_variants(reg_input)
    # Verbatim no-space form is always first (fast path when SRF accepts it).
    assert variants[0] == reg_input.replace(" ", "").upper()
    assert must_contain in variants


def test_reg_spacing_variants_dedup_and_case_insensitive():
    # A spaced, lower-case input still resolves the same ordered candidate set;
    # duplicates are removed.
    assert gp.reg_spacing_variants("yk73 wfn") == gp.reg_spacing_variants("YK73WFN")


# ── EV column auto-detection (pure) ──────────────────────────────────────────

_FULL_EV_COLUMNS = [
    "electricBatteryLevelPercent",
    "wheel_based_speed",
    "battery_pack_ac_watthours",
    "battery_pack_dc_watthours",
    "total_electric_energy_used_plugged_in_included",
    "electric_energy_wheelbased_speed_over_zero",
    "gross_combination_vehicle_weight",
    "gnss_altitude",
]


def test_detect_full_ev_columns():
    det = gp.detect_ev_columns(_FULL_EV_COLUMNS)
    assert det["has_soc"] is True
    assert det["speed_col"] == "wheel_based_speed"
    assert det["ac_col"] == "battery_pack_ac_watthours"
    assert det["dc_col"] == "battery_pack_dc_watthours"
    assert det["total_energy_col"] == "total_electric_energy_used_plugged_in_included"
    assert det["moving_energy_col"] == "electric_energy_wheelbased_speed_over_zero"
    assert det["mass_col"] == "gross_combination_vehicle_weight"
    assert det["altitude_col"] == "gnss_altitude"


def test_detect_counterless_ev_columns():
    # Mercedes/DAF style: 'speed' + 'total_electric_energy_used', no AC/DC/moving.
    cols = [
        "electricBatteryLevelPercent",
        "speed",
        "total_electric_energy_used",
        "gross_combination_vehicle_weight",
        "gnss_altitude",
    ]
    det = gp.detect_ev_columns(cols)
    assert det["has_soc"] is True
    assert det["speed_col"] == "speed"
    assert det["total_energy_col"] == "total_electric_energy_used"
    # Absent counters degrade to omission → run_segment_detection uses defaults /
    # the feature blanks.
    assert "ac_col" not in det
    assert "dc_col" not in det
    assert "moving_energy_col" not in det


def test_detect_socless_speed_only_columns():
    det = gp.detect_ev_columns(["wheel_based_speed", "odometer"])
    assert det["has_soc"] is False
    assert det["speed_col"] == "wheel_based_speed"


def test_detect_empty_columns():
    det = gp.detect_ev_columns([])
    assert det["has_soc"] is False
    # No field resolved beyond the has_soc flag.
    assert set(det) == {"has_soc"}


# ── Runtime-config assembly (pure) ───────────────────────────────────────────


def test_build_runtime_config_ev_full():
    cfg = gp.build_runtime_config(
        "XX99XXX",
        "XX99 XXX",
        fuel_type="EV",
        make="Volvo",
        model="FM Electric",
        srf_capacity_kwh=540.0,
        columns=_FULL_EV_COLUMNS,
    )
    assert cfg["srf_reg"] == "XX99 XXX"
    assert cfg["fuel_type"] == "EV"
    assert cfg["pipeline"] == "default_soc"  # SOC present → soc branch
    assert cfg["srf_capacity_kwh"] == 540.0
    assert cfg["nominal_kwh"] is None
    assert cfg["speed_col"] == "wheel_based_speed"
    assert cfg["total_energy_col"] == "total_electric_energy_used_plugged_in_included"
    assert gp.is_runtime_config(cfg) is True


def test_build_runtime_config_ev_speed_branch_when_no_soc():
    cfg = gp.build_runtime_config(
        "XX99XXX",
        "XX99 XXX",
        fuel_type="EV",
        columns=["wheel_based_speed", "gross_combination_vehicle_weight"],
    )
    assert cfg["pipeline"] == "default_speed"  # no SOC → speed branch
    assert cfg["srf_capacity_kwh"] is None


def test_build_runtime_config_diesel():
    cfg = gp.build_runtime_config(
        "WW70WWW",
        "WW70 WWW",
        fuel_type="DIESEL",
        make="DAF",
        model="XF 450",
        weight_class_t=44.0,
    )
    assert cfg["fuel_type"] == "DIESEL"
    assert cfg["pipeline"] == "daf_diesel_logger"
    assert cfg["leg_source"] == "SRFLOGGER_V1"
    assert cfg["weight_class_t"] == 44.0
    assert cfg["fuel_energy_col"] == "LFC engine total fuel used"
    assert gp.is_runtime_config(cfg) is True


def test_is_runtime_config_false_for_plain_dict():
    assert gp.is_runtime_config({"srf_reg": "AB12 CDE"}) is False
    assert gp.is_runtime_config(None) is False
    assert gp.is_runtime_config({}) is False


# ── Capacity post-processing tolerates an all-fallback report ────────────────


def _blank_soc_estimate_row():
    row = [float("nan")] * (len(HEADERS) - 1)
    row[_IDX_ESOURCE] = "soc_estimate"
    row[_IDX_SOC_CHANGE] = -12.0
    row[_IDX_DISTANCE] = 8.0
    row[_IDX_DURATION] = 0.05
    return row


def test_correct_effective_capacity_all_fallback_no_capacity_no_crash():
    """An un-onboarded EV with only soc_estimate legs and NO capacity source
    (fallback_kwh=None) must not crash on round(None); it returns cap=None and
    leaves the soc_estimate rows' capacity/energy as NaN."""
    r1 = _blank_soc_estimate_row()
    r2 = _blank_soc_estimate_row()
    r1[_IDX_START] = pd.Timestamp("2025-04-01T08:00:00Z")
    r2[_IDX_START] = pd.Timestamp("2025-04-02T08:00:00Z")
    rows = [r1, r2]

    out_rows, eff_cap, cap_source = _correct_effective_capacity(
        rows,
        _IDX_CAP,
        _IDX_ENERGY,
        _IDX_SOC_CHANGE,
        _IDX_EPERF,
        _IDX_DISTANCE,
        _IDX_ESOURCE,
        _IDX_BPOWER,
        _IDX_DURATION,
        _IDX_EPERF_CORR,
        _IDX_ELEV,
        _IDX_MASS,
        None,  # fallback_kwh — no srf/nominal capacity for the un-onboarded reg
        _IDX_EPERF_KIN,
        idx_start=_IDX_START,
        soc_fallback=None,
    )
    assert eff_cap is None
    assert cap_source == "fallback"
    # Capacity could not be attributed → stays NaN (not 0, not a crash).
    for row in out_rows:
        assert pd.isna(row[_IDX_CAP])


# ── Generator wiring: no persist for runtime cfg; empty report is writable ────


def _make_generator_without_srf():
    """Build a JOLTReportGenerator instance WITHOUT constructing an SRF client
    (its __init__ would need SRF_API_KEY)."""
    gen = JOLTReportGenerator.__new__(JOLTReportGenerator)
    gen.report_output_folder = "."
    gen.overwrite_existing_report = True
    gen.debug_mode = False
    gen.fast_mode = True  # skip the charger/logger patchers (need srf_data)
    gen.save_figures = False
    gen.srf_data = None
    return gen


def _write_outputs_args(out_dir, is_runtime_cfg):
    return dict(
        sorted_rows=[],  # empty → header-only report, exercises the empty path
        out_headers=HEADERS,
        reg="ZZ99ZZZ",
        ds=datetime.datetime(2025, 4, 1),
        de=datetime.datetime(2025, 4, 15),
        out_dir=out_dir,
        is_diesel=False,
        period_cap_kwh=None,
        period_n=0,
        period_src="fallback",
        period_key="20250401_20250415",
        charger_windows=[],
        logger_legs=[],
        logger_windows=[],
        time_start=perf_counter(),
        is_runtime_cfg=is_runtime_cfg,
    )


def test_write_outputs_skips_persist_for_runtime_config(tmp_path):
    gen = _make_generator_without_srf()
    persist = Mock()
    gen._persist_effective_capacity = persist  # shadow the class staticmethod

    path = gen._write_outputs(**_write_outputs_args(tmp_path, is_runtime_cfg=True))

    persist.assert_not_called()  # never invent a vehicles.json capacity entry
    from pathlib import Path

    assert Path(path).exists()


def test_write_outputs_persists_for_onboarded_config(tmp_path):
    # Control: with is_runtime_cfg=False the EV path DOES call the write-back.
    gen = _make_generator_without_srf()
    persist = Mock()
    gen._persist_effective_capacity = persist

    gen._write_outputs(**_write_outputs_args(tmp_path, is_runtime_cfg=False))

    persist.assert_called_once()


@pytest.mark.parametrize("headers", [HEADERS, DIESEL_HEADERS])
def test_empty_report_is_structurally_complete(tmp_path, headers):
    """The ultimate-degradation guarantee: a report with zero data rows still
    writes a structurally complete workbook (all sheets + the header row)."""
    import openpyxl

    out_path = tmp_path / "jolt_report_ZZ99ZZZ_20250401_20250415.xlsx"
    _write_excel_report(
        [],
        "ZZ99ZZZ",
        datetime.date(2025, 4, 1),
        datetime.date(2025, 4, 15),
        out_path,
        headers=headers,
    )
    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path, data_only=False)
    assert {"Report", "Graphs", "GraphsData", "Definitions"} <= set(wb.sheetnames)
    ws = wb["Report"]
    # Header row present, one data-less body.
    assert ws.cell(row=1, column=1).value == "Leg Number"
    assert ws.cell(row=1, column=2).value == headers[1]
    assert ws.max_row == 1  # header only, no data rows
