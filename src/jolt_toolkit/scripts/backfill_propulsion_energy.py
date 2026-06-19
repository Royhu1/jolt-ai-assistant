"""
backfill_propulsion_energy.py
==============================
一次性回填脚本：往已有的 ``excel_report_database/<version>/<REG>/jolt_report_*.xlsx`` 末尾
追加（或覆盖）一列 ``Propulsion Energy (kWh)``，**不重跑** SRF API / 分段算法。

数据来源：同一份报告目录下的 ``raw_telematics/raw_*.csv`` 中的
``electric_energy_propulsion`` 列（Wh，累计计数器）。算法与 v2.2.3 主管线一致 ——
对每个 trip 的 ``[Start Time, End Time]`` 时间窗，在最近的 RFMS 快照之间做线性
插值得到两端的累计 propulsion，差分后 / 1000 转 kWh。详见
:func:`jolt_toolkit.report_generator.report_builder._get_propulsion_energy`。

用法
----
::

    # 单车
    python -m jolt_toolkit.scripts.backfill_propulsion_energy --version 2.2.2 --veh YK73WFN

    # 整个版本目录下的全部 EV
    python -m jolt_toolkit.scripts.backfill_propulsion_energy --version 2.2.2

行为约定
--------
- 用 ``openpyxl`` 打开 xlsx，不重写 worksheet —— 现有格式 / 颜色 / 公式全部保留。
- 列已存在时（脚本被重复跑）直接覆盖该列，不重复追加。
- 仅对 trip / In Transit / Outbound / Return / In House / Round Trip 行计算；
  Stop / 充电行强制写 NaN（``=NA()``）。
- 找不到 raw_telematics CSV 或缺少 ``electric_energy_propulsion`` 列的报告
  会被跳过并 warn —— 不会损坏 xlsx。
- 仅处理 ``jolt_report_<REG>_<ds>_<de>.xlsx``，跳过 ``*_finetuned.xlsx``
  （finetune 链路自己负责保持一致）。
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

from jolt_toolkit.report_generator.report_builder import (
    HEADERS,
    _get_propulsion_energy,
    _PROPULSION_COL,
)

logger = logging.getLogger(__name__)

# Trip-类 leg type 的判定：用与 _generator 兜底 pass 一致的语义
_CHARGE_PREFIX_RE = re.compile(r'^(ac|dc|charge|mix|estimated)', re.IGNORECASE)


def _is_trip_row(leg_type) -> bool:
    """放电 / 行程行返回 True；Stop / charge / 缺失返回 False。"""
    if leg_type is None:
        return False
    if not isinstance(leg_type, str):
        return False
    s = leg_type.strip().lower()
    if not s:
        return False
    if s == 'stop':
        return False
    if _CHARGE_PREFIX_RE.match(s):
        return False
    return True


def _load_telematics_df(raw_dir: Path) -> pd.DataFrame | None:
    """拼接 raw_telematics 目录下所有 raw_*.csv。

    返回单一 DataFrame（dtype=str；交给 _get_propulsion_energy 内部 to_numeric）；
    目录不存在或无 CSV 时返回 None。
    """
    if not raw_dir.exists():
        return None
    csvs = sorted(raw_dir.glob('raw_*.csv'))
    if not csvs:
        return None
    frames = []
    for p in csvs:
        try:
            frames.append(pd.read_csv(p, dtype=str))
        except Exception as exc:
            logger.warning("读取 %s 失败：%s", p.name, exc)
    if not frames:
        return None
    df = pd.concat(frames, ignore_index=True)
    return df


def _find_or_append_column(ws, target_header: str) -> int:
    """在第 1 行查找 target_header；找到返回其 1-based 列号，找不到则在末尾新建。

    新建列时复制旧表头 cell 的字体（粗体居中）。
    """
    last_col = ws.max_column
    for col_idx in range(1, last_col + 1):
        v = ws.cell(1, col_idx).value
        if v == target_header:
            return col_idx
    new_col = last_col + 1
    cell = ws.cell(1, new_col)
    cell.value = target_header
    # 复制邻近表头的样式（粗体 + 居中）
    src = ws.cell(1, last_col)
    if src.font:
        cell.font = Font(
            bold=True,
            name=src.font.name,
            size=src.font.size,
            color=src.font.color,
        )
    if src.alignment:
        cell.alignment = Alignment(
            horizontal=src.alignment.horizontal or 'center',
            vertical=src.alignment.vertical,
        )
    # 设置列宽接近原报告其他列的视觉一致性
    ws.column_dimensions[cell.column_letter].width = max(len(target_header) + 4, 14)
    return new_col


def _row_format(ws, row_idx: int, last_col_before: int) -> tuple:
    """从该行已有的最后一列 cell 复制 fill / border / alignment / 数字格式，
    给新列上同样的格式（保持绿/红/白底背景一致）。
    """
    src = ws.cell(row_idx, last_col_before)
    return (src.fill, src.border, src.alignment)


def _backfill_xlsx(xlsx_path: Path, raw_dir: Path) -> dict:
    """处理一个 xlsx 文件。返回统计 dict。"""
    stats = {
        'file': xlsx_path.name,
        'trip_rows': 0,
        'nan_rows': 0,
        'non_trip_rows': 0,
        'mean_kwh': None,
        'median_kwh': None,
        'skipped_reason': None,
    }
    df_tel = _load_telematics_df(raw_dir)
    if df_tel is None:
        stats['skipped_reason'] = 'no raw_telematics CSV'
        return stats
    if _PROPULSION_COL not in df_tel.columns:
        stats['skipped_reason'] = f'missing column {_PROPULSION_COL!r} in raw CSV'
        return stats

    wb = load_workbook(xlsx_path)
    if 'Report' not in wb.sheetnames:
        stats['skipped_reason'] = "no 'Report' sheet"
        return stats
    ws = wb['Report']

    # 找到（或新建）目标列；记录 last_col_before 用于继承格式
    target_header = 'Propulsion Energy (kWh)'
    last_col_before = ws.max_column
    target_col = _find_or_append_column(ws, target_header)
    is_new_column = target_col > last_col_before
    fmt_source_col = last_col_before if is_new_column else (target_col - 1 if target_col > 1 else target_col)

    # 列索引 (1-based) —— Leg Type 第 2 列、Start 第 6、End 第 9（v2.2.2 EV HEADERS 固定）
    COL_LEG_TYPE = 2
    COL_START = 6
    COL_END = 9

    values_for_stats = []

    for row_idx in tqdm(range(2, ws.max_row + 1), desc=xlsx_path.name, leave=False):
        leg_type = ws.cell(row_idx, COL_LEG_TYPE).value
        cell = ws.cell(row_idx, target_col)

        # 继承同行末尾 cell 的视觉样式（绿/红/白底 + border + alignment）
        # openpyxl 的 fill / border / alignment 是 StyleProxy（不可直接赋值），
        # 必须通过 copy.copy() 取出可变副本再赋值。
        if is_new_column:
            src = ws.cell(row_idx, fmt_source_col)
            try:
                if src.fill is not None:
                    cell.fill = copy(src.fill)
                if src.border is not None:
                    cell.border = copy(src.border)
                if src.alignment is not None:
                    cell.alignment = copy(src.alignment)
            except Exception:
                # 单元格样式异常不应阻断回填；继续写值
                pass

        if not _is_trip_row(leg_type):
            stats['non_trip_rows'] += 1
            cell.value = '=NA()'
            continue

        t_s = ws.cell(row_idx, COL_START).value
        t_e = ws.cell(row_idx, COL_END).value
        if t_s is None or t_e is None:
            stats['trip_rows'] += 1
            stats['nan_rows'] += 1
            cell.value = '=NA()'
            continue

        try:
            ts_s = pd.Timestamp(t_s)
            ts_e = pd.Timestamp(t_e)
        except Exception:
            stats['trip_rows'] += 1
            stats['nan_rows'] += 1
            cell.value = '=NA()'
            continue
        if ts_s.tzinfo is None:
            ts_s = ts_s.tz_localize('UTC')
        if ts_e.tzinfo is None:
            ts_e = ts_e.tz_localize('UTC')

        prop_kwh = _get_propulsion_energy(df_tel, ts_s, ts_e)
        stats['trip_rows'] += 1
        if prop_kwh is None or (isinstance(prop_kwh, float) and np.isnan(prop_kwh)):
            stats['nan_rows'] += 1
            cell.value = '=NA()'
        else:
            cell.value = float(prop_kwh)
            values_for_stats.append(float(prop_kwh))

    if values_for_stats:
        stats['mean_kwh'] = round(float(np.mean(values_for_stats)), 2)
        stats['median_kwh'] = round(float(np.median(values_for_stats)), 2)

    wb.save(xlsx_path)
    return stats


def _iter_target_xlsx(reports_root: Path, veh_filter: str | None):
    """枚举 excel_report_database/<version>/<REG>/jolt_report_<REG>_*.xlsx（跳过 _finetuned）。"""
    if not reports_root.exists():
        return
    for veh_dir in sorted(reports_root.iterdir()):
        if not veh_dir.is_dir():
            continue
        reg = veh_dir.name
        if veh_filter and reg != veh_filter:
            continue
        raw_dir = veh_dir / 'raw_telematics'
        for xlsx in sorted(veh_dir.glob('jolt_report_*.xlsx')):
            if xlsx.stem.endswith('_finetuned'):
                continue
            yield reg, xlsx, raw_dir


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description='Backfill Propulsion Energy (kWh) column into existing JOLT xlsx reports.')
    parser.add_argument('--version', required=True,
                        help='Report version subdirectory (e.g. 2.2.2)')
    parser.add_argument('--veh', default=None,
                        help='Optional vehicle registration filter (e.g. YK73WFN)')
    parser.add_argument('--reports-dir', default='excel_report_database',
                        help='Reports root directory (default: ./excel_report_database)')
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Verbose logging')
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s: %(message)s',
    )

    reports_root = Path(args.reports_dir).resolve() / args.version
    if not reports_root.exists():
        logger.error('Reports directory not found: %s', reports_root)
        return 1

    targets = list(_iter_target_xlsx(reports_root, args.veh))
    if not targets:
        logger.warning('No target xlsx found under %s (veh filter=%r)',
                       reports_root, args.veh)
        return 1

    logger.info('Found %d xlsx file(s) to process under %s', len(targets), reports_root)

    fleet_stats = []
    for reg, xlsx, raw_dir in targets:
        try:
            s = _backfill_xlsx(xlsx, raw_dir)
        except Exception as exc:
            logger.exception('Failed: %s', xlsx)
            s = {'file': xlsx.name, 'skipped_reason': f'exception: {exc}'}
        s['reg'] = reg
        fleet_stats.append(s)
        if s.get('skipped_reason'):
            logger.warning('  [%s] SKIP %s — %s', reg, xlsx.name, s['skipped_reason'])
        else:
            logger.info(
                '  [%s] %s — trips=%d nan=%d non-trip=%d mean=%s kWh median=%s kWh',
                reg, xlsx.name, s['trip_rows'], s['nan_rows'],
                s['non_trip_rows'], s['mean_kwh'], s['median_kwh'])

    # 汇总
    print('\n=== Backfill summary ===')
    print(f'{"REG":<10} {"file":<55} {"trips":>6} {"nan":>5} {"mean":>8} {"median":>8} {"note"}')
    for s in fleet_stats:
        print(
            f'{s.get("reg",""):<10} {s.get("file","")[:55]:<55} '
            f'{s.get("trip_rows", 0):>6} {s.get("nan_rows", 0):>5} '
            f'{str(s.get("mean_kwh","-")):>8} {str(s.get("median_kwh","-")):>8} '
            f'{s.get("skipped_reason","")}'
        )
    return 0


if __name__ == '__main__':
    sys.exit(main())
