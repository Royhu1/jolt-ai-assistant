"""
patch_ep_exclude_aux.py
=======================
Post-hoc 补丁脚本：往已有的 ``excel_report_database/<version>/<REG>/jolt_report_*.xlsx`` 末尾
追加（或覆盖）一列 ``EP_exclude_aux``，**不重跑** SRF API / 分段算法。

定义（见 ``data_analysis_workspace/energy_balance_check/report.md``）::

    EP_exclude_aux = (propulsion − recuperation) / distance
                   = EP − auxiliary / distance        [kWh/km]

即每公里牵引能量扣除再生回收、再剔除 HVAC/驻车等辅助负载后的净值。只有同时报告
``electric_energy_propulsion`` 与 ``electric_energy_recuperation_watthours`` 两个
计数器的车才能算（EX74JXW / EX74JXY / YN25RSY / YN75NMA 计数器 NaN/缺失，
WU70GLV 为柴油车走 DIESEL_HEADERS）—— 这些车的 trip 行会写 ``=NA()``。

数据来源（两条路，优先用文件内已有列以避免重新拉 raw CSV）
--------------------------------------------------------------
1. **propulsion**：优先读 xlsx 的 ``Propulsion Energy (kWh)`` 列；该列缺失或某 trip
   行为空时，fallback 到同目录 ``raw_telematics/raw_*.csv`` 的
   ``electric_energy_propulsion`` 列，按 trip 时间窗插值差分
   （:func:`...report_builder._get_propulsion_energy`，与 v2.2.3 主管线同源）。
2. **recuperation**：优先读 xlsx 的 ``Recuperation Energy (kWh)`` 列；缺失/空时
   fallback 到 raw CSV 的 ``electric_energy_recuperation_watthours``
   （:func:`...report_builder._get_recuperation`，区间 max − min）。
3. **distance**：读 xlsx 的 ``Distance (km)`` 列。

用法
----
::

    # 单车
    python -m jolt_toolkit.scripts.patch_ep_exclude_aux --version 2.2.3 --veh YK73WFN

    # 整个版本目录下的全部 EV（位置参数亦可：REG ... ）
    python -m jolt_toolkit.scripts.patch_ep_exclude_aux --version 2.2.3
    python -m jolt_toolkit.scripts.patch_ep_exclude_aux --version 2.2.3 YK73WFN AV24LXJ

行为约定（镜像 backfill_propulsion_energy.py / LoggerPatcher）
--------------------------------------------------------------
- 用 ``openpyxl`` 打开 xlsx，不重写 worksheet —— 现有格式 / 颜色 / 公式全部保留。
- 列已存在时（脚本被重复跑）直接覆盖该列，不重复追加。
- 仅对 trip / In Transit / Outbound / Return / In House / Round Trip 行计算；
  Stop / 充电行强制写 ``=NA()``。
- 仅处理 ``jolt_report_<REG>_<ds>_<de>.xlsx``，**跳过 ``*_finetuned.xlsx``**
  （finetune 链路自己负责保持一致）。
- 计数器缺失/全空的车，trip 行写 ``=NA()`` 而不损坏 xlsx。
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from copy import copy
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from tqdm import tqdm

from jolt_toolkit.report_generator.report_builder import (
    _ep_exclude_aux,
    _get_propulsion_energy,
    _get_recuperation,
    _PROPULSION_COL,
    _RECUP_COL,
)

logger = logging.getLogger(__name__)

TARGET_HEADER = 'EP_exclude_aux'
PROP_HEADER = 'Propulsion Energy (kWh)'
RECUP_HEADER = 'Recuperation Energy (kWh)'
DIST_HEADER = 'Distance (km)'
LEG_TYPE_HEADER = 'Leg Type'
START_HEADER = 'Start Time (UTC)'
END_HEADER = 'End Time (UTC)'

# Trip-类 leg type 的判定：与 backfill_propulsion_energy / _generator 兜底 pass 一致
_CHARGE_PREFIX_RE = re.compile(r'^(ac|dc|charge|mix|estimated)', re.IGNORECASE)


def _is_trip_row(leg_type) -> bool:
    """放电 / 行程行返回 True；Stop / charge / 缺失返回 False。"""
    if not isinstance(leg_type, str):
        return False
    s = leg_type.strip().lower()
    if not s or s == 'stop':
        return False
    if _CHARGE_PREFIX_RE.match(s):
        return False
    return True


def _num(v):
    """把单元格值安全转为 float；非数 / ``#N/A`` / 空 → None。"""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if np.isnan(f):
        return None
    return f


def _load_telematics_df(raw_dir: Path) -> pd.DataFrame | None:
    """拼接 raw_telematics 目录下所有 raw_*.csv（dtype=str）。无 CSV 时返回 None。"""
    if not raw_dir.exists():
        return None
    csvs = sorted(raw_dir.glob('raw_*.csv'))
    if not csvs:
        return None
    frames = []
    for p in csvs:
        try:
            frames.append(pd.read_csv(p, dtype=str))
        except Exception as exc:
            logger.warning("读取 %s 失败：%s", p.name, exc)
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)


def _header_columns(ws) -> dict:
    """返回 {header_name: 1-based col index}（按第 1 行表头）。"""
    cols = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is not None:
            cols[str(v)] = col_idx
    return cols


def _find_or_append_column(ws, target_header: str, cols: dict) -> tuple[int, bool]:
    """查找 target_header；找到返回 (1-based 列号, False)，找不到则末尾新建返回 (col, True)。"""
    if target_header in cols:
        return cols[target_header], False
    last_col = ws.max_column
    new_col = last_col + 1
    cell = ws.cell(1, new_col)
    cell.value = target_header
    src = ws.cell(1, last_col)
    if src.font:
        cell.font = Font(bold=True, name=src.font.name, size=src.font.size,
                         color=src.font.color)
    if src.alignment:
        cell.alignment = Alignment(
            horizontal=src.alignment.horizontal or 'center',
            vertical=src.alignment.vertical)
    ws.column_dimensions[cell.column_letter].width = max(len(target_header) + 4, 14)
    return new_col, True


def _patch_xlsx(xlsx_path: Path, raw_dir: Path) -> dict:
    """处理一个 xlsx 文件。返回统计 dict。"""
    stats = {
        'file': xlsx_path.name,
        'trip_rows': 0,
        'computed_rows': 0,
        'nan_rows': 0,
        'non_trip_rows': 0,
        'prop_from_raw': 0,
        'recup_from_raw': 0,
        'median_kwh': None,
        'skipped_reason': None,
    }

    wb = load_workbook(xlsx_path)
    if 'Report' not in wb.sheetnames:
        stats['skipped_reason'] = "no 'Report' sheet"
        return stats
    ws = wb['Report']

    cols = _header_columns(ws)
    # 柴油报告没有 Distance/Recuperation 这套电量列结构 → 不适用，跳过
    if DIST_HEADER not in cols or LEG_TYPE_HEADER not in cols \
            or START_HEADER not in cols or END_HEADER not in cols:
        stats['skipped_reason'] = 'not an EV report (missing core columns)'
        return stats

    col_leg = cols[LEG_TYPE_HEADER]
    col_dist = cols[DIST_HEADER]
    col_start = cols[START_HEADER]
    col_end = cols[END_HEADER]
    col_prop = cols.get(PROP_HEADER)
    col_recup = cols.get(RECUP_HEADER)

    # 目标列（追加前先确定 last_col_before，用于继承格式 + fallback 取数前 raw CSV 懒加载）
    last_col_before = ws.max_column
    target_col, is_new = _find_or_append_column(ws, TARGET_HEADER, cols)
    fmt_source_col = last_col_before if is_new else max(target_col - 1, 1)

    # raw telematics 懒加载：仅当需要 fallback（缺 propulsion/recup 列或列内有空值）才读
    _df_cache: dict = {}

    def _get_telematics():
        if 'df' not in _df_cache:
            _df_cache['df'] = _load_telematics_df(raw_dir)
        return _df_cache['df']

    values_for_stats = []

    for row_idx in tqdm(range(2, ws.max_row + 1), desc=xlsx_path.name, leave=False):
        leg_type = ws.cell(row_idx, col_leg).value
        cell = ws.cell(row_idx, target_col)

        # 继承同行末尾 cell 的视觉样式（绿/红/白底 + border + alignment）
        if is_new:
            src = ws.cell(row_idx, fmt_source_col)
            try:
                if src.fill is not None:
                    cell.fill = copy(src.fill)
                if src.border is not None:
                    cell.border = copy(src.border)
                if src.alignment is not None:
                    cell.alignment = copy(src.alignment)
            except Exception:
                pass

        if not _is_trip_row(leg_type):
            stats['non_trip_rows'] += 1
            cell.value = '=NA()'
            continue

        stats['trip_rows'] += 1
        distance = _num(ws.cell(row_idx, col_dist).value)
        if distance is None or distance <= 0:
            stats['nan_rows'] += 1
            cell.value = '=NA()'
            continue

        # ── 时间窗（fallback 取数需要）──────────────────────────────
        t_s = ws.cell(row_idx, col_start).value
        t_e = ws.cell(row_idx, col_end).value
        ts_s = ts_e = None
        try:
            ts_s = pd.Timestamp(t_s)
            ts_e = pd.Timestamp(t_e)
            if ts_s.tzinfo is None:
                ts_s = ts_s.tz_localize('UTC')
            if ts_e.tzinfo is None:
                ts_e = ts_e.tz_localize('UTC')
        except Exception:
            ts_s = ts_e = None

        # ── propulsion：优先文件列，缺失/空 → raw CSV ─────────────
        prop = _num(ws.cell(row_idx, col_prop).value) if col_prop else None
        if prop is None and ts_s is not None:
            df = _get_telematics()
            if df is not None and _PROPULSION_COL in df.columns:
                prop = _num(_get_propulsion_energy(df, ts_s, ts_e))
                if prop is not None:
                    stats['prop_from_raw'] += 1

        # ── recuperation：优先文件列，缺失/空 → raw CSV ───────────
        recup = _num(ws.cell(row_idx, col_recup).value) if col_recup else None
        if recup is None and ts_s is not None:
            df = _get_telematics()
            if df is not None and _RECUP_COL in df.columns:
                recup = _num(_get_recuperation(df, ts_s, ts_e))
                if recup is not None:
                    stats['recup_from_raw'] += 1

        epx = _ep_exclude_aux(
            prop if prop is not None else float('nan'),
            recup if recup is not None else float('nan'),
            distance)
        if epx is None or (isinstance(epx, float) and np.isnan(epx)):
            stats['nan_rows'] += 1
            cell.value = '=NA()'
        else:
            cell.value = float(epx)
            stats['computed_rows'] += 1
            values_for_stats.append(float(epx))

    if values_for_stats:
        stats['median_kwh'] = round(float(np.median(values_for_stats)), 3)

    wb.save(xlsx_path)
    return stats


def _iter_target_xlsx(reports_root: Path, veh_filter: set | None):
    """枚举 excel_report_database/<version>/<REG>/jolt_report_<REG>_*.xlsx（跳过 _finetuned）。"""
    if not reports_root.exists():
        return
    for veh_dir in sorted(reports_root.iterdir()):
        if not veh_dir.is_dir():
            continue
        reg = veh_dir.name
        if veh_filter and reg not in veh_filter:
            continue
        raw_dir = veh_dir / 'raw_telematics'
        for xlsx in sorted(veh_dir.glob('jolt_report_*.xlsx')):
            if xlsx.stem.endswith('_finetuned'):
                continue
            yield reg, xlsx, raw_dir


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description='Patch EP_exclude_aux (kWh/km) column into existing JOLT xlsx reports.')
    parser.add_argument('--version', required=True,
                        help='Report version subdirectory (e.g. 2.2.3)')
    parser.add_argument('regs', nargs='*', default=None,
                        help='Optional vehicle registrations to process (default: all)')
    parser.add_argument('--veh', default=None,
                        help='Single vehicle registration filter (alias for a positional REG)')
    parser.add_argument('--reports-dir', default='excel_report_database',
                        help='Reports root directory (default: ./excel_report_database)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s: %(message)s')

    veh_filter: set | None = set()
    if args.veh:
        veh_filter.add(args.veh)
    for r in (args.regs or []):
        veh_filter.add(r)
    if not veh_filter:
        veh_filter = None

    reports_root = Path(args.reports_dir).resolve() / args.version
    if not reports_root.exists():
        logger.error('Reports directory not found: %s', reports_root)
        return 1

    targets = list(_iter_target_xlsx(reports_root, veh_filter))
    if not targets:
        logger.warning('No target xlsx found under %s (veh filter=%r)',
                       reports_root, veh_filter)
        return 1

    logger.info('Found %d xlsx file(s) to process under %s', len(targets), reports_root)

    fleet_stats = []
    for reg, xlsx, raw_dir in targets:
        try:
            s = _patch_xlsx(xlsx, raw_dir)
        except Exception as exc:
            logger.exception('Failed: %s', xlsx)
            s = {'file': xlsx.name, 'skipped_reason': f'exception: {exc}'}
        s['reg'] = reg
        fleet_stats.append(s)
        if s.get('skipped_reason'):
            logger.warning('  [%s] SKIP %s — %s', reg, xlsx.name, s['skipped_reason'])
        else:
            logger.info(
                '  [%s] %s — trips=%d computed=%d nan=%d non-trip=%d '
                'prop_from_raw=%d recup_from_raw=%d median=%s kWh/km',
                reg, xlsx.name, s['trip_rows'], s['computed_rows'], s['nan_rows'],
                s['non_trip_rows'], s['prop_from_raw'], s['recup_from_raw'],
                s['median_kwh'])

    print('\n=== EP_exclude_aux patch summary ===')
    print(f'{"REG":<10} {"file":<55} {"trips":>6} {"comp":>5} {"nan":>5} {"median":>8} {"note"}')
    for s in fleet_stats:
        print(
            f'{s.get("reg",""):<10} {s.get("file","")[:55]:<55} '
            f'{s.get("trip_rows", 0):>6} {s.get("computed_rows", 0):>5} '
            f'{s.get("nan_rows", 0):>5} {str(s.get("median_kwh","-")):>8} '
            f'{s.get("skipped_reason","")}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
