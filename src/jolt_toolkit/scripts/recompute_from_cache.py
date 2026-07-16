"""
recompute_from_cache
====================
Fast, SRF-free regeneration of the fleet at a **post-segmentation release** from
the cached artefacts of the previous version (``excel_report_database/<src>/<REG>/``).

Provenance: moved in v3.0.0 from
``report_generator/recompute_v227.py`` to ``scripts/recompute_from_cache.py`` (a
reusable one-off migration tool, not part of the deployed report-generation path).
The historical name ``recompute_v227`` is retained in the changelogs. Invoke via
``python -m jolt_toolkit.scripts.recompute_from_cache --src-ver <old> --dst-ver <new>``.

Despite the historical name, this is the **general** cached-recompute tool for any
release whose changes are all post-segmentation; it is ``--src-ver`` / ``--dst-ver``
parameterised. It was first used for **2.2.6 → 2.2.7** and is now the migration
path for **2.2.7 → 2.2.8**.

Why this exists
---------------
The v2.2.7 change (``fix/capacity-correction-counter-energy``) and the v2.2.8
change (this branch) are **both post-segmentation**:

* v2.2.7 — **MODE A step-2 gate** (an outlier IMPLIED capacity only rewrites the
  *capacity* column; counter-sourced legs keep their counter energy) +
  **ΔSOC-weighted donor capacity** (:func:`_soc_weighted_cap`).
* v2.2.8 — **per-vehicle SOC-energy fallback** inside ``_correct_effective_capacity``
  (opt-in ``soc_energy_fallback`` in vehicles.json; an outlier counter-sourced leg
  with a large, reliable ΔSOC and a far-off implied capacity is re-derived from
  ``SOC × effective capacity`` and marked ``'soc_fallback'``). The v2.2.8 charger
  fusion fix is applied **separately** via the ``charger_patcher`` backfill CLI, not
  here — this tool only carries the ``raw_charger/`` dir forward (see below).

The **segmentation is unchanged**, so a full regen (SRF re-pull + logger channel
loads at 1 Hz — slow and memory-heavy) is wasteful. This module instead re-runs
the **real** ``run_segment_detection`` on the locally-cached per-FPS-leg raw
telematics CSVs (``raw_telematics/raw_*.csv``), feeds the segments through the
**real** ``_seg_to_row`` and the **real** ``_correct_effective_capacity`` (with the
per-vehicle ``soc_fallback`` resolved from the config), then writes the report with
the real ``_write_excel_report`` (charts included). Because it re-uses the actual
pipeline functions on the identical raw, the result is provably identical to a full
regen (verified byte-exact on the 3 Knowles AV24LXJ/K/L for 8 of 9 quarters; the
most-recent quarter differs only because the cache holds slightly fewer donor legs
than a fresh SRF pull — a data-vintage effect, not a logic difference).

Everything the recompute cannot re-derive without SRF / logger / weather patchers
(weather columns, operator, charger/logger links, place names, logger-filled mass,
pedal histograms) is **preserved verbatim** from the src xlsx via a per-vehicle
``start_time`` lookup — so NO weather re-patch is needed. Kinetics-corrected EP is
reconstructed exactly from the energy-independent ``ke_per_d`` term carried in the
src ``(EP_corrected − EP_kinetics)`` difference.

What is written vs carried forward (per vehicle)
------------------------------------------------
* **Recomputed & written**: the ``.xlsx`` reports (energy / EP / capacity / SOC /
  distance / elevation / the embedded Graphs charts), one per **meteorological
  quarter** (DJF/MAM/JJA/SON, inclusive end, no overlap — which also dedups the
  boundary-day double-dumps in the cache).
* **Overlaid in-cell (verbatim from src)**: weather (6 cols), operator, place
  names, charger/logger link targets, pedal histograms, Vehicle Mass + CV.
* **Copied forward verbatim (dirs + files)** by :func:`_copy_forward_artifacts`:
  ``raw_telematics/``, ``validation_figures/``, ``raw_charger/`` and
  ``raw_logger*/``, plus the ``inspect_*.html`` viewers. **The first 2.2.7 run
  did NOT copy ``raw_charger/`` / ``raw_logger*/`` forward, which broke charger /
  logger visibility in the dashboard & data-collection-monitor — now fixed.**
  (``validation_figures`` / ``inspect_*.html`` reflect the src energy on the few
  ``soc_fallback``-rewritten legs; regenerating them needs an SRF debug run.)
* **Verified after every copy** by :func:`_verify_copy_forward`: for each carried
  artefact dir (``_AUX_COPY_DIRS_STATIC`` + every ``raw_logger*`` found in src) and
  for the ``inspect_*.html`` set, the recursive file count in dst must be **>=** the
  count in src. dst holding *more* files than src is fine and expected — the
  data-collection-monitor appends new days into the current version. A shortfall
  (dst < src), or any per-file copy failure (``n_failed > 0`` — a failed copy can
  leave a truncated dst file the count check would miss), is logged as an error,
  recorded in the per-vehicle result dict, aggregated across the fleet by
  :func:`main`, and forces a **non-zero exit**. *Why:* the copy itself is now fixed,
  but the first 2.2.7 run's omission went unnoticed and the 2.2.7→2.2.8 run silently
  inherited the damage — a migration must never be able to finish green while
  artefacts were lost, so completeness is now asserted, not assumed.

Verbatim-copy vehicles (NOT replayed — :func:`copy_verbatim`):
* **Diesel** (WU70GLV / YT21EFD) — no SOC/counter energy, so neither the v2.2.7 nor
  the v2.2.8 change applies; the report is unchanged across the release.
* **``prefer_logger_speed``** vehicles (config-driven; YN25RSY today) — their
  segmentation runs on **SRF-Logger speed**, which the replay does NOT reproduce (it
  reads only ``raw_telematics``), so a replay diverges from the canonical
  segmentation (YN25RSY: 37→33 trips, ~15 % start-time match) and the start-time
  overlay then loses weather. These are copied verbatim from src instead.
For both, dst is an unchanged copy of src (xlsx + artefact dirs + inspect HTML).

This is a migration tool; it does **not** modify any core pipeline code and does
**not** persist to ``vehicles.json`` (kept canonical — capacity persistence is
LOG-ONLY here). The rolled-up + quarterly ΔSOC-weighted effective capacity is
computed and logged for reference.
"""
from __future__ import annotations

import argparse
import glob
import logging
import os
import re
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from jolt_toolkit import __version__
from jolt_toolkit.report_generator.segment_algorithms import (
    run_segment_detection,
    resolve_mass_agg,
    VEHICLE_CONFIG,
    PIPELINE_CONFIGS,
    SOC_COL,
    _ANCHOR_PRIVATE_KEYS,
)
from jolt_toolkit.report_generator.report_builder import (
    HEADERS,
    _seg_to_row,
    _insert_stop_rows,
    _write_excel_report,
    _corrected_energy_perf,
)
from jolt_toolkit.report_generator._generator import (
    JOLTReportGenerator,
    _period_capacity_from_rows,
    _recompute_weighted_capacity,
    _resolve_soc_fallback,
    _IDX_CAP, _IDX_ENERGY, _IDX_SOC_CHANGE, _IDX_EPERF, _IDX_DISTANCE,
    _IDX_ESOURCE, _IDX_BPOWER, _IDX_DURATION, _IDX_EPERF_CORR, _IDX_ELEV,
    _IDX_MASS, _IDX_EPERF_KIN, _IDX_START, _IDX_LEG_TYPE, _IDX_EP_EXCL_AUX,
)

logger = logging.getLogger(__name__)

# ── row-tuple / xlsx column index helpers ────────────────────────────────────
# row tuple index = HEADERS.index(name) - 1 (row tuple excludes 'Leg Number').
def _ri(name: str) -> int:
    return HEADERS.index(name) - 1

_IDX_TELEM_LINK = _ri('Telematics Link')
_IDX_CHARGER_LINK = _ri('Charger Link')
_IDX_LOGGER_LINK = _ri('SRF Logger Link')
_IDX_MASS_CV = _ri('Vehicle Mass CV (reliability)')
_IDX_CUMDIST = _ri('Cumulative Distance (km)')

# Columns overlaid verbatim from the 2.2.6 report (the no-SRF/no-logger/no-weather
# recompute cannot reproduce these). All are plain VALUE cells.
_PRESERVE_VALUE_HEADERS = [
    'Origin Place', 'Destination Place',
    'Vehicle Mass (kg)', 'Vehicle Mass CV (reliability)',
    'Average Temperature (C)', 'Average Pressure (hPa)', 'Average Humidity (%)',
    'Average Wind Speed (m/s)', 'Average Wind Direction', 'Weather Type',
    'Histogram of Accelerator Pedal Position',
    'Histogram of Decelerator Pedal Position',
    'Operator',
]
# Link columns are HYPERLINK cells (value may be empty; the URL is the hyperlink
# target). Restored into the row tuple as URLs so _write_excel_report re-renders
# the blue "Link".
_PRESERVE_LINK_HEADERS = ['Charger Link', 'SRF Logger Link']

_CHARGE_PREFIX_RE = re.compile(r'^(ac|dc|charge|mix|estimated)', re.I)


def _norm_ts(v) -> pd.Timestamp:
    t = pd.Timestamp(v)
    return t.tz_convert(None) if t.tzinfo is not None else t


def _ts_key(v) -> str | None:
    if v is None or v == '':
        return None
    try:
        return str(_norm_ts(v))
    except Exception:
        return None


# ── meteorological-quarter split (mirrors batch_generate.split_into_periods) ──
# Vendored (2026-07-02) from .claude/skills/generate-excel-report/batch_generate.py
# — the skill dir is not on the package path. Boundaries 12-01/03-01/06-01/09-01;
# each segment = [q_start, next_q_start - 1 day] (inclusive end, no overlap).
def _met_quarter_start_on_or_before(ts: pd.Timestamp) -> pd.Timestamp:
    cands = [
        pd.Timestamp(ts.year - 1, 12, 1), pd.Timestamp(ts.year, 3, 1),
        pd.Timestamp(ts.year, 6, 1), pd.Timestamp(ts.year, 9, 1),
        pd.Timestamp(ts.year, 12, 1),
    ]
    return max(s for s in cands if s <= ts)


def _next_met_quarter_start(qstart: pd.Timestamp) -> pd.Timestamp:
    if qstart.month == 12:
        return pd.Timestamp(qstart.year + 1, 3, 1)
    return pd.Timestamp(qstart.year, qstart.month + 3, 1)


def split_into_meteorological_quarters(date_start: str, date_end: str):
    start = pd.Timestamp(date_start)
    end = pd.Timestamp(date_end)
    periods = []
    qstart = _met_quarter_start_on_or_before(start)
    while qstart <= end:
        nxt = _next_met_quarter_start(qstart)
        seg_start = max(qstart, start)
        seg_end = min(nxt - pd.Timedelta(days=1), end)
        if seg_start <= seg_end:
            periods.append((seg_start.strftime('%Y-%m-%d'), seg_end.strftime('%Y-%m-%d')))
        qstart = nxt
    return periods


# ── stage 1: replay segmentation from cached raw ─────────────────────────────
def replay_rows(reg: str, csv_dir: Path) -> tuple[list, dict]:
    """Re-run the real per-leg segmentation on every cached raw_*.csv.

    Returns ``(all_rows, cfg)`` where ``all_rows`` = list of
    ``(seg_start_time, row_list, mode)`` (pre-effective-capacity-correction).
    cumulative_km is left 0 here (fixed per quarter in :func:`_recompute_report`).
    """
    cfg = VEHICLE_CONFIG[reg]
    nominal = cfg.get('nominal_kwh')
    cap_lo = nominal * 0.5 if nominal else None
    cap_hi = nominal * 2.0 if nominal else None
    speed_col = cfg.get('speed_col', 'wheel_based_speed')
    altitude_col = cfg.get('altitude_col')
    mass_agg = resolve_mass_agg(reg, PIPELINE_CONFIGS.get(cfg.get('pipeline', 'default_soc')))
    files = sorted(glob.glob(str(Path(csv_dir) / 'raw_*.csv')))
    all_rows: list = []
    home_point = None
    for fp in files:
        try:
            df_leg = pd.read_csv(fp, dtype=str)
        except Exception as exc:
            logger.warning('read fail %s: %s', fp, exc)
            continue
        if df_leg.empty or SOC_COL not in df_leg.columns:
            continue
        leg_uri = 'https://data.csrf.ac.uk/api/legs/' + Path(fp).stem
        c_segs, d_segs = run_segment_detection(
            df_leg, reg=reg, suffix=Path(fp).stem, out_dir=None,
            generate_validation_fig=False, cap_lo=cap_lo, cap_hi=cap_hi)
        for seg in c_segs:
            sc = {k: v for k, v in seg.items() if k not in _ANCHOR_PRIVATE_KEYS}
            row, _ = _seg_to_row(sc, 'charge', leg_uri, [], [], df_leg, 0.0,
                                 home_point, srf_data=None, altitude_col=altitude_col,
                                 speed_col=speed_col, operator=None, mass_agg=mass_agg)
            all_rows.append((seg['start_time'], list(row), 'charge'))
        for seg in d_segs:
            sc = {k: v for k, v in seg.items() if k not in _ANCHOR_PRIVATE_KEYS}
            row, _ = _seg_to_row(sc, 'discharge', leg_uri, [], [], df_leg, 0.0,
                                 home_point, srf_data=None, altitude_col=altitude_col,
                                 speed_col=speed_col, has_logger=False, operator=None,
                                 mass_agg=mass_agg)
            all_rows.append((seg['start_time'], list(row), 'discharge'))
        if home_point is None and c_segs:
            from geopy import Point as GeoPoint
            for s in c_segs:
                la, lo = s.get('latitude'), s.get('longitude')
                if la is not None and lo is not None:
                    try:
                        home_point = GeoPoint(float(la), float(lo))
                        break
                    except Exception:
                        pass
    return all_rows, cfg


# ── stage 2: build the preserved-column lookup from all 2.2.6 xlsx ───────────
def build_preserved_lookup(reg: str, src_dir: Path) -> dict:
    """Read every 2.2.6 report for ``reg`` → ``{ts_key: {header: value, ...}}``.

    Non-read-only load so hyperlink targets (Charger/Logger Link) survive.
    Also stashes the 2.2.6 ``EP_corrected`` / ``EP_kinetics`` so kinetics EP can
    be reconstructed via the energy-independent ``ke_per_d`` term.
    """
    lookup: dict = {}
    kin_corr = 'Energy Performance Corrected by Elevation Difference (kWh/km)'
    kin_kin = 'Energy Performance Kinetics Corrected (kWh/km)'

    def _val(cell):
        """Cell value, but treat ``=NA()`` (and any formula) as None.

        The report is formula-driven: 'no data' cells are the xlsxwriter
        ``=NA()`` formula. In non-read-only + data_only mode openpyxl surfaces
        those as ``0`` (a cached value), which would silently overwrite a blank
        (e.g. no-GVM T88RNW/YN75NMA mass) with 0. Reading with data_only=False
        and skipping formula cells (``data_type == 'f'``) yields the real value
        for populated cells and None for ``=NA()`` blanks.
        """
        return cell.value if cell.data_type != 'f' else None

    for xlsx in sorted(src_dir.glob(f'jolt_report_{reg}_*.xlsx')):
        # data_only=False so =NA() surfaces as a formula (skipped), not cached 0;
        # non-read-only so hyperlink targets (Charger/Logger Link) survive.
        wb = load_workbook(str(xlsx), data_only=False)
        ws = wb['Report']
        hdr = [c.value for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(hdr)}  # 1-based
        i_start = col.get('Start Time (UTC)')
        for r in range(2, ws.max_row + 1):
            leg_no = ws.cell(r, 1).value
            if leg_no in (None, ''):
                continue
            k = _ts_key(ws.cell(r, i_start).value)
            if k is None:
                continue
            rec = {}
            for h in _PRESERVE_VALUE_HEADERS:
                if h in col:
                    rec[h] = _val(ws.cell(r, col[h]))
            for h in _PRESERVE_LINK_HEADERS:
                if h in col:
                    cell = ws.cell(r, col[h])
                    rec[h] = cell.hyperlink.target if cell.hyperlink is not None else None
            # kinetics reconstruction inputs
            if kin_corr in col and kin_kin in col:
                rec['_ep_corr226'] = _val(ws.cell(r, col[kin_corr]))
                rec['_ep_kin226'] = _val(ws.cell(r, col[kin_kin]))
            # first writer wins (boundary-day legs appear in two 2.2.6 reports
            # with identical preserved values, so this is deterministic).
            lookup.setdefault(k, rec)
        wb.close()
    return lookup


# ── stage 3: per-quarter recompute + write ───────────────────────────────────
def _num(v):
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def _is_valid(v) -> bool:
    if v is None:
        return False
    try:
        return not np.isnan(v)
    except (TypeError, ValueError):
        return False


def _recompute_report(reg, all_rows, cfg, qstart, qend, lookup, out_path):
    """Build one meteorological-quarter v2.2.7 report; returns (n_rows, pcap, pn, psrc)."""
    qs = pd.Timestamp(qstart)
    qe = pd.Timestamp(qend) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    # filter by seg start_time; dedup boundary-day double-dumps by (mode, start, leg_type, energy)
    seen, rows = set(), []
    for st, row, mode in sorted(all_rows, key=lambda x: _norm_ts(x[0])):
        t = _norm_ts(st)
        if not (qs <= t <= qe):
            continue
        e = row[_IDX_ENERGY]
        ekey = round(float(e), 3) if isinstance(e, (int, float)) and _is_valid(e) else None
        key = (mode, str(t), str(row[_IDX_LEG_TYPE]), ekey)
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
    if not rows:
        return 0, None, 0, 'empty'

    # overlay Vehicle Mass + CV from 2.2.6 BEFORE the correction so the
    # elevation-corrected EP recompute uses the authoritative (logger-filled) mass.
    for row in rows:
        k = _ts_key(row[_IDX_START])
        rec = lookup.get(k) if k else None
        if rec is None:
            continue
        if 'Vehicle Mass (kg)' in rec:
            mv = _num(rec['Vehicle Mass (kg)'])
            row[_IDX_MASS] = mv if mv is not None else float('nan')
        if 'Vehicle Mass CV (reliability)' in rec:
            cvv = _num(rec['Vehicle Mass CV (reliability)'])
            row[_IDX_MASS_CV] = cvv if cvv is not None else float('nan')

    # effective-capacity correction (MODE A gate + ΔSOC-weighted donors); v2.2.8
    # additionally applies the per-vehicle SOC-energy fallback resolved from cfg
    # (opt-in via vehicles.json 'soc_energy_fallback'; None → unchanged MODE A).
    soc_est_cap = (cfg.get('effective_capacity_kwh') or cfg.get('srf_capacity_kwh')
                   or cfg.get('nominal_kwh'))
    rows2, _eff, _src = JOLTReportGenerator._correct_effective_capacity(
        rows, _IDX_CAP, _IDX_ENERGY, _IDX_SOC_CHANGE, _IDX_EPERF, _IDX_DISTANCE,
        _IDX_ESOURCE, _IDX_BPOWER, _IDX_DURATION, _IDX_EPERF_CORR, _IDX_ELEV,
        _IDX_MASS, soc_est_cap, _IDX_EPERF_KIN, idx_start=_IDX_START,
        soc_fallback=_resolve_soc_fallback(cfg))

    # EP fallback pass (mirror _generate_report): non-discharge rows → NaN EP cols
    for row in rows2:
        lt = row[_IDX_LEG_TYPE]
        if isinstance(lt, str):
            ll = lt.strip().lower()
            if ll == 'stop' or _CHARGE_PREFIX_RE.match(ll):
                row[_IDX_EPERF] = float('nan')
                row[_IDX_EPERF_CORR] = float('nan')
                row[_IDX_EPERF_KIN] = float('nan')
                row[_IDX_EP_EXCL_AUX] = float('nan')

    # period ΔSOC-weighted donor capacity (same口径 as live + backfill)
    pcap, pn, psrc = _period_capacity_from_rows(
        rows2, _IDX_CAP, _IDX_SOC_CHANGE, _IDX_ESOURCE)

    # discharge post-pass: re-derive EP + EP-corrected from the FINAL
    # (energy, distance, elevation, overlaid-mass); reconstruct kinetics EP from
    # the energy-independent ke_per_d carried in the 2.2.6 (EP_corr - EP_kin).
    for row in rows2:
        lt = row[_IDX_LEG_TYPE]
        if not isinstance(lt, str):
            continue
        ll = lt.strip().lower()
        if ll == 'stop' or _CHARGE_PREFIX_RE.match(ll):
            continue
        e, d, h, m = row[_IDX_ENERGY], row[_IDX_DISTANCE], row[_IDX_ELEV], row[_IDX_MASS]
        if _is_valid(e) and _is_valid(d) and d > 0:
            row[_IDX_EPERF] = round(abs(e) / d, 4)
            if _is_valid(h) and _is_valid(m):
                row[_IDX_EPERF_CORR] = _corrected_energy_perf(e, d, h, m)
        # kinetics EP: EP_kin = EP_corr - ke_per_d, ke_per_d from 2.2.6 (energy-independent)
        k = _ts_key(row[_IDX_START])
        rec = lookup.get(k) if k else None
        if rec is not None:
            c226, kin226 = _num(rec.get('_ep_corr226')), _num(rec.get('_ep_kin226'))
            if c226 is not None and kin226 is not None and _is_valid(row[_IDX_EPERF_CORR]):
                ke_per_d = c226 - kin226
                row[_IDX_EPERF_KIN] = round(row[_IDX_EPERF_CORR] - ke_per_d, 4)

    # sort by start time, recompute per-quarter cumulative distance (discharge cumsum)
    rows2.sort(key=lambda r: _norm_ts(r[_IDX_START]))
    running = 0.0
    for row in rows2:
        lt = row[_IDX_LEG_TYPE]
        is_disch = isinstance(lt, str) and not (
            lt.strip().lower() == 'stop' or _CHARGE_PREFIX_RE.match(lt.strip().lower()))
        d = row[_IDX_DISTANCE]
        if is_disch and _is_valid(d) and d > 0:
            running += float(d)
            row[_IDX_CUMDIST] = round(running, 3)
        else:
            row[_IDX_CUMDIST] = float('nan')

    # overlay preserved VALUE + LINK columns from 2.2.6 (weather / operator /
    # place / links / pedal histograms). Mass already overlaid above.
    idx_map_val = {h: _ri(h) for h in _PRESERVE_VALUE_HEADERS
                   if h not in ('Vehicle Mass (kg)', 'Vehicle Mass CV (reliability)')}
    idx_map_link = {h: _ri(h) for h in _PRESERVE_LINK_HEADERS}
    for row in rows2:
        k = _ts_key(row[_IDX_START])
        rec = lookup.get(k) if k else None
        if rec is None:
            continue
        for h, idx in idx_map_val.items():
            if h in rec:
                row[idx] = rec[h]
        for h, idx in idx_map_link.items():
            if h in rec and rec[h]:
                row[idx] = rec[h]

    # insert Stop rows (after correction; they carry corrected SOC/mass neighbours)
    rows3 = _insert_stop_rows(rows2, headers=HEADERS)

    # overlay weather onto Stop rows (best-effort by start_time)
    weather_headers = ['Average Temperature (C)', 'Average Pressure (hPa)',
                       'Average Humidity (%)', 'Average Wind Speed (m/s)',
                       'Average Wind Direction', 'Weather Type']
    w_idx = {h: _ri(h) for h in weather_headers}
    for row in rows3:
        lt = row[_IDX_LEG_TYPE]
        if isinstance(lt, str) and lt.strip().lower() == 'stop':
            k = _ts_key(row[_IDX_START])
            rec = lookup.get(k) if k else None
            if rec is not None:
                for h, idx in w_idx.items():
                    if h in rec:
                        row[idx] = rec[h]

    ps = pd.Timestamp(qstart).date()
    pe = pd.Timestamp(qend).date()
    _write_excel_report([tuple(r) for r in rows3], reg, ps, pe, out_path, headers=HEADERS)
    return len(rows3), pcap, pn, psrc


# Non-recomputable per-vehicle artefact directories carried forward verbatim from
# the src version dir (the SRF-free recompute writes only the .xlsx). raw_logger*
# variants are discovered dynamically. See :func:`_copy_forward_artifacts`.
_AUX_COPY_DIRS_STATIC = ['raw_telematics', 'validation_figures', 'raw_charger']


def _same_size(src: Path, dst: Path) -> bool:
    """True iff ``dst`` exists and has the same byte size as ``src`` — the cheap
    skip-if-identical test used to resume a partially-copied dir without OneDrive
    churn. Any stat error is treated as 'not identical' (fall through to copy)."""
    try:
        return dst.exists() and dst.stat().st_size == src.stat().st_size
    except OSError:
        return False


def _safe_copy_file(src: Path, dst: Path) -> None:
    """Copy a file's CONTENT only (no metadata / mtime), tolerant of OneDrive.

    ``shutil.copy2`` (and hence ``copytree``) runs ``copystat`` → ``os.utime`` and,
    on Windows, a ``CopyFile2`` fast path — both raise ``OSError: [Errno 22]
    Invalid argument`` on OneDrive-synced files the sync client is mid-operation on.
    Artefact copies never need timestamps, so this copies bytes only: first the
    plain ``shutil.copyfile`` (no ``copystat``), and if even that raises, a manual
    streamed ``read → write`` (no fast path). Raises ``OSError`` only if BOTH
    byte-copy paths fail — callers catch it per file and continue."""
    try:
        shutil.copyfile(src, dst)
        return
    except OSError:
        pass
    with open(src, 'rb') as fi, open(dst, 'wb') as fo:
        shutil.copyfileobj(fi, fo)


def _copy_dir_content(src_dir: Path, dst_dir: Path,
                      overwrite: bool = False) -> tuple[int, int, int]:
    """Recursively copy ``src_dir`` → ``dst_dir`` with content-only, per-file-
    tolerant semantics (see :func:`_safe_copy_file`). Skips files that already
    exist at the same size (unless ``overwrite``), so a re-run resumes a partially
    copied dir cheaply. Returns ``(n_copied, n_skipped, n_failed)``; never raises
    for a per-file error — it logs a warning and moves on."""
    n_copied = n_skipped = n_failed = 0
    for root, _dirs, files in os.walk(src_dir):
        rel = Path(root).relative_to(src_dir)
        target_root = dst_dir / rel
        try:
            target_root.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            n_failed += len(files)
            logger.warning('copy-forward: cannot mkdir %s (%s)', target_root, exc)
            continue
        for fn in files:
            s = Path(root) / fn
            d = target_root / fn
            if not overwrite and _same_size(s, d):
                n_skipped += 1
                continue
            try:
                _safe_copy_file(s, d)
                n_copied += 1
            except OSError as exc:
                n_failed += 1
                logger.warning('copy-forward: skip %s (%s)', s, exc)
    return n_copied, n_skipped, n_failed


def _copy_forward_artifacts(src_dir: Path, dst_dir: Path,
                            overwrite: bool = False) -> tuple[list[str], int]:
    """Copy the non-recomputable per-vehicle artefact dirs + inspect HTML forward
    from the src version dir to the dst version dir.

    The recompute writes only the .xlsx (with embedded Graphs charts). Everything
    else the downstream tooling expects under a version dir is copied verbatim:

    * ``raw_telematics/`` — dashboard telematics-day detection + future replays;
    * ``validation_figures/`` — per-day validation PNGs;
    * ``raw_charger/`` and ``raw_logger*/`` — dashboard / data-collection-monitor
      charger + logger visibility. **The first 2.2.7 recompute run OMITTED these,
      which broke charger visibility in the dashboard/monitor — the reason this
      helper exists.**
    * ``inspect_*.html`` — the per-report HTML viewers.

    Staleness caveat: ``validation_figures`` / ``inspect_*.html`` embed the src
    version's energy on ``soc_fallback``-rewritten legs (segmentation is unchanged,
    only the energy/EP of a few outlier legs moves); regenerating them requires an
    SRF debug run. This is the same accepted trade-off as the preserved weather
    columns, and only affects debug-only artefacts.

    OneDrive tolerance: the repo lives under OneDrive, where ``shutil.copytree`` /
    ``shutil.copy2`` raise ``OSError: [Errno 22] Invalid argument`` from the
    ``copystat``/``utime`` / ``CopyFile2`` fast paths on files the sync client is
    touching. We therefore copy **CONTENT only** (no metadata — artefact copies
    never need mtimes), skip identical files, and **tolerate per-file failures**
    (log a warning and continue) so a single un-copyable file can never abort a
    whole-fleet run. Returns ``(summaries, total_failed)`` where ``summaries`` is a
    list of per-artefact summary strings and ``total_failed`` is the number of files
    that could not be copied (0 == every file copied or skipped-as-identical).
    """
    copied: list[str] = []
    total_failed = 0
    dir_names = list(_AUX_COPY_DIRS_STATIC)
    dir_names += sorted(p.name for p in src_dir.glob('raw_logger*') if p.is_dir())
    for name in dir_names:
        s = src_dir / name
        if not s.is_dir():
            continue
        nc, ns, nf = _copy_dir_content(s, dst_dir / name, overwrite)
        total_failed += nf
        summary = f'{name}/ ({nc} copied, {ns} skipped'
        summary += f', {nf} FAILED)' if nf else ')'
        copied.append(summary)
    for html in sorted(src_dir.glob('inspect_*.html')):
        d = dst_dir / html.name
        if not overwrite and _same_size(html, d):
            continue
        try:
            _safe_copy_file(html, d)
            copied.append(html.name)
        except OSError as exc:
            total_failed += 1
            logger.warning('copy-forward: skip %s (%s)', html, exc)
    if total_failed:
        logger.warning('copy-forward [%s]: %d file(s) could not be copied '
                       '(see warnings above)', dst_dir.name, total_failed)
    return copied, total_failed


def _count_files(d: Path) -> int:
    """Recursive count of regular files under ``d`` (0 if ``d`` does not exist)."""
    if not d.is_dir():
        return 0
    return sum(len(files) for _root, _dirs, files in os.walk(d))


def _verify_copy_forward(reg: str, src_dir: Path, dst_dir: Path,
                         n_failed: int) -> list[str]:
    """Assert the dst version dir ended up at least as complete as src for every
    carried-forward artefact, and surface any per-file copy failure.

    Completeness rule: for each artefact dir (``_AUX_COPY_DIRS_STATIC`` + every
    ``raw_logger*`` found in src) and for the ``inspect_*.html`` set, the recursive
    file count in dst must be **>=** that in src. dst holding *more* files than src
    is fine and expected — the data-collection-monitor appends new days into the
    current version. A shortfall (dst < src) means artefacts were lost in the
    migration, exactly the silent-loss failure this guard exists to catch.

    ``n_failed`` is the per-file copy-failure count from
    :func:`_copy_forward_artifacts`; > 0 is reported as its own problem because a
    failed copy can leave a truncated dst file that the count check would miss (the
    file exists) yet the artefact is still corrupt.

    Emits a ``logger.error`` per problem and returns the list of problem strings
    (empty == verified clean). The caller records these in the per-vehicle result
    dict so :func:`main` can aggregate them and force a non-zero exit.
    """
    problems: list[str] = []
    dir_names = list(_AUX_COPY_DIRS_STATIC)
    dir_names += sorted(p.name for p in src_dir.glob('raw_logger*') if p.is_dir())
    for name in dir_names:
        s = src_dir / name
        if not s.is_dir():
            continue
        n_src = _count_files(s)
        n_dst = _count_files(dst_dir / name)
        if n_dst < n_src:
            msg = f'{name}/ INCOMPLETE: dst {n_dst} file(s) < src {n_src}'
            problems.append(msg)
            logger.error('copy-verify [%s]: %s (%s -> %s)', reg, msg, s,
                         dst_dir / name)
    n_src_html = len(list(src_dir.glob('inspect_*.html')))
    n_dst_html = len(list(dst_dir.glob('inspect_*.html')))
    if n_dst_html < n_src_html:
        msg = f'inspect_*.html INCOMPLETE: dst {n_dst_html} < src {n_src_html}'
        problems.append(msg)
        logger.error('copy-verify [%s]: %s', reg, msg)
    if n_failed:
        msg = f'{n_failed} file(s) FAILED to copy (possible truncated dst files)'
        problems.append(msg)
        logger.error('copy-verify [%s]: %s', reg, msg)
    return problems


def recompute_vehicle(reg: str, db_root: Path, src_ver: str, dst_ver: str,
                      overwrite: bool = False) -> dict:
    """Recompute all meteorological-quarter reports for one EV, then carry the
    non-recomputable artefact dirs (raw_*/validation_figures/inspect HTML) forward."""
    src_dir = db_root / src_ver / reg
    dst_dir = db_root / dst_ver / reg
    dst_dir.mkdir(parents=True, exist_ok=True)
    csv_dir = src_dir / 'raw_telematics'
    logger.info('[%s] replaying segmentation from %s', reg, csv_dir)
    all_rows, cfg = replay_rows(reg, csv_dir)
    lookup = build_preserved_lookup(reg, src_dir)
    # target period list: met-quarters spanning the 2.2.6 report range
    srcs = sorted(src_dir.glob(f'jolt_report_{reg}_*.xlsx'))
    starts = [p.stem.split('_')[-2] for p in srcs]
    ends = [p.stem.split('_')[-1] for p in srcs]
    lo = min(starts)
    hi = max(ends)
    ds = f'{lo[:4]}-{lo[4:6]}-{lo[6:]}'
    de = f'{hi[:4]}-{hi[4:6]}-{hi[6:]}'
    periods = split_into_meteorological_quarters(ds, de)
    quarterly = {}
    made = []
    for qs, qe in periods:
        out_path = dst_dir / f'jolt_report_{reg}_{qs.replace("-", "")}_{qe.replace("-", "")}.xlsx'
        if out_path.exists() and not overwrite:
            logger.info('  [%s] exists, skip: %s', reg, out_path.name)
            continue
        n, pcap, pn, psrc = _recompute_report(reg, all_rows, cfg, qs, qe, lookup, out_path)
        if n == 0:
            logger.info('  [%s] %s..%s : no rows, skipped', reg, qs, qe)
            continue
        pk = f'{qs.replace("-", "")}_{qe.replace("-", "")}'
        if pcap is not None and pn:
            quarterly[pk] = {'kwh': round(float(pcap), 2), 'n': int(pn)}
        made.append(out_path.name)
        logger.info('  [%s] %s : %d rows, periodcap=%s n=%s src=%s',
                    reg, out_path.name, n,
                    None if pcap is None else round(pcap, 2), pn, psrc)
    aux, n_failed = _copy_forward_artifacts(src_dir, dst_dir, overwrite)
    if aux:
        logger.info('  [%s] carried forward %d artefact(s): %s',
                    reg, len(aux), ', '.join(aux))
    problems = _verify_copy_forward(reg, src_dir, dst_dir, n_failed)
    wavg, n_rel, n_sparse = _recompute_weighted_capacity(dict(quarterly))
    return {'reg': reg, 'reports': made, 'quarterly': quarterly,
            'rolled_up_kwh': wavg, 'n_reliable_q': n_rel, 'aux_copied': aux,
            'copy_problems': problems}


def copy_verbatim(reg: str, db_root: Path, src_ver: str, dst_ver: str,
                  overwrite: bool = False) -> dict:
    """dst == unchanged copy of src: copy every ``jolt_report_*.xlsx`` plus all
    non-recomputable artefact dirs (raw_*/validation_figures) and inspect HTML.

    Used for vehicles the SRF-free recompute **cannot** or **must not** replay:

    * **Diesel** (WU70GLV / YT21EFD) — no SOC/counter energy, so the SOC-fallback /
      capacity correction do not apply; the report is unchanged across the release.
    * **``prefer_logger_speed`` vehicles** (e.g. YN25RSY) — their segmentation runs
      on **SRF-Logger speed**, which the replay does NOT reproduce (it reads only
      ``raw_telematics``). Replaying from telematics speed diverges from the
      canonical segmentation (YN25RSY: 37→33 trips, ~15 % start-time match) and the
      start-time-keyed preserved-column overlay then loses weather etc. — so we copy
      the canonical report verbatim instead of producing an invalid replay.
    """
    src_dir = db_root / src_ver / reg
    dst_dir = db_root / dst_ver / reg
    dst_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for xlsx in sorted(src_dir.glob(f'jolt_report_{reg}_*.xlsx')):
        dst = dst_dir / xlsx.name
        if not overwrite and _same_size(xlsx, dst):
            continue
        # content-only copy — OneDrive rejects the copy2 metadata fast path (EINVAL)
        _safe_copy_file(xlsx, dst)
        copied.append(xlsx.name)
    # carry forward artefact dirs + inspect HTML.
    aux, n_failed = _copy_forward_artifacts(src_dir, dst_dir, overwrite)
    problems = _verify_copy_forward(reg, src_dir, dst_dir, n_failed)
    return {'reg': reg, 'copied': copied, 'aux_copied': aux,
            'copy_problems': problems}


# Backwards-compatible alias (the verbatim-copy path was originally diesel-only).
copy_diesel = copy_verbatim


# All 15 EV present in the report DB (the 3 Knowles AV24LX* were full-regen'd for
# 2.2.7 rather than recompute'd, so the original list omitted them; for the
# 2.2.7→2.2.8 recompute they must be covered too). + 2 diesel = 17 vehicles.
EV_FLEET = ['AV24LXJ', 'AV24LXK', 'AV24LXL', 'CMZ6260', 'EV73SAL', 'EX74JXW',
            'EX74JXY', 'KY24LHT', 'LN25NKE', 'N88GNW', 'T88RNW', 'TA70WTL',
            'YK73WFN', 'YN25RSY', 'YN75NMA']
DIESEL_FLEET = ['WU70GLV', 'YT21EFD']


def main(argv=None):
    ap = argparse.ArgumentParser(
        description='Fast SRF-free cached recompute for a post-segmentation release '
                    '(e.g. 2.2.7 -> 2.2.8). Reads src version dir, writes dst.')
    ap.add_argument('--db-root', default='./excel_report_database')
    ap.add_argument('--src-ver', default='2.2.7')
    ap.add_argument('--dst-ver', default='2.2.8')
    ap.add_argument('--veh', nargs='*', help='vehicles (default: 15 EV + 2 diesel)')
    ap.add_argument('--overwrite', action='store_true')
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    db_root = Path(args.db_root)
    regs = args.veh or (EV_FLEET + DIESEL_FLEET)
    results = []
    ok, failed = [], []
    # Per-vehicle failure isolation (mirrors the data-collection-monitor convention):
    # one vehicle's exception must NOT abort the fleet run — log an ERROR and carry
    # on. Only a total wipe-out (every vehicle failed) is a non-zero exit.
    for reg in regs:
        try:
            cfg = VEHICLE_CONFIG.get(reg, {})
            is_diesel = str(cfg.get('fuel_type', '')).upper() == 'DIESEL'
            # Vehicles whose segmentation runs on SRF-Logger speed cannot be replayed
            # from raw_telematics (logger speed is not cached), so their replay would
            # diverge from canonical — copy the canonical report verbatim instead.
            prefer_logger = bool(cfg.get('prefer_logger_speed'))
            if is_diesel or prefer_logger:
                res = copy_verbatim(reg, db_root, args.src_ver, args.dst_ver, args.overwrite)
                reason = ('DIESEL' if is_diesel else
                          'prefer_logger_speed (logger-speed segmentation cannot be '
                          'replayed from raw telematics)')
                logger.info('[%s] verbatim copy [%s]: %d reports', reg, reason,
                            len(res['copied']))
            else:
                res = recompute_vehicle(reg, db_root, args.src_ver, args.dst_ver, args.overwrite)
                logger.info('[%s] EV done: %d reports, rolled-up eff_cap=%s kWh',
                            reg, len(res['reports']), res.get('rolled_up_kwh'))
            results.append(res)
            ok.append(reg)
        except Exception as exc:  # noqa: BLE001 — isolate one vehicle's failure
            logger.error('[%s] FAILED: %s', reg, exc, exc_info=True)
            failed.append(reg)
    logger.info('=== recompute summary: %d OK, %d FAILED (of %d) ===',
                len(ok), len(failed), len(regs))
    if ok:
        logger.info('  OK: %s', ', '.join(ok))
    if failed:
        logger.warning('  FAILED: %s', ', '.join(failed))

    # Aggregate the per-vehicle copy-forward verification problems. A migration
    # that dropped artefacts (or failed a per-file copy) MUST NOT end green — this
    # is the guard against the silent 2.2.7 raw_charger/raw_logger loss recurring.
    copy_problems = [(res['reg'], p) for res in results
                     for p in res.get('copy_problems', [])]
    if copy_problems:
        n_veh = len({r for r, _ in copy_problems})
        logger.error('=== ARTEFACT COPY-FORWARD VERIFICATION FAILED: %d problem(s) '
                     'across %d vehicle(s) ===', len(copy_problems), n_veh)
        for reg, p in copy_problems:
            logger.error('  [%s] %s', reg, p)
        logger.error('The dst version dir is LESS complete than src — artefacts '
                     'were lost. Fix the copy and re-run before treating this '
                     'migration as done.')

    # Non-zero exit if EVERY vehicle failed (a partial run is still useful), or if
    # any artefact was lost / failed to copy (completeness is not optional).
    if failed and not ok:
        raise SystemExit(1)
    if copy_problems:
        raise SystemExit(2)
    return results


if __name__ == '__main__':
    main()
