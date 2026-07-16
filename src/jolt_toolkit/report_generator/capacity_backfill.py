"""
capacity_backfill.py
====================
不重跑报告，从既有 xlsx 报告库回填 vehicles.json 的
``effective_capacity_quarterly`` schema（v2.2.6+）。

对每辆电车（``fuel_type != "DIESEL"``）：

1. 扫描 ``<report_db>/<REG>/jolt_report_<REG>_<start>_<end>.xlsx``（按文件名严格
   匹配，自动排除 ``*_finetuned*`` 等非标准命名）。
2. 读 "Report" sheet，对每份报告复刻 :func:`_period_capacity_from_rows` 的
   donor-based 定义（充电优先：``Energy Source`` 非 ``soc_estimate`` 且
   ``SOC Change (%) > 0`` 的 leg 的 ``Battery Capacity (kWh)`` 均值；无则 discharge）
   算该周期 ``(kwh, n_donors)``。这与生成器 / 持久化路径**同一口径**，保证 backfill
   与重跑结果一致。
3. 把 ``(kwh, n)`` 写入 ``effective_capacity_quarterly[period_key]``，再用所有可靠
   季度（``n >= MIN_DONORS``）重算 donor 加权平均写回 ``effective_capacity_kwh``，
   稀疏季度 ``kwh`` 回填为该平均（见 :func:`_recompute_weighted_capacity`）。

柴油车整车跳过、不加任何 quarterly 字段。无 donor 的纯 soc_estimate 车（如 SOC-only
Mercedes）产生空 quarterly、不动既有标量。

用法（jolt env，从 repo 根运行）::

    PYTHONUTF8=1 D:/Anaconda/envs/jolt/python.exe \\
        -m jolt_toolkit.report_generator.capacity_backfill \\
        --report-db excel_report_database/2.2.6 [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import logging
import re
from pathlib import Path

from filelock import FileLock
from openpyxl import load_workbook

from jolt_toolkit.configs import get_config_path
from jolt_toolkit.report_generator._generator import (
    MIN_DONORS,
    _period_capacity_from_rows,
    _recompute_weighted_capacity,
)

# 仅匹配标准报告命名（结尾恰为 _<8digit>_<8digit>.xlsx）；finetuned / 其它后缀
# 自然不匹配而被跳过。
_REPORT_RE = re.compile(
    r'^jolt_report_(?P<reg>[A-Z0-9]+)_(?P<start>\d{8})_(?P<end>\d{8})\.xlsx$'
)


def _read_report_donor_capacity(xlsx_path: Path):
    """读单份报告 "Report" sheet，返回 ``(kwh|None, n_donors, source)``。

    用 ``data_only=True`` 读缓存值（Battery Capacity / SOC Change 是 xlsxwriter 写的
    纯数字，非 ``=NA()`` 公式，故安全）。复用 :func:`_period_capacity_from_rows`：把
    每行抽成 ``(cap, soc, src)`` 三元组传 idx 0/1/2，与 live 路径同一口径。
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb['Report'] if 'Report' in wb.sheetnames else wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return None, 0, 'fallback'
        try:
            i_cap = header.index('Battery Capacity (kWh)')
            i_soc = header.index('SOC Change (%)')
            i_src = header.index('Energy Source')
        except ValueError:
            # 柴油 / 旧列布局没有这些列
            return None, 0, 'fallback'
        triples = []
        for r in rows_iter:
            if r is None:
                continue
            cap = r[i_cap] if i_cap < len(r) else None
            soc = r[i_soc] if i_soc < len(r) else None
            src = r[i_src] if i_src < len(r) else None
            triples.append((cap, soc, src))
    finally:
        wb.close()
    return _period_capacity_from_rows(triples, 0, 1, 2)


def backfill_vehicle(reg: str, report_dir: Path, entry: dict,
                     min_donors: int = MIN_DONORS) -> dict:
    """回填单车 quarterly + 加权平均（就地改 ``entry``）。返回 summary dict。"""
    quarterly: dict = {}
    per_period = []  # [(period_key, kwh|None, n, source)]，含 fallback 周期供展示
    for xp in sorted(report_dir.glob(f'jolt_report_{reg}_*.xlsx')):
        m = _REPORT_RE.match(xp.name)
        if not m:
            continue  # 跳过 *_finetuned* 等非标准命名
        period_key = f"{m.group('start')}_{m.group('end')}"
        kwh, n, src = _read_report_donor_capacity(xp)
        per_period.append((period_key, kwh, n, src))
        if src == 'fallback' or kwh is None:
            continue  # 无 donor 的周期不入 quarterly
        quarterly[period_key] = {'kwh': round(float(kwh), 1), 'n': int(n)}

    summary = {
        'reg': reg,
        'per_period': per_period,
        'old_kwh': entry.get('effective_capacity_kwh'),
        'quarterly': quarterly,
        'new_kwh': entry.get('effective_capacity_kwh'),
        'n_reliable': 0,
        'n_sparse': 0,
        'wrote': False,
    }
    if not quarterly:
        return summary

    wavg, n_rel, n_sparse = _recompute_weighted_capacity(quarterly, min_donors)
    entry['effective_capacity_quarterly'] = quarterly
    if wavg is not None:
        entry['effective_capacity_kwh'] = wavg
    summary.update(new_kwh=entry.get('effective_capacity_kwh'),
                   n_reliable=n_rel, n_sparse=n_sparse, wrote=wavg is not None)
    return summary


def _print_summary(s: dict, min_donors: int) -> None:
    reg = s['reg']
    if not s['per_period']:
        print(f"[no reports]  {reg}")
        return
    arrow = f"{s['old_kwh']} -> {s['new_kwh']}"
    print(f"\n== {reg} ==  effective_capacity_kwh: {arrow}  "
          f"(reliable={s['n_reliable']}, sparse={s['n_sparse']})")
    for period_key, kwh, n, src in s['per_period']:
        if src == 'fallback' or kwh is None:
            print(f"    {period_key}  (no donor, src=fallback)  -- skipped")
            continue
        stored = s['quarterly'].get(period_key, {})
        tag = 'SPARSE->avg' if n < min_donors else 'reliable'
        print(f"    {period_key}  raw_kwh={round(float(kwh),1):>6}  n={n:>3}  "
              f"src={src:<9}  stored_kwh={stored.get('kwh'):>6}  [{tag}]")


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Backfill effective_capacity_quarterly from existing xlsx reports")
    ap.add_argument('--report-db', required=True,
                    help="report database version dir, e.g. excel_report_database/2.2.6")
    ap.add_argument('--min-donors', type=int, default=MIN_DONORS,
                    help=f"reliability threshold on donor count (default {MIN_DONORS})")
    ap.add_argument('--dry-run', action='store_true',
                    help="compute + print, but do NOT write vehicles.json")
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.WARNING, format='%(message)s')

    db = Path(args.report_db)
    path = get_config_path('vehicles.json')
    # Guard the read-modify-write against a concurrent report-gen capacity
    # write-back (same lock file as _persist_effective_capacity).
    with FileLock(str(path) + '.lock'):
        with open(path, encoding='utf-8') as f:
            all_cfg = json.load(f)

        summaries = []
        for reg, cfg in all_cfg.items():
            if str(cfg.get('fuel_type', '')).upper() == 'DIESEL':
                print(f"[skip diesel] {reg}")
                continue
            rdir = db / reg
            if not rdir.is_dir():
                print(f"[no reports]  {reg}")
                continue
            s = backfill_vehicle(reg, rdir, cfg, args.min_donors)
            summaries.append(s)
            _print_summary(s, args.min_donors)

        if args.dry_run:
            print("\n[dry-run] vehicles.json NOT written")
        else:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(all_cfg, f, indent=2, ensure_ascii=False)
                f.write('\n')
            print(f"\nvehicles.json written: {path}")
    return summaries


if __name__ == '__main__':
    main()
