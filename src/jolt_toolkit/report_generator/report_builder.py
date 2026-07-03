"""
report_builder.py
==================
т░є segment_algorithms УЙЊтЄ║уџётЁЁТћЙућхтѕєТ«хтГЌтЁИУйгТЇбСИ║ Excel ТіЦтЉіУАї№╝ї
т╣ХтєЎтЁЦТа╝т╝Јтїќ Excel ТќЄС╗Х + тЈ»жђЅ HTML жфїУ»ЂтЏЙТЪЦуюІтЎесђѓ

С╗ј segment_dev уџётјЪтДІт«ъуј░У┐ЂуД╗УђїТЮЦ№╝ї
СйюСИ║ jolt_toolkit.report_generator тїЁуџётєЁжЃеТеАтЮЌСй┐ућесђѓ
"""

from __future__ import annotations

import json
import re
import logging
from pathlib import Path
from datetime import date, timedelta, timezone
from urllib.parse import urlencode, urljoin
from math import ceil, nan

import numpy as np
import pandas as pd
import xlsxwriter

from srf_client import paging
from srf_client import filter as srf_filter
from geopy import Point as GeoPoint
from geopy.distance import Distance, geodesic

from jolt_toolkit.report_generator.segment_algorithms import (
    run_segment_detection,
    SOC_COL,
    TIME_COL,
    MOVING_SPEED_THRESHOLD_KMH,
    _ANCHOR_PRIVATE_KEYS,
    _agg_mass,
    VEHICLE_CONFIG,
)
from jolt_toolkit.report_generator.pedal_histogram import (
    compute_pedal_histogram,
    MIN_DISTANCE_FOR_PEDAL_KM,
    EEC2_COL,
    EBC1_COL,
)

logger = logging.getLogger(__name__)

# РћђРћђ Excel ТіЦтЉітѕЌтц┤№╝ѕућхтіеУйд№╝ЏСИјУђЂуЅѕ JOLT LegRecord тГЌТ«хжА║т║Јт«їтЁеСИђУЄ┤№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђРћђ
HEADERS = (
    "Leg Number",
    "Leg Type",
    "Telematics Link",
    "Charger Link",
    "SRF Logger Link",
    "Start Time (UTC)",
    "Origin (Lat, Lon)",
    "Origin Place",
    "End Time (UTC)",
    "Destination (Lat, Lon)",
    "Destination Place",
    "Duration (HH:MM:SS)",
    "Distance (km)",
    "Average Speed (km/h)",
    "Elevation Difference (m)",
    "Vehicle Mass (kg)",
    "Vehicle Mass CV (reliability)",
    "Recuperation Energy (kWh)",
    "Start SOC (%)",
    "End SOC (%)",
    "SOC Change (%)",
    "Energy Change (kWh)",
    "Energy Charged AC (kWh)",
    "Energy Charged DC (kWh)",
    "CO2 level (g/kWh)",
    "Cumulative Distance (km)",
    "CO2 for event (g)",
    "Cumulative CO2 (g)",
    "Battery Power (kW)",
    "Energy Performance (kWh/km)",
    "Energy Performance Corrected by Elevation Difference (kWh/km)",
    "Battery Capacity (kWh)",
    "Energy Output from Charger (kWh)",
    "Wire Energy Efficiency (kWh/kWh)",
    "Peak Charging (kW)",
    "Average Charging (kW)",
    "Energy based on motor power (kWh)",
    "Average Temperature (C)",
    "Average Pressure (hPa)",
    "Average Humidity (%)",
    "Average Wind Speed (m/s)",
    "Average Wind Direction",
    "Weather Type",
    "Histogram of Accelerator Pedal Position",
    "Histogram of Decelerator Pedal Position",
    # Тќ░уЅѕжбЮтцќтѕЌ
    "Energy Source",
    "Energy Performance Kinetics Corrected (kWh/km)",
    # v2.2.3 Тќ░тбъ№╝џtrip ТЌХТ«хтєЁуџёућхТю║жЕ▒тіеТђ╗УЃйжЄЈ№╝ѕkWh№╝Ѕ№╝їТЮЦУЄф telematics
    # `electric_energy_propulsion` у┤»У«АУ«АТЋ░тЎе№╝ѕWh№╝ЅТїЅТЌХжЌ┤уфЌТЈњтђ╝ти«тѕєсђѓ
    # СИј 'Energy Change (kWh)' тї║тѕФ№╝џpropulsion **СИЇТЅБ** тєЇућЪтѕХтіетЏъТћХ№╝їтЈфу╗ЪУ«А
    # ТГБтљЉжЕ▒тіе№╝ЏтИИућеС║јтЈЇу«Ќ ╬и_BMсђѓС╗Ё EV№╝ЏТЪ┤Т▓╣УйдУх░ DIESEL_HEADERS СИЇтљФТГцтѕЌсђѓ
    "Propulsion Energy (kWh)",
    # v2.2.4 Тќ░тбъ№╝џтј╗ТјЅУЙЁтіЕ/жЕ╗УйдУ┤ЪУйй№╝ѕHVACсђЂСйјтјІу│╗у╗ЪуГЅ№╝ЅтљјуџётЄђуЅхт╝ЋУЃйУђЌТЋѕујЄ
    # (kWh/km)сђѓт«џС╣ЅУДЂ data_analysis_workspace/energy_balance_check/report.md№╝џ
    #   EP_exclude_aux = (propulsion Рѕњ recuperation) / distance = EP Рѕњ auxiliary/distance
    # уГЅС╗иТјет»╝№╝џSRF ТЂњуГЅт╝Ј total = propulsion + auxiliary Рѕњ recuperation№╝їТћЙућх
    # trip уџё EP = |Energy Change| / dist РЅѕ total / dist№╝їТЋЁ
    #   EP Рѕњ aux/dist = (total Рѕњ aux)/dist = (propulsion Рѕњ recuperation)/distсђѓ
    # уЏ┤ТјЦуће propulsion СИј recuperation СИцСИф trip у║ДжЄЈ№╝їСИЇСЙЮУхќ aux уџёУДБу«Ќсђѓ
    # С╗ЁтйЊСИЅСИфУ«АТЋ░тЎе№╝ѕpropulsion / recuperation№╝ЅУ»Ц trip жЃйжЮъуЕ║ТЌХтЈ»у«Ќ№╝їтљдтѕЎ NaN
    # РђћРђћ У┐ЎтцЕуёХТіі EX74JXW / EX74JXY / YN25RSY / YN75NMA№╝ѕУ«АТЋ░тЎе NaN/у╝║тц▒№╝Ѕуй«уЕ║сђѓ
    # тіатюе HEADERS **ТюФт░Й**№╝џLoggerPatcher / WeatherPatcher уџёуАгу╝ќуаЂтѕЌу┤бт╝Ћ
    # №╝ѕtemp=38 / wind=41 / link=5 / mass=16 / kin=47№╝ЅтЮЄ РЅц 47№╝їСИЇтЈЌтй▒тЊЇсђѓ
    # ТЪ┤Т▓╣УйдУх░ DIESEL_HEADERS СИЇтљФТГцтѕЌсђѓ
    "EP_exclude_aux",
    # v2.2.5 Тќ░тбъ№╝џтЇЋУйдтЇЋ leg уџёУ┐љУљЦтЋєС╗БуаЂ№╝ѕproject operator CODE№╝ЅсђѓТЮЦТ║љу║ДУЂћУДЂ
    # report_generator/operators.py№╝ѕSRF СИ║СИ╗№╝џround-robin тЈќ leg.trip.trial.
    # description№╝їСИЊт▒ъУйдтЈќ vehicle.organisation.name№╝Џvehicles.json СИ║тЁют║Ћ№╝Ѕсђѓ
    # тіатюе HEADERS **ТюФт░Й**№╝їСИЇуД╗тіеС╗╗СйЋТЌбТюЅтѕЌу┤бт╝Ћ№╝ѕLoggerPatcher / WeatherPatcher
    # уџёуАгу╝ќуаЂтѕЌу┤бт╝ЋсђЂ_generator уџё _IDX_* тЮЄ РЅц 48№╝їСИЇтЈЌтй▒тЊЇ№╝Ѕсђѓ
    "Operator",
)

# РћђРћђ ТЪ┤Т▓╣УйдСИЊућетѕЌтц┤№╝ѕv2.2.2 ТЅЕт▒Ћ№╝џСИЇтєЇтцЇућеућхУйд HEADERS№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
# тЈфС┐ЮуЋЎт»╣ТЪ┤Т▓╣ТюЅуЅЕуљєТёЈС╣ЅуџётГЌТ«х№╝џСИЇтЄ║уј░ SOCсђЂAC/DC тЁЁућхУЃйжЄЈсђЂBattery CapacityсђЂ
# Energy Performance (kWh/km) уГЅСИјућхжЄЈуЏИтЁ│уџётѕЌсђѓуЄЃТ▓╣ТХѕУђЌуће L тњї L/100km УАеУЙЙсђѓ
DIESEL_HEADERS = (
    "Leg Number",
    "Leg Type",
    "SRF Logger Link",
    "Start Time (UTC)",
    "Origin (Lat, Lon)",
    "Origin Place",
    "End Time (UTC)",
    "Destination (Lat, Lon)",
    "Destination Place",
    "Duration (HH:MM:SS)",
    "Distance (km)",
    "Average Speed (km/h)",
    "Elevation Difference (m)",
    "Vehicle Mass (kg)",
    "Vehicle Mass CV (reliability)",
    "Cumulative Distance (km)",
    "Fuel Used (L)",
    "Fuel Consumption (L/100km)",
    "Average Temperature (C)",
    "Average Pressure (hPa)",
    "Average Humidity (%)",
    "Average Wind Speed (m/s)",
    "Average Wind Direction",
    "Weather Type",
    "Energy Source",
    # v2.2.5 Тќ░тбъ№╝џУ┐љУљЦтЋєС╗БуаЂ№╝їТЪ┤Т▓╣УйдСИјућхУйдтѕЌжЏєт»╣жйљсђѓтіатюе **ТюФт░Й**№╝ѕdiesel row
    # ТюФт░ЙУ┐йтіа№╝їжЋ┐т║дТќГУеђ len(row) == len(DIESEL_HEADERS) - 1 УЄфтіеУиЪжџЈ№╝Ѕсђѓ
    "Operator",
)

# =============================================================================
# Graphs worksheet chart specifications (single source of truth)
# =============================================================================
# One spec per scatter chart on the ``Graphs`` worksheet. Both rendering paths
# share this config so EV / Diesel reports look identical across every file:
#   - the generator (xlsxwriter, ``_write_excel_report``) reads it to set fixed
#     axis bounds and to point each series at a pre-filtered helper block;
#   - the in-place patch script (openpyxl) imports the very same list to rebuild
#     the ``Graphs`` sheet on existing 2.2.3 workbooks.
#
# Every axis is a FIXED constant (never per-file autoscaling), so the same chart
# type uses the exact same scale on all vehicles and all reports. Basis for each
# bound is documented inline:
#   - Energy Performance y-axis 0РђЊ3, Vehicle Mass x-axis 0РђЊ45000 and the ambient
#     Temperature x-axis Рѕњ5..30 follow the FIXED PDF-briefing / paper conventions
#     (PERF_YLIM / MASS_XLIM / TEMP_XLIM in generate_pdf_report + plot-figure
#     SKILL), so the Excel charts use the same scales the PDF briefing does.
#   - The remaining axes (speed, fuel consumption) were fixed from robust
#     fleet-wide percentiles over the entire 2.2.3 report set (1st/99th percentile
#     of driving-leg values, then rounded outwards to a clean tick), so ~99 % of
#     real data fits in-frame and the rare telematics spikes (e.g. speed 5299 km/h,
#     power 96 MW) are clipped rather than blowing up the scale.
#
# Each spec is a dict with:
#   x_hdr / y_hdr           Рђћ column headers (must exist in HEADERS/DIESEL_HEADERS).
#                             These are the DATA references: the series / categories
#                             and the filter logic both look the column up by this
#                             exact Report-sheet header, so it must never be renamed
#                             to a display alias.
#   x_title / y_title       Рђћ optional axis DISPLAY title, decoupled from the data
#                             header above. When absent the axis title falls back to
#                             x_hdr / y_hdr. Use this to show a friendlier / paper-
#                             consistent axis label (e.g. "Vehicle gross weight (kg)")
#                             while the series keeps pointing at the real column
#                             ("Vehicle Mass (kg)"). It changes ONLY the printed
#                             label Рђћ not the data, axis range or any other style.
#   x_min/x_max/x_major     Рђћ fixed x-axis bounds + major unit
#   y_min/y_max/y_major     Рђћ fixed y-axis bounds + major unit
#   x_filter / y_filter     Рђћ optional (lo, hi) inclusive value windows applied
#                             before a point is written to the helper block
#                             (None РЄњ no value window for that axis)
#   driving_only            Рђћ if True, charge / Stop rows are excluded (used for
#                             every Energy-Performance / Fuel-Consumption chart)
#   title                   Рђћ short chart title that occupies the top band only
#                             (kept distinct from the verbose "y vs x" so it does
#                             not crowd the plot / grid lines)
#   series_name             Рђћ concise legend label (NOT identical to the title) so
#                             the right-aligned legend stays compact
#
# Visual style is intentionally NOT per-spec: every chart shares ``CHART_STYLE``
# (single steelblue scatter colour + alpha, marker size, legend position, chart
# size, whether the fit equation / R┬▓ are shown). Both rendering paths
# (xlsxwriter generator + openpyxl patch) read these same fields, so the two can
# never drift apart.
#
# Filtering note: xlsxwriter / openpyxl charts reference cell ranges and cannot
# "filter" in place, so the generator and patcher both write the filtered
# (x, y) pairs to a hidden ``GraphsData`` block and point the scatter series AND
# the linear trendline at that clean block Рђћ keeping the fit consistent with the
# paperтЈБтЙё (fit computed on clean data only).

# Shared chart aesthetic for every Graphs chart, EV and diesel alike. Keeping a
# single source of truth here means the xlsxwriter generator and the openpyxl
# patch script render an identical look. Colours are RGB hex WITHOUT the leading
# '#': openpyxl wants the 6-char form, and xlsxwriter accepts '#RRGGBB' so the
# generator prepends '#'.
CHART_STYLE: dict = {
    "scatter_rgb": "1F77B4",  # steelblue Рђћ the muted single colour for all points
    "scatter_opacity": 72,  # marker fill OPACITY (%); soft like the paper.
    # openpyxl uses it directly as the DrawingML alpha
    # (alpha = opacity ├Ќ 1000); xlsxwriter wants the
    # complementary transparency (100 Рѕњ opacity).
    "marker_size": 6,  # points are slightly larger now the charts are bigger
    "legend_pos_xlsxwriter": "top_right",  # xlsxwriter legend position keyword
    "legend_pos_openpyxl": "tr",  # openpyxl LegendPosition (top-right)
    # Chart size Рђћ enlarged so the bigger fonts (below) have room and the axis
    # titles sit clear of the tick labels instead of overlapping them. BOTH render
    # paths now size the chart explicitly from ``chart_width_cm`` /
    # ``chart_height_cm`` (xlsxwriter via ``set_size``, openpyxl via ``chart.width``
    # / ``chart.height``) instead of the old xlsxwriter ``x_scale`` / ``y_scale``,
    # so the two paths render at the SAME physical size and one shared row step
    # (``chart_row_step``) places them identically without overlap.
    "chart_width_cm": 24.0,  # explicit chart width  (both paths)
    "chart_height_cm": 13.0,  # explicit chart height (both paths)
    # Vertical gap (in worksheet rows) left BELOW each chart before the next one
    # starts. The per-chart row step = ceil(chart height in rows) + this gap; see
    # ``chart_row_step``. A 13 cm chart is РЅѕ 24.6 default rows, so a gap of 2 rows
    # gives a clean ~2-row strip between charts. Bumping the chart height or font
    # sizes only requires re-deriving the step from these constants Рђћ never hand-
    # tuning per-path anchors again (the old 22-row xlsxwriter / 24-row openpyxl
    # anchors were smaller than the chart and made adjacent charts overlap).
    "chart_gap_rows": 2,
    # Plot-area manual layout Рђћ fractions of the CHART area, shared by both render
    # paths so the inner plot box sits in the same place everywhere. Without a
    # manual layout Excel auto-sizes the plot area as large as it can; at the
    # enlarged plot-figure font sizes (axis-title 14 / tick 12) that pushed the
    # rotated y-axis title on top of the y tick numbers and the x-axis title on
    # top of the x tick numbers. Reserving fixed margins keeps every axis title
    # OUTSIDE its tick labels: ``plot_x`` (left) clears the rotated y title + y
    # ticks, ``1 - plot_y - plot_h`` (bottom) clears the x title + x ticks,
    # ``plot_y`` (top) clears the two-line title band, ``1 - plot_x - plot_w``
    # (right) leaves a little room for the top-right legend. xlsxwriter
    # ``set_plotarea({'layout': {x, y, width, height}})`` and openpyxl
    # ``ManualLayout(xMode='edge', yMode='edge', x, y, w, h)`` use the same
    # edge-anchored fraction semantics, so one set of numbers drives both.
    "plot_x": 0.11,  # left margin: rotated y-axis title + y tick labels
    "plot_y": 0.17,  # top margin: two-line title band (title + fit subtitle)
    "plot_w": 0.85,  # plot width  (right margin 1РѕњxРѕњw РЅѕ 0.04 for the legend)
    "plot_h": 0.71,  # plot height (bottom margin 1РѕњyРѕњh РЅѕ 0.12 for x title+ticks)
    # Font sizes (pt) Рђћ aligned with the plot-figure skill Style constants so the
    # Excel charts read at the same scale as the published matplotlib figures.
    "title_font_size": 14,  # pt Рђћ chart title band (FS_TITLE)
    "subtitle_font_size": 10,  # pt Рђћ the fit-equation / R┬▓ subtitle line
    "axis_title_font_size": 14,  # pt Рђћ x / y axis titles (FS_LABEL)
    "tick_font_size": 12,  # pt Рђћ axis tick labels (FS_TICK)
    "legend_font_size": 9,  # pt Рђћ legend text (FS_LEGEND)
    # Major gridlines Рђћ light grey + high transparency so the plot area reads
    # cleanly (the paper uses GRID_ALPHA = 0.3). ``grid_rgb`` is the 6-char hex
    # WITHOUT '#'; ``grid_opacity`` is the line OPACITY (%) Рђћ openpyxl writes it as
    # the DrawingML alpha (alpha = opacity ├Ќ 1000) and xlsxwriter as the
    # complementary transparency (100 Рѕњ opacity). Minor gridlines are switched off.
    "grid_rgb": "D9D9D9",  # light grey
    "grid_opacity": 30,  # 30 % opacity РЅѕ paper GRID_ALPHA 0.3
    "grid_width_emu": 9525,  # 0.75 pt hairline (EMU); xlsxwriter uses pt below
    "show_fit_in_subtitle": True,  # put "y = ... , R┬▓ = ..." in a 2nd title line,
    # NOT as overlapping floating data labels
    # Graceful-degradation panel Рђћ when a chart has ZERO filtered data points (e.g.
    # a vehicle whose ``mass_col`` is null for the whole period, so the EP-vs-Mass
    # chart can plot nothing), the empty scatter frame is replaced by a centred
    # text note occupying the chart's footprint. The note is a merged-cell panel
    # (NOT an xlsxwriter textbox / openpyxl drawing) because merged cells are the
    # ONE primitive both render paths write identically Рђћ keeping the two paths in
    # lockstep. These keys style that panel; colours are 6-char hex WITHOUT '#'.
    "empty_note_font_size": 12,  # pt Рђћ note text size
    "empty_note_rgb": "808080",  # grey italic note text
    "empty_note_fill_rgb": "F2F2F2",  # very light grey panel background
    "empty_note_border_rgb": "BFBFBF",  # thin grey panel border
}

# Default Excel row height is 15 pt; 1 pt = 12 700 EMU, so one un-resized row is
# 190 500 EMU tall. A chart of ``chart_height_cm`` (cm Рєњ EMU via 360 000 EMU/cm)
# therefore spans ``chart_height_cm * 360000 / 190500`` rows. Both the Graphs
# sheets keep the default row height, so this constant converts the shared chart
# height into the worksheet-row span used to stack the charts. The matching
# column constant (default column РЅѕ 64 px = 609 600 EMU for Calibri 11 at width
# 8.43) converts the shared chart WIDTH into a worksheet-column span Рђћ used only
# to size the graceful-degradation note panel so it roughly fills the footprint a
# real chart would occupy.
_EMU_PER_CM = 360_000
_EMU_PER_DEFAULT_ROW = 15 * 12_700  # 15 pt default row height, in EMU
_EMU_PER_DEFAULT_COL = 64 * 9_525  # 64 px default column width, in EMU


def _chart_height_rows() -> int:
    """Worksheet-row span of one Graphs chart (its own height, excluding the gap)."""
    return ceil(CHART_STYLE["chart_height_cm"] * _EMU_PER_CM / _EMU_PER_DEFAULT_ROW)


def chart_col_span() -> int:
    """Worksheet-column span of one Graphs chart.

    Used only to size the graceful-degradation note panel (a merged-cell block)
    so it roughly matches the footprint a real chart would occupy. Derived from
    the shared ``chart_width_cm`` and the default column width, so both render
    paths size the panel identically.
    """
    return ceil(CHART_STYLE["chart_width_cm"] * _EMU_PER_CM / _EMU_PER_DEFAULT_COL)


def chart_row_step() -> int:
    """Return the worksheet-row step between successive Graphs charts.

    The step is ``ceil(chart height in default rows) + chart_gap_rows`` using the
    shared ``CHART_STYLE`` chart-size / gap constants, so the i-th chart is
    anchored at row ``i * chart_row_step()`` (0-based) and the next chart always
    starts in the clear strip BELOW it. Both render paths (the xlsxwriter
    generator here and the openpyxl ``patch_graphs_2_2_3.py``) import this single
    helper, so adjacent charts can never overlap and the two paths stay in
    lockstep. Returns a plain ``int`` (xlsxwriter / openpyxl anchors are 1-based
    cell references built from it).
    """
    return _chart_height_rows() + CHART_STYLE["chart_gap_rows"]


def empty_chart_note(spec: dict) -> str:
    """Text shown in place of a chart that has zero filtered data points.

    Returns the spec's ``empty_note`` (e.g. the EP/Fuel-vs-Mass charts say "No
    vehicle mass data available for this vehicle") or a generic fallback. Shared
    by both render paths so the degraded panel reads the same message everywhere.
    """
    return spec.get("empty_note", "No data available for this chart")


def empty_note_extent(gi: int) -> tuple[int, int, int, int]:
    """0-based ``(first_row, first_col, last_row, last_col)`` for chart ``gi``'s
    no-data note panel.

    The panel occupies the same row band and roughly the same width the real
    chart would (anchor ``gi * chart_row_step()``, height ``_chart_height_rows()``
    rows, width ``chart_col_span()`` columns), so replacing a chart with a note
    leaves the remaining charts' positions untouched. Both render paths build
    their merge range from this single helper.
    """
    first_row = gi * chart_row_step()
    last_row = first_row + _chart_height_rows() - 1
    first_col = 0
    last_col = chart_col_span() - 1
    return first_row, first_col, last_row, last_col


def _chart_subtitle(pts) -> str:
    """Build the "y = a┬иx + b   ┬и   R┬▓ = r" subtitle from filtered points.

    Returns an empty string when fewer than two points (a line cannot be fit).
    Computing the least-squares fit here lets us render the equation + R┬▓ as a
    single non-overlapping title sub-line instead of two floating labels that
    Excel stacks on top of each other over the data.
    """
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    n = len(xs)
    if n < 2:
        return ""
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    sxx = sum((x - mean_x) ** 2 for x in xs)
    if sxx == 0:
        return ""
    sxy = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    slope = sxy / sxx
    intercept = mean_y - slope * mean_x
    syy = sum((y - mean_y) ** 2 for y in ys)
    r2 = (sxy * sxy / (sxx * syy)) if syy > 0 else 0.0
    sign = "+" if intercept >= 0 else "Рѕњ"
    return f"y = {slope:.4g}x {sign} {abs(intercept):.4g}" f"    ┬и    R┬▓ = {r2:.4f}"


# EV charts. Only Energy-Performance-vs-{Mass, Temperature, Speed} are kept; the
# former Battery-Power-vs-Energy-Change and SOC-Change-vs-Energy-Change pairs were
# dropped (v2.2.4 visual pass) so EV and diesel report the same three-panel
# "{performance metric} vs {Mass / Temp / Speed}" set, matching the paperтЈБтЙё.
CHART_SPECS_EV: tuple[dict, ...] = (
    {
        "x_hdr": "Vehicle Mass (kg)",
        # Display the x-axis as "Vehicle gross weight (kg)" to match the paper
        # figures; the series still reads the real "Vehicle Mass (kg)" column.
        "x_title": "Vehicle gross weight (kg)",
        "y_hdr": "Energy Performance (kWh/km)",
        "title": "EP vs Vehicle Mass",
        "series_name": "EP",
        # Shown instead of an empty chart frame when this vehicle has no mass data
        # at all (e.g. its ``mass_col`` is null for the whole period).
        "empty_note": "No vehicle mass data available for this vehicle",
        # x: paper MASS_XLIM 0РђЊ45000 kg. y: paper PERF_YLIM 0РђЊ3 kWh/km.
        "x_min": 0,
        "x_max": 45000,
        "x_major": 5000,
        "y_min": 0,
        "y_max": 3,
        "y_major": 0.5,
        "x_filter": (1.0, None),
        "y_filter": (0.1, 3.0),
        "driving_only": True,
    },
    {
        "x_hdr": "Average Temperature (C)",
        "y_hdr": "Energy Performance (kWh/km)",
        "title": "EP vs Average Temperature (┬░C)",
        "series_name": "EP",
        "empty_note": "No temperature data available for this vehicle",
        # x: ambient temperature Рѕњ5..30 ┬░C Рђћ the fixed PDF-briefing convention
        # (generate_pdf_report TEMP_XLIM = (-5, 30)), so the Excel and PDF temp
        # axes match. Fleet temp p1РЅѕ-0.1 / p99РЅѕ26.9 (rare spikes above 30 ┬░C are
        # clipped to the frame, exactly as the PDF does via set_xlim).
        "x_min": -5,
        "x_max": 30,
        "x_major": 5,
        "y_min": 0,
        "y_max": 3,
        "y_major": 0.5,
        "x_filter": None,
        "y_filter": (0.1, 3.0),
        "driving_only": True,
    },
    {
        "x_hdr": "Average Speed (km/h)",
        "y_hdr": "Energy Performance (kWh/km)",
        "title": "EP vs Average Speed",
        "series_name": "EP",
        "empty_note": "No speed data available for this vehicle",
        # x: fleet speed p1РЅѕ3.8 / p99РЅѕ76.8 (max 5299 is a telematics spike) Рєњ 0..90.
        "x_min": 0,
        "x_max": 90,
        "x_major": 10,
        "y_min": 0,
        "y_max": 3,
        "y_major": 0.5,
        "x_filter": (0.0, 90.0),
        "y_filter": (0.1, 3.0),
        "driving_only": True,
    },
)

# Diesel charts. Fuel-Consumption-vs-{Mass, Temperature, Speed} only; the former
# Distance-vs-Fuel-Used pair was dropped to mirror the EV three-panel set.
CHART_SPECS_DIESEL: tuple[dict, ...] = (
    {
        "x_hdr": "Vehicle Mass (kg)",
        # Display the x-axis as "Vehicle gross weight (kg)" to match the paper
        # figures; the series still reads the real "Vehicle Mass (kg)" column.
        "x_title": "Vehicle gross weight (kg)",
        "y_hdr": "Fuel Consumption (L/100km)",
        "title": "Fuel cons. vs Vehicle Mass",
        "series_name": "Fuel cons.",
        "empty_note": "No vehicle mass data available for this vehicle",
        # x: paper MASS_XLIM 0РђЊ45000 kg (fleet mass p99РЅѕ44000).
        # y: fleet fuel-cons p1РЅѕ25 / p99РЅѕ50 (max 64.5) Рєњ 0..60.
        "x_min": 0,
        "x_max": 45000,
        "x_major": 5000,
        "y_min": 0,
        "y_max": 60,
        "y_major": 10,
        "x_filter": (1.0, None),
        "y_filter": (0.0, 60.0),
        "driving_only": True,
    },
    {
        "x_hdr": "Average Temperature (C)",
        "y_hdr": "Fuel Consumption (L/100km)",
        "title": "Fuel cons. vs Average Temperature (┬░C)",
        "series_name": "Fuel cons.",
        "empty_note": "No temperature data available for this vehicle",
        # x: ambient temperature Рѕњ5..30 ┬░C Рђћ shared with the EV temp axis and the
        # fixed PDF-briefing convention (TEMP_XLIM = (-5, 30)). Fleet temp
        # p1РЅѕ-0.8 / p99РЅѕ29.7; rare spikes above 30 ┬░C are clipped to the frame.
        "x_min": -5,
        "x_max": 30,
        "x_major": 5,
        "y_min": 0,
        "y_max": 60,
        "y_major": 10,
        "x_filter": None,
        "y_filter": (0.0, 60.0),
        "driving_only": True,
    },
    {
        "x_hdr": "Average Speed (km/h)",
        "y_hdr": "Fuel Consumption (L/100km)",
        "title": "Fuel cons. vs Average Speed",
        "series_name": "Fuel cons.",
        "empty_note": "No speed data available for this vehicle",
        # x: fleet speed p1РЅѕ7 / p99РЅѕ70.5 (max 81) Рєњ 0..90 (shared with EV speed axis).
        "x_min": 0,
        "x_max": 90,
        "x_major": 10,
        "y_min": 0,
        "y_max": 60,
        "y_major": 10,
        "x_filter": (0.0, 90.0),
        "y_filter": (0.0, 60.0),
        "driving_only": True,
    },
)


def chart_specs_for(headers: tuple) -> tuple[dict, ...]:
    """Return the chart-spec group matching a report's ``headers`` layout.

    Diesel reports are identified by the presence of the
    ``Fuel Consumption (L/100km)`` column (the same discriminator used elsewhere
    in this module); everything else is treated as an EV report.
    """
    return (
        CHART_SPECS_DIESEL
        if "Fuel Consumption (L/100km)" in headers
        else CHART_SPECS_EV
    )


_CHARGE_LEG_RE = re.compile(r"^(AC|DC|Charge|Mix|estimated)", re.IGNORECASE)


def _leg_is_charge(leg_type) -> bool:
    """True if a Leg Type string denotes a charge segment (red row)."""
    return bool(
        leg_type and isinstance(leg_type, str) and _CHARGE_LEG_RE.match(leg_type)
    )


def _leg_is_stop(leg_type) -> bool:
    """True if a Leg Type string denotes a Stop segment (white row)."""
    return bool(
        leg_type and isinstance(leg_type, str) and leg_type.strip().lower() == "stop"
    )


def is_trip_leg(leg_type) -> bool:
    """True if a Leg Type denotes a driving / trip segment.

    The single shared definition of a "trip" (driving) row: a non-blank Leg Type
    that is neither a charge segment (:func:`_leg_is_charge`) nor a Stop
    (:func:`_leg_is_stop`) Рђћ i.e. one of "In House" / "Round Trip" / "Outbound" /
    "Return" / "In Transit" (see :func:`_get_leg_type`). Public because the
    weather patchers import it: weather is backfilled on trip rows ONLY (charge
    and Stop rows do not need weather and would only waste OpenWeather quota), so
    the patchers and the chart ``driving_only`` filter agree on what counts as a
    driving row.
    """
    if not leg_type or not isinstance(leg_type, str) or not leg_type.strip():
        return False
    return not _leg_is_charge(leg_type) and not _leg_is_stop(leg_type)


def _filtered_chart_points(rows, col_idx_by_header: dict, spec: dict):
    """Return the cleaned ``(x, y)`` pairs for one chart spec.

    ``rows`` is the list of report-row tuples WITHOUT the leading Leg-Number
    column (i.e. ``row[0]`` is ``Leg Type``); ``col_idx_by_header`` maps a
    header name to its index within those tuples. Applies the spec's
    ``driving_only`` rule (drop charge / Stop rows), the per-axis value windows,
    and finally requires both x and y to be finite numbers. This is the single
    filtering routine shared by the xlsxwriter generator and the openpyxl patch
    script, so both render scatter + trendline on identical clean data.
    """
    xi = col_idx_by_header.get(spec["x_hdr"])
    yi = col_idx_by_header.get(spec["y_hdr"])
    if xi is None or yi is None:
        return []
    x_filter = spec.get("x_filter")
    y_filter = spec.get("y_filter")
    driving_only = spec.get("driving_only", False)

    def _in_window(val, window) -> bool:
        if window is None:
            return True
        lo, hi = window
        if lo is not None and val < lo:
            return False
        if hi is not None and val > hi:
            return False
        return True

    pts: list[tuple[float, float]] = []
    for row in rows:
        if not row:
            continue
        leg_type = row[0]
        if _leg_is_stop(leg_type):
            continue
        if driving_only and _leg_is_charge(leg_type):
            continue
        if xi >= len(row) or yi >= len(row):
            continue
        xv, yv = row[xi], row[yi]
        if not isinstance(xv, (int, float)) or not isinstance(yv, (int, float)):
            continue
        if _is_nan(xv) or _is_nan(yv):
            continue
        if not _in_window(xv, x_filter) or not _in_window(yv, y_filter):
            continue
        pts.append((float(xv), float(yv)))
    return pts


_WEIGHT_COL = "gross_combination_vehicle_weight"
# v2.2.4: ућхУйджЂЦТхІжђЪт║дтѕЌ№╝ѕkm/h№╝Ѕсђѓper-leg У┤ежЄЈтЮЄтђ╝/CV С╝ўтЁѕтЈфућеУАїжЕХСИГТаиТюг№╝ї
# ТјњжЎцжЮЎТГбТЌХСИЇтЈ»жЮауџё GCVW т╣┐ТњГ№╝ЏтѕЌу╝║тц▒ТЌХжђђтЏъТЌДуџётЁежЃе (> 0) ТаиТюгУАїСИ║сђѓ
_SPEED_COL = "wheel_based_speed"
_RECUP_COL = "electric_energy_recuperation_watthours"
# v2.2.3: у┤»У«АућхТю║жЕ▒тіеУЃйжЄЈУ«АТЋ░тЎе№╝ѕWh№╝їsince vehicle inception№╝Ѕсђѓtrip УхиТГбТЌХжЌ┤жђџУ┐Є
# у║┐ТђДТЈњтђ╝тюеТюђУ┐Љуџё RFMS т┐ФуЁДС╣ІжЌ┤тЙЌтѕ░ ╬ћ№╝їтєЇжЎцС╗Ц 1000 Уйг kWh№╝їтєЎтЁЦТќ░уџё
# `Propulsion Energy (kWh)` тѕЌсђѓТЪ┤Т▓╣УйдТ▓АТюЅТГцтѕЌсђѓ
_PROPULSION_COL = "electric_energy_propulsion"

# =============================================================================
# URL Тъёт╗║тиЦтЁи
# =============================================================================


def _ts_iso(t) -> str:
    """т░є pd.Timestamp Тѕќ datetime УйгТЇбСИ║ ISO тГЌугдСИ▓№╝ѕС┐ЮуЋЎТЌХтї║№╝Ѕсђѓ"""
    ts = pd.Timestamp(t)
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return str(ts)


def _build_telematics_url(leg_uri: str, t_start, t_end) -> str | None:
    """Тъёт╗║жЂЦТхІтЈ»УДєтїќ URLсђѓ"""
    if not leg_uri:
        return None
    query = urlencode(
        {
            "start": _ts_iso(t_start),
            "end": _ts_iso(t_end),
            "type": (96, "HSS2"),
            "resourceUri": leg_uri,
        },
        doseq=True,
    )
    return urljoin(leg_uri, "/explore/graphics/plots") + "?" + query


def _build_charger_url(charger_uri: str, t_start, t_end) -> str | None:
    """Тъёт╗║тЁЁућхТАЕтЈ»УДєтїќ URLсђѓ"""
    if not charger_uri:
        return None
    query = urlencode(
        {
            "start": _ts_iso(t_start),
            "end": _ts_iso(t_end),
            "resourceUri": charger_uri,
        }
    )
    return urljoin(charger_uri, "/explore/graphics/usage") + "?" + query


def _build_logger_url(logger_uri: str, t_start, t_end) -> str | None:
    """Тъёт╗║ SRF Logger тЈ»УДєтїќ URLсђѓ"""
    if not logger_uri:
        return None
    query = urlencode(
        {
            "start": _ts_iso(t_start),
            "end": _ts_iso(t_end),
            "resourceUri": logger_uri,
        }
    )
    return urljoin(logger_uri, "/explore/graphics/plots") + "?" + query


def _find_overlap(windows: list, t_start, t_end, tol_min: float = 5) -> str | None:
    """
    тюеТЌХжЌ┤уфЌтЈБтѕЌУАеСИГТЪЦТЅЙСИј [t_start, t_end] жЄЇтЈауџёуггСИђСИфТЮАуЏ«сђѓ

    windows: list of (start, end, uri)
    У┐ћтЏъ URI Тѕќ Noneсђѓ
    """
    tol = pd.Timedelta(minutes=tol_min)
    t_s = pd.Timestamp(t_start)
    t_e = pd.Timestamp(t_end)
    if t_s.tzinfo is None:
        t_s = t_s.tz_localize("UTC")
    if t_e.tzinfo is None:
        t_e = t_e.tz_localize("UTC")
    ext_s = t_s - tol
    ext_e = t_e + tol
    for entry in windows:
        try:
            ws, we, uri = entry[0], entry[1], entry[2]
            ws = pd.Timestamp(ws)
            we = pd.Timestamp(we)
            if ws.tzinfo is None:
                ws = ws.tz_localize("UTC")
            if we.tzinfo is None:
                we = we.tz_localize("UTC")
            if ws <= ext_e and we >= ext_s:
                return uri
        except Exception:
            continue
    return None


# =============================================================================
# жЂЦТхІТЋ░ТЇ«УЙЁтіЕтЄйТЋ░
# =============================================================================


def _to_utc_series(df: pd.DataFrame) -> pd.Series:
    """т░є TIME_COL тѕЌУйгСИ║ UTC tz-aware datetime Seriesсђѓ"""
    return pd.to_datetime(df[TIME_COL], errors="coerce", utc=True)


def _seg_mask(df: pd.DataFrame, t_start, t_end) -> pd.Series:
    """У┐ћтЏъУљйтюе [t_start, t_end] уфЌтЈБтєЁуџёУАїТјЕуаЂсђѓ"""
    ts = _to_utc_series(df)
    t_s = pd.Timestamp(t_start)
    t_e = pd.Timestamp(t_end)
    if t_s.tzinfo is None:
        t_s = t_s.tz_localize("UTC")
    if t_e.tzinfo is None:
        t_e = t_e.tz_localize("UTC")
    return (ts >= t_s) & (ts <= t_e)


def _get_vehicle_mass(
    df: pd.DataFrame,
    t_start,
    t_end,
    speed_col: str = _SPEED_COL,
    speed_threshold_kmh: float = MOVING_SPEED_THRESHOLD_KMH,
    method: str = "mean",
) -> tuple[float, float]:
    """У┐ћтЏъ (mass_kg, cv) Тѕќ (nan, nan)сђѓ

    v2.2.4: С╝ўтЁѕтЈфуће**УАїжЕХСИГ** (speed > ``speed_threshold_kmh``) уџёУ┤ежЄЈТаиТюгУ«Ау«Ќ
    leg тЮЄтђ╝/CV РђћРђћ жЮЎТГбТЌХуџё J1939 GCVW т╣┐ТњГ№╝ѕУБЁтЇИУ┤ДуъгТђЂ / ж╗ўУ«цтђ╝№╝ЅСИЇтЈ»жЮасђѓ
    тйЊуфЌтЈБтєЁТюЅ РЅЦ2 СИфУАїжЕХТаиТюгТЌХућеУАїжЕХТаиТюг№╝ЏтљдтѕЎтЏъжђђтѕ░ТЌДУАїСИ║№╝ѕуфЌтЈБтєЁтЁежЃе > 0 ТаиТюг№╝Ѕсђѓ
    УІЦжђЪт║дтѕЌу╝║тц▒№╝їУАїСИ║СИјТЌДуЅѕт«їтЁеСИђУЄ┤сђѓ

    v2.2.6: У┐ЄТ╗цтЈБтЙёСИЇтЈў№╝їСйєТюФТ«хУЂџтљѕТћ╣ућ▒ ``_agg_mass`` ТїЅ ``method`` т«їТѕљ
    №╝ѕ``mean`` / ``median`` / ``iqr_median`` / ``mad_mean`` / ``mad_tw_mean`` Рђд№╝Ѕ№╝ї
    СЙЏт╝ѓтИИжЄЇжЄЈт░ќт│░уџёуе│тЂЦС╝░У«Асђѓж╗ўУ«ц ``mean`` СИјТЌДт«ъуј░жђљтђ╝СИђУЄ┤№╝ѕ``sel.mean()`` /
    ``std/mean``№╝Ѕсђѓ``mad_tw_mean`` жбЮтцќжюђУдЂСИј ``sel`` т»╣жйљуџёТЌХжЌ┤Уй┤№╝їТЋЁС╗ЁтюеУ»ЦТќ╣Т│Ћ
    СИІС╗ј ``TIME_COL`` ТъёжђаТЌХжЌ┤Тѕ│С╝атЁЦ№╝ЏтЁХСйЎТќ╣Т│ЋУи»тЙёжђљтђ╝СИЇтЈўсђѓ
    """
    if _WEIGHT_COL not in df.columns:
        return nan, nan
    mask = _seg_mask(df, t_start, t_end)
    vals = pd.to_numeric(df.loc[mask, _WEIGHT_COL], errors="coerce")
    # У┐ЄТ╗ц J1939 default 0 тђ╝№╝ѕжЮЎТГбТЌХ broadcast ж╗ўУ«цтђ╝№╝їжЮъуюЪт«ъУ»╗ТЋ░№╝Ѕ
    # тЈѓУђЃ diesel_pipeline.py line ~549 т»╣ CVW тљїТаиСй┐уће m > 0 У┐ЄТ╗ц
    valid = vals.notna() & (vals > 0)

    # С╝ўтЁѕУАїжЕХСИГТаиТюг№╝ѕspeed > жўѕтђ╝№╝їNaN speed УДєСИ║жЮъУАїжЕХ№╝Ѕ
    if speed_col in df.columns:
        spd = pd.to_numeric(df.loc[mask, speed_col], errors="coerce")
        moving = valid & spd.notna() & (spd > speed_threshold_kmh)
        if moving.sum() >= 2:
            valid = moving

    sel = vals[valid]
    # ``mad_tw_mean`` needs a per-sample time axis aligned to ``sel``; build it
    # only for that method so every other method's path is byte-identical.
    timestamps = None
    if (method or "").lower() == "mad_tw_mean":
        timestamps = _to_utc_series(df).loc[sel.index]
    return _agg_mass(sel, method, timestamps=timestamps)


def _get_recuperation(df: pd.DataFrame, t_start, t_end) -> float:
    """У┐ћтЏътї║жЌ┤тєЁућхтѕХтіетЏъТћХУЃйжЄЈ№╝ѕkWh№╝Ѕсђѓу┤»У«АУ«АТЋ░тЎе№╝їтЈќ max - minсђѓ"""
    if _RECUP_COL not in df.columns:
        return nan
    mask = _seg_mask(df, t_start, t_end)
    vals = pd.to_numeric(df.loc[mask, _RECUP_COL], errors="coerce").dropna()
    if len(vals) < 2:
        return nan
    return round(float(vals.max() - vals.min()) / 1000.0, 3)


def _propulsion_at(
    t_target: pd.Timestamp, times: np.ndarray, values: np.ndarray
) -> float:
    """тюе `electric_energy_propulsion` у┤»У«АУ«АТЋ░тЎет║ЈтѕЌСИіт»╣ t_target тЂџу║┐ТђДТЈњтђ╝сђѓ

    times: 1D ns ТЌХжЌ┤Тѕ│ТЋ░у╗ё№╝ѕти▓Тјњт║Ј№╝Ѕ
    values: тљїжЋ┐т║дуџёу┤»У«А propulsion тђ╝ (Wh)
    У┐ћтЏъ t_target тцёуџёТЈњтђ╝ propulsion (Wh)№╝Џt_target УљйтюеТаиТюгУїЃтЏ┤С╣ІтцќТЌХтЈќТюђУ┐ЉуФ»уѓ╣сђѓ
    """
    t_ns = np.int64(pd.Timestamp(t_target).value)
    if t_ns <= times[0]:
        return float(values[0])
    if t_ns >= times[-1]:
        return float(values[-1])
    # np.interp т»╣тЇЋУ░Ѓжђњтбъ x тЂџу║┐ТђДТЈњтђ╝
    return float(np.interp(t_ns, times, values))


def _get_propulsion_energy(df: pd.DataFrame, t_start, t_end) -> float:
    """У«Ау«ЌСИђСИф trip ТЌХжЌ┤уфЌ [t_start, t_end] тєЁуџё propulsion тбъжЄЈ (kWh)сђѓ

    ТЋ░ТЇ«Т║љ№╝џraw telematics уџё `electric_energy_propulsion` тѕЌ№╝ѕWh№╝їу┤»У«АУ«АТЋ░тЎе№╝ї
    тЇЋУ░Ѓжђњтбъ№╝ЅсђѓRFMS т┐ФуЁДуеђуќЈ№╝ѕ~15 тѕєжњЪСИђТЮА№╝Ѕ№╝їТЅђС╗ЦСИђСИф trip тєЁтЈ»УЃйтЈфТюЅ 2РђЊ20 СИф
    жЄЄТаиуѓ╣№╝їжюђУдЂтюеуфЌтЈБУЙ╣уЋїтЂџу║┐ТђДТЈњтђ╝сђѓ

    у«ЌТ│Ћ№╝џ
    1. ТЈљтЈќТЋ┤ leg уџё (timestamp, propulsion) т║ЈтѕЌ№╝їтј╗жЄЇ + Тјњт║Јсђѓ
    2. С╝ўтЁѕ№╝џ[t_start, t_end] СИЦТа╝тїЁтљФтюеТаиТюгУїЃтЏ┤тєЁ Рєњ тюеСИцуФ»тљётЂџСИђТгАТЈњтђ╝ Рєњ ╬ћ
    3. fallback№╝џt_start тюеУїЃтЏ┤тєЁСйє t_end тюеУїЃтЏ┤тцќ№╝ѕТѕќтЈЇС╣І№╝ЅРєњ тЈќуфЌтЈБтєЁждќ/ТюФ
       ТаиТюгуџёти«№╝їтЇ│ partial-coverage С╝░У«Асђѓ
    4. жЃйСИЇтЈ»уће№╝ѕуфЌтЈБтюеТаиТюгУїЃтЏ┤тцќТѕќТаиТюг < 2№╝ЅРєњ NaNсђѓ

    У┐ћтЏъ ╬ћ propulsion (kWh)№╝ЏтѕЌу╝║тц▒ТѕќТаиТюгСИЇУХ│ТЌХУ┐ћтЏъ NaNсђѓ
    """
    if _PROPULSION_COL not in df.columns:
        return nan
    # тЄєтцЄТаиТюгт║ЈтѕЌ
    ts = _to_utc_series(df)
    vals = pd.to_numeric(df[_PROPULSION_COL], errors="coerce")
    keep = (~ts.isna()) & (~vals.isna())
    if keep.sum() < 1:
        return nan
    sub = pd.DataFrame({"t": ts[keep], "v": vals[keep]}).drop_duplicates("t")
    sub = sub.sort_values("t")
    if len(sub) < 1:
        return nan
    # Pandas 2.x уџё pd.to_datetime С╝џТа╣ТЇ«УЙЊтЁЦу▓Йт║дТјеТќГ dtype тЇЋСйЇ№╝ѕтдѓ 'datetime64[us, UTC]'№╝Ѕ№╝ї
    # ТГцТЌХ .asi8 / .astype('int64') У┐ћтЏъ microseconds УђїСИЇТў» nanoseconds№╝їСИј
    # pd.Timestamp.value№╝ѕТЂњСИ║ ns№╝ЅтЇЋСйЇСИЇтї╣жЁЇС╝џт»╝УЄ┤Т»ћУЙЃтЁежћЎсђѓт╝║тѕХ .as_unit('ns')сђѓ
    times = pd.DatetimeIndex(sub["t"]).as_unit("ns").asi8
    values = sub["v"].values

    # УДёУїЃтїќуфЌтЈБУЙ╣уЋїСИ║ UTC tz-aware
    t_s = pd.Timestamp(t_start)
    t_e = pd.Timestamp(t_end)
    if t_s.tzinfo is None:
        t_s = t_s.tz_localize("UTC")
    if t_e.tzinfo is None:
        t_e = t_e.tz_localize("UTC")
    t_s_ns = np.int64(t_s.value)
    t_e_ns = np.int64(t_e.value)

    s_min, s_max = times[0], times[-1]

    # РћђРћђ С╝ўтЁѕ№╝џbracketed ТЈњтђ╝№╝ѕуфЌтЈБт«їтЁетюеТаиТюгУїЃтЏ┤тєЁ№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    if t_s_ns >= s_min and t_e_ns <= s_max:
        v_s = _propulsion_at(t_s, times, values)
        v_e = _propulsion_at(t_e, times, values)
        delta_wh = v_e - v_s
        if delta_wh < 0:  # у┤»У«АУ«АТЋ░тЎеуљєУ«║СИітЇЋУ░Ѓжђњтбъ№╝їУ┤Ътђ╝УДєСИ║тЎфтБ░
            return nan
        return round(delta_wh / 1000.0, 3)

    # РћђРћђ fallback№╝џуфЌтЈБтєЁждќ/ТюФТаиТюгуџёти«№╝ѕpartial coverage№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђ
    in_window = (times >= t_s_ns) & (times <= t_e_ns)
    if in_window.sum() >= 2:
        idx = np.where(in_window)[0]
        delta_wh = float(values[idx[-1]] - values[idx[0]])
        if delta_wh < 0:
            return nan
        return round(delta_wh / 1000.0, 3)

    # РћђРћђ т«їтЁеТ▓АТюЅтЈ»ућеТаиТюг РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    return nan


def _ep_exclude_aux(
    propulsion_kwh: float, recuperation_kwh: float, distance_km: float
) -> float:
    """тЄђуЅхт╝Ћ№╝ѕтј╗УЙЁтіЕУ┤ЪУйй№╝ЅУЃйУђЌТЋѕујЄ (kWh/km)сђѓ

    т«џС╣ЅУДЂ ``data_analysis_workspace/energy_balance_check/report.md``№╝џ

        EP_exclude_aux = (propulsion Рѕњ recuperation) / distance

    тЇ│Т»ЈтЁгжЄїуџёуЅхт╝ЋУЃйжЄЈТЅБжЎцтєЇућЪтЏъТћХтљјсђЂтєЇтЅћжЎц HVAC/жЕ╗УйдуГЅУЙЁтіЕУ┤ЪУййуџётЄђтђ╝сђѓ
    propulsion / recuperation С╗╗СИђСИ║ NaN№╝ѕУ«АТЋ░тЎеу╝║тц▒ТѕќтЁеуЕ║№╝Ѕ№╝їТѕќ distance РЅц 0
    ТЌХУ┐ћтЏъ NaN РђћРђћ тЏаТГцтЈфТюЅтљїТЌХТіЦтЉі propulsion + recuperation СИцСИфУ«АТЋ░тЎеуџёУйд
    ТЅЇС╝џтЙЌтѕ░ТюЅТЋѕтђ╝№╝їСИј report.md уџётЈ»у«ЌУйджўЪСИђУЄ┤сђѓ
    """
    if any(
        v is None or np.isnan(v)
        for v in (propulsion_kwh, recuperation_kwh, distance_km)
    ):
        return nan
    if distance_km <= 0:
        return nan
    return round((propulsion_kwh - recuperation_kwh) / distance_km, 4)


def _get_elevation_diff(
    df: pd.DataFrame, t_start, t_end, altitude_col: str | None = None
) -> float:
    """У┐ћтЏътї║жЌ┤тєЁуџёТхиТІћти«№╝ѕу▒│№╝Ѕ№╝џТюФт░ЙжФўт║д - УхитДІжФўт║дсђѓ"""
    if altitude_col is None or altitude_col not in df.columns:
        return nan
    mask = _seg_mask(df, t_start, t_end)
    vals = pd.to_numeric(df.loc[mask, altitude_col], errors="coerce").dropna()
    if len(vals) < 2:
        return nan
    return round(float(vals.iloc[-1] - vals.iloc[0]), 1)


_G = 9.81  # m/s┬▓


def _corrected_energy_perf(
    energy_kwh: float, distance_km: float, elevation_m: float, mass_kg: float
) -> float:
    """
    ТхиТІћС┐«ТГБтљјуџёУЃйжЄЈТЋѕујЄ (kWh/km)сђѓ

    тЁгт╝Ј№╝џE_gravity = m * g * ╬ћh / 3,600,000 (kWh)
    СИітЮА (╬ћh > 0) ТЌХТЅБжЎцжЄЇтіЏтЂџтіЪ№╝їСИІтЮА (╬ћh < 0) ТЌХтіаСИітЏъТћХсђѓ
    corrected = (|delta_energy| - E_gravity) / distance
    """
    if any(np.isnan(v) for v in (energy_kwh, distance_km, elevation_m, mass_kg)):
        return nan
    if distance_km <= 0:
        return nan
    e_gravity_kwh = mass_kg * _G * elevation_m / 3_600_000.0
    corrected = abs(energy_kwh) - e_gravity_kwh
    return round(corrected / distance_km, 4)


ETA_DT = 0.90  # С╝атіеТЋѕујЄ ╬и№╝ѕућхТ▒аРєњУйдУй«№╝їтљФућхТю║ 0.95 ├Ќ тЈўжђЪтЎе 0.95№╝Ѕ
ETA_REGEN = 0.90  # тєЇућЪтѕХтіеТЋѕујЄ ╬и№╝ѕУйдУй«РєњућхТ▒а№╝їСИј ╬и_dt т»╣уД░№╝Ѕ
_V_MAX_KMH = 100.0  # GPS жђЪт║дСИіжЎљТѕфТќГ№╝ѕkm/h№╝Ѕ№╝їUK жЄЇтЇАжЎљжђЪ 56 mph РЅѕ 90 km/h


def _kinetics_corrected_energy_perf(
    energy_kwh: float,
    distance_km: float,
    elevation_m: float,
    mass_kg: float,
    speed_array_kmh,
) -> float:
    """
    ТхиТІћ + тіеУЃйС┐«ТГБтљјуџёУЃйжЄЈТЋѕујЄ (kWh/km)сђѓ

    тЪ║С║ј Sherborne (2024) Vehicle Model C№╝ѕPhD thesis, Section 2.2.4, Eq. 2.6РђЊ2.10№╝Ѕ№╝џ
    Сй┐уће Logger 1Hz жђЪт║дТЋ░ТЇ«жђљуДњУ«Ау«ЌтєЇућЪУ░ЃТЋ┤тіеУЃйтЈўтїќ ke№╝ї
    тєЇУйгТЇбСИ║ућхТ▒ау║ДУЃйжЄЈТХѕУђЌС╗ЦТаАТГБУАїжЕХтиЦтєхти«т╝ѓсђѓ

    жђљуДњтіеУЃйтЈўтїќ№╝џ
      ╬ћ_i = ┬й(v┬▓_{i+1} - v┬▓_i)            [J/kg]
    тєЇућЪУ░ЃТЋ┤ТЮЃжЄЇ№╝џ
      W_i = 1     if ╬ћ_i > 0№╝ѕтіажђЪ№╝Ѕ
      W_i = ╬и┬▓    if ╬ћ_i РЅц 0№╝ѕтЄЈжђЪ№╝ї╬и┬▓ СИ║тЙђУ┐ћТЋѕујЄ№╝Ѕ
    ућхТ▒ау║ДтЄђтіеУЃйТХѕУђЌ№╝џ
      E_KE_bat = (1/╬и) ├Ќ m ├Ќ ╬Б(╬ћ_i ├Ќ W_i)
              = ╬Б(╬ћKE>0)/╬и_dt - ╬и_regen ├Ќ ╬Б(|╬ћKE<0|)

    ╬и┬▓ уџёуЅЕуљєтљФС╣Ѕ№╝џтЄЈжђЪТЌХтіеУЃйу╗ЈућхТю║РєњжђєтЈўтЎеРєњућхТ▒атГўтѓе№╝ѕТЋѕујЄ ╬и№╝Ѕ№╝ї
    СИІСИђТгАтіажђЪТЌХтєЇу╗ЈућхТ▒аРєњжђєтЈўтЎеРєњућхТю║УйгтЏътіеУЃй№╝ѕТЋѕујЄ ╬и№╝Ѕ№╝ї
    тЙђУ┐ћТЋѕујЄ = ╬и ├Ќ ╬и = ╬и┬▓сђѓ

    Args:
        speed_array_kmh: numpy array№╝їLogger 1Hz жђЪт║дтђ╝ (km/h)
    """
    if any(np.isnan(v) for v in (energy_kwh, distance_km, mass_kg)):
        return nan
    if distance_km <= 0:
        return nan
    if speed_array_kmh is None or len(speed_array_kmh) < 2:
        return nan
    # ТхиТІћС┐«ТГБ
    e_gravity_kwh = 0.0
    if not np.isnan(elevation_m):
        e_gravity_kwh = mass_kg * _G * elevation_m / 3_600_000.0
    # GPS жђЪт║джбётцёуљє№╝џТѕфТќГт╝ѓтИИтђ╝ + 3 уѓ╣СИГСйЇТЋ░Т╗цТ│бТіЉтѕХт░ќтѕ║
    v_kmh = np.asarray(speed_array_kmh, dtype=float)
    v_kmh = np.clip(v_kmh, 0.0, _V_MAX_KMH)
    if len(v_kmh) >= 3:
        v_kmh = pd.Series(v_kmh).rolling(3, center=True, min_periods=1).median().values
    # жђљуДњтіеУЃйтЈўтїќ№╝ѕSherborne Eq. 2.6РђЊ2.9№╝Ѕ
    v_ms = v_kmh / 3.6  # km/h Рєњ m/s
    delta_ke_j = 0.5 * mass_kg * np.diff(v_ms**2)  # Joules
    accel_energy_j = delta_ke_j[delta_ke_j > 0].sum()
    braking_energy_j = np.abs(delta_ke_j[delta_ke_j < 0]).sum()
    # ућхТ▒ау║ДтЄђтіеУЃйТХѕУђЌ = тіажђЪУђЌУЃй/╬и_dt - ╬и_regen ├Ќ тѕХтіетЏъТћХ
    net_ke_kwh = (accel_energy_j / ETA_DT - ETA_REGEN * braking_energy_j) / 3_600_000.0
    # т«ЅтЁеТБђТЪЦ№╝џТаАТГБжЄЈСИЇт║ћУХЁУ┐Єт«ъжЎЁућхТ▒аУЃйжЄЈуџё 80%
    if abs(net_ke_kwh) > 0.8 * abs(energy_kwh):
        return nan
    corrected = abs(energy_kwh) - e_gravity_kwh - net_ke_kwh
    return round(corrected / distance_km, 4)


def _get_trip_speed_array(logger_speed_df, t_start, t_end):
    """ТЈљтЈќ Logger 1Hz жђЪт║дТЋ░у╗ё№╝ѕућеС║јтіеУЃйС┐«ТГБУ«Ау«Ќ№╝Ѕсђѓ

    Args:
        logger_speed_df: pd.DataFrame№╝їу┤бт╝ЋСИ║ UTC ТЌХжЌ┤Тѕ│№╝їтѕЌ 'logger_speed' (km/h)
        t_start, t_end: УАїуеІТЌХжЌ┤уфЌтЈБУЙ╣уЋї

    Returns:
        numpy array (km/h) Тѕќ None№╝ѕТЋ░ТЇ«СИЇУХ│ТЌХ№╝Ѕ
    """
    if logger_speed_df is None or logger_speed_df.empty:
        return None
    try:
        sliced = logger_speed_df.loc[t_start:t_end, "logger_speed"].dropna()
    except Exception:
        return None
    if len(sliced) < 2:
        return None
    return sliced.values  # numpy array№╝їтЇЋСйЇ km/h


def _point_str(lat, lon) -> str | None:
    """Та╝т╝ЈтїќСИ║ 'Point(lat lon)'№╝ѕСИјУђЂуЅѕ JOLT СИђУЄ┤№╝Ѕсђѓ"""
    if lat is None or lon is None:
        return None
    try:
        return f"Point({float(lat):.6f} {float(lon):.6f})"
    except Exception:
        return None


# =============================================================================
# Postcode reverse geocoding
# =============================================================================

_POSTCODE_CACHE_PATH = (
    Path(__file__).resolve().parent.parent.parent.parent
    / "cache"
    / "postcode_cache.json"
)


def _load_postcode_cache() -> dict:
    """С╗јуБЂуЏўтіаУйй postcode у╝ЊтГўсђѓ"""
    if _POSTCODE_CACHE_PATH.exists():
        try:
            with open(_POSTCODE_CACHE_PATH, "r") as f:
                raw = json.load(f)
            # JSON жћ«СИ║тГЌугдСИ▓тдѓ "(52.0368, -0.6572)"№╝їУйгтЏъ tuple
            return {
                tuple(map(float, k.strip("()").split(", "))): v for k, v in raw.items()
            }
        except Exception:
            return {}
    return {}


def _save_postcode_cache():
    """С┐ЮтГў postcode у╝ЊтГўтѕ░уБЂуЏўсђѓ"""
    try:
        _POSTCODE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        serializable = {str(k): v for k, v in _postcode_cache.items()}
        with open(_POSTCODE_CACHE_PATH, "w") as f:
            json.dump(serializable, f, indent=2)
    except Exception:
        pass


_postcode_cache: dict = _load_postcode_cache()


def _get_postcode(lat, lon, srf_data=None) -> str | None:
    """Reverse geocode (lat, lon) Рєњ UK postcode via SRF location API. Results cached."""
    if lat is None or lon is None:
        return None
    try:
        lat_f, lon_f = float(lat), float(lon)
    except (TypeError, ValueError):
        return None
    if lat_f == 0.0 and lon_f == 0.0:
        return None
    key = (round(lat_f, 4), round(lon_f, 4))
    if key in _postcode_cache:
        return _postcode_cache[key]
    if srf_data is None:
        return None
    postcode = None
    try:
        pt = GeoPoint(lat_f, lon_f)
        places = srf_data.locations.find_all(
            point=srf_filter.near(pt, Distance(meters=500))
        )
        if places.total == 1:
            postcode = getattr(places.items[0], "post_code", None)
        elif places.total > 1:
            dis = [(p, geodesic(pt, p.point)) for p in paging.paged_items(places)]
            closest, _ = min(dis, key=lambda pair: pair[1])
            postcode = getattr(closest, "post_code", None)
    except Exception:
        pass
    _postcode_cache[key] = postcode
    _save_postcode_cache()
    return postcode


# =============================================================================
# Home detection & leg-type classification
# =============================================================================

HOME_DETECTION_KM = 0.5
ROUND_TRIP_MIN_KM = (
    5.0  # trips starting AND ending at depot but longer than this Рєњ "Round Trip"
)


def _is_home(lat, lon, home_point) -> bool:
    """Return True if (lat, lon) is within HOME_DETECTION_KM of home_point."""
    if lat is None or lon is None or home_point is None:
        return False
    try:
        return (
            geodesic(home_point, GeoPoint(float(lat), float(lon))).km
            < HOME_DETECTION_KM
        )
    except Exception:
        return False


def _get_leg_type(mode: str, seg: dict, energy_ac, energy_dc, home_point) -> str:
    """
    Classify a segment into one of the old-JOLT leg types:
      Charge: "AC Home" / "DC Home" / "AC/DC Home" / "Charge Home"
              "AC Away" / "DC Away" / "AC/DC Away" / "Charge Away"
      Trip:   "In House" / "Round Trip" / "Outbound" / "Return" / "In Transit"
              (Round Trip = circular delivery starting AND ending at depot, dist > ROUND_TRIP_MIN_KM)
    """
    if mode == "charge":
        lat = seg.get("latitude")
        lon = seg.get("longitude")
        # тЁЁућхТ«х№╝џтЈфУдЂтЮљТаЄТјЦУ┐Љ home тЇ│тѕцт«џСИ║ Home
        place = "Home" if _is_home(lat, lon, home_point) else "Away"
        ac_ok = (
            abs(float(energy_ac)) > 1
            if (energy_ac is not None and not np.isnan(float(energy_ac)))
            else False
        )
        dc_ok = (
            abs(float(energy_dc)) > 1
            if (energy_dc is not None and not np.isnan(float(energy_dc)))
            else False
        )
        if ac_ok and dc_ok:
            return f"AC/DC {place}"
        elif ac_ok:
            return f"AC {place}"
        elif dc_ok:
            return f"DC {place}"
        else:
            return f"Charge {place}"
    else:  # discharge / trip
        is_home_s = _is_home(seg.get("lat_start"), seg.get("lon_start"), home_point)
        is_home_e = _is_home(seg.get("lat_end"), seg.get("lon_end"), home_point)
        if is_home_s and is_home_e:
            # Distinguish depot-to-depot circular delivery from a short local move
            try:
                odo_s = seg.get("odo_start_km")
                odo_e = seg.get("odo_end_km")
                if odo_s is not None and odo_e is not None:
                    dist = float(odo_e) - float(odo_s)
                    if dist > ROUND_TRIP_MIN_KM:
                        return "Round Trip"
            except (TypeError, ValueError):
                pass
            return "In House"
        elif is_home_s:
            return "Outbound"
        elif is_home_e:
            return "Return"
        else:
            return "In Transit"


# =============================================================================
# Stop leg type Рђћ gap-fill between consecutive trip/charging rows
# =============================================================================

# Minimum gap (seconds) between two segments that qualifies as a Stop.
# Anything shorter is assumed to be a segment-algorithm boundary artefact.
STOP_MIN_GAP_SECONDS = 60.0


def _row_col_index(col_name: str, headers: tuple = HEADERS) -> int:
    """Return the 0-based index of a column inside the row tuple.

    ``headers[0]`` is 'Leg Number' which is written separately by the Excel
    writer and does **not** appear in the row tuple Рђћ so the row tuple index is
    ``headers.index(col_name) - 1``. Pass ``DIESEL_HEADERS`` to get the layout
    used by diesel rows.
    """
    return headers.index(col_name) - 1


def _stop_row_from_neighbours(
    prev_row: list, next_row: list, headers: tuple = HEADERS
) -> list:
    """Build a Stop row from the two surrounding (trip/charge) rows.

    The stop spans the gap between ``prev_row``'s End Time and ``next_row``'s
    Start Time. We pull the vehicle's location from the end of ``prev_row``
    (= start of ``next_row``, modulo noise), the mass from ``prev_row``'s
    end value, and the SOC endpoints from both neighbours to expose any
    auxiliary/standby drain during the stop.

    ``headers`` chooses the row layout: pass ``DIESEL_HEADERS`` for diesel rows;
    diesel Stop rows skip SOC bookkeeping since diesel rows have no SOC columns.
    """
    row = [nan] * (len(headers) - 1)

    def _i(h):
        return _row_col_index(h, headers)

    def _has(h):
        return h in headers

    def _get(r, h):
        if not _has(h):
            return nan
        try:
            return r[_i(h)]
        except (IndexError, KeyError):
            return nan

    # РћђРћђ Core identity РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    row[_i("Leg Type")] = "Stop"
    for link_col in ("Telematics Link", "Charger Link", "SRF Logger Link"):
        if _has(link_col):
            row[_i(link_col)] = None

    # РћђРћђ Time window РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    t_start = _get(prev_row, "End Time (UTC)")
    t_end = _get(next_row, "Start Time (UTC)")
    row[_i("Start Time (UTC)")] = t_start
    row[_i("End Time (UTC)")] = t_end

    try:
        ts_start = pd.Timestamp(t_start)
        ts_end = pd.Timestamp(t_end)
        if ts_start.tzinfo is None:
            ts_start = ts_start.tz_localize("UTC")
        if ts_end.tzinfo is None:
            ts_end = ts_end.tz_localize("UTC")
        duration_days = max((ts_end - ts_start).total_seconds(), 0) / 86400.0
    except Exception:
        duration_days = nan
    row[_i("Duration (HH:MM:SS)")] = duration_days

    # РћђРћђ Location РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # At a Stop, origin and destination are the same vehicle location Рђћ
    # take the previous row's destination as the canonical stop location.
    stop_location = _get(prev_row, "Destination (Lat, Lon)")
    stop_place = _get(prev_row, "Destination Place")
    # Fallback to next row's origin if previous destination is missing
    if not stop_location or (
        isinstance(stop_location, float) and np.isnan(stop_location)
    ):
        stop_location = _get(next_row, "Origin (Lat, Lon)")
        stop_place = _get(next_row, "Origin Place")
    row[_i("Origin (Lat, Lon)")] = stop_location
    row[_i("Origin Place")] = stop_place
    row[_i("Destination (Lat, Lon)")] = stop_location
    row[_i("Destination Place")] = stop_place

    # РћђРћђ Motion fields РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # The vehicle doesn't move during a Stop
    row[_i("Distance (km)")] = 0.0
    row[_i("Average Speed (km/h)")] = 0.0
    if _has("Elevation Difference (m)"):
        row[_i("Elevation Difference (m)")] = 0.0

    # РћђРћђ Mass (carry over from previous leg's value) РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    row[_i("Vehicle Mass (kg)")] = _get(prev_row, "Vehicle Mass (kg)")
    if _has("Vehicle Mass CV (reliability)"):
        row[_i("Vehicle Mass CV (reliability)")] = _get(
            prev_row, "Vehicle Mass CV (reliability)"
        )

    # РћђРћђ Cumulative Distance (carry over so the xlsx shows running total) Рћђ
    if _has("Cumulative Distance (km)"):
        row[_i("Cumulative Distance (km)")] = _get(prev_row, "Cumulative Distance (km)")

    # РћђРћђ SOC endpoints Рђћ expose any auxiliary drain during the stop РћђРћђРћђРћђРћђРћђРћђ
    # Diesel rows have no SOC columns, so this block is a no-op there.
    if _has("Start SOC (%)") and _has("End SOC (%)"):
        row[_i("Start SOC (%)")] = _get(prev_row, "End SOC (%)")
        row[_i("End SOC (%)")] = _get(next_row, "Start SOC (%)")
        if _has("SOC Change (%)"):
            try:
                start_soc = float(row[_i("Start SOC (%)")])
                end_soc = float(row[_i("End SOC (%)")])
                row[_i("SOC Change (%)")] = end_soc - start_soc
            except (TypeError, ValueError):
                pass  # leave as nan

    # РћђРћђ Operator Рђћ carry from the neighbouring leg (same vehicle, same stop) РћђРћђ
    if _has("Operator"):
        op = _get(prev_row, "Operator")
        if op is None or (isinstance(op, float) and np.isnan(op)):
            op = _get(next_row, "Operator")
        row[_i("Operator")] = op

    return row


def _insert_stop_rows(
    sorted_rows: list,
    min_gap_seconds: float = STOP_MIN_GAP_SECONDS,
    headers: tuple = HEADERS,
) -> list:
    """Walk sorted rows and insert Stop rows in every non-trivial gap.

    A Stop is added whenever ``next.start_time - prev.end_time > min_gap_seconds``.
    The input list is not mutated; a new list is returned.
    """
    if not sorted_rows:
        return sorted_rows

    result: list = []
    end_col = _row_col_index("End Time (UTC)", headers)
    start_col = _row_col_index("Start Time (UTC)", headers)

    for i, row in enumerate(sorted_rows):
        result.append(row)
        if i == len(sorted_rows) - 1:
            break
        prev_end = row[end_col]
        next_start = sorted_rows[i + 1][start_col]
        try:
            ts_prev = pd.Timestamp(prev_end)
            ts_next = pd.Timestamp(next_start)
            if ts_prev.tzinfo is None:
                ts_prev = ts_prev.tz_localize("UTC")
            if ts_next.tzinfo is None:
                ts_next = ts_next.tz_localize("UTC")
            gap_s = (ts_next - ts_prev).total_seconds()
        except Exception:
            continue
        if gap_s > min_gap_seconds:
            result.append(
                _stop_row_from_neighbours(row, sorted_rows[i + 1], headers=headers)
            )
    return result


# =============================================================================
# Segment Рєњ Excel УАїТъёт╗║
# =============================================================================


def _seg_to_row(
    seg: dict,
    mode: str,
    leg_uri: str,
    charger_windows: list,
    logger_windows: list,
    df_leg: pd.DataFrame,
    cumulative_km: float,
    home_point=None,
    srf_data=None,
    altitude_col: str | None = None,
    speed_col: str = "wheel_based_speed",
    has_logger: bool = False,
    logger_speed_all: pd.DataFrame | None = None,
    logger_acc_pedal_all: pd.DataFrame | None = None,
    logger_dec_pedal_all: pd.DataFrame | None = None,
    operator: str | None = None,
    mass_agg: str = "mean",
) -> tuple:
    """
    т░єСИђСИф segment dict УйгТЇбСИ║СИђУАї Excel ТЋ░ТЇ«№╝ѕHEADERS жА║т║Ј№╝Ѕсђѓ

    mode: 'charge' | 'discharge'
    logger_speed_all:      Logger 1Hz жђЪт║д DataFrame№╝ѕу┤бт╝ЋСИ║ UTC ТЌХжЌ┤Тѕ│№╝їтѕЌ 'logger_speed'№╝Ѕ№╝ї
                           ућеС║јжђљуДњтіеУЃйС┐«ТГБУ«Ау«Ќсђѓ
    logger_acc_pedal_all:  Logger EEC2 Т▓╣жЌеУИЈТЮ┐СйЇуй« DataFrame№╝ѕу┤бт╝ЋСИ║ UTC ТЌХжЌ┤Тѕ│№╝Ѕсђѓ
    logger_dec_pedal_all:  Logger EBC1 тѕХтіеУИЈТЮ┐СйЇуй« DataFrame№╝ѕу┤бт╝ЋСИ║ UTC ТЌХжЌ┤Тѕ│№╝Ѕсђѓ
    """
    t_s = pd.Timestamp(seg["start_time"])
    t_e = pd.Timestamp(seg["end_time"])
    # у╗ЪСИђТЌХтї║№╝џуА«С┐ЮСИцУђЁжЃйТў» UTC ТѕќжЃйТў» tz-naive
    if t_s.tzinfo is not None and t_e.tzinfo is None:
        t_e = t_e.tz_localize("UTC")
    elif t_s.tzinfo is None and t_e.tzinfo is not None:
        t_s = t_s.tz_localize("UTC")
    duration = t_e - t_s

    # РћђРћђ URLs РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    telem_url = _build_telematics_url(leg_uri, t_s, t_e)
    # Logger Link С╗ЁућеС║јТћЙућх/УАїуеІТ«х№╝ЏтЁЁућхТ«хСИЇт║ћТюЅ Logger жЊЙТјЦ
    logger_url = None
    if mode == "discharge":
        logger_uri = _find_overlap(logger_windows, t_s, t_e, tol_min=5)
        logger_url = _build_logger_url(logger_uri, t_s, t_e) if logger_uri else None
    charger_url = None
    charger_energy_kwh = nan
    if mode == "charge":
        charger_uri = _find_overlap(charger_windows, t_s, t_e, tol_min=4)
        charger_url = _build_charger_url(charger_uri, t_s, t_e) if charger_uri else None
        # ТЈљтЈќтї╣жЁЇуџётЁЁућхТАЕУЃйжЄЈТЋ░ТЇ«
        if charger_uri:
            for entry in charger_windows:
                if len(entry) >= 4 and entry[2] == charger_uri and entry[3] is not None:
                    charger_energy_kwh = round(entry[3], 3)
                    break

    # РћђРћђ Leg type РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    energy_ac_delta = (
        (float(seg.get("ac_end_wh", 0) or 0) - float(seg.get("ac_start_wh", 0) or 0))
        / 1000.0
        if (seg.get("ac_end_wh") is not None and seg.get("ac_start_wh") is not None)
        else float("nan")
    )
    energy_dc_delta = (
        (float(seg.get("dc_end_wh", 0) or 0) - float(seg.get("dc_start_wh", 0) or 0))
        / 1000.0
        if (seg.get("dc_end_wh") is not None and seg.get("dc_start_wh") is not None)
        else float("nan")
    )
    leg_type = _get_leg_type(mode, seg, energy_ac_delta, energy_dc_delta, home_point)

    # РћђРћђ Location РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    if mode == "charge":
        lat = seg.get("latitude")
        lon = seg.get("longitude")
        origin = _point_str(lat, lon)
        destination = origin
        origin_pc = _get_postcode(lat, lon, srf_data)
        dest_pc = origin_pc
    else:
        origin = _point_str(seg.get("lat_start"), seg.get("lon_start"))
        destination = _point_str(seg.get("lat_end"), seg.get("lon_end"))
        origin_pc = _get_postcode(seg.get("lat_start"), seg.get("lon_start"), srf_data)
        dest_pc = _get_postcode(seg.get("lat_end"), seg.get("lon_end"), srf_data)

    # РћђРћђ Distance & speed РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    odo_s = seg.get("odo_start_km")
    odo_e = seg.get("odo_end_km")
    distance = nan
    if odo_s is not None and odo_e is not None:
        d = float(odo_e) - float(odo_s)
        if d > 0:
            distance = round(d, 3)

    dur_h = duration.total_seconds() / 3600.0
    # РћђРћђ Average Speed РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # ж╗ўУ«ц№╝ѕfirst_motion anchor№╝Ѕ№╝џdistance / уФ»уѓ╣ти«
    # zero_speed anchor ТеАт╝ЈСИІ№╝їуФ»уѓ╣ти▓УбФтцќТЅЕтѕ░жЏХжђЪТаиТюг№╝їtrip уфЌтЈБтљФжЏХжђЪт░Йти┤ Рєњ
    # тѕєТ»ЇТћ╣уће trip тєЁ v > speed_threshold тГљтї║жЌ┤уџёу┤»У«АТЌХжЋ┐№╝їС┐ЮТїЂжђЪт║дуЅЕуљєТёЈС╣Ѕсђѓ
    # ућ▒ find_discharge_segments_by_speed() тюе zero_speed ТеАт╝ЈтєЎтЁЦ seg['motion_duration_s']№╝Џ
    # first_motion ТеАт╝Ј / charge Т«х У»ЦтГЌТ«хСИ║ None ТѕќСИЇтГўтюесђѓ
    _motion_s = seg.get("motion_duration_s") if mode == "discharge" else None
    if _motion_s is not None and _motion_s > 0 and not np.isnan(distance):
        avg_speed = round(distance / (_motion_s / 3600.0), 2)
    else:
        avg_speed = (
            round(distance / dur_h, 2)
            if (not np.isnan(distance) and dur_h > 0)
            else nan
        )

    # РћђРћђ Vehicle mass РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    veh_mass, veh_mass_cv = _get_vehicle_mass(df_leg, t_s, t_e, method=mass_agg)
    recuperation = _get_recuperation(df_leg, t_s, t_e)
    elevation_diff = _get_elevation_diff(df_leg, t_s, t_e, altitude_col)

    # РћђРћђ Propulsion energy (kWh) Рђћ С╗Ё trip Т«х№╝ѕv2.2.3№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # тЁЁућхТ«х / жЮЎТГбТ«х propulsion т┐ЁСИ║ 0№╝їтєЎ NaN СИјуј░ТюЅ EP тѕЌтюе charge/stop
    # УАїуџётцёуљєСИђУЄ┤№╝ЏтЈфт»╣ discharge У«Ау«ЌТЈњтђ╝ти«тѕєсђѓ
    propulsion_kwh = nan
    if mode == "discharge":
        propulsion_kwh = _get_propulsion_energy(df_leg, t_s, t_e)

    # РћђРћђ SOC / energy РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    start_soc = seg.get("start_soc", nan)
    end_soc = seg.get("end_soc", nan)
    soc_change = seg.get("delta_soc_pct", nan)  # signed
    energy_change = seg.get("delta_energy_kwh", nan)  # signed
    battery_cap = seg.get("effective_capacity_kwh", nan)
    energy_source = seg.get("energy_source", None)

    # РћђРћђ Charge-specific РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    energy_ac = nan
    energy_dc = nan
    battery_power = nan
    if mode == "charge":
        energy_ac = round(energy_ac_delta, 3) if not np.isnan(energy_ac_delta) else nan
        energy_dc = round(energy_dc_delta, 3) if not np.isnan(energy_dc_delta) else nan
        if not np.isnan(energy_change) and dur_h > 0:
            battery_power = round(energy_change / dur_h, 3)

    # РћђРћђ Trip-specific РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    energy_perf = nan
    energy_perf_corrected = nan
    energy_perf_kinetics = nan
    ep_exclude_aux = nan
    if mode == "discharge" and not np.isnan(distance) and distance > 0:
        energy_perf = round(abs(energy_change) / distance, 4)
        energy_perf_corrected = _corrected_energy_perf(
            energy_change, distance, elevation_diff, veh_mass
        )
        # тј╗УЙЁтіЕУ┤ЪУййтЄђуЅхт╝ЋТЋѕујЄ№╝џ(propulsion Рѕњ recuperation) / distanceсђѓ
        # propulsion / recuperation С╗╗СИђ NaN№╝ѕУ«АТЋ░тЎеу╝║тц▒№╝ЅРєњ NaNсђѓ
        ep_exclude_aux = _ep_exclude_aux(propulsion_kwh, recuperation, distance)
        if logger_speed_all is not None:
            speed_arr = _get_trip_speed_array(logger_speed_all, t_s, t_e)
            if speed_arr is not None:
                energy_perf_kinetics = _kinetics_corrected_energy_perf(
                    energy_change, distance, elevation_diff, veh_mass, speed_arr
                )

    # РћђРћђ УИЈТЮ┐СйЇуй«уЏ┤Тќ╣тЏЙ№╝ѕС╗ЁТћЙућхТ«х№╝їУиЮуд╗ > 10 km№╝ЅРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    acc_hist = None
    dec_hist = None
    if (
        mode == "discharge"
        and not np.isnan(distance)
        and distance > MIN_DISTANCE_FOR_PEDAL_KM
    ):
        if logger_acc_pedal_all is not None:
            try:
                acc_slice = logger_acc_pedal_all.loc[t_s:t_e].dropna()
                if len(acc_slice) > 0:
                    acc_hist = compute_pedal_histogram(acc_slice, value_col=EEC2_COL)
            except Exception:
                pass
        if logger_dec_pedal_all is not None:
            try:
                dec_slice = logger_dec_pedal_all.loc[t_s:t_e].dropna()
                if len(dec_slice) > 0:
                    dec_hist = compute_pedal_histogram(dec_slice, value_col=EBC1_COL)
            except Exception:
                pass

    # РћђРћђ Cumulative distance РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    if mode == "discharge" and not np.isnan(distance):
        cumulative_km += distance
    cumulative_km_out = cumulative_km if mode == "discharge" else nan

    # РћђРћђ Duration as fractional days (Excel format) РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    dur_days = duration.total_seconds() / 86400.0

    # Build row in HEADERS order (excluding 'Leg Number' which is added by writer)
    row = (
        leg_type,  # Leg Type
        telem_url,  # Telematics Link
        charger_url,  # Charger Link
        logger_url,  # SRF Logger Link
        pd.Timestamp(t_s),  # Start Time (UTC)
        origin,  # Origin (Lat, Lon)
        origin_pc,  # Origin Place (postcode)
        pd.Timestamp(t_e),  # End Time (UTC)
        destination,  # Destination (Lat, Lon)
        dest_pc,  # Destination Place (postcode)
        dur_days,  # Duration (HH:MM:SS) Рђћ stored as fractional days
        distance,  # Distance (km)
        avg_speed,  # Average Speed (km/h)
        elevation_diff,  # Elevation Difference (m)
        veh_mass,  # Vehicle Mass (kg)
        veh_mass_cv,  # Vehicle Mass CV
        recuperation,  # Recuperation Energy (kWh)
        start_soc,  # Start SOC (%)
        end_soc,  # End SOC (%)
        soc_change,  # SOC Change (%) Рђћ signed
        energy_change,  # Energy Change (kWh) Рђћ signed
        energy_ac,  # Energy Charged AC (kWh)
        energy_dc,  # Energy Charged DC (kWh)
        nan,  # CO2 level (g/kWh)
        cumulative_km_out,  # Cumulative Distance (km)
        nan,  # CO2 for event (g)
        nan,  # Cumulative CO2 (g)
        battery_power,  # Battery Power (kW)
        energy_perf,  # Energy Performance (kWh/km)
        energy_perf_corrected,  # Energy Performance Corrected
        battery_cap,  # Battery Capacity (kWh)
        charger_energy_kwh,  # Energy Output from Charger (kWh)
        nan,  # Wire Energy Efficiency
        nan,  # Peak Charging (kW)
        nan,  # Average Charging (kW)
        nan,  # Energy based on motor power
        nan,  # Average Temperature
        nan,  # Average Pressure
        nan,  # Average Humidity
        nan,  # Average Wind Speed
        nan,  # Average Wind Direction
        nan,  # Weather Type
        acc_hist,  # Histogram Acc Pedal
        dec_hist,  # Histogram Dec Pedal
        energy_source,  # Energy Source
        energy_perf_kinetics,  # Energy Performance Kinetics Corrected (kWh/km)
        propulsion_kwh,  # Propulsion Energy (kWh)  [v2.2.3]
        ep_exclude_aux,  # EP_exclude_aux (kWh/km)  [v2.2.4]
        operator,  # Operator (project code)  [v2.2.5]
    )
    return row, cumulative_km


# =============================================================================
# Excel ТіЦтЉіућЪТѕљ
# =============================================================================


def _write_na(ws, ri: int, ci: int, cell_format) -> None:
    """Write the ``=NA()`` "no data" formula with an EMPTY cached result.

    xlsxwriter's ``write`` / ``write_formula`` default the cached formula result
    to ``0``. Readers that do not recalculate the workbook Рђћ openpyxl
    ``data_only=True`` and ``pandas.read_excel`` Рђћ then surface a blank cell as
    ``0`` instead of NaN. For the ``Vehicle Mass (kg)`` column that
    mis-classifies a no-GVM leg (T88RNW / YN75NMA Рђд) as mass-bearing, and more
    generally makes every empty numeric cell read as a spurious ``0``.

    Passing an empty string as the cached result makes those readers see the
    cell as blank (``None`` Рєњ NaN), which matches (a) Excel's own recalculated
    ``#N/A`` and (b) the convention of every report that has been round-tripped
    through openpyxl afterwards (weather / logger / charger patchers, the chart
    re-patch) Рђћ those re-saves strip xlsxwriter's cached ``0`` to blank, which is
    why the full-regen path never exhibited the ``0`` but the single-pass
    SRF-free recompute did. Excel still recalculates ``NA()`` to ``#N/A`` on
    load, so the on-screen value and the (data_only=False) formula are unchanged.
    """
    ws.write_formula(ri, ci, "=NA()", cell_format, "")


def _write_excel_report(
    rows: list[tuple],
    reg: str,
    period_start: date,
    period_end: date,
    out_path: Path,
    headers: tuple = HEADERS,
) -> None:
    """
    т░єТіЦтЉіУАїтєЎтЁЦ Excel№╝їТа╝т╝ЈСИјУђЂуЅѕ JOLT СИђУЄ┤№╝џ
      - у╗┐УЅ▓УЃїТЎ»№╝џУАїуеІ№╝ѕTrip№╝Ѕ
      - у║бУЅ▓УЃїТЎ»№╝џтЁЁућх№╝ѕCharge№╝Ѕ
      - уЎйУЅ▓УЃїТЎ»№╝џтЂюУйд№╝ѕStop№╝Ѕ
      - ТЌХжЌ┤Тѕ│Та╝т╝Ј№╝џyyyy-mm-dd hh:mm:ss
      - ТЌХжЋ┐Та╝т╝Ј№╝џ[hh]:mm:ss№╝ѕС╗Цт░ЈТЋ░тцЕтГўтѓе№╝Ѕ
      - УХЁжЊЙТјЦ№╝џУЊЮУЅ▓СИІтѕњу║┐
      - уггСИЅт╝атиЦСйюУАе№╝џтГЌТ«хУ»┤Тўј

    ``headers`` ТјДтѕХСй┐ућетЊфтЦЌтѕЌтИЃт▒ђ№╝џућхтіеУйдС╝а ``HEADERS``№╝ѕж╗ўУ«ц№╝Ѕ№╝їТЪ┤Т▓╣УйдС╝а
    ``DIESEL_HEADERS``сђѓСИцтЦЌ headers уџёуггСИђтѕЌжЃйТў» 'Leg Number'№╝їтљју╗ГтГЌТ«хСИј row
    tuple уџёжА║т║ЈСИђСИђт»╣т║ћсђѓ
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(out_path), {"nan_inf_to_errors": True})
    ws = workbook.add_worksheet("Report")

    # РћђРћђ Formats РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # Green = trip / discharge, Red = charging, White = stop (between events).
    _FMT = {}
    for color_name, bg in [
        ("green", "#C6EFCE"),
        ("red", "#FFC7CE"),
        ("white", "#FFFFFF"),
    ]:
        _FMT[color_name] = {
            "ts": workbook.add_format(
                {
                    "align": "center",
                    "bg_color": bg,
                    "num_format": "yyyy-mm-dd hh:mm:ss",
                    "border": 1,
                    "border_color": "#D0D0D0",
                }
            ),
            "dur": workbook.add_format(
                {
                    "align": "center",
                    "bg_color": bg,
                    "num_format": "[hh]:mm:ss",
                    "border": 1,
                    "border_color": "#D0D0D0",
                }
            ),
            "url": workbook.add_format(
                {
                    "align": "center",
                    "bg_color": bg,
                    "font_color": "#0000FF",
                    "underline": True,
                    "border": 1,
                    "border_color": "#D0D0D0",
                }
            ),
            "def": workbook.add_format(
                {
                    "align": "center",
                    "bg_color": bg,
                    "border": 1,
                    "border_color": "#D0D0D0",
                }
            ),
        }
    header_fmt = workbook.add_format({"bold": True, "align": "center"})

    # РћђРћђ Headers РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    for ci, h in enumerate(headers):
        ws.write(0, ci, h, header_fmt)
        if h in (
            "Leg Type",
            "Origin (Lat, Lon)",
            "Origin Place",
            "Destination (Lat, Lon)",
            "Destination Place",
            "Telematics Link",
            "Charger Link",
            "SRF Logger Link",
        ):
            ws.set_column(ci, ci, 30)
        elif h in (
            "Energy Performance Corrected by Elevation Difference (kWh/km)",
            "Energy Performance Kinetics Corrected (kWh/km)",
            "Histogram of Accelerator Pedal Position",
            "Histogram of Decelerator Pedal Position",
        ):
            ws.set_column(ci, ci, len(h) + 4)
        else:
            ws.set_column(ci, ci, max(len(h) + 2, 12))

    # РћђРћђ Data rows РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    for ri, row in enumerate(rows, start=1):
        leg_type = row[0]  # first field after Leg Number
        is_stop = bool(
            leg_type
            and isinstance(leg_type, str)
            and leg_type.strip().lower() == "stop"
        )
        is_charge = (not is_stop) and bool(
            leg_type
            and re.match(r"^(AC|DC|Charge|Mix|estimated)", leg_type, re.IGNORECASE)
        )
        if is_stop:
            cname = "white"
        elif is_charge:
            cname = "red"
        else:
            cname = "green"
        fmt = _FMT[cname]

        ws.write(ri, 0, ri, fmt["def"])  # Leg Number

        for ci_offset, val in enumerate(row):
            ci = ci_offset + 1  # +1 for Leg Number column
            col_name = headers[ci]
            if (
                col_name in ("Telematics Link", "Charger Link", "SRF Logger Link")
                and val
            ):
                ws.write_url(ri, ci, val, string="Link", cell_format=fmt["url"])
            elif col_name == "Duration (HH:MM:SS)":
                if val is not None and not _is_nan(val):
                    ws.write(ri, ci, val, fmt["dur"])
                else:
                    _write_na(ws, ri, ci, fmt["dur"])
            elif col_name in ("Start Time (UTC)", "End Time (UTC)") and isinstance(
                val, pd.Timestamp
            ):
                # xlsxwriter needs a naive datetime value for timestamp formats
                naive = val.tz_localize(None) if val.tzinfo else val
                ws.write_datetime(ri, ci, naive.to_pydatetime(), fmt["ts"])
            elif _is_nan(val):
                _write_na(ws, ri, ci, fmt["def"])
            elif val is None:
                ws.write(ri, ci, "", fmt["def"])
            else:
                ws.write(ri, ci, val, fmt["def"])

    ws.freeze_panes(1, 0)

    # РћђРћђ Graphs worksheet РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    # Charts are driven by the shared CHART_SPECS config so every report uses
    # identical fixed axes and identical data filtering. Because xlsxwriter
    # charts reference cell ranges and cannot filter in place, the cleaned
    # (x, y) pairs for each chart are written to a hidden ``GraphsData`` sheet;
    # both the scatter series and its linear trendline point at that clean
    # block (fit on filtered data only, matching the paperтЈБтЙё).
    graphs_ws = workbook.add_worksheet("Graphs")
    data_ws = workbook.add_worksheet("GraphsData")
    data_ws.hide()

    specs = chart_specs_for(headers)
    # Row-relative index map: ``rows`` tuples omit the leading 'Leg Number'
    # column, so a header's position in a row is its HEADERS index minus one.
    row_col_idx = {h: i - 1 for i, h in enumerate(headers) if i >= 1}

    # Format for the graceful-degradation note panel (created once; harmless if a
    # report never has an empty chart). Centred grey italic text on a light panel
    # with a thin border, so a chart that can plot nothing reads as a deliberate
    # "no data" note rather than a broken empty chart frame.
    empty_note_fmt = workbook.add_format(
        {
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "italic": True,
            "font_size": CHART_STYLE["empty_note_font_size"],
            "font_color": "#" + CHART_STYLE["empty_note_rgb"],
            "bg_color": "#" + CHART_STYLE["empty_note_fill_rgb"],
            "border": 1,
            "border_color": "#" + CHART_STYLE["empty_note_border_rgb"],
        }
    )

    data_col = 0  # next free column pair on GraphsData
    for gi, spec in enumerate(specs):
        x_hdr, y_hdr = spec["x_hdr"], spec["y_hdr"]
        if x_hdr not in row_col_idx or y_hdr not in row_col_idx:
            continue
        pts = _filtered_chart_points(rows, row_col_idx, spec)

        # Write the filtered pairs to the hidden helper block (with a header row
        # so the range is self-describing if anyone unhides the sheet).
        xcol, ycol = data_col, data_col + 1
        data_ws.write(0, xcol, x_hdr)
        data_ws.write(0, ycol, y_hdr)
        for di, (xv, yv) in enumerate(pts, start=1):
            data_ws.write_number(di, xcol, xv)
            data_ws.write_number(di, ycol, yv)
        n_pts = len(pts)
        # Advance the GraphsData column pair now so it stays in lockstep whether
        # this chart is drawn or replaced by a note.
        data_col += 2

        # Graceful degradation: with no plottable points (e.g. a vehicle whose
        # mass column is null for the whole period Рєњ EP-vs-Mass has nothing to
        # plot) we replace the empty scatter frame with a centred text note in the
        # chart's footprint, leaving the other charts' positions untouched. A
        # merged cell is used (not a textbox) because it is the one primitive the
        # openpyxl patch path can write identically.
        if n_pts == 0:
            fr, fc, lr, lc = empty_note_extent(gi)
            graphs_ws.merge_range(
                fr, fc, lr, lc, empty_chart_note(spec), empty_note_fmt
            )
            continue

        # Shared single-colour scatter look. A scatter series with no explicit
        # fill makes Excel auto-vary the per-point colour (the "five colours"
        # artefact); pinning one steelblue solidFill + alpha + no border gives
        # the soft uniform look. The fit equation + R┬▓ go into a title sub-line
        # (see ``_chart_subtitle``) rather than overlapping floating data labels.
        scatter_hex = "#" + CHART_STYLE["scatter_rgb"]
        title = spec.get("title", f"{y_hdr} vs {x_hdr}")
        subtitle = _chart_subtitle(pts) if CHART_STYLE["show_fit_in_subtitle"] else ""

        chart = workbook.add_chart({"type": "scatter"})
        if n_pts > 0:
            chart.add_series(
                {
                    "name": spec.get("series_name", y_hdr),
                    "categories": ["GraphsData", 1, xcol, n_pts, xcol],
                    "values": ["GraphsData", 1, ycol, n_pts, ycol],
                    "trendline": {
                        "type": "linear",
                        "display_equation": False,
                        "display_r_squared": False,
                        "line": {"color": "#404040", "width": 1.25},
                    },
                    "marker": {
                        "type": "circle",
                        "size": CHART_STYLE["marker_size"],
                        # xlsxwriter transparency = 100 Рѕњ opacity (so the openpyxl and
                        # xlsxwriter paths render the same soft fill).
                        "fill": {
                            "color": scatter_hex,
                            "transparency": 100 - CHART_STYLE["scatter_opacity"],
                        },
                        "border": {"none": True},
                    },
                }
            )
        # Light-grey, high-transparency major gridlines on BOTH axes; minor
        # gridlines off. xlsxwriter wants the complementary transparency
        # (100 Рѕњ opacity) and a pt line width, so the openpyxl patch and this
        # path render the same soft grid.
        grid_hex = "#" + CHART_STYLE["grid_rgb"]
        major_gridlines = {
            "visible": True,
            "line": {
                "color": grid_hex,
                "transparency": 100 - CHART_STYLE["grid_opacity"],
                "width": CHART_STYLE["grid_width_emu"] / 12700.0,
            },  # EMU Рєњ pt
        }
        minor_gridlines = {"visible": False}
        chart.set_x_axis(
            {
                # Axis DISPLAY title is decoupled from the data header: the series
                # still points at ``x_hdr``; only the printed label may differ.
                "name": spec.get("x_title", x_hdr),
                "name_font": {
                    "size": CHART_STYLE["axis_title_font_size"],
                    "bold": False,
                },
                "num_font": {"size": CHART_STYLE["tick_font_size"]},
                "min": spec["x_min"],
                "max": spec["x_max"],
                "major_unit": spec["x_major"],
                "major_gridlines": major_gridlines,
                "minor_gridlines": minor_gridlines,
            }
        )
        chart.set_y_axis(
            {
                # Axis DISPLAY title decoupled from the data header (mirrors x).
                "name": spec.get("y_title", y_hdr),
                "name_font": {
                    "size": CHART_STYLE["axis_title_font_size"],
                    "bold": False,
                },
                "num_font": {"size": CHART_STYLE["tick_font_size"]},
                "min": spec["y_min"],
                "max": spec["y_max"],
                "major_unit": spec["y_major"],
                "major_gridlines": major_gridlines,
                "minor_gridlines": minor_gridlines,
            }
        )
        # Title on its own top band; fit equation + R┬▓ as a smaller 2nd line so
        # the two no longer overlap each other or sit on top of the points.
        title_name = f"{title}\n{subtitle}" if subtitle else title
        chart.set_title(
            {
                "name": title_name,
                "name_font": {"size": CHART_STYLE["title_font_size"], "bold": True},
            }
        )
        # Legend in the top-right corner, not overlaying the plot.
        chart.set_legend(
            {
                "position": CHART_STYLE["legend_pos_xlsxwriter"],
                "font": {"size": CHART_STYLE["legend_font_size"]},
            }
        )
        # Pin the inner plot area with fixed margins (shared CHART_STYLE
        # fractions) so the axis titles always sit outside their tick labels Рђћ
        # without this Excel auto-grows the plot box and the big-font axis
        # titles overlap the tick numbers.
        chart.set_plotarea(
            {
                "layout": {
                    "x": CHART_STYLE["plot_x"],
                    "y": CHART_STYLE["plot_y"],
                    "width": CHART_STYLE["plot_w"],
                    "height": CHART_STYLE["plot_h"],
                },
            }
        )
        # Size the chart EXPLICITLY from the shared cm constants (px = cm ├Ќ 360000
        # EMU/cm ├и 9525 EMU/px) instead of the old ``x_scale`` / ``y_scale``, so
        # this path renders at the same physical size as the openpyxl patch path
        # and one shared ``chart_row_step`` stacks them identically.
        chart.set_size(
            {
                "width": round(CHART_STYLE["chart_width_cm"] * 360_000 / 9525),
                "height": round(CHART_STYLE["chart_height_cm"] * 360_000 / 9525),
            }
        )
        # Anchor each chart a fixed number of rows below the previous one Рђћ the
        # step (chart height in rows + gap) comes from the shared ``chart_row_step``
        # helper, so charts never overlap and both render paths match.
        graphs_ws.insert_chart(f"A{gi * chart_row_step() + 1}", chart)

    # РћђРћђ Definitions worksheet РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    defs_ws = workbook.add_worksheet("Definitions")
    def_fmt = workbook.add_format({"text_wrap": True, "valign": "top"})
    if "Fuel Consumption (L/100km)" in headers:
        # Diesel report definitions
        def_texts = [
            'Leg Type: "In Transit" = trip (green); "Stop" = parked/idling gap between trips (white). '
            "Diesel vehicles have no charging events.",
            "Vehicle Mass (kg): Gross combination vehicle weight (GCVW) read from the SRF Logger "
            '"CVW gross combination vehicle weight" channel at 1 Hz; reported value is the per-trip median.',
            "Fuel Used (L): Per-trip fuel consumption, computed from cumulative differences of the "
            '"LFC engine total fuel used" channel over the trip window.',
            "Fuel Consumption (L/100km): Fuel Used (L) / Distance (km) ├Ќ 100.",
            'Energy Source: "lfc_fuel" = trip energy reconstructed from LFC fuel used ├Ќ diesel LHV (default 10 kWh/L). '
            "Kept for cross-fuel comparability with the EV report.",
            "SRF Logger Link: URL to the SRF Logger visualisation on the SRF platform.",
            "Average Temperature (C): Per-trip mean of AMB ambient air temperature (Logger 1 Hz).",
            'Elevation Difference (m): End Рѕњ start of Channel 2 "2 altitude" (GPS altitude) over the trip.',
        ]
    else:
        # Electric report definitions
        def_texts = [
            "SOC: State of Charge. The percentage of the battery capacity that is currently charged.",
            'Energy Source: "ac_dc" = charging energy from AC+DC telematics counters; '
            + '"total_energy" = discharge energy from total_electric_energy_used_plugged_in_included; '
            + '"moving_energy" = discharge energy from electric_energy_wheelbased_speed_over_zero; '
            + '"soc_estimate" = energy estimated from SOC change ├Ќ nominal capacity.',
            "Energy Performance Kinetics Corrected (kWh/km): Elevation + per-second kinetic energy "
            + "corrected energy performance. Uses Logger 1Hz speed data to compute ╬ћKE per second, "
            + "with 90% regenerative braking efficiency (╬и_regen = 0.90). "
            + "Net KE = ╬Б(accel ╬ћKE) Рѕњ ╬и_regen ├Ќ ╬Б(|braking ╬ћKE|).",
            "Charger Link / SRF Logger Link: URL to the add-on sensor visualisation on the SRF platform. "
            + "Derived metrics (Peak Charging, Wire Efficiency, Weather, etc.) are reserved for future work.",
            "Battery Capacity (kWh): Effective capacity estimated by the segment algorithm from "
            + "delta_energy_kwh / (delta_soc / 100). Not the nominal manufacturer value.",
            "Energy Performance (kWh/km): |delta_energy_kwh| / distance for discharge trips "
            + "with distance > 0. Only meaningful for discharge segments.",
        ]
    for dr, dt in enumerate(def_texts):
        defs_ws.write(dr, 0, dt, def_fmt)
    defs_ws.set_column(0, 0, max(len(t) for t in def_texts) * 0.9, def_fmt)

    workbook.set_properties(
        {
            "title": f"JOLT Segment Report Рђћ {reg}",
            "subject": f"{period_start} to {period_end}",
            "author": "Centre for Sustainable Road Freight",
            "comments": "Generated by jolt_toolkit.report_generator",
        }
    )

    workbook.close()
    print(f"  Excel report: {out_path.name}")


def _compute_active_dates_from_xlsx(xlsx_path: Path) -> set[str]:
    """У»╗тЈќ xlsx уџё ``Report`` sheet№╝їУ┐ћтЏъТюЅ trip Тѕќ charge Т┤╗тіеуџёТЌЦТюЪжЏєтљѕсђѓ

    Stop УАї№╝ѕ`Leg Type == 'Stop'`№╝ЅтњїуЕ║УАїУбФТјњжЎц№╝ЏтЁХт«ЃС╗╗СйЋ leg type
    №╝ѕ`In Transit / Round Trip / Outbound / Return / In House / AC Charge /
    DC Charge / Mix Charge / estimated Charge` уГЅ№╝ЅжЃйУДєСИ║Т┤╗тіесђѓ

    У┐ћтЏътђ╝Тў» ``{'YYYY-MM-DD', ...}`` тГЌугдСИ▓жЏєтљѕ№╝їСИј figure ТќЄС╗ХтљЇжЄїуџё
    ТЌЦТюЪтЅЇу╝ђ№╝ѕ``validation_<REG>_<DATE>_<idx>.png``№╝ЅС┐ЮТїЂтљїТаиТа╝т╝ЈС╗ЦСЙ┐Т»ћт»╣сђѓ
    У»╗тЈќтц▒У┤ЦТѕќ xlsx СИЇтГўтюеТЌХУ┐ћтЏъуЕ║ set РђћРђћ viewer С╝џтЏъжђђтѕ░тЁежЃеТЮАуЏ«жЃйУбФУДєСИ║
    "active"№╝їС┐ЮТїЂтљЉтљјтЁ╝т«╣сђѓ
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return set()
    if not xlsx_path.exists():
        return set()
    try:
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    except Exception:
        return set()
    try:
        ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return set()
        # уће header тљЇтГЌт«џСйЇ Leg Type / Start Time тѕЌ№╝їжЂ┐тЁЇуАгу╝ќуаЂ EV/Diesel тЂЈуД╗
        col_leg_type = None
        col_start = None
        for idx, name in enumerate(header_row):
            if name == "Leg Type":
                col_leg_type = idx
            elif name == "Start Time (UTC)":
                col_start = idx
        if col_leg_type is None or col_start is None:
            return set()
        active: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(col_leg_type, col_start):
                continue
            leg_type = row[col_leg_type]
            if leg_type is None:
                continue
            leg_type_str = str(leg_type).strip()
            if not leg_type_str or leg_type_str == "Stop":
                continue
            start = row[col_start]
            if start is None:
                continue
            # Start Time тЈ»УЃйТў» datetime т»╣У▒АТѕќ ISO тГЌугдСИ▓
            if hasattr(start, "strftime"):
                date_str = start.strftime("%Y-%m-%d")
            else:
                date_str = str(start)[:10]
            if re.match(r"\d{4}-\d{2}-\d{2}$", date_str):
                active.add(date_str)
        return active
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _group_paths_by_date(
    paths, date_re: "re.Pattern[str]"
) -> dict[str, list[Path]]:
    """Group per-leg raw-CSV paths by the ``YYYY-MM-DD`` captured by ``date_re``.

    Both regenerate paths (EV ``raw_<date>_<idx>.csv`` and diesel
    ``logger_<date>_<idx>.csv``) split a single calendar day into many short
    legs. To draw one validation figure per day we first bucket each day's legs
    together. ``date_re`` must expose the date in capture group 1
    (``re.search`` is used, so the pattern need only match the filename's date
    token). The returned dict preserves chronological order (paths are sorted
    first, dict insertion order is stable on Python 3.7+), and each day's list
    is in filename Рђћ i.e. leg-index, i.e. time Рђћ order.
    """
    groups: dict[str, list[Path]] = {}
    for p in sorted(paths):
        m = date_re.search(p.name)
        if not m:
            continue
        groups.setdefault(m.group(1), []).append(p)
    return groups


def _clear_day_validation_figures(fig_dir: Path, reg: str) -> int:
    """Delete stale validation figures + overlay sidecars before a per-day repaint.

    The one-figure-per-day regeneration writes ``validation_<reg>_<date>.png``;
    earlier builds wrote one figure per leg (``validation_<reg>_<date>_<NNNN>.png``).
    Both match the ``validation_<reg>_*`` glob the inspect viewer lists, so without
    this sweep a re-painted directory would show BOTH the consolidated day figure
    and every stale per-leg fragment. Clearing the whole non-finetuned set up front
    also drops figures for days that no longer segment to any trip.

    ``*_finetuned.*`` artefacts are left untouched Рђћ they belong to the
    report-finetuner flow and must survive a base repaint. Returns the number of
    files removed.
    """
    if not fig_dir.exists():
        return 0
    removed = 0
    for pattern in (
        f"validation_{reg}_*.png",
        f"validation_{reg}_*.boxes.json",
        f"validation_{reg}_*.dsoc.json",
    ):
        for p in fig_dir.glob(pattern):
            if "_finetuned" in p.name:
                continue
            try:
                p.unlink()
                removed += 1
            except OSError:
                pass
    return removed


def _write_html_viewer(
    out_dir: Path, reg: str, period_start, period_end, report_name: str
) -> None:
    """Generate an HTML viewer that shows all validation figures for this vehicle/period."""
    fig_dir = out_dir / "validation_figures"
    all_figs = (
        sorted(fig_dir.glob(f"validation_{reg}_*.png")) if fig_dir.exists() else []
    )
    if not all_figs:
        return

    # С╗ЁС┐ЮуЋЎТЌЦТюЪтюе period_start ~ period_end УїЃтЏ┤тєЁуџёжфїУ»ЂтЏЙсђѓ
    # ТЌЦТюЪ token тљјуће ``[._]`` тЁ╝т«╣СИцуДЇтЉйтљЇ№╝џТќ░уџёСИђТЌЦСИђтЏЙ ``validation_<reg>_<date>.png``
    # №╝ѕТЌЦТюЪтљјуЏ┤ТјЦТў»ТЅЕт▒ЋтљЇ ``.``№╝ЅСИјтјєтЈ▓ / finetuned уџё per-leg
    # ``validation_<reg>_<date>_<NNNN>.png``№╝ѕТЌЦТюЪтљјТў» ``_``№╝Ѕсђѓ
    _date_re = re.compile(
        r"validation_" + re.escape(reg) + r"_(\d{4}-\d{2}-\d{2})[._]"
    )
    ds_str = str(period_start)
    de_str = str(period_end)
    figs = []
    fig_dates: list[str] = []
    for p in all_figs:
        m = _date_re.match(p.name)
        if m and ds_str <= m.group(1) <= de_str:
            figs.append(p)
            fig_dates.append(m.group(1))
    if not figs:
        return

    # У»╗ xlsx Т▒ѓтйЊтЉеТюЪтєЁТюЅ trip/charge Т┤╗тіеуџёТЌЦТюЪжЏєтљѕ№╝їућеС║јСЙДТаЈу║бтГЌ + У┐ЄТ╗ц
    active_dates = _compute_active_dates_from_xlsx(out_dir / report_name)
    is_active = [d in active_dates for d in fig_dates]

    # Relative paths from out_dir to validation_figures/
    rel = [f"validation_figures/{p.name}" for p in figs]
    labels = [p.stem for p in figs]  # e.g. validation_AV24LXK_2024-10-01_0000

    # Per-figure interactive annotation overlay sidecar. A figure produced with
    # ``export_dsoc_overlay=True`` writes ``<stem>.boxes.json`` next to the PNG
    # (figure-fraction coordinates, origin top-left) carrying every panel's
    # rounded-bbox data label Рђћ dSOC, energy/charger-meter deltas, recuperation
    # deltas, mass labels. Two shapes are accepted (the viewer auto-detects):
    #   * **flat list** Рђћ legacy / diesel figures; rendered as ghosted-on-hover
    #     ``.annot-box`` divs (the original behaviour, kept for back-compat with
    #     not-yet-re-painted vehicles).
    #   * **dict** ``{boxes, segments, soc_axis}`` (v2.2.6, EV figures) Рђћ drives
    #     the hover redesign: default triangles-only, per-segment hover hotzones,
    #     a pinned info box at the SOC panel's bottom-left, and label de-collision.
    # Figures without a sidecar (legacy / baked-in text) yield an empty list and
    # simply show no overlay.
    boxes_per_fig: list = []
    for p in figs:
        sidecar = fig_dir / (p.stem + ".boxes.json")
        # Backward compat: figures painted by earlier builds (before the overlay
        # rename, i.e. pre the ``.boxes.json`` sidecar) emit the older
        # ``.dsoc.json`` (Panel-1 dSOC only). Fall back to it so an HTML re-render
        # of a not-yet-re-painted vehicle still shows its dSOC overlay.
        if not sidecar.exists():
            sidecar = fig_dir / (p.stem + ".dsoc.json")
        data: list | dict = []
        if sidecar.exists():
            try:
                with open(sidecar, encoding="utf-8") as fh:
                    loaded = json.load(fh)
                if isinstance(loaded, (list, dict)):
                    data = loaded
            except (json.JSONDecodeError, OSError):
                data = []
        boxes_per_fig.append(data)

    html_name = "inspect_" + report_name.replace(".xlsx", ".html")
    html_path = out_dir / html_name

    # Build sidebar items & image tags as JS arrays
    imgs_js = "[\n" + ",\n".join(f'        "{r}"' for r in rel) + "\n    ]"
    labels_js = "[\n" + ",\n".join(f'        "{l}"' for l in labels) + "\n    ]"
    active_js = (
        "[\n" + ",\n".join(f"        {str(a).lower()}" for a in is_active) + "\n    ]"
    )
    # Inline the box data (avoids a runtime fetch, which the file:// protocol
    # blocks when the HTML is opened directly from disk).
    boxes_js = json.dumps(boxes_per_fig, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>JOLT Inspection Рђћ {reg} {period_start} РђЊ {period_end}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ display: flex; height: 100vh; font-family: Arial, sans-serif; }}
  /* v2.2.4: viewer chrome font sizes doubled for readability; sidebar widened
     to keep the longer day labels on one line. */
  #sidebar {{
    width: 380px; min-width: 240px; background: #1e1e2e; color: #cdd6f4;
    overflow-y: auto; flex-shrink: 0; padding: 12px 0;
  }}
  #sidebar h2 {{ font-size: 26px; padding: 14px 18px 8px; color: #89b4fa;
                 border-bottom: 1px solid #313244; margin-bottom: 6px; }}
  /* жАХжЃеТ┤╗тіетцЕУ┐ЄТ╗цтЎе№╝џтІЙжђЅтљјжџљУЌЈТ▓АТюЅ trip/charge уџёуЕ║тцЕ */
  #filter-box {{
    padding: 10px 18px 12px; font-size: 22px; color: #a6adc8;
    border-bottom: 1px solid #313244; margin-bottom: 6px;
    display: flex; align-items: center; gap: 10px;
  }}
  #filter-box input {{ cursor: pointer; width: 18px; height: 18px; }}
  #filter-box label {{ cursor: pointer; user-select: none; }}
  #filter-count {{ color: #7f849c; font-size: 18px; margin-left: auto; }}
  #sidebar ul {{ list-style: none; }}
  #sidebar li {{
    padding: 9px 18px; cursor: pointer; font-size: 23px;
    border-left: 5px solid transparent;
  }}
  #sidebar li:hover {{ background: #313244; }}
  /* Т┤╗тіетцЕ№╝џтйЊтцЕТюЅ trip Тѕќ charge Рєњ у║бтГЌ№╝ѕТџЌу║б #f38ba8№╝їУиЪ Catppuccin
     СИ╗жбўжЄїуџё red СИђУЄ┤№╝їТ»ћу║»у║бТЪћтњїсђЂСИјТџЌУЅ▓УЃїТЎ»т»╣Т»ћТИЁТЎ░№╝Ѕсђѓ */
  #sidebar li.active-day {{ color: #f38ba8; }}
  /* тйЊтЅЇжђЅСИГТЮАуЏ«№╝ѕСИјТў»тљдТ┤╗тіетцЕТЌатЁ│№╝їТ▓┐ућеУЊЮУЅ▓жФўС║«№╝Ѕ */
  #sidebar li.active {{
    background: #313244; border-left-color: #89b4fa; color: #89b4fa;
  }}
  /* жђЅСИГТђЂС╝ўтЁѕС║јТ┤╗тіетцЕуЮђУЅ▓№╝џactive жФўС║«УЊЮУЅ▓УдєуЏќу║бУЅ▓ */
  #sidebar li.active.active-day {{ color: #89b4fa; }}
  /* жџљУЌЈуЕ║тцЕуџёт«ъжЎЁ class№╝ѕућ▒ JS тюе checkbox тІЙжђЅТЌХтѕЄТЇб№╝Ѕ */
  #sidebar li.hidden-day {{ display: none; }}
  #main {{
    flex: 1; display: flex; flex-direction: column;
    background: #11111b; overflow: hidden;
  }}
  #toolbar {{
    display: flex; align-items: center; gap: 16px;
    padding: 12px 22px; background: #181825; border-bottom: 1px solid #313244;
  }}
  #toolbar button {{
    padding: 8px 22px; background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 5px; cursor: pointer; font-size: 24px;
  }}
  #toolbar button:hover {{ background: #45475a; }}
  #counter {{ color: #a6adc8; font-size: 24px; }}
  #label {{ color: #89b4fa; font-size: 22px; font-family: monospace;
            flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  #img-wrap {{ flex: 1; display: flex; align-items: center;
               justify-content: center; overflow: auto; padding: 12px; }}
  /* The stage is sized in JS to the displayed image rect so the overlay can be
     positioned by simple figure-fraction ├Ќ pixel maths (no object-fit guessing). */
  #fig-stage {{ position: relative; line-height: 0; }}
  #fig {{ display: block; }}
  #overlay {{ position: absolute; left: 0; top: 0; width: 100%; height: 100%;
              pointer-events: none; }}
  /* Interactive annotation box (dSOC + every other panel's data label): ghosted
     by default (so the data lines dominate) yet discoverable, solid on hover. */
  .annot-box {{
    position: absolute; font-family: monospace; font-size: 22px; font-weight: bold;
    line-height: 1.15; white-space: pre; padding: 2px 7px; border-radius: 6px;
    background: rgba(255, 255, 255, 0); border: 1px solid transparent;
    opacity: 0.22; pointer-events: auto; cursor: default; z-index: 5;
    transition: opacity 0.12s ease, background 0.12s ease, box-shadow 0.12s ease,
                border-color 0.12s ease;
  }}
  .annot-box:hover {{
    opacity: 1; background: rgba(255, 255, 255, 0.97);
    border-color: rgba(0, 0, 0, 0.18); box-shadow: 0 2px 10px rgba(0, 0, 0, 0.35);
    z-index: 30;
  }}
  /* РћђРћђ v2.2.6 hover redesign (dict-schema sidecars) РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
     Default view = baked triangles + baked legends/titles only. Every per-segment
     data label is hidden until its full-height hover hotzone is entered. */
  /* Transparent per-segment hover catcher spanning the segment's x-range ├Ќ the
     full stage height. The ONLY element with pointer-events in this mode, so
     value labels / the pinned box never steal the hover (no flicker). */
  .hotzone {{
    position: absolute; background: transparent; pointer-events: auto;
    z-index: 1; cursor: pointer;
  }}
  /* Per-segment value label (Panel-2/3/4 energy + mass deltas): hidden by default,
     revealed solid on hover. Non-interactive (the hotzone beneath handles hover). */
  .val-box {{
    position: absolute; font-family: monospace; font-size: 22px; font-weight: bold;
    line-height: 1.15; white-space: pre; padding: 2px 7px; border-radius: 6px;
    background: rgba(255, 255, 255, 0.97); border: 1px solid rgba(0, 0, 0, 0.18);
    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.35); pointer-events: none; z-index: 20;
  }}
  /* Pinned info box at the SOC panel's bottom-left Рђћ filled with the hovered
     segment's dSOC / energy / capacity / EP block. */
  .pinned-box {{
    position: absolute; font-family: monospace; font-size: 22px; font-weight: bold;
    line-height: 1.3; white-space: pre; padding: 8px 12px; border-radius: 8px;
    background: rgba(255, 255, 255, 0.98); border: 1px solid rgba(0, 0, 0, 0.25);
    box-shadow: 0 3px 14px rgba(0, 0, 0, 0.4); pointer-events: none; z-index: 25;
  }}
  /* Session-level label with no owning segment (e.g. the charger-meter total):
     always visible but ghosted, so it does not depend on a hover. */
  .static-box {{
    position: absolute; font-family: monospace; font-size: 22px; font-weight: bold;
    line-height: 1.15; white-space: pre; padding: 2px 7px; border-radius: 6px;
    background: rgba(255, 255, 255, 0.55); pointer-events: none; z-index: 6;
    opacity: 0.6;
  }}
</style>
</head>
<body>
<div id="sidebar">
  <h2>{reg} &nbsp;{period_start} РђЊ {period_end}</h2>
  <div id="filter-box">
    <input type="checkbox" id="filter-active">
    <label for="filter-active">Show active days only</label>
    <span id="filter-count"></span>
  </div>
  <ul id="list"></ul>
</div>
<div id="main">
  <div id="toolbar">
    <button onclick="go(-1)">&#9664; Prev</button>
    <button onclick="go(1)">Next &#9654;</button>
    <span id="counter"></span>
    <span id="label"></span>
  </div>
  <div id="img-wrap">
    <div id="fig-stage">
      <img id="fig" src="" alt="validation figure">
      <div id="overlay"></div>
    </div>
  </div>
</div>
<script>
const imgs     = {imgs_js};
const labels   = {labels_js};
const isActive = {active_js};
const annotBoxes = {boxes_js};
const STORAGE_KEY = "jolt_inspect_filter_active_only";
let cur = 0;
const listEl     = document.getElementById("list");
const figEl      = document.getElementById("fig");
const stageEl    = document.getElementById("fig-stage");
const overlayEl  = document.getElementById("overlay");
const imgWrapEl  = document.getElementById("img-wrap");
const counterEl  = document.getElementById("counter");
const labelEl    = document.getElementById("label");
const filterEl   = document.getElementById("filter-active");
const filterCntEl = document.getElementById("filter-count");

// ТИ▓ТЪЊСЙДТаЈсђѓтљїТЌХу╗Ў"Т┤╗тіетцЕ"тіа .active-day Рєњ у║бтГЌсђѓ
labels.forEach((lb, i) => {{
  const li = document.createElement("li");
  li.textContent = lb.replace(/^validation_[^_]+_/, "");
  if (isActive[i]) {{
    li.classList.add("active-day");
  }}
  li.onclick = () => show(i);
  listEl.appendChild(li);
}});

const activeCount = isActive.filter(Boolean).length;
const idleCount   = isActive.length - activeCount;
filterCntEl.textContent = activeCount + " active / " + idleCount + " idle";

function applyFilter() {{
  const onlyActive = filterEl.checked;
  document.querySelectorAll("#list li").forEach((el, j) => {{
    el.classList.toggle("hidden-day", onlyActive && !isActive[j]);
  }});
  try {{ localStorage.setItem(STORAGE_KEY, onlyActive ? "1" : "0"); }} catch (e) {{}}
}}

// У┐ўтјЪСИіТгАуџёУ┐ЄТ╗цуіХТђЂ
try {{
  if (localStorage.getItem(STORAGE_KEY) === "1") {{
    filterEl.checked = true;
  }}
}} catch (e) {{}}
filterEl.addEventListener("change", () => {{
  applyFilter();
  // тѕЄТЇб filter тљј№╝їУІЦтйЊтЅЇжђЅСИГти▓УбФжџљУЌЈ Рєњ Уи│тѕ░СИІСИђСИфтЈ»УДЂуџё
  const curEl = document.querySelectorAll("#list li")[cur];
  if (curEl && curEl.classList.contains("hidden-day")) {{
    go(1);
  }}
}});
applyFilter();

// Size the stage to the displayed image rect (preserve aspect ratio, fit the
// available area) so the overlay can be positioned by figure-fraction ├Ќ pixels.
function layout() {{
  const natW = figEl.naturalWidth, natH = figEl.naturalHeight;
  if (!natW || !natH) return;
  const cs = getComputedStyle(imgWrapEl);
  const padX = parseFloat(cs.paddingLeft) + parseFloat(cs.paddingRight);
  const padY = parseFloat(cs.paddingTop) + parseFloat(cs.paddingBottom);
  const availW = Math.max(1, imgWrapEl.clientWidth - padX);
  const availH = Math.max(1, imgWrapEl.clientHeight - padY);
  const scale = Math.min(availW / natW, availH / natH);
  const dispW = Math.max(1, Math.round(natW * scale));
  const dispH = Math.max(1, Math.round(natH * scale));
  figEl.style.width = dispW + "px";
  figEl.style.height = dispH + "px";
  stageEl.style.width = dispW + "px";
  stageEl.style.height = dispH + "px";
  renderBoxes(dispW, dispH);
}}

// Render the current figure's annotation overlay. Dispatches on the sidecar
// shape: a flat list Рєњ legacy ghosted-on-hover boxes (back-compat with diesel /
// not-yet-re-painted vehicles); a dict Рєњ the v2.2.6 hover redesign.
function renderBoxes(dispW, dispH) {{
  overlayEl.innerHTML = "";
  const entry = annotBoxes[cur];
  if (!entry) return;
  if (Array.isArray(entry)) {{ renderLegacy(entry, dispW, dispH); return; }}
  renderInteractive(entry, dispW, dispH);
}}

// Legacy / diesel sidecar (flat list). Coordinates are figure-fraction with origin
// top-left; ha/va decide the anchor transform. Each div is ghosted by default and
// solid on hover, then clamped back inside the [0..dispW] ├Ќ [0..dispH] image rect.
function renderLegacy(boxes, dispW, dispH) {{
  for (const b of boxes) {{
    const d = document.createElement("div");
    d.className = "annot-box";
    d.textContent = b.text;
    d.style.color = b.color;
    let left = b.x * dispW;
    let top  = b.y * dispH;
    d.style.left = left + "px";
    d.style.top  = top + "px";
    let tx = "-50%", ty = "0";          // ha=center, va=top defaults
    if (b.ha === "left") tx = "0";
    else if (b.ha === "right") tx = "-100%";
    if (b.va === "bottom") ty = "-100%";
    else if (b.va === "center") ty = "-50%";
    d.style.transform = "translate(" + tx + ", " + ty + ")";
    overlayEl.appendChild(d);
    const w = d.offsetWidth, h = d.offsetHeight;
    const txFrac = (b.ha === "right") ? -1 : (b.ha === "left") ? 0 : -0.5;
    const tyFrac = (b.va === "bottom") ? -1 : (b.va === "center") ? -0.5 : 0;
    const rx = left + txFrac * w;       // rendered left edge
    const ry = top  + tyFrac * h;       // rendered top edge
    const newRx = Math.max(0, Math.min(rx, dispW - w));
    const newRy = Math.max(0, Math.min(ry, dispH - h));
    if (newRx !== rx) {{ left += newRx - rx; d.style.left = left + "px"; }}
    if (newRy !== ry) {{ top  += newRy - ry; d.style.top  = top  + "px"; }}
  }}
}}

// v2.2.6 hover redesign (dict sidecar {{boxes, segments, soc_axis}}). Default view
// shows only the baked triangles + legends/titles. A full-height hover hotzone per
// segment reveals (a) the pinned info box at the SOC panel's bottom-left with that
// segment's dSOC block, and (b) that segment's value labels (Panel-2/3/4 energy +
// mass deltas) at their own coords Рђћ de-collided so none overlap or spill out.
function renderInteractive(entry, dispW, dispH) {{
  const boxes   = entry.boxes || [];
  const segs    = entry.segments || [];
  const socAxis = entry.soc_axis || null;
  const PAD = 6;

  // A box's anchored pixel top-left from its figure-fraction (x,y) + ha/va.
  function anchorPx(b, w, h) {{
    const left = b.x * dispW, top = b.y * dispH;
    const txFrac = (b.ha === "right") ? -1 : (b.ha === "left") ? 0 : -0.5;
    const tyFrac = (b.va === "bottom") ? -1 : (b.va === "center") ? -0.5 : 0;
    return [left + txFrac * w, top + tyFrac * h];
  }}
  // Position a visible box at its anchor, clamped into the image rect.
  function place(el) {{
    const w = el.offsetWidth, h = el.offsetHeight;
    const xy = anchorPx(el._box, w, h);
    el.style.left = Math.max(0, Math.min(xy[0], dispW - w)) + "px";
    el.style.top  = Math.max(0, Math.min(xy[1], dispH - h)) + "px";
  }}

  // Build value-label divs (hidden) + collect each segment's info-block text.
  const segBoxes = {{}};      // seg -> [el, ...]  (value labels)
  const infoBySeg = {{}};     // seg -> {{text, color}}  (Panel-1 dSOC block)
  for (const b of boxes) {{
    if (b.role === "info" && b.seg != null) {{
      infoBySeg[b.seg] = {{ text: b.text, color: b.color }};
      continue;  // info routes to the pinned box, not an inline div
    }}
    const d = document.createElement("div");
    d.textContent = b.text;
    d.style.color = b.color;
    d._box = b;
    if (b.seg == null) {{
      // Session-level label (e.g. charger-meter total): always visible, ghosted.
      d.className = "static-box";
      overlayEl.appendChild(d);
      place(d);
    }} else {{
      d.className = "val-box";
      d.style.display = "none";
      overlayEl.appendChild(d);
      (segBoxes[b.seg] = segBoxes[b.seg] || []).push(d);
    }}
  }}

  // One reusable pinned info box (filled per hover).
  const pinned = document.createElement("div");
  pinned.className = "pinned-box";
  pinned.style.display = "none";
  overlayEl.appendChild(pinned);

  // Greedy vertical de-collision: push later boxes below earlier overlapping ones,
  // then clamp every box back inside the image rect.
  function decollide(els) {{
    const items = els.map(el => ({{
      el: el,
      left: parseFloat(el.style.left) || 0,
      top:  parseFloat(el.style.top)  || 0,
      w: el.offsetWidth, h: el.offsetHeight,
    }}));
    items.sort((a, b) => (a.top - b.top) || (a.left - b.left));
    const GAP = 2;
    for (let i = 0; i < items.length; i++) {{
      for (let j = 0; j < i; j++) {{
        const a = items[j], b = items[i];
        const ox = a.left < b.left + b.w && b.left < a.left + a.w;
        const oy = a.top  < b.top + b.h  && b.top  < a.top + a.h;
        if (ox && oy) b.top = a.top + a.h + GAP;
      }}
    }}
    for (const it of items) {{
      it.left = Math.max(0, Math.min(it.left, dispW - it.w));
      it.top  = Math.max(0, Math.min(it.top,  dispH - it.h));
      it.el.style.left = it.left + "px";
      it.el.style.top  = it.top + "px";
    }}
  }}

  function showSeg(seg) {{
    const shown = [];
    for (const el of (segBoxes[seg] || [])) {{
      el.style.display = ""; place(el); shown.push(el);
    }}
    const info = infoBySeg[seg];
    if (info && socAxis) {{
      pinned.textContent = info.text;
      pinned.style.color = info.color;
      pinned.style.display = "";
      const w = pinned.offsetWidth, h = pinned.offsetHeight;
      let l = socAxis.x0 * dispW + PAD;
      let t = socAxis.y1 * dispH - PAD - h;   // bottom-aligned to the SOC panel bottom
      pinned.style.left = Math.max(0, Math.min(l, dispW - w)) + "px";
      pinned.style.top  = Math.max(0, Math.min(t, dispH - h)) + "px";
      shown.push(pinned);
    }}
    decollide(shown);
  }}
  function hideSeg(seg) {{
    for (const el of (segBoxes[seg] || [])) el.style.display = "none";
    pinned.style.display = "none";
  }}

  // Full-height hover hotzones (one per segment). Base (d/c) appended before
  // overlay (od/oc) so a finetuned figure surfaces the finetuned segment on top.
  for (const s of segs) {{
    const hz = document.createElement("div");
    hz.className = "hotzone";
    hz.style.left   = (s.x0 * dispW) + "px";
    hz.style.width  = Math.max(1, (s.x1 - s.x0) * dispW) + "px";
    hz.style.top    = "0px";
    hz.style.height = dispH + "px";
    const segId = s.seg;
    hz.addEventListener("mouseenter", () => showSeg(segId));
    hz.addEventListener("mouseleave", () => hideSeg(segId));
    overlayEl.appendChild(hz);
  }}
}}

figEl.addEventListener("load", layout);
window.addEventListener("resize", layout);

function show(i) {{
  cur = i;
  overlayEl.innerHTML = "";   // clear stale boxes until the new image lays out
  figEl.src = imgs[i];
  counterEl.textContent = (i + 1) + " / " + imgs.length;
  labelEl.textContent   = labels[i];
  document.querySelectorAll("#list li").forEach((el, j) =>
    el.classList.toggle("active", j === i));
  document.querySelectorAll("#list li")[i]
    .scrollIntoView({{block: "nearest"}});
  // Cached images may not fire 'load'; lay out immediately if already decoded.
  if (figEl.complete && figEl.naturalWidth) layout();
}}
function go(d) {{
  // Уи│У┐ЄУбФ filter жџљУЌЈуџёТЮАуЏ«№╝їтљдтѕЎТїЅ next С╝џТўЙуц║"уюІСИЇУДЂуџё"тЏЙ
  const n = imgs.length;
  let next = cur;
  for (let step = 0; step < n; step++) {{
    next = (next + d + n) % n;
    const el = document.querySelectorAll("#list li")[next];
    if (!el.classList.contains("hidden-day")) {{
      show(next);
      return;
    }}
  }}
}}
document.addEventListener("keydown", e => {{
  if (e.key === "ArrowLeft"  || e.key === "ArrowUp")   go(-1);
  if (e.key === "ArrowRight" || e.key === "ArrowDown")  go(1);
}});
// тѕЮтДІТўЙуц║№╝џУІЦ filter ти▓тІЙСИіСИћугг 0 т╝аТў» idle№╝їУи│тѕ░уггСИђСИф active
let initIdx = 0;
if (filterEl.checked) {{
  for (let i = 0; i < imgs.length; i++) {{
    if (isActive[i]) {{ initIdx = i; break; }}
  }}
}}
show(initIdx);
</script>
</body>
</html>
"""
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML viewer: {html_name}  ({len(figs)} figures)")


def _is_nan(v) -> bool:
    """Safe NaN check."""
    if v is None:
        return False
    try:
        return bool(np.isnan(v))
    except (TypeError, ValueError):
        return False
