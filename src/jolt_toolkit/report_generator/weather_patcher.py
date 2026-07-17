"""
weather_patcher.py
==================
Standalone weather data backfill tool.

Reads a generated Excel report, fetches historical weather data via the
OpenWeather API, and backfills the report's missing weather columns
(temperature, pressure, humidity, wind speed, wind direction).

Usage:
    from jolt_toolkit.report_generator.weather_patcher import WeatherPatcher
    patcher = WeatherPatcher()
    patcher.patch_folder("excel_report_database/1.0.0/KY24LHT/")
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from tqdm import tqdm

from jolt_toolkit.report_generator.paths import get_cache_dir
from jolt_toolkit.report_generator.report_builder import (
    HEADERS,
    is_trip_leg,
)
from jolt_toolkit.report_generator.weather_fetcher.openweather import (
    KeyManager,
    WeatherCache,
    WeatherFetcher,
)

logger = logging.getLogger(__name__)

# ── Excel column indices (1-based, openpyxl convention) ─────────────────
_COL_LEG_TYPE = 2
_COL_START_TIME = 6
_COL_ORIGIN = 7
_COL_END_TIME = 9
_COL_DEST = 10
_COL_TEMP = 38
_COL_PRESSURE = 39
_COL_HUMIDITY = 40
_COL_WIND_SPEED = 41
_COL_WIND_DIR = 42
_COL_WEATHER_TYPE = 43

_WEATHER_COLS = (
    _COL_TEMP,
    _COL_PRESSURE,
    _COL_HUMIDITY,
    _COL_WIND_SPEED,
    _COL_WIND_DIR,
    _COL_WEATHER_TYPE,
)

# Cheap sanity check (mirrors charger_patcher): the hard-coded 1-based Excel
# columns above must stay in step with report_builder.HEADERS (Excel col ==
# HEADERS.index(name) + 1). If HEADERS is ever reordered this fails loudly at
# import instead of silently patching the wrong cell. EV layout only — the
# diesel-layout guard in patch_file refuses DIESEL_HEADERS workbooks.
assert _COL_LEG_TYPE == HEADERS.index("Leg Type") + 1
assert _COL_START_TIME == HEADERS.index("Start Time (UTC)") + 1
assert _COL_ORIGIN == HEADERS.index("Origin (Lat, Lon)") + 1
assert _COL_END_TIME == HEADERS.index("End Time (UTC)") + 1
assert _COL_DEST == HEADERS.index("Destination (Lat, Lon)") + 1
assert _COL_TEMP == HEADERS.index("Average Temperature (C)") + 1
assert _COL_PRESSURE == HEADERS.index("Average Pressure (hPa)") + 1
assert _COL_HUMIDITY == HEADERS.index("Average Humidity (%)") + 1
assert _COL_WIND_SPEED == HEADERS.index("Average Wind Speed (m/s)") + 1
assert _COL_WIND_DIR == HEADERS.index("Average Wind Direction") + 1
assert _COL_WEATHER_TYPE == HEADERS.index("Weather Type") + 1


# ── Utility functions ─────────────────────────────────────────────────────


def _parse_point(point_str) -> tuple[float | None, float | None]:
    """Parse a coordinate string in 'Point(lat lon)' format."""
    if not point_str or not isinstance(point_str, str):
        return None, None
    m = re.match(r"Point\(([+-]?\d+\.?\d*)\s+([+-]?\d+\.?\d*)\)", point_str)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _deg_to_cardinal(deg: float) -> str:
    """Degrees → 8-point compass wind direction."""
    directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
    idx = round(deg / 45) % 8
    return directions[idx]


def _cell_needs_patch(cell) -> bool:
    """Return whether a cell needs weather data backfilled (empty value or the =NA() formula)."""
    v = cell.value
    if v is None:
        return True
    if isinstance(v, str) and (v.strip() == "" or v.strip().upper() == "=NA()"):
        return True
    return False


def _to_unix_utc(dt_val) -> int | None:
    """Convert a datetime value read by openpyxl to a UTC unix timestamp."""
    if dt_val is None:
        return None
    try:
        if isinstance(dt_val, datetime):
            # A datetime read by openpyxl is usually naive; the report labels it UTC
            if dt_val.tzinfo is None:
                dt_val = dt_val.replace(tzinfo=timezone.utc)
            return int(dt_val.timestamp())
        return None
    except (AttributeError, TypeError, ValueError):
        return None


def _is_ev_layout(ws) -> bool:
    """True iff the worksheet's header row matches the EV ``HEADERS`` layout.

    The ``_COL_*`` write indices assume the EV column order; a diesel report
    uses ``DIESEL_HEADERS`` (fewer columns, different order), so comparing the
    header row against ``HEADERS`` before writing prevents silently patching the
    wrong cells of a diesel / unknown workbook.
    """
    header = [ws.cell(1, c).value for c in range(1, len(HEADERS) + 1)]
    return header == list(HEADERS)


# ── API key management + weather cache ─────────────────────────────────────
# _KeyManager / _WeatherCache now live in weather_fetcher.openweather (shared
# with FineGrainedWeatherPatcher). The coarse cache writes no metadata block on
# init (metadata=None), preserving cache/.weather_cache.json's prior format.


# ── Main class ────────────────────────────────────────────────────────────


class WeatherPatcher:
    """
    Weather data backfill tool.

    Reads a generated xlsx report file, fetches the origin's and destination's
    historical weather data via the OpenWeather timemachine API, and writes the
    average into the report.

    Backfills **driving / trip rows only** (judged by ``is_trip_leg``): charge
    segments and Stop rows do not need weather, and querying them would only waste
    OpenWeather quota, so any non-trip row is skipped before its coordinates are
    ever collected (never enters the unique-location set at all).

    Args:
        cache_file:    cache file path (default ./cache/.weather_cache.json)
        precision:     coordinate cache precision (decimal places, default 2, ~1 km
                       grid). Together with ``time_bucket_s`` it lets nearby /
                       same-hour points share a cache key, greatly improving the hit
                       rate of cross-day / cross-vehicle backfill (the same cache key
                       as FineGrainedWeatherPatcher).
        time_bucket_s: time-bucket size (seconds, default 3600, aggregated by hour,
                       because the OpenWeather timemachine historical data is itself
                       at hourly granularity).
        max_workers:   number of concurrent requests (default 30)
    """

    def __init__(
        self,
        cache_file: str | Path | None = None,
        precision: int = 2,
        time_bucket_s: int = 3600,
        max_workers: int = 30,
    ):
        cache_file = cache_file or os.environ.get(
            "WEATHER_CACHE_FILE", str(get_cache_dir() / ".weather_cache.json")
        )
        self._cache = WeatherCache(cache_file, precision, time_bucket_s)
        self._keys = KeyManager("WeatherPatcher")
        self._fetcher = WeatherFetcher(self._keys, max_workers)

    # ── Public interface ──────────────────────────────────────────────────

    def patch_file(self, xlsx_path: str | Path) -> int:
        """
        Backfill a single xlsx report's weather data.

        Reads the coordinates and times of the **driving / trip rows** in the
        Report worksheet (charge / Stop rows are skipped, see the class docstring),
        queries the OpenWeather API, and writes temperature/pressure/humidity/wind
        speed/wind direction into the corresponding columns.

        Returns:
            The number of rows backfilled.
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.error(f"File not found: {xlsx_path}")
            return 0

        logger.info(f"Patching: {xlsx_path.name}")
        wb = load_workbook(str(xlsx_path))
        if "Report" not in wb.sheetnames:
            logger.error(f"  'Report' sheet not found in {xlsx_path.name}")
            wb.close()
            return 0
        ws = wb["Report"]

        # Diesel-layout guard: this patcher's _COL_* indices are the EV HEADERS
        # layout. A diesel report uses DIESEL_HEADERS (different column order),
        # so patching one here would silently write weather into the wrong
        # cells. Detect it by comparing the header row against HEADERS and skip.
        if not _is_ev_layout(ws):
            logger.error(
                "  %s: diesel/unknown layout, skipping (coarse weather patcher "
                "supports the EV layout only)",
                xlsx_path.name,
            )
            wb.close()
            return 0

        # 1. Collect the rows needing backfill
        tasks = []  # (row_idx, o_lat, o_lon, o_dt, d_lat, d_lon, d_dt)
        total_rows = ws.max_row - 1  # minus the header row
        for row_idx in tqdm(
            range(2, ws.max_row + 1),
            desc="Scanning weather rows",
            total=total_rows,
            leave=False,
        ):  # skip the header row
            # Weather is backfilled on driving / trip rows ONLY. Charge and Stop
            # rows do not need weather, and querying them would only waste
            # OpenWeather quota, so skip any non-trip row before its coordinates
            # are ever collected (kept out of the unique-location set entirely).
            # ``is_trip_leg`` is the shared trip definition used by the chart
            # ``driving_only`` filter, so both agree on what a trip is.
            if not is_trip_leg(ws.cell(row_idx, _COL_LEG_TYPE).value):
                continue
            if not any(_cell_needs_patch(ws.cell(row_idx, c)) for c in _WEATHER_COLS):
                continue

            origin_str = ws.cell(row_idx, _COL_ORIGIN).value
            dest_str = ws.cell(row_idx, _COL_DEST).value
            o_lat, o_lon = _parse_point(origin_str)
            d_lat, d_lon = _parse_point(dest_str)
            if o_lat is None or d_lat is None:
                continue

            o_dt = _to_unix_utc(ws.cell(row_idx, _COL_START_TIME).value)
            d_dt = _to_unix_utc(ws.cell(row_idx, _COL_END_TIME).value)
            if o_dt is None or d_dt is None:
                continue

            tasks.append((row_idx, o_lat, o_lon, o_dt, d_lat, d_lon, d_dt))

        if not tasks:
            logger.info(f"  No rows need weather patching in {xlsx_path.name}")
            wb.close()
            return 0

        logger.info(f"  {len(tasks)} rows need weather data")

        # 2. Collect all unique (lat, lon, dt) locations
        loc_set: set[tuple] = set()
        for _, o_lat, o_lon, o_dt, d_lat, d_lon, d_dt in tasks:
            loc_set.add((o_lat, o_lon, o_dt))
            loc_set.add((d_lat, d_lon, d_dt))
        all_locs = list(loc_set)

        # 3. Look up the cache
        weather_map, missing = self._cache.get_batch(all_locs)
        logger.info(
            f"  {len(all_locs)} unique locations: "
            f"{len(weather_map)} cached, {len(missing)} need API"
        )

        # 4. Fetch the missing data from the API
        if missing:
            fetched = self._fetcher.fetch_batch(
                missing, desc="Fetching weather API", warn_on_failure=True
            )
            weather_map.update(fetched)
            if fetched:
                self._cache.put_batch(fetched)
                logger.info(f"  Fetched and cached {len(fetched)} new locations")

        # 5. Write into Excel
        patched = 0
        for row_idx, o_lat, o_lon, o_dt, d_lat, d_lon, d_dt in tasks:
            o_w = weather_map.get((o_lat, o_lon, o_dt))
            d_w = weather_map.get((d_lat, d_lon, d_dt))
            if o_w is None or d_w is None:
                continue

            avg_temp = round((o_w[0] + d_w[0]) / 2, 1)
            avg_press = round((o_w[1] + d_w[1]) / 2, 1)
            avg_humid = round((o_w[2] + d_w[2]) / 2, 1)
            avg_wind = round((o_w[3] + d_w[3]) / 2, 1)
            avg_dir = (o_w[4] + d_w[4]) / 2
            cardinal = _deg_to_cardinal(avg_dir)
            # Origin's weather type (compatible with the old 5-element cache)
            weather_type = o_w[5] if len(o_w) > 5 else None

            ws.cell(row_idx, _COL_TEMP).value = avg_temp
            ws.cell(row_idx, _COL_PRESSURE).value = avg_press
            ws.cell(row_idx, _COL_HUMIDITY).value = avg_humid
            ws.cell(row_idx, _COL_WIND_SPEED).value = avg_wind
            ws.cell(row_idx, _COL_WIND_DIR).value = cardinal
            ws.cell(row_idx, _COL_WEATHER_TYPE).value = weather_type
            patched += 1

        if patched > 0:
            wb.save(str(xlsx_path))
            logger.info(f"  Patched {patched} rows, saved {xlsx_path.name}")
        else:
            logger.info(f"  No weather data available for any rows")

        wb.close()

        summary = self._keys.summary()
        logger.info(
            f"  API keys: {summary['active']}/{summary['total_keys']} active, "
            f"{summary['total_usage']} calls this session"
        )

        return patched

    def patch_folder(self, folder_path: str | Path) -> dict[str, int]:
        """
        Backfill the weather data of all jolt_report_*.xlsx reports under a folder.

        Returns:
            A {file name: number of rows backfilled} dict.
        """
        folder = Path(folder_path)
        if not folder.is_dir():
            logger.error(f"Folder not found: {folder}")
            return {}

        xlsx_files = sorted(folder.glob("jolt_report_*.xlsx"))
        if not xlsx_files:
            logger.info(f"No jolt_report_*.xlsx files in {folder}")
            return {}

        logger.info(f"Found {len(xlsx_files)} report(s) in {folder}")
        results = {}
        for fp in tqdm(xlsx_files, desc="Weather patch files"):
            results[fp.name] = self.patch_file(fp)
        return results
