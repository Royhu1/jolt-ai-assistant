"""
fine_grained_patcher.py
=======================
Fine-grained weather data backfill tool (FineGrainedWeatherPatcher).

Differences from ``weather_patcher.WeatherPatcher``:
    - The old patcher uses only 2 sample points (trip origin / destination), too coarse for long trips.
    - This patcher selects all ``(eventDatetime, latitude, longitude)`` triples in
      the raw_telematics CSV by the trip time window, downsamples them by
      ``min_sample_interval_s`` (default 60 s), queries the OpenWeather History API
      for each sample point, and finally aggregates per column:
        * numeric columns (temp / pressure / humidity / wind_speed / wind_deg) → average
          (wind_deg is converted to an 8-point cardinal string before writing, consistent with the old patcher)
        * text columns (Weather Type / Description) → mode

Cache design:
    Same JSON schema as the old patcher, but by default at
    ``cache/weather/.weather_cache_fine.json`` (to avoid polluting the old main cache).
    The key is still ``"{lat:.{precision}f},{lon:.{precision}f},{dt}"``, with configurable precision:
        precision=2  → ~1 km grid (about 0.01°), time bucket quantised by hour
        precision=4  → ~10 m grid (still uses the hour time bucket by default)
    The time bucket is controlled by the constructor parameter ``time_bucket_s``
    (default 3600, i.e. aggregated by hour, because the OpenWeather timemachine API
    is at hourly granularity).

Usage:
    from jolt_toolkit.report_generator.weather_fetcher.fine_grained_patcher \
        import FineGrainedWeatherPatcher
    patcher = FineGrainedWeatherPatcher(
        raw_telematics_dir="excel_report_database/<version>/YK73WFN/raw_telematics",
        min_sample_interval_s=60,
    )
    patcher.patch_file("excel_report_database/<version>/YK73WFN/jolt_report_*.xlsx")

Notes:
    - The patcher is a standalone post-processing tool, **not embedded** in the
      ``generate_report()`` flow, consistent with the old ``WeatherPatcher``.
    - Deleting the cache file does not break recoverability (the cache is only a
      local copy of the API results).
"""

from __future__ import annotations

import logging
import os
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from jolt_toolkit.report_generator.paths import get_cache_dir
from jolt_toolkit.report_generator.report_builder import (
    DIESEL_HEADERS,
    HEADERS,
    is_trip_leg,
)
from jolt_toolkit.report_generator.weather_fetcher.openweather import (
    KeyManager,
    WeatherCache,
    WeatherFetcher,
)

logger = logging.getLogger(__name__)

# ── Excel column-name set (aligned with HEADERS / DIESEL_HEADERS, 1-based indices derived dynamically) ──
_TEMP_COL_NAME = "Average Temperature (C)"
_PRESS_COL_NAME = "Average Pressure (hPa)"
_HUMID_COL_NAME = "Average Humidity (%)"
_WIND_SPEED_COL_NAME = "Average Wind Speed (m/s)"
_WIND_DIR_COL_NAME = "Average Wind Direction"
_WEATHER_TYPE_NAME = "Weather Type"
_LEG_TYPE_COL_NAME = "Leg Type"
_START_TIME_COL_NAME = "Start Time (UTC)"
_END_TIME_COL_NAME = "End Time (UTC)"
_ORIGIN_COL_NAME = "Origin (Lat, Lon)"
_DEST_COL_NAME = "Destination (Lat, Lon)"


# ── Generic utilities ─────────────────────────────────────────────────────


def _resolve_col_indices(headers: tuple) -> dict[str, int]:
    """Derive the 1-based column indices dynamically from the HEADERS tuple."""
    return {
        "leg_type": headers.index(_LEG_TYPE_COL_NAME) + 1,
        "start_time": headers.index(_START_TIME_COL_NAME) + 1,
        "end_time": headers.index(_END_TIME_COL_NAME) + 1,
        "origin": headers.index(_ORIGIN_COL_NAME) + 1,
        "destination": headers.index(_DEST_COL_NAME) + 1,
        "temp": headers.index(_TEMP_COL_NAME) + 1,
        "pressure": headers.index(_PRESS_COL_NAME) + 1,
        "humidity": headers.index(_HUMID_COL_NAME) + 1,
        "wind_speed": headers.index(_WIND_SPEED_COL_NAME) + 1,
        "wind_dir": headers.index(_WIND_DIR_COL_NAME) + 1,
        "weather_type": headers.index(_WEATHER_TYPE_NAME) + 1,
    }


def _parse_point(point_str) -> tuple[float | None, float | None]:
    """Parse a coordinate string in 'Point(lat lon)' format."""
    if not point_str or not isinstance(point_str, str):
        return None, None
    m = re.match(r"Point\(([+-]?\d+\.?\d*)\s+([+-]?\d+\.?\d*)\)", point_str)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _deg_to_cardinal(deg: float) -> str:
    """Degrees → 8-point cardinal."""
    directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
    idx = round(deg / 45) % 8
    return directions[idx]


def _to_utc_dt(dt_val) -> datetime | None:
    """A date value read by openpyxl → UTC datetime."""
    if dt_val is None:
        return None
    if isinstance(dt_val, datetime):
        if dt_val.tzinfo is None:
            dt_val = dt_val.replace(tzinfo=timezone.utc)
        return dt_val
    return None


def _cell_needs_patch(cell) -> bool:
    """A cell that is empty / =NA() / NaN is treated as needing backfill."""
    v = cell.value
    if v is None:
        return True
    if isinstance(v, str) and (v.strip() == "" or v.strip().upper() == "=NA()"):
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    return False


# ── API key management + cache ──────────────────────────────────────────
# _KeyManager / _WeatherCache now live in weather_fetcher.openweather (shared
# with the coarse WeatherPatcher). This patcher passes the fine cache's
# metadata block + init log so the freshly-created
# cache/weather/.weather_cache_fine.json keeps its prior format.


# ── raw_telematics CSV cache ─────────────────────────────────────────────


class _RawTelematicsIndex:
    """
    raw_telematics directory index: lazily loads CSVs on demand, extracts the
    ``(timestamp_utc, latitude, longitude)`` sequence, and caches it in memory.

    Each CSV file covers one day (named ``raw_YYYY-MM-DD_NNNN.csv``).
    """

    def __init__(self, raw_dir: Path):
        self._raw_dir = Path(raw_dir)
        # date(str YYYY-MM-DD) → DataFrame[ts, lat, lon]
        self._cache: dict[str, pd.DataFrame] = {}
        # File names indexed by date prefix
        self._date_to_file: dict[str, Path] = {}
        if self._raw_dir.is_dir():
            for fp in self._raw_dir.glob("raw_*.csv"):
                m = re.match(r"raw_(\d{4}-\d{2}-\d{2})_\d+\.csv$", fp.name)
                if m:
                    self._date_to_file[m.group(1)] = fp

    @property
    def available(self) -> bool:
        return bool(self._date_to_file)

    # The two lat/lon schemas of the raw_telematics CSV (in priority order):
    #   Early CSVs (about ≤2025-10) have both gnss_latitude/longitude and
    #     latitude/longitude, with identical values (measured max abs diff = 0).
    #   Recent CSVs (from about 2025-11) have only latitude/longitude, with the
    #     gnss_* columns removed.
    # Prefer gnss_* (consistent with the canonical history), falling back to
    # latitude/longitude when missing, so both schemas can be read and both take
    # the true multi-point sampling path.
    _LAT_CANDIDATES = ("gnss_latitude", "latitude")
    _LON_CANDIDATES = ("gnss_longitude", "longitude")

    def _load_day(self, date_str: str) -> Optional[pd.DataFrame]:
        if date_str in self._cache:
            return self._cache[date_str]
        fp = self._date_to_file.get(date_str)
        if fp is None or not fp.is_file():
            self._cache[date_str] = pd.DataFrame(columns=["ts", "lat", "lon"])
            return self._cache[date_str]

        # Probe the actual header first, then pick usecols by the columns that
        # exist; missing gnss_* no longer hard-raises "Usecols do not match
        # columns" (which made recent CSVs unreadable and degrade to two endpoints).
        try:
            available = set(pd.read_csv(fp, nrows=0).columns)
        except Exception as exc:
            logger.warning(f"Failed to read header of {fp.name}: {exc}")
            self._cache[date_str] = pd.DataFrame(columns=["ts", "lat", "lon"])
            return self._cache[date_str]

        lat_cols = [c for c in self._LAT_CANDIDATES if c in available]
        lon_cols = [c for c in self._LON_CANDIDATES if c in available]
        if "eventDatetime" not in available or not lat_cols or not lon_cols:
            logger.warning(
                f"{fp.name}: missing eventDatetime / lat / lon columns "
                f"(lat={lat_cols}, lon={lon_cols}); skipping"
            )
            self._cache[date_str] = pd.DataFrame(columns=["ts", "lat", "lon"])
            return self._cache[date_str]

        usecols = ["eventDatetime", *lat_cols, *lon_cols]
        try:
            df = pd.read_csv(fp, usecols=usecols, low_memory=False)
        except Exception as exc:
            logger.warning(f"Failed to read {fp.name}: {exc}")
            self._cache[date_str] = pd.DataFrame(columns=["ts", "lat", "lon"])
            return self._cache[date_str]

        # Prefer gnss_* (candidate columns already sorted by priority), falling back to latitude/longitude for missing values
        lat = pd.to_numeric(df[lat_cols[0]], errors="coerce")
        for c in lat_cols[1:]:
            lat = lat.fillna(pd.to_numeric(df[c], errors="coerce"))
        lon = pd.to_numeric(df[lon_cols[0]], errors="coerce")
        for c in lon_cols[1:]:
            lon = lon.fillna(pd.to_numeric(df[c], errors="coerce"))
        ts = pd.to_datetime(df["eventDatetime"], utc=True, errors="coerce")

        df_out = pd.DataFrame({"ts": ts, "lat": lat, "lon": lon}).dropna()
        df_out = df_out.sort_values("ts").reset_index(drop=True)
        self._cache[date_str] = df_out
        return df_out

    def slice_trip(self, t_start: datetime, t_end: datetime) -> pd.DataFrame:
        """Extract all GPS points within the [t_start, t_end] time window (merging across days)."""
        if t_start.tzinfo is None:
            t_start = t_start.replace(tzinfo=timezone.utc)
        if t_end.tzinfo is None:
            t_end = t_end.replace(tzinfo=timezone.utc)

        ts_start = pd.Timestamp(t_start).tz_convert("UTC")
        ts_end = pd.Timestamp(t_end).tz_convert("UTC")

        dfs = []
        cur = ts_start.normalize()
        last = ts_end.normalize()
        while cur <= last:
            date_str = cur.strftime("%Y-%m-%d")
            df_day = self._load_day(date_str)
            if df_day is not None and not df_day.empty:
                dfs.append(df_day)
            cur = cur + pd.Timedelta(days=1)

        if not dfs:
            return pd.DataFrame(columns=["ts", "lat", "lon"])

        df_all = pd.concat(dfs, ignore_index=True)
        mask = (df_all["ts"] >= ts_start) & (df_all["ts"] <= ts_end)
        return df_all.loc[mask].reset_index(drop=True)


def _downsample_by_interval(df: pd.DataFrame, min_interval_s: int) -> pd.DataFrame:
    """
    Downsample by time interval: keep the first row, and keep subsequent rows
    only when the timestamp differs from the last kept row by ≥ ``min_interval_s``.
    """
    if df.empty:
        return df
    if min_interval_s <= 0:
        return df.reset_index(drop=True)

    keep_idx = [0]
    last_ts = df["ts"].iloc[0]
    interval = pd.Timedelta(seconds=min_interval_s)
    for i in range(1, len(df)):
        ts = df["ts"].iloc[i]
        if ts - last_ts >= interval:
            keep_idx.append(i)
            last_ts = ts
    return df.iloc[keep_idx].reset_index(drop=True)


# ── Main class ────────────────────────────────────────────────────────────


class FineGrainedWeatherPatcher:
    """
    Fine-grained weather data backfill tool.

    Args:
        raw_telematics_dir:     the directory holding raw_*.csv (usually
                                ``excel_report_database/{ver}/{REG}/raw_telematics/``).
                                If None, tries to auto-locate the raw_telematics directory next to the xlsx.
        min_sample_interval_s:  minimum sampling interval within a trip (seconds, default 60).
        cache_file:             cache JSON file path, default
                                ``cache/weather/.weather_cache_fine.json``.
        precision:              coordinate quantisation precision (default 2, ~1 km grid).
                                The OpenWeather timemachine itself returns historical
                                data by hour, and temperature/wind-speed variation
                                within ~1 km is negligible, so precision=2 preserves
                                accuracy while greatly improving cross-vehicle cache hits.
        time_bucket_s:          time-bucket size (default 3600, aggregated by hour).
        max_workers:            number of concurrent API requests (default 20).
        headers:                column structure, default the EV ``HEADERS``; pass ``DIESEL_HEADERS`` for diesel.
    """

    def __init__(
        self,
        raw_telematics_dir: str | Path | None = None,
        min_sample_interval_s: int = 60,
        cache_file: str | Path | None = None,
        precision: int = 2,
        time_bucket_s: int = 3600,
        max_workers: int = 20,
        headers: tuple = HEADERS,
    ):
        self._raw_dir = Path(raw_telematics_dir) if raw_telematics_dir else None
        self._raw_index: Optional[_RawTelematicsIndex] = None
        self._min_interval = max(int(min_sample_interval_s), 0)
        self._headers = headers
        self._col_idx = _resolve_col_indices(headers)

        cache_file = cache_file or os.environ.get(
            "WEATHER_CACHE_FILE_FINE",
            str(get_cache_dir() / "weather" / ".weather_cache_fine.json"),
        )
        # Fine cache keeps its metadata block + init log (coarse writes neither).
        cache_metadata = {
            "version": "1.0",
            "precision": precision,
            "time_bucket_s": max(int(time_bucket_s), 1),
            "description": (
                "FineGrainedWeatherPatcher cache "
                "(value = [temp, pressure, humidity, wind_speed, "
                "wind_deg, weather_type])"
            ),
        }
        self._cache = WeatherCache(
            cache_file,
            precision=precision,
            time_bucket_s=time_bucket_s,
            metadata=cache_metadata,
            init_log=f"Initialized fine-grained cache: {Path(cache_file)}",
        )
        self._keys = KeyManager("FineGrainedWeatherPatcher")
        self._fetcher = WeatherFetcher(self._keys, max_workers)

        # Statistics counters (reset each patch_file); api_calls / failures are recorded by the fetcher
        self._stat_cache_hits = 0

    # ── Public interface ──────────────────────────────────────────────────

    def patch_file(
        self, xlsx_path: str | Path, overwrite: bool = True, force_repatch: bool = False
    ) -> dict:
        """
        Backfill a single xlsx report.

        Args:
            xlsx_path:     the target xlsx file.
            overwrite:     True overwrites the original file directly; False writes ``*_fineweather.xlsx``.
            force_repatch: when True, ignore the ``_cell_needs_patch`` decision and
                           rewrite the weather columns for all trip-like rows (used
                           for a fine-grained recompute to overwrite old coarse results).

        Returns:
            A statistics dict: ``{patched_rows, total_samples, api_calls, cache_hits, failures}``.
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.error(f"File not found: {xlsx_path}")
            return {
                "patched_rows": 0,
                "total_samples": 0,
                "api_calls": 0,
                "cache_hits": 0,
                "failures": 0,
            }

        # Auto-locate raw_telematics
        raw_dir = self._raw_dir or (xlsx_path.parent / "raw_telematics")
        self._raw_index = _RawTelematicsIndex(raw_dir)
        if not self._raw_index.available:
            logger.error(
                f"raw_telematics directory not found or empty: {raw_dir}. "
                f"FineGrainedWeatherPatcher requires --debug raw CSVs."
            )
            return {
                "patched_rows": 0,
                "total_samples": 0,
                "api_calls": 0,
                "cache_hits": 0,
                "failures": 0,
            }

        logger.info(
            f"Fine-grained patching: {xlsx_path.name} "
            f"(min_interval={self._min_interval}s)"
        )
        wb = load_workbook(str(xlsx_path))
        if "Report" not in wb.sheetnames:
            logger.error(f"  'Report' sheet not found")
            wb.close()
            return {
                "patched_rows": 0,
                "total_samples": 0,
                "api_calls": 0,
                "cache_hits": 0,
                "failures": 0,
            }
        ws = wb["Report"]

        # Reset statistics (api_calls / failures are counted on the fetcher)
        self._fetcher.reset_stats()
        self._stat_cache_hits = 0

        # 1. Scan the rows needing backfill: backfill **driving / trip rows only**
        #    (is_trip_leg); charge / Stop rows do not need weather and querying them
        #    would only waste OpenWeather quota, so skip them directly.
        weather_cols = (
            self._col_idx["temp"],
            self._col_idx["pressure"],
            self._col_idx["humidity"],
            self._col_idx["wind_speed"],
            self._col_idx["wind_dir"],
            self._col_idx["weather_type"],
        )
        tasks: list[dict] = []
        total_rows = ws.max_row - 1
        for row_idx in range(2, ws.max_row + 1):
            if not force_repatch:
                if not any(
                    _cell_needs_patch(ws.cell(row_idx, c)) for c in weather_cols
                ):
                    continue

            leg_type = ws.cell(row_idx, self._col_idx["leg_type"]).value
            # Backfill driving / trip rows only; charge segments and Stop rows are
            # skipped before sample points are collected. is_trip_leg is the trip
            # definition shared with the chart's driving_only filter.
            if not is_trip_leg(leg_type):
                continue
            t_s = _to_utc_dt(ws.cell(row_idx, self._col_idx["start_time"]).value)
            t_e = _to_utc_dt(ws.cell(row_idx, self._col_idx["end_time"]).value)
            if t_s is None or t_e is None:
                continue

            # Within trip rows there are two sampling modes: those labelled
            # "Trip"/"Transit" have a continuous GPS track and take multi-point
            # sampling; the remaining trip rows (Outbound/Return/In House) fall back
            # to the origin/dest endpoints (charge/Stop were already skipped above
            # and never reach here).
            is_moving = isinstance(leg_type, str) and (
                "Trip" in leg_type or "Transit" in leg_type
            )

            origin_pt = _parse_point(ws.cell(row_idx, self._col_idx["origin"]).value)
            dest_pt = _parse_point(ws.cell(row_idx, self._col_idx["destination"]).value)

            tasks.append(
                {
                    "row": row_idx,
                    "t_s": t_s,
                    "t_e": t_e,
                    "is_moving": bool(is_moving),
                    "origin": origin_pt,
                    "dest": dest_pt,
                }
            )

        if not tasks:
            logger.info(f"  No rows need weather patching")
            wb.close()
            return {
                "patched_rows": 0,
                "total_samples": 0,
                "api_calls": 0,
                "cache_hits": 0,
                "failures": 0,
            }

        logger.info(f"  {len(tasks)} rows need weather data (of {total_rows} total)")

        # 2. Collect sample points (lat, lon, dt_unix) for each task
        task_samples: dict[int, list[tuple[float, float, int]]] = {}
        all_locs: set[tuple] = set()
        total_samples = 0

        for t in tasks:
            samples = self._collect_samples_for_trip(t)
            task_samples[t["row"]] = samples
            for loc in samples:
                all_locs.add(loc)
            total_samples += len(samples)

        logger.info(
            f"  Collected {total_samples} samples across {len(tasks)} rows, "
            f"{len(all_locs)} unique (after cache-key quantization)"
        )

        # 3. Look up the cache + fetch the API
        all_locs_list = list(all_locs)
        weather_map, missing = self._cache.get_batch(all_locs_list)
        self._stat_cache_hits = len(weather_map)
        logger.info(
            f"  {self._stat_cache_hits}/{len(all_locs_list)} cache hits, "
            f"{len(missing)} need API"
        )

        if missing:
            fetched = self._fetcher.fetch_batch(
                missing, desc="OpenWeather API", warn_on_failure=False
            )
            weather_map.update(fetched)
            if fetched:
                self._cache.put_batch(fetched)
                logger.info(f"  Fetched and cached {len(fetched)} new entries")

        # 4. Aggregate + write back to xlsx
        patched = 0
        for t in tasks:
            samples = task_samples[t["row"]]
            agg = self._aggregate_weather(samples, weather_map)
            if agg is None:
                continue
            temp, press, humid, wind_s, wind_deg, w_type = agg
            ws.cell(t["row"], self._col_idx["temp"]).value = temp
            ws.cell(t["row"], self._col_idx["pressure"]).value = press
            ws.cell(t["row"], self._col_idx["humidity"]).value = humid
            ws.cell(t["row"], self._col_idx["wind_speed"]).value = wind_s
            ws.cell(t["row"], self._col_idx["wind_dir"]).value = _deg_to_cardinal(
                wind_deg
            )
            ws.cell(t["row"], self._col_idx["weather_type"]).value = w_type
            patched += 1

        # 5. Save
        if patched > 0:
            out_path = (
                xlsx_path
                if overwrite
                else xlsx_path.with_name(
                    xlsx_path.stem + "_fineweather" + xlsx_path.suffix
                )
            )
            wb.save(str(out_path))
            logger.info(f"  Patched {patched} rows, saved {out_path.name}")
        else:
            logger.info(f"  No weather data available for any rows")
        wb.close()

        summary = self._keys.summary()
        logger.info(
            f"  Summary: {patched} rows patched, {total_samples} samples, "
            f"{self._fetcher.api_calls} API calls, {self._stat_cache_hits} cache hits, "
            f"{self._fetcher.failures} failures"
        )
        logger.info(
            f"  API keys: {summary['active']}/{summary['total_keys']} active, "
            f"{summary['total_usage']} total calls this session"
        )

        return {
            "patched_rows": patched,
            "total_samples": total_samples,
            "api_calls": self._fetcher.api_calls,
            "cache_hits": self._stat_cache_hits,
            "failures": self._fetcher.failures,
        }

    def patch_folder(
        self,
        folder_path: str | Path,
        overwrite: bool = True,
        force_repatch: bool = False,
    ) -> dict[str, dict]:
        """Backfill all ``jolt_report_*.xlsx`` under a folder (excluding ``*_finetuned.xlsx``)."""
        folder = Path(folder_path)
        if not folder.is_dir():
            logger.error(f"Folder not found: {folder}")
            return {}
        xlsx_files = sorted(
            fp
            for fp in folder.glob("jolt_report_*.xlsx")
            if not fp.stem.endswith("_finetuned")
        )
        if not xlsx_files:
            logger.info(f"No jolt_report_*.xlsx files in {folder}")
            return {}

        results: dict[str, dict] = {}
        for fp in tqdm(xlsx_files, desc="fine-grained weather"):
            results[fp.name] = self.patch_file(
                fp,
                overwrite=overwrite,
                force_repatch=force_repatch,
            )
        return results

    # ── Sample collection ─────────────────────────────────────────────────

    def _collect_samples_for_trip(self, task: dict) -> list[tuple[float, float, int]]:
        """
        Collect the sample points for one trip.

        Moving legs (labelled Trip / Transit): slice the [t_s, t_e] time window
        from raw_telematics and downsample by ``min_sample_interval_s``. If there is
        not a single point, fall back to the origin / dest endpoints.

        The remaining trip rows (Outbound / Return / In House, lacking the
        Trip/Transit label): use the origin / dest endpoints directly (if any).
        Charge / Stop rows never reach here — they were already skipped by
        ``is_trip_leg`` during the ``patch_file`` scan stage.
        """
        t_s, t_e = task["t_s"], task["t_e"]
        origin = task["origin"]
        dest = task["dest"]

        samples: list[tuple[float, float, int]] = []

        if task["is_moving"] and self._raw_index is not None:
            df = self._raw_index.slice_trip(t_s, t_e)
            df = _downsample_by_interval(df, self._min_interval)
            for _, r in df.iterrows():
                lat = float(r["lat"])
                lon = float(r["lon"])
                ts = int(r["ts"].timestamp())
                samples.append((lat, lon, ts))

        # fallback: origin / destination
        if not samples:
            o_lat, o_lon = origin
            d_lat, d_lon = dest
            if o_lat is not None and o_lon is not None:
                samples.append((o_lat, o_lon, int(t_s.timestamp())))
            if d_lat is not None and d_lon is not None:
                samples.append((d_lat, d_lon, int(t_e.timestamp())))

        return samples

    # ── Aggregation ───────────────────────────────────────────────────────

    @staticmethod
    def _aggregate_weather(samples: list[tuple], weather_map: dict) -> Optional[tuple]:
        """
        Aggregate the weather tuples of several (lat, lon, dt) samples into a single trip-level value.

        Returns: (temp, pressure, humidity, wind_speed, wind_deg, weather_type)
                 or None (if no sample obtained weather).
        """
        if not samples:
            return None
        temps, presses, humids, winds, degs, types = [], [], [], [], [], []
        for loc in samples:
            w = weather_map.get(loc)
            if w is None:
                continue
            # Compatible with 5-element (old cache) / 6-element
            temps.append(float(w[0]))
            presses.append(float(w[1]))
            humids.append(float(w[2]))
            winds.append(float(w[3]))
            degs.append(float(w[4]))
            if len(w) > 5 and w[5] is not None:
                types.append(str(w[5]))
        if not temps:
            return None

        temp = round(float(np.mean(temps)), 1)
        press = round(float(np.mean(presses)), 1)
        humid = round(float(np.mean(humids)), 1)
        wind_s = round(float(np.mean(winds)), 1)
        # Wind direction: average via sin/cos then back-compute (avoids 359°/1° averaging back to 180°)
        sin_mean = float(np.mean([np.sin(np.deg2rad(d)) for d in degs]))
        cos_mean = float(np.mean([np.cos(np.deg2rad(d)) for d in degs]))
        wind_deg = (np.rad2deg(np.arctan2(sin_mean, cos_mean)) + 360) % 360

        if types:
            cnt = Counter(types).most_common()
            w_type = cnt[0][0]  # take the first on a tie
        else:
            w_type = None

        return temp, press, humid, wind_s, wind_deg, w_type
