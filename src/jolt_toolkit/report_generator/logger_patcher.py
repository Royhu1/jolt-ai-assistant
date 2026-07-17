"""
logger_patcher.py
=================
Standalone SRF Logger data backfill tool.

Reads a generated Excel report, fetches Logger legs from the SRF API, and
backfills the report's missing Logger Link, weather columns and mass data.

Usage:
    from jolt_toolkit.report_generator.logger_patcher import LoggerPatcher
    patcher = LoggerPatcher()
    patcher.patch_file("excel_report_database/1.0.0/YK73WFN/jolt_report_YK73WFN_20250820_20250822.xlsx")
"""

from __future__ import annotations

import datetime
import logging
import os
import re
from pathlib import Path

import numpy as np
import pandas as pd
import srf_client
from openpyxl import load_workbook
from openpyxl.styles import Font
from srf_client import paging
from tqdm import tqdm

from jolt_toolkit.report_generator.paths import get_cache_dir
from jolt_toolkit.report_generator.pedal_histogram import (
    EBC1_COL,
    EEC2_COL,
    MIN_DISTANCE_FOR_PEDAL_KM,
    compute_pedal_histogram,
)
from jolt_toolkit.report_generator.report_builder import (
    HEADERS,
    _build_logger_url,
    _find_overlap,
    _get_trip_speed_array,
    _kinetics_corrected_energy_perf,
)
from jolt_toolkit.report_generator.segment_algorithms import VEHICLE_CONFIG
from jolt_toolkit.report_generator.xlsx_patch_common import (
    _cell_is_empty,
    _parse_report_filename,
    _to_timestamp,
    make_srf_client,
)

logger = logging.getLogger(__name__)

# ── Excel column indices (1-based, openpyxl convention) ─────────────────
_COL_LEG_TYPE = 2  # Leg Type
_COL_LOGGER_LINK = 5  # SRF Logger Link
_COL_START_TIME = 6
_COL_END_TIME = 9
_COL_MASS = 16  # Vehicle Mass (kg)
_COL_MASS_CV = 17  # Vehicle Mass CV (reliability)
_COL_TEMP = 38
_COL_PRESSURE = 39
_COL_HUMIDITY = 40
_COL_WIND_SPEED = 41
_COL_WIND_DIR = 42

_COL_DISTANCE = 13  # Distance (km)
_COL_ELEV_DIFF = 15  # Elevation Difference (m)
_COL_ENERGY = 22  # Energy Change (kWh)
_COL_EPERF_KIN = 47  # Energy Performance Kinetics Corrected (kWh/km)
_COL_HIST_ACC = 44  # Histogram of Accelerator Pedal Position
_COL_HIST_DEC = 45  # Histogram of Decelerator Pedal Position

_WEATHER_COLS = (
    _COL_TEMP,
    _COL_PRESSURE,
    _COL_HUMIDITY,
    _COL_WIND_SPEED,
    _COL_WIND_DIR,
)

# Cheap sanity check (mirrors charger_patcher): the hard-coded 1-based Excel
# columns above must stay in step with report_builder.HEADERS (Excel col ==
# HEADERS.index(name) + 1). Any HEADERS reorder that moves one of these fields
# fails loudly at import instead of silently patching the wrong cell. This only
# guards the EV layout — diesel goes through DIESEL_HEADERS and never reaches the
# LoggerPatcher.
assert _COL_LEG_TYPE == HEADERS.index("Leg Type") + 1
assert _COL_LOGGER_LINK == HEADERS.index("SRF Logger Link") + 1
assert _COL_START_TIME == HEADERS.index("Start Time (UTC)") + 1
assert _COL_END_TIME == HEADERS.index("End Time (UTC)") + 1
assert _COL_MASS == HEADERS.index("Vehicle Mass (kg)") + 1
assert _COL_MASS_CV == HEADERS.index("Vehicle Mass CV (reliability)") + 1
assert _COL_TEMP == HEADERS.index("Average Temperature (C)") + 1
assert _COL_PRESSURE == HEADERS.index("Average Pressure (hPa)") + 1
assert _COL_HUMIDITY == HEADERS.index("Average Humidity (%)") + 1
assert _COL_WIND_SPEED == HEADERS.index("Average Wind Speed (m/s)") + 1
assert _COL_WIND_DIR == HEADERS.index("Average Wind Direction") + 1
assert _COL_DISTANCE == HEADERS.index("Distance (km)") + 1
assert _COL_ELEV_DIFF == HEADERS.index("Elevation Difference (m)") + 1
assert _COL_ENERGY == HEADERS.index("Energy Change (kWh)") + 1
assert (
    _COL_EPERF_KIN
    == HEADERS.index("Energy Performance Kinetics Corrected (kWh/km)") + 1
)
assert _COL_HIST_ACC == HEADERS.index("Histogram of Accelerator Pedal Position") + 1
assert _COL_HIST_DEC == HEADERS.index("Histogram of Decelerator Pedal Position") + 1

# ── Logger Channel 7 column names ─────────────────────────────────────────
_W_TEMP = "7 temperature"
_W_PRESS = "7 pressure"
_W_HUMID = "7 humidity"
_W_WIND_S = "7 wind speed"
_W_WIND_D = "7 wind direction"

# ── Logger CVW column name ───────────────────────────────────────────────
_CVW_COL = "CVW gross combination vehicle weight"

_CARDINALS = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]

# ── Charge-type regex (consistent with report_builder) ─────────────────────
_CHARGE_RE = re.compile(r"^(AC|DC|Charge|Mix|estimated)", re.IGNORECASE)


# ── Utility functions ─────────────────────────────────────────────────────
# _parse_report_filename / _cell_is_empty / _to_timestamp are shared with the
# charger patcher — see report_generator.xlsx_patch_common.


def _cell_needs_weather(cell) -> bool:
    """Return whether a weather cell needs backfilling."""
    v = cell.value
    if v is None:
        return True
    if isinstance(v, str) and (v.strip() == "" or v.strip().upper() == "=NA()"):
        return True
    return False


# ── Main class ────────────────────────────────────────────────────────────


class LoggerPatcher:
    """
    SRF Logger data backfill tool.

    Reads a generated xlsx report file, fetches Logger legs from the SRF API, and
    backfills the Logger Link URL and weather columns by time-window matching.

    What is backfilled:
    - SRF Logger Link — a clickable Logger visualisation URL
    - Average Temperature / Pressure / Humidity / Wind Speed / Wind Direction
      — averaged over the time window from the Logger Channel 7 weather data
    - Vehicle Mass (kg) + CV — the time-window median from the Logger CVW channel
      (backfilled only when the report row has no valid mass data, never overwriting an existing value)
    - Energy Performance Kinetics Corrected (kWh/km) — the kinetics-corrected
      energy performance computed second by second from the Logger 1Hz speed data
      (90% regenerative-braking efficiency)

    Args:
        srf_data:    optional existing SRF client instance (share the connection to avoid recreating it)
        cache_dir:   SRF API cache directory
    """

    def __init__(self, srf_data=None, cache_dir: str | None = None):
        if cache_dir is None:
            cache_dir = str(get_cache_dir())
        if srf_data is not None:
            self._srf_data = srf_data
        else:
            api_key = os.environ.get("SRF_API_KEY")
            if not api_key:
                logger.warning("LoggerPatcher: SRF_API_KEY is not set")
            self._srf_data = make_srf_client(cache_dir, api_key=api_key)

    # ── Public interface ──────────────────────────────────────────────────

    def patch_file(
        self,
        xlsx_path: str | Path,
        *,
        logger_legs: list | None = None,
        logger_windows: list | None = None,
    ) -> int:
        """
        Backfill a single xlsx report's Logger Link and weather data.

        Args:
            xlsx_path:       report file path
            logger_legs:     optional preloaded list of Logger leg objects
            logger_windows:  optional preloaded list of Logger time windows [(start, end, uri), ...]
                             if both logger_legs and logger_windows are None, fetch from the SRF API

        Returns:
            The number of rows backfilled (counted if at least one of Link or weather columns is backfilled).
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.error("File does not exist: %s", xlsx_path)
            return 0

        # 1. Obtain the logger data
        if logger_legs is None or logger_windows is None:
            fetched = self._fetch_logger_data(xlsx_path)
            if fetched is None:
                return 0
            if logger_legs is None:
                logger_legs = fetched[0]
            if logger_windows is None:
                logger_windows = fetched[1]

        if not logger_legs and not logger_windows:
            logger.info("LoggerPatcher: no Logger data, skipping %s", xlsx_path.name)
            return 0

        # 2. Preload Channel 7 weather data + CVW mass data + CCVS speed data + pedal data
        weather_df = self._load_weather_data(logger_legs)
        mass_df = self._load_mass_data(logger_legs)
        speed_df = self._load_speed_data(logger_legs)
        acc_pedal_df = self._load_pedal_data(logger_legs, channel="EEC2")
        dec_pedal_df = self._load_pedal_data(logger_legs, channel="EBC1")

        # 3. Open the xlsx
        logger.info("LoggerPatcher: backfilling %s", xlsx_path.name)
        wb = load_workbook(str(xlsx_path))
        if "Report" not in wb.sheetnames:
            logger.error("  'Report' worksheet not found: %s", xlsx_path.name)
            wb.close()
            return 0
        ws = wb["Report"]

        # 4. Iterate rows, backfilling Link and weather (discharge/trip segments only, skip charge segments)
        patched = 0
        total_rows = ws.max_row - 1  # minus the header row
        for row_idx in tqdm(
            range(2, ws.max_row + 1),
            desc="Logger backfill rows",
            total=total_rows,
            leave=False,
        ):
            # Charge segments should have no Logger data
            leg_type_val = ws.cell(row_idx, _COL_LEG_TYPE).value
            if (
                leg_type_val
                and isinstance(leg_type_val, str)
                and _CHARGE_RE.match(leg_type_val)
            ):
                continue

            t_s = _to_timestamp(ws.cell(row_idx, _COL_START_TIME).value)
            t_e = _to_timestamp(ws.cell(row_idx, _COL_END_TIME).value)
            if t_s is None or t_e is None:
                continue

            row_patched = False

            # 4a. Logger Link
            if _cell_is_empty(ws.cell(row_idx, _COL_LOGGER_LINK)):
                uri = _find_overlap(logger_windows, t_s, t_e, tol_min=5)
                if uri:
                    url = _build_logger_url(uri, t_s, t_e)
                    if url:
                        cell = ws.cell(row_idx, _COL_LOGGER_LINK)
                        cell.value = "Link"
                        cell.hyperlink = url
                        cell.font = Font(color="0000FF", underline="single")
                        row_patched = True

            # 4b. Weather data
            if weather_df is not None and any(
                _cell_needs_weather(ws.cell(row_idx, c)) for c in _WEATHER_COLS
            ):
                matched = weather_df.loc[t_s:t_e]
                if not matched.empty:
                    if _W_TEMP in matched.columns:
                        ws.cell(row_idx, _COL_TEMP).value = round(
                            float(matched[_W_TEMP].mean()), 1
                        )
                    if _W_PRESS in matched.columns:
                        ws.cell(row_idx, _COL_PRESSURE).value = round(
                            float(matched[_W_PRESS].mean()), 1
                        )
                    if _W_HUMID in matched.columns:
                        ws.cell(row_idx, _COL_HUMIDITY).value = round(
                            float(matched[_W_HUMID].mean()), 1
                        )
                    if _W_WIND_S in matched.columns:
                        ws.cell(row_idx, _COL_WIND_SPEED).value = round(
                            float(matched[_W_WIND_S].mean()), 1
                        )
                    if _W_WIND_D in matched.columns:
                        wind_d = matched[_W_WIND_D].dropna()
                        if not wind_d.empty:
                            avg_deg = float(wind_d.mean())
                            ws.cell(row_idx, _COL_WIND_DIR).value = _CARDINALS[
                                round(avg_deg / 45) % 8
                            ]
                    row_patched = True

            # 4c. Mass data (backfilled only when the report has no valid mass)
            if mass_df is not None and _cell_needs_weather(ws.cell(row_idx, _COL_MASS)):
                matched_m = mass_df.loc[t_s:t_e]
                valid_m = matched_m[_CVW_COL].dropna()
                valid_m = valid_m[valid_m > 0]
                if len(valid_m) >= 5:
                    median_mass = float(np.median(valid_m.values))
                    cv = (
                        float(valid_m.std() / valid_m.mean())
                        if valid_m.mean() > 0
                        else None
                    )
                    ws.cell(row_idx, _COL_MASS).value = round(median_mass, 0)
                    if cv is not None:
                        ws.cell(row_idx, _COL_MASS_CV).value = round(cv, 3)
                    row_patched = True

            # 4d. Kinetics-corrected energy performance (computed second by second from the Logger 1Hz speed data)
            if speed_df is not None and _cell_needs_weather(
                ws.cell(row_idx, _COL_EPERF_KIN)
            ):
                energy_val = ws.cell(row_idx, _COL_ENERGY).value
                dist_val = ws.cell(row_idx, _COL_DISTANCE).value
                elev_val = ws.cell(row_idx, _COL_ELEV_DIFF).value
                mass_val = ws.cell(row_idx, _COL_MASS).value
                if (
                    energy_val is not None
                    and dist_val is not None
                    and mass_val is not None
                ):
                    try:
                        energy_kwh = float(energy_val)
                        distance_km = float(dist_val)
                        elevation_m = (
                            float(elev_val) if elev_val is not None else float("nan")
                        )
                        mass_kg = float(mass_val)
                        speed_arr = _get_trip_speed_array(speed_df, t_s, t_e)
                        if speed_arr is not None and distance_km > 0:
                            ep_kin = _kinetics_corrected_energy_perf(
                                energy_kwh, distance_km, elevation_m, mass_kg, speed_arr
                            )
                            if not np.isnan(ep_kin):
                                ws.cell(row_idx, _COL_EPERF_KIN).value = ep_kin
                                row_patched = True
                    except (TypeError, ValueError):
                        pass

            # 4e. Pedal-position histograms (discharge segments only, distance > 10 km)
            dist_val = ws.cell(row_idx, _COL_DISTANCE).value
            try:
                dist_km = float(dist_val) if dist_val is not None else 0.0
            except (TypeError, ValueError):
                dist_km = 0.0

            if dist_km > MIN_DISTANCE_FOR_PEDAL_KM:
                # Accelerator-pedal histogram
                if acc_pedal_df is not None and _cell_is_empty(
                    ws.cell(row_idx, _COL_HIST_ACC)
                ):
                    try:
                        acc_slice = acc_pedal_df.loc[t_s:t_e].dropna()
                        hist_str = compute_pedal_histogram(
                            acc_slice, value_col=EEC2_COL
                        )
                        if hist_str is not None:
                            ws.cell(row_idx, _COL_HIST_ACC).value = hist_str
                            row_patched = True
                    except Exception:
                        pass
                # Brake-pedal histogram
                if dec_pedal_df is not None and _cell_is_empty(
                    ws.cell(row_idx, _COL_HIST_DEC)
                ):
                    try:
                        dec_slice = dec_pedal_df.loc[t_s:t_e].dropna()
                        hist_str = compute_pedal_histogram(
                            dec_slice, value_col=EBC1_COL
                        )
                        if hist_str is not None:
                            ws.cell(row_idx, _COL_HIST_DEC).value = hist_str
                            row_patched = True
                    except Exception:
                        pass

            if row_patched:
                patched += 1

        if patched > 0:
            wb.save(str(xlsx_path))
            logger.info(
                "  Backfilled %d rows (Logger Link + weather + mass + kinetics correction + pedal histograms), saved",
                patched,
            )
        else:
            logger.info("  No Logger data to backfill")

        wb.close()
        return patched

    def patch_folder(self, folder_path: str | Path) -> dict[str, int]:
        """Backfill the Logger data of all jolt_report_*.xlsx under a folder."""
        folder = Path(folder_path)
        if not folder.is_dir():
            logger.error("Folder does not exist: %s", folder)
            return {}

        xlsx_files = sorted(folder.glob("jolt_report_*.xlsx"))
        if not xlsx_files:
            logger.info("LoggerPatcher: no report files under %s", folder)
            return {}

        results = {}
        for fp in tqdm(xlsx_files, desc="Logger patch files"):
            results[fp.name] = self.patch_file(fp)
        return results

    # ── Internal methods ──────────────────────────────────────────────────

    def _fetch_logger_data(self, xlsx_path: Path) -> tuple[list, list] | None:
        """Fetch Logger legs and time windows from the SRF API. Returns (legs, windows) or None."""
        parsed = _parse_report_filename(xlsx_path)
        if parsed is None:
            logger.error(
                "  Cannot parse vehicle info from the file name: %s", xlsx_path.name
            )
            return None

        reg, ds_str, de_str = parsed
        cfg = VEHICLE_CONFIG.get(reg)
        if cfg is None:
            logger.error("  Vehicle %s is not registered in vehicles.json", reg)
            return None

        reg_srf = cfg["srf_reg"]
        ds = datetime.datetime.strptime(ds_str, "%Y%m%d")
        de = datetime.datetime.strptime(de_str, "%Y%m%d")

        logger.info(
            "  Fetching Logger legs from the SRF API: %s  %s ~ %s",
            reg_srf,
            ds_str,
            de_str,
        )
        params = {
            "start_time": srf_client.filter.between(
                datetime.datetime.combine(ds, datetime.time.min, datetime.timezone.utc),
                datetime.datetime.combine(de, datetime.time.max, datetime.timezone.utc),
            ),
            "sort": srf_client.sort.asc("startTime"),
        }

        legs = []
        windows = []
        try:
            for leg in paging.paged_items(
                self._srf_data.legs.find_all(
                    **params, **{"trip.vehicle.registration": reg_srf}
                )
            ):
                try:
                    src = leg.trip.source or ""
                    if src.startswith("SRFLOGGER"):
                        legs.append(leg)
                        windows.append((leg.start_time, leg.end_time, leg.uri))
                except AttributeError:
                    pass
        except Exception as exc:
            logger.warning("  Logger legs fetch failed: %s", exc)
            return [], []

        logger.info("  Fetched %d Logger legs", len(legs))
        return legs, windows

    def _load_weather_data(self, logger_legs: list) -> pd.DataFrame | None:
        """Load weather data from Channel 7 of all Logger legs."""
        if not logger_legs:
            return None

        dfs = []
        for leg in tqdm(logger_legs, desc="Loading Logger weather", leave=False):
            try:
                if "7" not in leg.types:
                    continue
                df_w = leg.get_data_frame("7", resolution="1s")
                if df_w is not None and not df_w.empty:
                    dfs.append(df_w)
            except Exception as exc:
                logger.debug("  Logger weather extraction failed: %s", exc)

        if not dfs:
            logger.info("  No Channel 7 weather data in the Logger")
            return None

        weather_df = pd.concat(dfs).sort_index()
        if weather_df.index.tz is None:
            weather_df.index = weather_df.index.tz_localize("UTC")
        else:
            weather_df.index = weather_df.index.tz_convert("UTC")
        logger.info("  Extracted %d weather readings from the Logger", len(weather_df))
        return weather_df

    def _load_mass_data(self, logger_legs: list) -> pd.DataFrame | None:
        """Load mass data from the CVW channel of all Logger legs."""
        if not logger_legs:
            return None

        dfs = []
        for leg in tqdm(logger_legs, desc="Loading Logger CVW", leave=False):
            try:
                if "CVW" not in leg.types:
                    continue
                df_m = leg.get_data_frame("CVW", resolution="1s")
                if df_m is not None and not df_m.empty and _CVW_COL in df_m.columns:
                    dfs.append(df_m[[_CVW_COL]])
            except Exception as exc:
                logger.debug("  Logger CVW extraction failed: %s", exc)

        if not dfs:
            logger.info("  No CVW mass data in the Logger")
            return None

        mass_df = pd.concat(dfs).sort_index()
        if mass_df.index.tz is None:
            mass_df.index = mass_df.index.tz_localize("UTC")
        else:
            mass_df.index = mass_df.index.tz_convert("UTC")
        logger.info("  Extracted %d CVW mass readings from the Logger", len(mass_df))
        return mass_df

    def _load_speed_data(self, logger_legs: list) -> pd.DataFrame | None:
        """Load wheel-based speed data from the CCVS channel of all Logger legs (for the kinetics correction)."""
        if not logger_legs:
            return None

        dfs = []
        for leg in tqdm(logger_legs, desc="Loading Logger speed", leave=False):
            try:
                avail = leg.types
                if "CCVS" in avail:
                    df_spd = leg.get_data_frame("CCVS", resolution="1s")
                    col = "CCVS wheel based vehicle speed"
                    if col in df_spd.columns:
                        dfs.append(df_spd[[col]].rename(columns={col: "logger_speed"}))
                elif "2" in avail:
                    df_spd = leg.get_data_frame("2", resolution="1s")
                    col = "2 speed"
                    if col in df_spd.columns:
                        dfs.append(df_spd[[col]].rename(columns={col: "logger_speed"}))
            except Exception as exc:
                logger.debug("  Logger CCVS speed extraction failed: %s", exc)

        if not dfs:
            logger.info("  No CCVS speed data in the Logger")
            return None

        speed_df = pd.concat(dfs).sort_index()
        if speed_df.index.tz is None:
            speed_df.index = speed_df.index.tz_localize("UTC")
        else:
            speed_df.index = speed_df.index.tz_convert("UTC")
        logger.info("  Extracted %d CCVS speed readings from the Logger", len(speed_df))
        return speed_df

    def _load_pedal_data(
        self, logger_legs: list, channel: str = "EEC2"
    ) -> pd.DataFrame | None:
        """Load pedal-position data from all Logger legs.

        Args:
            logger_legs: list of Logger leg objects
            channel: 'EEC2' (accelerator pedal) or 'EBC1' (brake pedal)

        Returns:
            DataFrame (indexed by UTC timestamp, column = pedal-position percentage) or None
        """
        if not logger_legs:
            return None

        col_map = {
            "EEC2": EEC2_COL,
            "EBC1": EBC1_COL,
        }
        target_col = col_map.get(channel)
        if target_col is None:
            return None

        dfs = []
        for leg in tqdm(logger_legs, desc=f"Loading Logger {channel}", leave=False):
            try:
                if channel not in leg.types:
                    continue
                df_p = leg.get_data_frame(channel, resolution="1s")
                if df_p is not None and not df_p.empty and target_col in df_p.columns:
                    dfs.append(df_p[[target_col]])
            except Exception as exc:
                logger.debug("  Logger %s extraction failed: %s", channel, exc)

        if not dfs:
            logger.info("  No %s pedal data in the Logger", channel)
            return None

        pedal_df = pd.concat(dfs).sort_index()
        if pedal_df.index.tz is None:
            pedal_df.index = pedal_df.index.tz_localize("UTC")
        else:
            pedal_df.index = pedal_df.index.tz_convert("UTC")
        logger.info(
            "  Extracted %d %s pedal readings from the Logger", len(pedal_df), channel
        )
        return pedal_df
