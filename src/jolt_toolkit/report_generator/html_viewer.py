"""
report_generator.html_viewer
============================
Builds the ``inspect_*.html`` validation-figure viewer for a vehicle/period and
the per-day figure bookkeeping helpers (active-date extraction from the xlsx,
per-day path grouping, stale-figure sweep). The viewer's HTML/CSS/JS body lives
in the packaged data file ``assets/inspect_viewer_template.html`` and is filled
via ``str.format`` (byte-identical to the former inline f-string).

Split out of report_builder.py in v3.0.0 (pure move + template externalization).
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

# The viewer template is the exact source text of the former inline f-string
# (doubled ``{{``/``}}`` preserved); ``str.format`` reproduces the f-string
# output byte-for-byte. Loaded once at import.
_INSPECT_TEMPLATE = (
    Path(__file__).resolve().parent / "assets" / "inspect_viewer_template.html"
).read_text(encoding="utf-8")


def _compute_active_dates_from_xlsx(xlsx_path: Path) -> set[str]:
    """Read the ``Report`` sheet of the xlsx and return the set of dates with trip or charge activity.

    Stop rows (`Leg Type == 'Stop'`) and blank rows are excluded; any other leg
    type (`In Transit / Round Trip / Outbound / Return / In House / AC Charge /
    DC Charge / Mix Charge / estimated Charge`, etc.) counts as activity.

    The return value is a ``{'YYYY-MM-DD', ...}`` string set, kept in the same
    format as the date prefix in the figure file names
    (``validation_<REG>_<DATE>_<idx>.png``) so they can be compared. Returns an
    empty set on read failure or when the xlsx does not exist — the viewer then
    falls back to treating every entry as "active", preserving backward
    compatibility.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return set()
    if not xlsx_path.exists():
        return set()
    try:
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    except Exception:
        return set()
    try:
        ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return set()
        # Locate the Leg Type / Start Time columns by header name, avoiding a hard-coded EV/Diesel offset
        col_leg_type = None
        col_start = None
        for idx, name in enumerate(header_row):
            if name == "Leg Type":
                col_leg_type = idx
            elif name == "Start Time (UTC)":
                col_start = idx
        if col_leg_type is None or col_start is None:
            return set()
        active: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(col_leg_type, col_start):
                continue
            leg_type = row[col_leg_type]
            if leg_type is None:
                continue
            leg_type_str = str(leg_type).strip()
            if not leg_type_str or leg_type_str == "Stop":
                continue
            start = row[col_start]
            if start is None:
                continue
            # Start Time may be a datetime object or an ISO string
            if hasattr(start, "strftime"):
                date_str = start.strftime("%Y-%m-%d")
            else:
                date_str = str(start)[:10]
            if re.match(r"\d{4}-\d{2}-\d{2}$", date_str):
                active.add(date_str)
        return active
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _group_paths_by_date(paths, date_re: "re.Pattern[str]") -> dict[str, list[Path]]:
    """Group per-leg raw-CSV paths by the ``YYYY-MM-DD`` captured by ``date_re``.

    Both regenerate paths (EV ``raw_<date>_<idx>.csv`` and diesel
    ``logger_<date>_<idx>.csv``) split a single calendar day into many short
    legs. To draw one validation figure per day we first bucket each day's legs
    together. ``date_re`` must expose the date in capture group 1
    (``re.search`` is used, so the pattern need only match the filename's date
    token). The returned dict preserves chronological order (paths are sorted
    first, dict insertion order is stable on Python 3.7+), and each day's list
    is in filename — i.e. leg-index, i.e. time — order.
    """
    groups: dict[str, list[Path]] = {}
    for p in sorted(paths):
        m = date_re.search(p.name)
        if not m:
            continue
        groups.setdefault(m.group(1), []).append(p)
    return groups


def _clear_day_validation_figures(fig_dir: Path, reg: str) -> int:
    """Delete stale validation figures + overlay sidecars before a per-day repaint.

    The one-figure-per-day regeneration writes ``validation_<reg>_<date>.png``;
    earlier builds wrote one figure per leg (``validation_<reg>_<date>_<NNNN>.png``).
    Both match the ``validation_<reg>_*`` glob the inspect viewer lists, so without
    this sweep a re-painted directory would show BOTH the consolidated day figure
    and every stale per-leg fragment. Clearing the whole non-finetuned set up front
    also drops figures for days that no longer segment to any trip.

    ``*_finetuned.*`` artefacts are left untouched — they belong to the
    report-finetuner flow and must survive a base repaint. Returns the number of
    files removed.
    """
    if not fig_dir.exists():
        return 0
    removed = 0
    for pattern in (
        f"validation_{reg}_*.png",
        f"validation_{reg}_*.boxes.json",
        f"validation_{reg}_*.dsoc.json",
    ):
        for p in fig_dir.glob(pattern):
            if "_finetuned" in p.name:
                continue
            try:
                p.unlink()
                removed += 1
            except OSError:
                pass
    return removed


def _write_html_viewer(
    out_dir: Path, reg: str, period_start, period_end, report_name: str
) -> None:
    """Generate an HTML viewer that shows all validation figures for this vehicle/period."""
    fig_dir = out_dir / "validation_figures"
    all_figs = (
        sorted(fig_dir.glob(f"validation_{reg}_*.png")) if fig_dir.exists() else []
    )
    if not all_figs:
        return

    # Keep only the validation figures whose date lies in period_start ~ period_end.
    # The ``[._]`` after the date token accommodates both naming schemes: the new
    # one-figure-per-day ``validation_<reg>_<date>.png`` (extension ``.`` directly
    # after the date) and the historical / finetuned per-leg
    # ``validation_<reg>_<date>_<NNNN>.png`` (``_`` after the date).
    _date_re = re.compile(r"validation_" + re.escape(reg) + r"_(\d{4}-\d{2}-\d{2})[._]")
    ds_str = str(period_start)
    de_str = str(period_end)
    figs = []
    fig_dates: list[str] = []
    for p in all_figs:
        m = _date_re.match(p.name)
        if m and ds_str <= m.group(1) <= de_str:
            figs.append(p)
            fig_dates.append(m.group(1))
    if not figs:
        return

    # Read the xlsx to find the set of dates with trip/charge activity in the period, used for the sidebar red text + filtering
    active_dates = _compute_active_dates_from_xlsx(out_dir / report_name)
    is_active = [d in active_dates for d in fig_dates]

    # Relative paths from out_dir to validation_figures/
    rel = [f"validation_figures/{p.name}" for p in figs]
    labels = [p.stem for p in figs]  # e.g. validation_AV24LXK_2024-10-01_0000

    # Per-figure interactive annotation overlay sidecar. A figure produced with
    # ``export_dsoc_overlay=True`` writes ``<stem>.boxes.json`` next to the PNG
    # (figure-fraction coordinates, origin top-left) carrying every panel's
    # rounded-bbox data label — dSOC, energy/charger-meter deltas, recuperation
    # deltas, mass labels. Two shapes are accepted (the viewer auto-detects):
    #   * **flat list** — legacy / diesel figures; rendered as ghosted-on-hover
    #     ``.annot-box`` divs (the original behaviour, kept for back-compat with
    #     not-yet-re-painted vehicles).
    #   * **dict** ``{boxes, segments, soc_axis}`` (v2.2.6, EV figures) — drives
    #     the hover redesign: default triangles-only, per-segment hover hotzones,
    #     a pinned info box at the SOC panel's bottom-left, and label de-collision.
    # Figures without a sidecar (legacy / baked-in text) yield an empty list and
    # simply show no overlay.
    boxes_per_fig: list = []
    for p in figs:
        sidecar = fig_dir / (p.stem + ".boxes.json")
        # Backward compat: figures painted by earlier builds (before the overlay
        # rename, i.e. pre the ``.boxes.json`` sidecar) emit the older
        # ``.dsoc.json`` (Panel-1 dSOC only). Fall back to it so an HTML re-render
        # of a not-yet-re-painted vehicle still shows its dSOC overlay.
        if not sidecar.exists():
            sidecar = fig_dir / (p.stem + ".dsoc.json")
        data: list | dict = []
        if sidecar.exists():
            try:
                with open(sidecar, encoding="utf-8") as fh:
                    loaded = json.load(fh)
                if isinstance(loaded, (list, dict)):
                    data = loaded
            except (json.JSONDecodeError, OSError):
                data = []
        boxes_per_fig.append(data)

    html_name = "inspect_" + report_name.replace(".xlsx", ".html")
    html_path = out_dir / html_name

    # Build sidebar items & image tags as JS arrays
    imgs_js = "[\n" + ",\n".join(f'        "{r}"' for r in rel) + "\n    ]"
    labels_js = "[\n" + ",\n".join(f'        "{l}"' for l in labels) + "\n    ]"
    active_js = (
        "[\n" + ",\n".join(f"        {str(a).lower()}" for a in is_active) + "\n    ]"
    )
    # Inline the box data (avoids a runtime fetch, which the file:// protocol
    # blocks when the HTML is opened directly from disk).
    boxes_js = json.dumps(boxes_per_fig, ensure_ascii=False)

    html = _INSPECT_TEMPLATE.format(
        reg=reg,
        period_start=period_start,
        period_end=period_end,
        imgs_js=imgs_js,
        labels_js=labels_js,
        active_js=active_js,
        boxes_js=boxes_js,
    )
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info("  HTML viewer: %s  (%d figures)", html_name, len(figs))
