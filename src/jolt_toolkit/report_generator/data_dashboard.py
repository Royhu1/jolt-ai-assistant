"""Data-availability dashboard generator.

Scans a report-database version directory (``excel_report_database/<version>/``)
and renders a single self-contained ``data_dashboard.html`` — a **three-panel,
interactive, offline** dashboard (no CDN / external assets; opens by
double-click):

- **Left panel** — a scrollable vehicle selector (one row per registration that
  has reports). Clicking a registration switches the middle + right panels.
- **Middle panel** — the selected vehicle's computed stat block (type / make /
  model, battery capacity or weight class, data date range, total distance
  driven, trip / charge / stop counts, and per-category day counts) plus an
  **Operator** block: the trial type ("Fixed" / "Round-robin" / "—") and the
  operator assignment — a single coloured operator for fixed vehicles, or the
  compact dated period sequence for round-robin vehicles (see *Operator overlay*
  below).
- **Right panel** — a traditional month-calendar availability view: a
  vertically-scrolling wrap grid of "Month YYYY" blocks (Monday-first Mo–Su
  weekday headers + a day grid) rendered on a *fleet-wide* month axis (the
  global min→max month across all vehicles, identical for every vehicle). Day
  cells show the day-of-month number — light grey when empty, filled with the
  category vertical colour bands (white number) when data is present, and a
  3px operator-coloured border on data-bearing cells (see *Operator overlay*).
  A segmented **Events / Raw data** toggle near the top switches the calendar
  between the two availability bases (see below) without re-flowing the month
  blocks (the fleet axis spans both bases, so only cell colours change).

Two independent availability bases
----------------------------------
Every vehicle carries **two** day→category maps, and the toggle picks which one
the calendar renders:

- **Events** (``avail``, default) — days with *processed legs* in the generated
  Excel reports (the report-based detection described below). This is the
  original behaviour, unchanged.
- **Raw data** (``avail_raw``) — days with *raw files fetched from SRF* on disk,
  scanned per vehicle directory (``raw_telematics/``, ``raw_logger_v1/`` +
  ``raw_logger_v2/``, ``raw_charger/charger_transactions.csv``). This is a pure
  raw-file view — no union with report data — so the two bases can be compared
  directly (e.g. a logger day with raw files but no processed legs shows up in
  Raw mode but not Events mode, and vice-versa). Raw dates are parsed from
  *filenames* (``raw_YYYY-MM-DD_*.csv`` / ``logger_YYYY-MM-DD_*.csv``) for
  speed; the sole file actually read is the charger transactions CSV (its
  ``start_time`` column). Only the four named sub-directories directly under a
  reg dir are scanned — dot-dirs, ``*.bak_*`` backups, ``validation_figures*``
  and ``raw_telematics_stub`` are ignored — and parsed dates outside
  ``[2024-01-01, today+1]`` are dropped (kills the 2000-01-01 telematics stub).

Three data categories are tracked per calendar day:

- **telematics** — the FPS / SRF telematics feed every leg derives from;
- **logger** — high-rate SRF Logger (CAN) data;
- **charger** — charge-point energy / power readings.

Availability + stats are derived **from the generated reports** (the ``Report``
sheet, sheet index 0, of each ``jolt_report_<REG>_<start>_<end>.xlsx``), one leg
per row. The ``Charger Link`` hyperlink column is deliberately **not** used (in
EV reports it is unreliable); charger availability uses energy/power data columns
instead. The ``SRF Logger Link`` column, by contrast, *is* used as one of the
logger signals (see below).

Category detection (per leg, aggregated per calendar date):

- **telematics**: any leg row exists for that date (every leg comes from
  telematics).
- **logger** (EV + diesel, unified): the **union** of two complementary signals
  on any leg that date — (1) the ``SRF Logger Link`` cell carries a hyperlink
  **or** a non-empty value (the display value is often blank while only the
  hyperlink target is set, so the hyperlink must be checked, which requires a
  non-read-only workbook), and (2) a non-empty ``Histogram of Accelerator Pedal
  Position`` OR ``Histogram of Decelerator Pedal Position`` OR ``Energy based on
  motor power (kWh)`` (these EV columns are absent from ``DIESEL_HEADERS``). The
  two signals are complementary, not redundant.
- **charger** (EV only): a non-empty / non-zero ``Energy Charged AC (kWh)`` OR
  ``Energy Charged DC (kWh)`` OR ``Energy Output from Charger (kWh)`` OR ``Peak
  Charging (kW)`` OR ``Average Charging (kW)`` on any leg that date.

Leg classification (for the stat block) mirrors :mod:`finetune`'s strict
semantics: ``Stop`` rows are stops; ``AC*`` / ``DC*`` / ``Charge*`` / ``Mix*`` /
``estimated*`` rows are charging events; everything else (``In Transit`` /
``Outbound`` / ``Return`` / ``Round Trip`` / ``In House`` / diesel ``In
Transit`` …) is a driving trip.

Both header layouts are handled (EV ``HEADERS`` / diesel ``DIESEL_HEADERS``);
per-row lookups go through the live header row of each workbook (robust to
column reordering) and missing columns are skipped gracefully. A vehicle's
multiple report files (different periods, sometimes overlapping — e.g. a short
diagnostic file fully contained in a full-range file, or consecutive files that
share a boundary day) are merged into a single timeline by **de-duplicating legs
on ``(start, end, leg-class)``**, so event counts and distances are not
double-counted.

Operator overlay
----------------
Operator / trial-type information is **data-driven** from the reports' ``Operator``
column (the SRF ``leg.trip.trial.description`` / ``vehicle.organisation.name``
value resolved by :mod:`report_generator.operators`), so the overlay tracks
operator handovers automatically with no manual config edit. Per vehicle, the
non-blank per-leg operator codes are collapsed to one operator per day (daily
majority), and consecutive days carrying the same operator are grouped into dated
periods. A single distinct operator over the whole timeline ⇒ ``trial_type``
"Fixed" (one open-ended whole-axis period); two or more ⇒ "Round-robin" (one
dated period per run). The locally-curated ``configs/plot_config.json``
``company_assignment`` (``simple`` / ``round_robin``) is kept only as a
**fallback** for a vehicle whose reports carry no operator at all, so nothing
regresses. ``colors.company`` still supplies each operator's CSS colour; any
operator absent from the config (e.g. ``HTL``) is given a stable, deterministic
fallback colour. Display names show underscores as spaces (e.g.
``WELCH_TRANSPORT`` → "WELCH TRANSPORT"), matching the report generator / PDF
briefing.

Each vehicle is embedded with a compact ``periods`` list (``[{c, s, e}]`` with
ISO bounds; ``s``/``e`` ``None`` = open-ended whole-axis period for fixed
vehicles) and a ``trial_type``; the fleet-level ``operatorColors`` /
``operatorNames`` maps and the neutral grey are embedded once. The JS resolves a
date's operator by scanning the (few) periods — no per-day expansion. On the
calendar, every **data-bearing** day cell (in the active mode) gets a 3px border
in the operator's colour for that date; a round-robin day inside a gap between
periods gets the neutral grey (``#94a3b8``) so gaps stay visible; vehicles with
no operator at all get no border. The per-vehicle operator legend (hollow rings,
to distinguish it from the solid-fill category swatches) lists only the relevant
operators plus the neutral entry when applicable, recomputed on vehicle change
and mode toggle.

Usage
-----
::

    python -m jolt_toolkit.report_generator.data_dashboard            # installed package version
    python -m jolt_toolkit.report_generator.data_dashboard --version 2.2.8
    python -m jolt_toolkit.report_generator.data_dashboard --out /tmp/dash.html
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import logging
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from jolt_toolkit import __version__
from jolt_toolkit.report_generator.report_builder import DIESEL_HEADERS, HEADERS

LOG = logging.getLogger("data_dashboard")

# ── Category model ───────────────────────────────────────────────────────────
# Canonical ordering drives both the legend and the in-cell stripe order.
TELEMATICS = "telematics"
LOGGER = "logger"
CHARGER = "charger"
CATEGORY_ORDER = (TELEMATICS, LOGGER, CHARGER)

# Three clear, distinct colours (blue / green / amber).
CATEGORY_COLOURS = {
    TELEMATICS: "#2563EB",  # blue
    LOGGER: "#16A34A",      # green
    CHARGER: "#F59E0B",     # amber
}
CATEGORY_LABELS = {
    TELEMATICS: "Telematics",
    LOGGER: "Logger",
    CHARGER: "Charger",
}
# Short single-letter key used to encode a day's category combination compactly
# (e.g. {telematics, logger} -> "tl"). Canonical order is preserved.
_CATEGORY_KEY = {TELEMATICS: "t", LOGGER: "l", CHARGER: "c"}

# GitHub-style "no data" square colour.
EMPTY_COLOUR = "#ebedf0"

# Neutral border colour for a round-robin day that falls in a gap between
# operator-assignment periods (so coverage gaps stay visible on the calendar).
OPERATOR_NEUTRAL_COLOUR = "#94a3b8"

# ── Report-sheet column references ───────────────────────────────────────────
# Imported from report_builder so a header rename surfaces here at import time
# (see the membership assertions below) rather than silently mis-detecting.
START_COL = "Start Time (UTC)"
END_COL = "End Time (UTC)"
LEG_TYPE_COL = "Leg Type"
DISTANCE_COL = "Distance (km)"
# Logger signal #1 — present in BOTH layouts; carries a hyperlink (and often an
# empty display value), so detection must inspect ``cell.hyperlink`` too.
LOGGER_LINK_COL = "SRF Logger Link"
# Logger signal #2 — EV-only columns (absent from DIESEL_HEADERS). The logger
# rule is the UNION of these two signals, applied uniformly to EV + diesel.
EV_LOGGER_COLS = (
    "Histogram of Accelerator Pedal Position",
    "Histogram of Decelerator Pedal Position",
    "Energy based on motor power (kWh)",
)
CHARGER_SIGNAL_COLS = (
    "Energy Charged AC (kWh)",
    "Energy Charged DC (kWh)",
    "Energy Output from Charger (kWh)",
    "Peak Charging (kW)",
    "Average Charging (kW)",
)

# Per-event metric columns — read per leg row so the drill-down detail pages can
# label each trip / charge event (inspect.html style: dSOC / dkWh / C / EP).
# These live in EV HEADERS only; DIESEL_HEADERS omits them, so for diesel reports
# the column lookup misses and the metric simply stays blank on the label.
SOC_CHANGE_COL = "SOC Change (%)"
ENERGY_CHANGE_COL = "Energy Change (kWh)"
ENERGY_AC_COL = "Energy Charged AC (kWh)"
ENERGY_DC_COL = "Energy Charged DC (kWh)"
ENERGY_PERF_COL = "Energy Performance (kWh/km)"
BATTERY_CAP_COL = "Battery Capacity (kWh)"

# Per-leg operator CODE (last column of both HEADERS and DIESEL_HEADERS since
# v2.2.5; resolved from SRF by ``report_generator.operators``). Drives the
# data-driven operator overlay. Looked up via the live header row (idx.get) — NOT
# asserted against HEADERS — so the dashboard still reads older artefacts whose
# reports predate the column (those vehicles just fall back to plot_config).
OPERATOR_COL = "Operator"

# Defensive: keep the signal column names in lock-step with report_builder.
assert START_COL in HEADERS and START_COL in DIESEL_HEADERS
assert END_COL in HEADERS and END_COL in DIESEL_HEADERS
assert LEG_TYPE_COL in HEADERS and LEG_TYPE_COL in DIESEL_HEADERS
assert DISTANCE_COL in HEADERS and DISTANCE_COL in DIESEL_HEADERS
assert LOGGER_LINK_COL in HEADERS and LOGGER_LINK_COL in DIESEL_HEADERS
assert all(c in HEADERS for c in EV_LOGGER_COLS)
assert all(c in HEADERS for c in CHARGER_SIGNAL_COLS)
assert all(
    c in HEADERS
    for c in (
        SOC_CHANGE_COL, ENERGY_CHANGE_COL, ENERGY_AC_COL,
        ENERGY_DC_COL, ENERGY_PERF_COL, BATTERY_CAP_COL,
    )
)

REPORT_SHEET = "Report"

# Charging-leg prefix (mirrors finetune._CHARGE_LEG_PREFIX_RE): "AC Home",
# "DC Away", "AC/DC …", "Charge Home", "Mix …", "estimated …".
_CHARGE_PREFIX_RE = re.compile(r"^(AC|DC|Charge|Mix|estimated)", re.IGNORECASE)

# ── Raw-file (on-disk) availability references ───────────────────────────────
# The second availability basis is scanned directly from the raw files SRF
# fetched into each reg directory. ONLY these four named sub-directories are
# read (anything else — dot-dirs, ``*.bak_*`` backups, ``validation_figures*``,
# ``raw_telematics_stub`` — is ignored by construction).
RAW_TELEMATICS_DIR = "raw_telematics"
RAW_LOGGER_DIRS = ("raw_logger_v1", "raw_logger_v2")  # either / both / neither
RAW_CHARGER_DIR = "raw_charger"
RAW_CHARGER_CSV = "charger_transactions.csv"

# Raw dates come from FILENAMES (fast); the charger CSV is the one file read.
_TELEMATICS_FNAME_RE = re.compile(r"^raw_(\d{4}-\d{2}-\d{2})_.*\.csv$", re.IGNORECASE)
_LOGGER_FNAME_RE = re.compile(r"^logger_(\d{4}-\d{2}-\d{2})_.*\.csv$", re.IGNORECASE)

# Sanity window for parsed raw dates: drop anything before 2024-01-01 or after
# today+1 (kills the WU70GLV 2000-01-01 telematics stub regardless of location).
_RAW_DATE_MIN = date(2024, 1, 1)


# ── Cell-value helpers ───────────────────────────────────────────────────────
def _is_present(value) -> bool:
    """Return ``True`` if a cell value counts as real (non-empty / non-zero) data.

    Treats ``None``, blanks, ``#N/A`` / ``=NA()`` strings, zero numbers and NaN
    as absent. openpyxl reads xlsxwriter ``=NA()`` formula cells as ``None`` (no
    cached value) under ``data_only=True``, so most empties already arrive as
    ``None``; the string / zero / NaN guards cover the remaining cases.
    """
    if value is None:
        return False
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return False
        upper = s.upper()
        return s not in ("0", "#N/A") and not upper.startswith("=NA")
    if isinstance(value, (int, float)):
        f = float(value)
        if math.isnan(f):
            return False
        return f != 0.0
    return True


def _link_present(cell) -> bool:
    """Return ``True`` if a ``SRF Logger Link`` cell carries logger data.

    The cell counts as a logger signal if it has either a hyperlink target or a
    non-empty display value. In many reports the display value is blank while
    only the hyperlink is set, so ``cell.hyperlink`` must be checked — this only
    works on a non-read-only workbook, where openpyxl populates per-cell
    hyperlinks (read-only cells expose no ``.hyperlink`` attribute).
    """
    if getattr(cell, "hyperlink", None) is not None:
        return True
    return _is_present(cell.value)


def _to_date(value):
    """Coerce a ``Start Time (UTC)`` cell to a ``date`` (or ``None``)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date()
        except ValueError:
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
    return None


def _iso_key(value) -> str:
    """Stable string key for a datetime/date/other cell (for leg de-duplication)."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return str(value).strip()


def _cell_num(row, i, n) -> float | None:
    """Return a finite ``float`` for cell ``i`` of ``row`` (else ``None``).

    Used to lift per-event metric columns (``SOC Change`` / ``Energy Change`` …)
    off a report row. A missing index (``i is None`` — column absent, e.g. diesel),
    an out-of-range cell, a non-numeric value, or ``NaN`` all yield ``None`` so the
    drill-down label can simply omit that line.
    """
    if i is None or i >= n:
        return None
    v = row[i].value
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        return float(v)
    return None


def _classify_leg(leg_type) -> str:
    """Classify a ``Leg Type`` value into ``"trip"`` / ``"charge"`` / ``"stop"``.

    Mirrors :mod:`finetune`: ``Stop`` -> stop; an ``AC`` / ``DC`` / ``Charge`` /
    ``Mix`` / ``estimated`` prefix -> charge; everything else (the discharge /
    driving leg types) -> trip. Non-string values default to ``"trip"`` (never
    occurs in real reports, where every leg carries a type string).
    """
    if not isinstance(leg_type, str):
        return "trip"
    s = leg_type.strip()
    if s.lower() == "stop":
        return "stop"
    if _CHARGE_PREFIX_RE.match(s):
        return "charge"
    return "trip"


def _combo_key(cats) -> str:
    """Stable single-string key for a category combination (canonical order)."""
    return "".join(_CATEGORY_KEY[c] for c in CATEGORY_ORDER if c in cats)


def _combo_background(cats) -> str:
    """CSS ``background`` value for a category combination.

    One present category -> a solid colour; multiple -> equal horizontal bands
    stacked top-to-bottom in canonical order via a hard-stop linear gradient.
    Used here only for the static legend "Multiple" swatch; the per-cell
    backgrounds are built by the equivalent ``comboBackground`` JS function.
    """
    present = [c for c in CATEGORY_ORDER if c in cats]
    colours = [CATEGORY_COLOURS[c] for c in present]
    if not colours:
        return EMPTY_COLOUR
    if len(colours) == 1:
        return colours[0]
    n = len(colours)
    stops = []
    for i, colour in enumerate(colours):
        a = 100.0 * i / n
        b = 100.0 * (i + 1) / n
        stops.append(f"{colour} {a:.4g}% {b:.4g}%")
    return f"linear-gradient(to bottom, {', '.join(stops)})"


# ── Config loading ───────────────────────────────────────────────────────────
def _load_vehicles_cfg() -> dict:
    """Load ``configs/vehicles.json`` (make / model / capacity / fuel-type)."""
    cfg_path = Path(__file__).resolve().parent.parent / "configs" / "vehicles.json"
    try:
        with cfg_path.open(encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:  # noqa: BLE001
        LOG.warning("Could not load vehicles.json (%s); stats will be sparse", exc)
        return {}


def _load_plot_config() -> dict:
    """Load ``configs/plot_config.json`` (operator assignment + company colours).

    This is the canonical, locally-curated operator map maintained at
    vehicle-onboarding, so the dashboard needs no SRF API call to know who
    operated each vehicle and when. A missing / unreadable file degrades the
    operator overlay gracefully (no borders, "—" trial type).
    """
    cfg_path = Path(__file__).resolve().parent.parent / "configs" / "plot_config.json"
    try:
        with cfg_path.open(encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:  # noqa: BLE001
        LOG.warning(
            "Could not load plot_config.json (%s); operator info will be sparse", exc
        )
        return {}


def _prettify_company(code: str) -> str:
    """Prettify an operator code for display: underscores become spaces.

    Operator codes are upper-case project tokens (``WELCH_TRANSPORT`` /
    ``DP_WORLD`` / ``PORT_EXPRESS_DAIMLER``); the only transformation is replacing
    underscores with spaces (case preserved), matching the report generator / PDF
    briefing convention — so ``WELCH_TRANSPORT`` -> ``WELCH TRANSPORT`` and
    ``DP_WORLD`` -> ``DP WORLD``.
    """
    return code.replace("_", " ")


def _yyyymmdd_to_iso(value: str) -> str | None:
    """Coerce a compact ``YYYYMMDD`` assignment date to ISO ``YYYY-MM-DD``."""
    try:
        return datetime.strptime(value, "%Y%m%d").date().isoformat()
    except (ValueError, TypeError):
        return None


def _clean_operator(value) -> str | None:
    """Normalise a raw ``Operator`` cell to a project code (or ``None``).

    Treats ``None``, NaN, blanks and the literal ``nan`` / ``none`` / ``#N/A``
    placeholders as "no operator" so those legs are simply skipped. Otherwise the
    code is returned stripped (case preserved — codes are UPPER_SNAKE tokens).
    """
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "#n/a") or s.upper().startswith("=NA"):
        return None
    return s


# Stable fallback palette for operators with no ``colors.company`` entry (e.g.
# ``HTL``): a deterministic md5-indexed pick keeps the colour identical across
# runs and visually distinct from the neutral grey + the named company colours.
_FALLBACK_OPERATOR_PALETTE = (
    "#0EA5E9", "#22C55E", "#EF4444", "#A855F7", "#F97316",
    "#14B8A6", "#EAB308", "#EC4899", "#6366F1", "#84CC16",
    "#06B6D4", "#F43F5E",
)


def _stable_operator_colour(code: str) -> str:
    """Deterministic CSS colour for an operator code absent from the config."""
    h = int(hashlib.md5(code.encode("utf-8")).hexdigest(), 16)
    return _FALLBACK_OPERATOR_PALETTE[h % len(_FALLBACK_OPERATOR_PALETTE)]


def _operator_days_from_legs(legs: list[dict]) -> dict[date, str]:
    """Collapse a leg list into ``{date: operator_code}`` (one operator per day).

    Each leg's ``operator`` (set by :func:`read_report_legs`) is bucketed by its
    calendar day; blank operators are ignored. A day with legs from more than one
    operator (a handover day) resolves to the **daily majority** (ties keep the
    first-seen operator), so the per-day operator is stable and single-valued.
    """
    by_day: dict[date, Counter] = defaultdict(Counter)
    for leg in legs:
        op = leg.get("operator")
        if op:
            by_day[leg["day"]][op] += 1
    return {d: c.most_common(1)[0][0] for d, c in by_day.items() if c}


def _periods_from_operator_days(
    op_days: dict[date, str]
) -> tuple[str, list[dict]] | None:
    """Derive ``(trial_type, periods)`` from a ``{date: operator}`` map.

    Consecutive days (in date order) carrying the same operator are grouped into
    one dated run; a new run starts whenever the operator changes. A single
    distinct operator across the whole timeline ⇒ ``("Fixed", [open-ended
    period])`` (so the border applies to every data day and the middle panel
    shows the clean single-operator line); two or more distinct operators ⇒
    ``("Round-robin", [dated period per run])``. Returns ``None`` when the map is
    empty (no data-driven operator → caller falls back to plot_config).
    """
    if not op_days:
        return None
    runs: list[list] = []  # [code, first_date, last_date]
    for d in sorted(op_days):
        code = op_days[d]
        if runs and runs[-1][0] == code:
            runs[-1][2] = d
        else:
            runs.append([code, d, d])
    distinct = {r[0] for r in runs}
    if len(distinct) == 1:
        # Single operator over the whole timeline → one open-ended period.
        return "Fixed", [{"c": runs[0][0], "s": None, "e": None}]
    periods = [
        {"c": code, "s": s.isoformat(), "e": e.isoformat()} for code, s, e in runs
    ]
    return "Round-robin", periods


def _config_operator_vehicles(plot_cfg: dict) -> dict[str, dict]:
    """Build the plot_config-derived per-vehicle operator periods (fallback layer).

    Mirrors the pre-v2.2.6 behaviour: ``company_assignment.simple`` →
    fixed-operator vehicles (one open-ended period), ``company_assignment
    .round_robin`` → one dated period per assignment span (ISO bounds parsed from
    the ``YYYYMMDD`` config dates). Used only for a vehicle whose reports carry no
    operator at all.
    """
    ca = plot_cfg.get("company_assignment", {}) or {}
    simple = ca.get("simple", {}) or {}
    round_robin = ca.get("round_robin", {}) or {}

    out: dict[str, dict] = {}
    for reg, company in simple.items():
        out[reg] = {
            "trial_type": "Fixed",
            "periods": [{"c": company, "s": None, "e": None}],
        }
    for reg, spans in round_robin.items():
        periods = []
        for span in spans:
            company = span.get("company")
            if company is None:
                continue
            periods.append(
                {
                    "c": company,
                    "s": _yyyymmdd_to_iso(str(span.get("date_start", ""))),
                    "e": _yyyymmdd_to_iso(str(span.get("date_end", ""))),
                }
            )
        out[reg] = {"trial_type": "Round-robin", "periods": periods}
    return out


def build_operator_assignment(
    plot_cfg: dict, op_days_by_reg: dict[str, dict[date, str]] | None = None
) -> dict:
    """Build per-vehicle operator assignment + fleet colour / name maps.

    **Data-driven**: ``op_days_by_reg`` (``{REG: {date: operator_code}}``, derived
    from the reports' ``Operator`` column by :func:`_operator_days_from_legs`) is
    the primary source — each vehicle's per-day operators are grouped into dated
    periods (see :func:`_periods_from_operator_days`). The locally-curated
    ``plot_cfg`` ``company_assignment`` is only a **fallback** for a vehicle that
    has no operator in the data (or when ``op_days_by_reg`` is omitted, e.g. an
    older caller), so nothing regresses.

    Returns ``{"vehicles": {REG: {"trial_type", "periods"}}, "operatorColors",
    "operatorNames", "neutralColour"}``. ``periods`` is a compact list
    ``[{"c": company, "s": iso|None, "e": iso|None}]`` — ``s`` / ``e`` ``None``
    means open-ended (the single whole-axis period of a fixed-operator vehicle).
    The renderer's JS resolves a date's operator by scanning these few periods
    (no per-day expansion). Each operator gets its ``colors.company`` colour where
    present, else a stable :func:`_stable_operator_colour`.
    """
    op_days_by_reg = op_days_by_reg or {}
    company_colours = (plot_cfg.get("colors", {}) or {}).get("company", {}) or {}
    config_vehicles = _config_operator_vehicles(plot_cfg)

    vehicles: dict[str, dict] = {}
    used: set[str] = set()

    # Data-driven primary + plot_config fallback, per vehicle. The union of both
    # key sets so config-only vehicles (no data) still surface, and data-only
    # vehicles (absent from the config, e.g. WU70GLV) are picked up automatically.
    for reg in set(config_vehicles) | set(op_days_by_reg):
        derived = _periods_from_operator_days(op_days_by_reg.get(reg, {}))
        if derived is not None:
            trial_type, periods = derived
            vehicles[reg] = {"trial_type": trial_type, "periods": periods}
        elif reg in config_vehicles:
            vehicles[reg] = config_vehicles[reg]
        else:
            continue
        for p in vehicles[reg]["periods"]:
            if p.get("c"):
                used.add(p["c"])

    operator_colours = {
        c: (company_colours[c] if company_colours.get(c) else _stable_operator_colour(c))
        for c in used
    }
    operator_names = {c: _prettify_company(c) for c in used}
    return {
        "vehicles": vehicles,
        "operatorColors": operator_colours,
        "operatorNames": operator_names,
        "neutralColour": OPERATOR_NEUTRAL_COLOUR,
    }


# ── Scanning ─────────────────────────────────────────────────────────────────
def read_report_legs(xlsx_path: Path) -> list[dict]:
    """Read one report workbook → a list of per-leg records.

    Each record is ``{"day": date, "sig": (start, end, class), "cats": set,
    "leg_class": str, "distance_km": float|None}``. The ``sig`` tuple is used to
    de-duplicate the same physical leg appearing in overlapping report files.

    Returns an empty list (and logs) for workbooks without a ``Report`` sheet or
    a parseable ``Start Time (UTC)`` column.
    """
    legs: list[dict] = []
    # Non-read-only so ``cell.hyperlink`` is populated: the ``SRF Logger Link``
    # logger signal lives in the hyperlink, which read-only mode discards.
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        if REPORT_SHEET not in wb.sheetnames:
            LOG.warning("Skip (no '%s' sheet): %s", REPORT_SHEET, xlsx_path.name)
            return []
        ws = wb[REPORT_SHEET]
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            return []
        idx = {c.value: i for i, c in enumerate(header_row) if c.value is not None}
        if START_COL not in idx:
            LOG.warning("Skip (no '%s' column): %s", START_COL, xlsx_path.name)
            return []

        start_i = idx[START_COL]
        end_i = idx.get(END_COL)
        type_i = idx.get(LEG_TYPE_COL)
        dist_i = idx.get(DISTANCE_COL)
        # Logger = UNION of two complementary signals (uniform for EV + diesel):
        #   (1) the SRF Logger Link cell (value OR hyperlink) — both layouts;
        #   (2) the EV pedal-histogram / motor-power columns — EV reports only
        #       (absent from DIESEL_HEADERS, so naturally skipped for diesel).
        link_i = idx.get(LOGGER_LINK_COL)
        pedal_i = [idx[c] for c in EV_LOGGER_COLS if c in idx]
        # Charger columns only exist in EV reports; absent for diesel.
        charger_i = [idx[c] for c in CHARGER_SIGNAL_COLS if c in idx]
        # Per-event metric columns (EV reports only; ``idx.get`` -> None for
        # diesel, so _cell_num returns None and the metric is simply omitted).
        soc_i = idx.get(SOC_CHANGE_COL)
        echg_i = idx.get(ENERGY_CHANGE_COL)
        ac_i = idx.get(ENERGY_AC_COL)
        dc_i = idx.get(ENERGY_DC_COL)
        ep_i = idx.get(ENERGY_PERF_COL)
        cap_i = idx.get(BATTERY_CAP_COL)
        # Operator code — drives the data-driven operator overlay. ``idx.get`` ->
        # None for artefacts predating the v2.2.5 column, so it degrades to the
        # plot_config fallback (operator simply stays None on every leg).
        op_i = idx.get(OPERATOR_COL)

        for row in ws.iter_rows(min_row=2):
            n = len(row)
            if start_i >= n:
                continue
            start_cell = row[start_i]
            day = _to_date(start_cell.value)
            if day is None:
                continue

            # Categories.
            cats = {TELEMATICS}  # a leg row exists for this date
            logger_link = (
                link_i is not None and link_i < n and _link_present(row[link_i])
            )
            if logger_link or any(
                i < n and _is_present(row[i].value) for i in pedal_i
            ):
                cats.add(LOGGER)
            if any(i < n and _is_present(row[i].value) for i in charger_i):
                cats.add(CHARGER)

            # Leg class.
            leg_type = row[type_i].value if (type_i is not None and type_i < n) else None
            leg_class = _classify_leg(leg_type)

            # Distance (driving legs only carry a real value).
            distance_km = None
            if dist_i is not None and dist_i < n:
                v = row[dist_i].value
                if isinstance(v, (int, float)) and not (
                    isinstance(v, float) and math.isnan(v)
                ):
                    distance_km = float(v)

            end_val = row[end_i].value if (end_i is not None and end_i < n) else None
            sig = (_iso_key(start_cell.value), _iso_key(end_val), leg_class)
            operator = (
                _clean_operator(row[op_i].value)
                if (op_i is not None and op_i < n)
                else None
            )
            # Raw per-event metrics carried through to build_day_segments, which
            # turns them into the drill-down info-label fields (dsoc/dkwh/cap/ep).
            metrics = {
                "dsoc": _cell_num(row, soc_i, n),
                "energy_change": _cell_num(row, echg_i, n),
                "ac": _cell_num(row, ac_i, n),
                "dc": _cell_num(row, dc_i, n),
                "ep": _cell_num(row, ep_i, n),
                "cap": _cell_num(row, cap_i, n),
            }
            legs.append(
                {
                    "day": day,
                    "sig": sig,
                    "cats": cats,
                    "leg_class": leg_class,
                    "distance_km": distance_km,
                    "metrics": metrics,
                    "operator": operator,
                }
            )
    finally:
        wb.close()
    return legs


def read_report_availability(xlsx_path: Path) -> dict[date, set[str]]:
    """Read one report workbook → ``{date: set_of_categories}`` (public helper)."""
    avail: dict[date, set[str]] = defaultdict(set)
    for leg in read_report_legs(xlsx_path):
        avail[leg["day"]].update(leg["cats"])
    return dict(avail)


def _collect_legs_by_reg(db_root: Path, version: str) -> dict[str, list[dict]]:
    """Scan ``<db_root>/<version>/<REG>/`` → ``{REG: [deduped leg records]}``.

    Merges each vehicle's multiple period files and de-duplicates legs on
    ``(start, end, leg-class)`` so overlapping files (e.g. a short diagnostic
    file contained in a full-range file, or boundary days shared by consecutive
    files) are counted once. ``*_finetuned.xlsx`` files are skipped (they cover
    the same period as a base file, already merged by the de-duplication).
    """
    root = db_root / version
    if not root.exists():
        raise FileNotFoundError(f"Report database directory not found: {root}")

    out: dict[str, list[dict]] = {}
    for reg_dir in sorted(root.iterdir()):
        if not reg_dir.is_dir():
            continue
        reg = reg_dir.name
        files = [
            f
            for f in sorted(reg_dir.glob("jolt_report_*.xlsx"))
            if not f.stem.endswith("_finetuned")
        ]
        if not files:
            continue
        merged: dict[tuple, dict] = {}
        for f in files:
            try:
                legs = read_report_legs(f)
            except Exception as exc:  # noqa: BLE001 — report which file is locked/bad
                LOG.error("FAIL reading %s: %s", f, exc)
                continue
            for leg in legs:
                sig = leg["sig"]
                if sig in merged:
                    merged[sig]["cats"].update(leg["cats"])
                    if merged[sig]["distance_km"] is None:
                        merged[sig]["distance_km"] = leg["distance_km"]
                    # Backfill any metric the first occurrence lacked (same
                    # physical leg, but a short diagnostic file may omit columns).
                    m0 = merged[sig].get("metrics") or {}
                    for k, v in (leg.get("metrics") or {}).items():
                        if m0.get(k) is None and v is not None:
                            m0[k] = v
                    merged[sig]["metrics"] = m0
                    # Same physical leg → backfill the operator if the first
                    # occurrence (e.g. a short diagnostic file) lacked it.
                    if merged[sig].get("operator") is None and leg.get("operator"):
                        merged[sig]["operator"] = leg["operator"]
                else:
                    merged[sig] = leg
        if merged:
            out[reg] = list(merged.values())
        else:
            LOG.warning("%s: no legs found", reg)
    return out


def _avail_from_legs(legs: list[dict]) -> dict[date, set[str]]:
    """Collapse a leg list into ``{date: set_of_categories}``."""
    avail: dict[date, set[str]] = defaultdict(set)
    for leg in legs:
        avail[leg["day"]].update(leg["cats"])
    return dict(avail)


def _parse_iso_dt(value) -> datetime | None:
    """Parse a leg ``sig`` start/end value (datetime or ISO string) to datetime."""
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str) or not value.strip():
        return None
    s = value.strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        try:
            return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
        except ValueError:
            return None


def build_day_segments(legs: list[dict]) -> dict[str, list[dict]]:
    """Regroup already-parsed report legs into per-date trip / charge segments.

    Reuses the leg records produced by :func:`read_report_legs` so segment
    semantics stay single-sourced with the dashboard's leg classification: the
    ISO start / end timestamps already live in ``leg["sig"][0]`` / ``[1]`` and
    the class in ``leg["leg_class"]``. ``Stop`` legs are dropped.

    Each segment is ``{"type": "trip"|"charge", "s": int, "e": int}`` plus the
    finite per-event metrics for the drill-down info labels (inspect.html style;
    any NaN field is omitted):

    - ``dsoc`` — ``SOC Change (%)`` (signed; both classes).
    - ``cap``  — ``Battery Capacity (kWh)``, the effective capacity ``C`` (both).
    - ``dkwh`` — signed energy: a trip uses ``Energy Change (kWh)``; a charge uses
      ``Energy Charged AC + DC (kWh)`` (the energy added).
    - ``ep``   — ``Energy Performance (kWh/km)`` (trip only; NaN for charges).

    ``s`` / ``e`` are seconds from that date's midnight — the same time base the
    detail-page channel series use. A leg whose end crosses midnight is clamped
    to ``86400`` on its start date. Imported by ``data_dashboard_detail`` so the
    drill-down pages and the calendar share one definition of a segment.
    """
    out: dict[str, list[dict]] = defaultdict(list)
    for leg in legs:
        cls = leg.get("leg_class")
        if cls not in ("trip", "charge"):
            continue
        start_dt = _parse_iso_dt(leg["sig"][0])
        if start_dt is None:
            continue
        end_dt = _parse_iso_dt(leg["sig"][1]) or start_dt
        day = start_dt.date()
        midnight = datetime(day.year, day.month, day.day, tzinfo=start_dt.tzinfo)
        s = int((start_dt - midnight).total_seconds())
        e = int((end_dt - midnight).total_seconds())
        s = max(0, min(s, 86400))
        e = max(s, min(e, 86400))
        seg = {"type": cls, "s": s, "e": e}
        # Attach the finite per-event metrics (omit NaN fields). Trip shows the
        # signed Energy Change + EP; charge shows the AC+DC energy added. dSOC and
        # effective capacity (C) apply to both classes.
        m = leg.get("metrics") or {}
        if m.get("dsoc") is not None:
            seg["dsoc"] = m["dsoc"]
        if m.get("cap") is not None:
            seg["cap"] = m["cap"]
        if cls == "charge":
            ac, dc = m.get("ac"), m.get("dc")
            if ac is not None or dc is not None:
                seg["dkwh"] = (ac or 0.0) + (dc or 0.0)
        else:
            if m.get("energy_change") is not None:
                seg["dkwh"] = m["energy_change"]
            if m.get("ep") is not None:
                seg["ep"] = m["ep"]
        out[day.isoformat()].append(seg)
    return dict(out)


# ── Raw-file (on-disk) availability scanning ─────────────────────────────────
def _raw_date_sane(d: date) -> bool:
    """Return ``True`` if a parsed raw date is within ``[2024-01-01, today+1]``."""
    return _RAW_DATE_MIN <= d <= (date.today() + timedelta(days=1))


def _dates_from_filenames(dir_path: Path, pattern: re.Pattern) -> set[date]:
    """Collect sane dates parsed from the filenames in one raw sub-directory.

    Files whose name does not match ``pattern`` (group 1 = ``YYYY-MM-DD``) are
    skipped, as are sub-directories and dates outside the sanity window. A
    missing / unreadable directory yields an empty set (logged for the latter).
    """
    out: set[date] = set()
    if not dir_path.is_dir():
        return out
    try:
        entries = list(dir_path.iterdir())
    except OSError as exc:  # noqa: BLE001 — report which directory is unreadable
        LOG.error("FAIL listing %s: %s", dir_path, exc)
        return out
    for f in entries:
        if not f.is_file():
            continue
        m = pattern.match(f.name)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            continue
        if _raw_date_sane(d):
            out.add(d)
    return out


def _raw_charger_dates(reg_dir: Path) -> set[date]:
    """Dates from ``raw_charger/charger_transactions.csv`` ``start_time`` column.

    Only 5 vehicles have this file; its absence yields an empty set (no charger
    days in raw mode for the rest). ``start_time`` is ISO with a tz offset
    (e.g. ``2025-12-01 04:31:47+00:00``); the leading 10 chars give the date.
    """
    out: set[date] = set()
    csv_path = reg_dir / RAW_CHARGER_DIR / RAW_CHARGER_CSV
    if not csv_path.is_file():
        return out
    try:
        with csv_path.open(encoding="utf-8", newline="") as fh:
            for row in csv.DictReader(fh):
                ts = (row.get("start_time") or "").strip()
                if not ts:
                    continue
                d = _to_date(ts[:10])
                if d is not None and _raw_date_sane(d):
                    out.add(d)
    except OSError as exc:  # noqa: BLE001 — report which file is locked / unreadable
        LOG.error("FAIL reading charger CSV %s: %s", csv_path, exc)
    return out


def scan_raw_availability(reg_dir: Path) -> dict[date, set[str]]:
    """Scan one reg directory's raw files → ``{date: set_of_categories}``.

    Pure raw-file view (no union with report data): telematics from
    ``raw_telematics/`` filenames, logger from the union of ``raw_logger_v1/`` +
    ``raw_logger_v2/`` filenames, charger from the transactions CSV.
    """
    avail: dict[date, set[str]] = defaultdict(set)
    for d in _dates_from_filenames(reg_dir / RAW_TELEMATICS_DIR, _TELEMATICS_FNAME_RE):
        avail[d].add(TELEMATICS)
    logger_dates: set[date] = set()
    for sub in RAW_LOGGER_DIRS:
        logger_dates |= _dates_from_filenames(reg_dir / sub, _LOGGER_FNAME_RE)
    for d in logger_dates:
        avail[d].add(LOGGER)
    for d in _raw_charger_dates(reg_dir):
        avail[d].add(CHARGER)
    return dict(avail)


def scan_report_database(
    db_root: Path, version: str
) -> dict[str, dict[date, set[str]]]:
    """Scan the database → ``{REG: {date: set_of_categories}}`` (public helper)."""
    return {
        reg: _avail_from_legs(legs)
        for reg, legs in _collect_legs_by_reg(db_root, version).items()
    }


def compute_vehicle_stats(
    legs: list[dict],
    avail: dict[date, set[str]],
    cfg: dict,
    avail_raw: dict[date, set[str]] | None = None,
) -> dict:
    """Compute the per-vehicle stat block from its deduped legs + config entry.

    ``avail`` drives the report-based (Events) day counts; the optional
    ``avail_raw`` (on-disk raw scan) drives a parallel set of raw day counts so
    both availability bases are visible regardless of the active calendar mode.
    """
    is_diesel = str(cfg.get("fuel_type", "")).upper() == "DIESEL"
    dates = sorted(avail.keys())
    n_trips = sum(1 for leg in legs if leg["leg_class"] == "trip")
    n_charges = sum(1 for leg in legs if leg["leg_class"] == "charge")
    n_stops = sum(1 for leg in legs if leg["leg_class"] == "stop")
    dist_km = sum(
        leg["distance_km"]
        for leg in legs
        if leg["leg_class"] == "trip" and leg["distance_km"]
    )
    days = {
        c: sum(1 for cats in avail.values() if c in cats) for c in CATEGORY_ORDER
    }
    raw = avail_raw or {}
    raw_days = {
        c: sum(1 for cats in raw.values() if c in cats) for c in CATEGORY_ORDER
    }
    return {
        "fuel": "Diesel" if is_diesel else "EV",
        "make": cfg.get("make"),
        "model": cfg.get("model"),
        "nominal_kwh": cfg.get("nominal_kwh"),
        "effective_kwh": cfg.get("effective_capacity_kwh"),
        "weight_class_t": cfg.get("weight_class_t"),
        "date_first": dates[0].isoformat() if dates else None,
        "date_last": dates[-1].isoformat() if dates else None,
        "active_days": len(dates),
        "dist_km": round(dist_km, 1),
        "n_trips": n_trips,
        "n_charges": n_charges,
        "n_stops": n_stops,
        "days_telematics": days[TELEMATICS],
        "days_logger": days[LOGGER],
        "days_charger": days[CHARGER],
        "raw_days_telematics": raw_days[TELEMATICS],
        "raw_days_logger": raw_days[LOGGER],
        "raw_days_charger": raw_days[CHARGER],
    }


def scan_report_database_full(
    db_root: Path, version: str, legs_by_reg: dict[str, list[dict]] | None = None
) -> dict[str, dict]:
    """Scan the database → ``{REG: {"avail", "avail_raw", "stats"}}``.

    Reads each workbook once (Events basis) and additionally scans each reg
    directory's raw files (Raw basis), returning both per-date availability maps
    and the computed stat block per vehicle (the structure the renderer embeds).
    ``legs_by_reg`` may be supplied (already read elsewhere, e.g. for the
    drill-down detail pages) to avoid re-reading every workbook.
    """
    cfg = _load_vehicles_cfg()
    root = db_root / version
    if legs_by_reg is None:
        legs_by_reg = _collect_legs_by_reg(db_root, version)
    out: dict[str, dict] = {}
    for reg, legs in legs_by_reg.items():
        avail = _avail_from_legs(legs)
        avail_raw = scan_raw_availability(root / reg)
        out[reg] = {
            "avail": avail,
            "avail_raw": avail_raw,
            # Per-day operator (data-driven overlay source); not embedded in the
            # HTML directly — consumed by build_operator_assignment to build the
            # compact periods list. Empty for diesel/EV reports with no operator.
            "operator_days": _operator_days_from_legs(legs),
            "stats": compute_vehicle_stats(legs, avail, cfg.get(reg, {}), avail_raw),
        }
    return out


# ── Rendering ────────────────────────────────────────────────────────────────
# Static stylesheet (no dynamic values; injected verbatim). The right panel is a
# traditional month-calendar grid: each month block is a 7-column day grid whose
# column width (.cal-month-days / .cal-month-wd) sets the block width; the JS that
# builds the blocks need only agree on the Monday-first weekday order.
_CSS = """
* { box-sizing: border-box; }
html, body { height: 100%; }
body {
  margin: 0; height: 100vh; display: flex; flex-direction: column;
  font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
  color: #1f2937; background: #f3f4f6; font-size: 14px;
}
header.topbar {
  padding: 14px 22px; background: #111827; color: #f9fafb;
  display: flex; align-items: baseline; gap: 18px; flex-wrap: wrap;
}
header.topbar h1 { margin: 0; font-size: 18px; font-weight: 650; }
header.topbar .meta { font-size: 12.5px; color: #cbd5e1; line-height: 1.6; }
header.topbar .meta code {
  background: #1f2937; color: #e5e7eb; padding: 1px 6px; border-radius: 4px;
}
.dashboard { flex: 1; display: flex; min-height: 0; }
.panel { display: flex; flex-direction: column; min-height: 0; background: #fff; }
.panel-left { width: 210px; border-right: 1px solid #e5e7eb; }
.panel-mid { width: 350px; border-right: 1px solid #e5e7eb; }
.panel-right { flex: 1; min-width: 0; }
.panel-head {
  padding: 12px 16px; font-size: 12.5px; font-weight: 700; color: #374151;
  text-transform: uppercase; letter-spacing: .05em;
  border-bottom: 1px solid #eef2f7; background: #fafbfc; flex: 0 0 auto;
}
.panel-body { padding: 12px; overflow: auto; flex: 1; min-height: 0; }
.panel-left .panel-body { padding: 8px; }

/* left list */
.veh-item {
  display: flex; align-items: center; justify-content: space-between; gap: 8px;
  padding: 9px 12px; border-radius: 8px; cursor: pointer; margin-bottom: 3px;
  border: 1px solid transparent;
}
.veh-item:hover { background: #f3f4f6; }
.veh-item.active { background: #eef2ff; border-color: #c7d2fe; }
.veh-reg { font-weight: 600; font-size: 14px; letter-spacing: .02em; }
.veh-badge {
  font-size: 10.5px; padding: 1px 8px; border-radius: 999px; font-weight: 600;
}

/* fuel pills */
.pill { font-size: 12px; padding: 2px 11px; border-radius: 999px; font-weight: 600; }
.pill-ev { background: #dcfce7; color: #166534; }
.pill-dz { background: #fef3c7; color: #92400e; }

/* middle */
.info-hero { padding: 4px 6px 14px; border-bottom: 1px solid #eef2f7; margin-bottom: 14px; }
.info-reg { font-size: 25px; font-weight: 700; letter-spacing: .03em; }
.info-hero .pill { margin-top: 8px; display: inline-block; }
.info-sub { margin-top: 9px; color: #6b7280; font-size: 13.5px; }
.cards { display: flex; gap: 8px; margin-bottom: 16px; }
.card {
  flex: 1; background: #f8fafc; border: 1px solid #eef2f7; border-radius: 10px;
  padding: 11px 8px; text-align: center;
}
.card-v { font-size: 19px; font-weight: 700; color: #111827; }
.card-l { font-size: 11px; color: #6b7280; margin-top: 3px; }
.info-section { margin-bottom: 16px; }
/* Coverage group that matches the active calendar mode gets a subtle accent. */
.info-section.mode-active {
  background: #eef2ff; border: 1px solid #c7d2fe; border-radius: 8px;
  padding: 8px 8px 4px; margin-left: -8px; margin-right: -8px;
}
.info-section.mode-active .info-title { color: #4f46e5; }
.info-title {
  font-size: 11.5px; font-weight: 700; color: #9ca3af; text-transform: uppercase;
  letter-spacing: .05em; margin: 0 0 6px 2px;
}
.info-title .mode-tag {
  margin-left: 6px; font-size: 9.5px; font-weight: 700; color: #4f46e5;
  background: #e0e7ff; padding: 1px 6px; border-radius: 999px;
  letter-spacing: .03em; vertical-align: middle;
}
.kv {
  display: flex; justify-content: space-between; align-items: center;
  padding: 6px 8px; border-radius: 6px; font-size: 13.5px;
}
.kv:nth-child(even) { background: #fafbfc; }
.kv .k { color: #6b7280; display: flex; align-items: center; gap: 8px; }
.kv .v { font-weight: 600; color: #111827; text-align: right; }
.dot { width: 11px; height: 11px; border-radius: 3px; display: inline-block; }

/* right / calendar — traditional month-grid view */
/* Calendar panel head carries the title + the Events/Raw segmented toggle. */
.cal-head {
  display: flex; align-items: center; justify-content: space-between; gap: 12px;
  padding-top: 8px; padding-bottom: 8px;
}
.seg {
  display: inline-flex; border: 1px solid #d1d5db; border-radius: 8px;
  overflow: hidden; flex: 0 0 auto;
}
.seg-btn {
  border: 0; background: #fff; color: #6b7280; font: inherit; font-size: 12px;
  font-weight: 600; text-transform: none; letter-spacing: .02em;
  padding: 5px 14px; cursor: pointer;
}
.seg-btn + .seg-btn { border-left: 1px solid #d1d5db; }
.seg-btn:hover { background: #f3f4f6; }
.seg-btn.active { background: #111827; color: #fff; }
/* Semantics caption above the legend. */
.cal-caption {
  font-size: 12px; color: #6b7280; margin: 2px 4px 10px; flex: 0 0 auto;
  line-height: 1.5;
}
.cal-caption strong { color: #374151; font-weight: 650; }
.legend {
  display: flex; flex-wrap: wrap; gap: 16px; align-items: center;
  font-size: 12.5px; color: #374151; margin: 2px 4px 14px; flex: 0 0 auto;
}
.legend-item { display: inline-flex; align-items: center; gap: 6px; }
.swatch {
  width: 14px; height: 14px; border-radius: 3px; display: inline-block;
  border: 1px solid rgba(0,0,0,.08);
}
/* Operator legend — hollow rings (border colour only) so it reads as a distinct
   key from the solid-fill category swatches even when the colours overlap. */
.operator-legend { margin-top: -6px; }
.legend-label {
  font-size: 11.5px; font-weight: 700; color: #6b7280; text-transform: uppercase;
  letter-spacing: .05em; margin-right: 2px;
}
.swatch-ring {
  width: 14px; height: 14px; border-radius: 3px; display: inline-block;
  border: 2px solid #94a3b8; background: transparent; box-sizing: border-box;
}
/* Operator assignment list in the middle panel (round-robin period sequence). */
.op-periods { margin-top: 2px; }
.op-period {
  display: flex; align-items: center; gap: 8px;
  padding: 5px 8px; border-radius: 6px; font-size: 13px;
}
.op-period:nth-child(even) { background: #fafbfc; }
.op-name { font-weight: 600; color: #111827; }
.op-range {
  color: #6b7280; margin-left: auto; font-size: 12px;
  font-variant-numeric: tabular-nums; white-space: nowrap;
}
/* The right panel body stacks a fixed legend over a vertically-scrolling grid. */
.cal-panel-body { display: flex; flex-direction: column; overflow: hidden; }
.cal-scroll-area { flex: 1; min-height: 0; overflow-y: auto; padding: 2px 4px 10px; }
.cal-grid {
  display: flex; flex-wrap: wrap; gap: 22px 26px; align-content: flex-start;
}
.cal-month-block { flex: 0 0 auto; }
.cal-month-title {
  font-size: 14.5px; font-weight: 700; color: #1f2937; margin: 0 0 8px 2px;
}
.cal-month-wd {
  display: grid; grid-template-columns: repeat(7, 34px); gap: 4px; margin-bottom: 5px;
}
.cal-month-wd > div {
  font-size: 11px; font-weight: 600; color: #6b7280; text-align: center;
  line-height: 16px;
}
.cal-month-days { display: grid; grid-template-columns: repeat(7, 34px); gap: 4px; }
.cal-day {
  height: 34px; border-radius: 6px; display: flex;
  align-items: center; justify-content: center;
  font-size: 12.5px; font-weight: 600;
}
.cal-day.empty-day { background: transparent; }
.cal-day.no-data { background: #f1f5f9; color: #475569; }
.cal-day.has-data { color: #fff; text-shadow: 0 1px 2px rgba(0,0,0,.45); }
/* Drill-down anchor wrapping a data-bearing cell (keeps the cell's own look). */
.cal-link { text-decoration: none; color: inherit; display: inline-block; cursor: pointer; }
.cal-link:hover .cal-day { outline: 2px solid rgba(17,24,39,.55); outline-offset: 1px; }
.empty-note { color: #9ca3af; font-size: 13px; padding: 14px; }
"""

# Static client-side render logic (no f-string interpolation: literal JS braces).
# The DATA / CAT / CAT_LABELS / EMPTY consts are injected as a prelude above this.
_JS = r"""
const ORDER = ["t", "l", "c"];
const MONTHS_FULL = ["January", "February", "March", "April", "May", "June",
                     "July", "August", "September", "October", "November", "December"];
// Monday-first weekday headers (matches the reference Windows-10 month calendar).
const WD_MIN = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"];

// Calendar render state: the selected vehicle + the active availability basis.
let CUR_REG = null;
let MODE = "events";  // "events" (report-based) | "raw" (on-disk raw files)

function esc(s) {
  return (s == null ? "" : String(s))
    .replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;")
    .replace(/"/g, "&quot;");
}
function fmtNum(x, d) {
  if (x == null || x === "" || isNaN(x)) return "—";
  return Number(x).toLocaleString("en-GB", {maximumFractionDigits: (d == null ? 0 : d)});
}
function val(x) { return (x == null || x === "") ? "—" : x; }

// Vertical multi-colour cell background (matches Python _combo_background).
function comboBackground(code) {
  if (!code) return EMPTY;
  const cols = ORDER.filter(k => code.indexOf(k) >= 0).map(k => CAT[k]);
  if (cols.length === 0) return EMPTY;
  if (cols.length === 1) return cols[0];
  const n = cols.length, stops = [];
  for (let i = 0; i < n; i++) {
    stops.push(cols[i] + " " + (100 * i / n).toFixed(4) + "% "
               + (100 * (i + 1) / n).toFixed(4) + "%");
  }
  return "linear-gradient(to bottom, " + stops.join(", ") + ")";
}

// Resolve the operator colour governing a given ISO day (YYYY-MM-DD) for a
// vehicle, or null when the vehicle has no operator assignment at all (→ no
// border). A fixed vehicle carries one open-ended period (s=e=null) so every
// day matches; a round-robin day that falls in a gap between periods returns the
// neutral grey so coverage gaps stay visible. ISO date strings compare
// lexicographically, so plain >= / <= bounds checks are correct.
function operatorColourOn(k, v) {
  const periods = v.periods || [];
  if (!periods.length) return null;
  for (const p of periods) {
    if ((p.s == null || k >= p.s) && (p.e == null || k <= p.e)) {
      return DATA.operatorColors[p.c] || DATA.neutralColour;
    }
  }
  return DATA.neutralColour;
}

// Inclusive list of {y, m} (m is 0-based) from "YYYY-MM" first → "YYYY-MM" last.
// The range is fleet-wide, so every vehicle is rendered on the same month axis.
function monthList(first, last) {
  const fp = first.split("-"), lp = last.split("-");
  let y = +fp[0], m = +fp[1] - 1;
  const ey = +lp[0], em = +lp[1] - 1;
  const out = [];
  while (y < ey || (y === ey && m <= em)) {
    out.push({y: y, m: m});
    if (++m > 11) { m = 0; y++; }
  }
  return out;
}

// One Windows-10-style month block: title, Mo-first weekday row, day grid with
// leading/trailing blanks so the 1st lands under the correct weekday. When the
// vehicle has a drill-down page (v.hasDetail), each data-bearing cell is wrapped
// in an <a> linking to detail_<reg>.html?date=…&mode=… (MODE is the current
// basis; the calendar re-renders on toggle so the link stays in sync).
function monthBlock(y, m, avail, v, reg) {
  // JS getDay(): 0=Sun..6=Sat → Monday-first index 0=Mon..6=Sun.
  const lead = (new Date(y, m, 1).getDay() + 6) % 7;
  const nDays = new Date(y, m + 1, 0).getDate();
  let h = '<div class="cal-month-block">';
  h += '<div class="cal-month-title">' + MONTHS_FULL[m] + ' ' + y + '</div>';
  h += '<div class="cal-month-wd">';
  WD_MIN.forEach(w => { h += '<div>' + w + '</div>'; });
  h += '</div><div class="cal-month-days">';
  for (let i = 0; i < lead; i++) h += '<div class="cal-day empty-day"></div>';
  for (let d = 1; d <= nDays; d++) {
    const k = y + "-" + String(m + 1).padStart(2, "0") + "-" + String(d).padStart(2, "0");
    const code = avail[k];
    if (!code) {
      // Empty (no-data) cells never get an operator border.
      h += '<div class="cal-day no-data" title="' + k + '">' + d + '</div>';
    } else {
      let title = k + ": " + code.split("").map(c => CAT_LABELS[c]).join(", ");
      // Operator border = who operated the vehicle on this (data-bearing) date.
      // box-sizing: border-box keeps the 34px cell size stable, so the 3px
      // border never re-flows the grid. Unknown operator → no border.
      const opc = operatorColourOn(k, v);
      let border = "";
      if (opc) {
        border = ";border:3px solid " + opc;
        const named = (v.periods || []).find(
          p => (p.s == null || k >= p.s) && (p.e == null || k <= p.e));
        title += " · " + (named
          ? (DATA.operatorNames[named.c] || named.c) : "Outside assignment");
      }
      let cell = '<div class="cal-day has-data" style="background:'
        + comboBackground(code) + border + '" title="' + esc(title) + '">'
        + d + '</div>';
      // Drill-down link: only when the vehicle has a detail page. The cell keeps
      // its style / title / operator border; the anchor just adds the click.
      if (v.hasDetail) {
        const href = 'detail_' + reg + '.html?date=' + k + '&mode=' + MODE;
        cell = '<a class="cal-link" href="' + href + '">' + cell + '</a>';
      }
      h += cell;
    }
  }
  const trail = (7 - ((lead + nDays) % 7)) % 7;
  for (let i = 0; i < trail; i++) h += '<div class="cal-day empty-day"></div>';
  return h + '</div></div>';
}

function card(v, l) {
  return '<div class="card"><div class="card-v">' + v
    + '</div><div class="card-l">' + esc(l) + '</div></div>';
}
function section(title, rows, active) {
  const cls = active ? "info-section mode-active" : "info-section";
  const tag = active ? '<span class="mode-tag">active</span>' : "";
  let h = '<div class="' + cls + '"><div class="info-title">' + esc(title) + tag + '</div>';
  for (const [k, v] of rows) {
    h += '<div class="kv"><span class="k">' + k + '</span><span class="v">' + v + '</span></div>';
  }
  return h + '</div>';
}

// Operator section for the middle panel: trial type + the operator assignment.
// Fixed vehicles show a single dot + name; round-robin vehicles show the period
// sequence with a colour dot and a compact, year-elided date range per span.
function operatorSection(v) {
  const tt = v.trial_type || "—";
  const periods = v.periods || [];
  let h = '<div class="info-section"><div class="info-title">Operator</div>';
  h += '<div class="kv"><span class="k">Trial type</span><span class="v">'
    + esc(tt) + '</span></div>';
  if (!periods.length) {
    // Operator unknown (vehicle absent from both assignment maps).
    h += '<div class="kv"><span class="k">Operator</span><span class="v">—</span></div>';
    return h + '</div>';
  }
  // Fixed = one open-ended whole-axis period → a single operator line.
  if (periods.length === 1 && periods[0].s == null && periods[0].e == null) {
    const p = periods[0];
    const col = DATA.operatorColors[p.c] || DATA.neutralColour;
    const nm = DATA.operatorNames[p.c] || p.c;
    h += '<div class="kv"><span class="k">Operator</span><span class="v">'
      + '<span class="dot" style="background:' + col + '"></span> ' + esc(nm)
      + '</span></div>';
    return h + '</div>';
  }
  // Round-robin: stacked period list. The date range elides the year whenever it
  // repeats the previously shown year (e.g. 2025-08-27→09-29, 09-30→2026-01-09).
  h += '<div class="op-periods">';
  let lastYear = null;
  const fmtD = iso => {
    if (!iso) return "…";
    const yr = iso.slice(0, 4);
    if (yr === lastYear) return iso.slice(5);   // MM-DD
    lastYear = yr;
    return iso;                                 // YYYY-MM-DD
  };
  periods.forEach(p => {
    const col = DATA.operatorColors[p.c] || DATA.neutralColour;
    const nm = DATA.operatorNames[p.c] || p.c;
    const s = fmtD(p.s), e = fmtD(p.e);
    h += '<div class="op-period"><span class="dot" style="background:' + col
      + '"></span><span class="op-name">' + esc(nm)
      + '</span><span class="op-range">' + esc(s) + '&rarr;' + esc(e)
      + '</span></div>';
  });
  return h + '</div></div>';
}

function renderInfo(reg) {
  const v = DATA.vehicles[reg];
  const s = v.stats;
  const isEV = s.fuel !== "Diesel";
  let h = "";
  h += '<div class="info-hero">';
  h += '<div class="info-reg">' + esc(reg) + '</div>';
  h += '<span class="pill ' + (isEV ? "pill-ev" : "pill-dz") + '">' + esc(s.fuel) + '</span>';
  h += '<div class="info-sub">' + esc(val(s.make)) + ' ' + esc(val(s.model)) + '</div>';
  h += '</div>';

  h += '<div class="cards">';
  h += card(fmtNum(s.active_days), "Active days");
  h += card(fmtNum(s.dist_km), "km driven");
  h += card(fmtNum(s.n_trips), "Trips");
  h += '</div>';

  h += section("Vehicle", [
    ["Make", esc(val(s.make))],
    ["Model", esc(val(s.model))],
    ["Type", esc(s.fuel)],
  ]);
  h += operatorSection(v);
  if (isEV) {
    h += section("Battery", [
      ["Nominal capacity", s.nominal_kwh != null ? fmtNum(s.nominal_kwh) + " kWh" : "—"],
      ["Effective capacity", s.effective_kwh != null ? fmtNum(s.effective_kwh, 1) + " kWh" : "—"],
    ]);
  } else {
    h += section("Class", [
      ["Weight class", s.weight_class_t != null ? fmtNum(s.weight_class_t, 1) + " t" : "—"],
    ]);
  }
  h += section("Activity", [
    ["Data range", esc(val(s.date_first)) + " → " + esc(val(s.date_last))],
    ["Active days", fmtNum(s.active_days)],
    ["Total distance", fmtNum(s.dist_km, 1) + " km"],
  ]);
  h += section("Events", [
    ["Trips", fmtNum(s.n_trips)],
    ["Charging events", fmtNum(s.n_charges)],
    ["Stops", fmtNum(s.n_stops)],
  ]);
  // Two parallel coverage groups (both always shown); the group matching the
  // active calendar mode is highlighted.
  const evActive = (MODE === "events");
  h += section("Coverage · Events (report)", [
    ['<span class="dot" style="background:' + CAT.t + '"></span>Telematics', fmtNum(s.days_telematics) + " d"],
    ['<span class="dot" style="background:' + CAT.l + '"></span>Logger', fmtNum(s.days_logger) + " d"],
    ['<span class="dot" style="background:' + CAT.c + '"></span>Charger', fmtNum(s.days_charger) + " d"],
  ], evActive);
  h += section("Coverage · Raw data (files)", [
    ['<span class="dot" style="background:' + CAT.t + '"></span>Telematics', fmtNum(s.raw_days_telematics) + " d"],
    ['<span class="dot" style="background:' + CAT.l + '"></span>Logger', fmtNum(s.raw_days_logger) + " d"],
    ['<span class="dot" style="background:' + CAT.c + '"></span>Charger', fmtNum(s.raw_days_charger) + " d"],
  ], !evActive);
  document.getElementById("info").innerHTML = h;
}

function renderCalendar(reg) {
  const v = DATA.vehicles[reg];
  // Pick the availability basis selected by the toggle (Events vs Raw data).
  const avail = (MODE === "raw") ? (v.avail_raw || {}) : v.avail;
  const el = document.getElementById("calendar");
  // Fleet-wide month axis (spans BOTH bases of all vehicles): identical month
  // blocks for every vehicle and every mode, so toggling never re-flows the
  // blocks — only cell colours change. Days with no data render light-grey.
  const months = monthList(DATA.firstMonth, DATA.lastMonth);
  if (!months.length) { el.innerHTML = '<div class="empty-note">No data.</div>'; return; }
  let h = '<div class="cal-grid">';
  months.forEach(ym => { h += monthBlock(ym.y, ym.m, avail, v, reg); });
  el.innerHTML = h + '</div>';
}

// Per-vehicle operator legend (hollow rings), shown next to the category legend.
// Only the operators relevant to the selected vehicle are listed (1 for fixed,
// N for round-robin, in assignment order); the neutral "Outside assignment"
// entry is added only when some data day in the ACTIVE mode falls in a gap, so
// it is recomputed on both vehicle change and Events/Raw toggle. Vehicles with
// no operator assignment hide the legend entirely.
function renderOperatorLegend(reg) {
  const v = DATA.vehicles[reg];
  const el = document.getElementById("op-legend");
  const periods = v.periods || [];
  if (!periods.length) { el.style.display = "none"; el.innerHTML = ""; return; }
  el.style.display = "";
  // Unique companies in assignment order.
  const seen = {}, companies = [];
  periods.forEach(p => { if (!seen[p.c]) { seen[p.c] = 1; companies.push(p.c); } });
  let h = '<span class="legend-label">Operator</span>';
  companies.forEach(c => {
    const col = DATA.operatorColors[c] || DATA.neutralColour;
    const nm = DATA.operatorNames[c] || c;
    h += '<span class="legend-item"><span class="swatch-ring" style="border-color:'
      + col + '"></span>' + esc(nm) + '</span>';
  });
  // Neutral "Outside assignment" — only for round-robin vehicles with at least
  // one data day (active mode) in a gap between periods.
  if (v.trial_type === "Round-robin") {
    const avail = (MODE === "raw") ? (v.avail_raw || {}) : v.avail;
    let hasGap = false;
    for (const k in avail) {
      if (operatorColourOn(k, v) === DATA.neutralColour) { hasGap = true; break; }
    }
    if (hasGap) {
      h += '<span class="legend-item"><span class="swatch-ring" style="border-color:'
        + DATA.neutralColour + '"></span>Outside assignment</span>';
    }
  }
  el.innerHTML = h;
}

// Switch the availability basis and re-render the active vehicle in place.
function setMode(m) {
  if (m === MODE) return;
  MODE = m;
  document.querySelectorAll(".seg-btn").forEach(
    b => b.classList.toggle("active", b.dataset.mode === m));
  if (CUR_REG) {
    renderInfo(CUR_REG); renderCalendar(CUR_REG); renderOperatorLegend(CUR_REG);
  }
}

function selectReg(reg) {
  CUR_REG = reg;
  document.querySelectorAll(".veh-item").forEach(
    n => n.classList.toggle("active", n.dataset.reg === reg));
  renderInfo(reg);
  renderCalendar(reg);
  renderOperatorLegend(reg);
}

function init() {
  let h = "";
  DATA.regs.forEach(reg => {
    const s = DATA.vehicles[reg].stats;
    const isEV = s.fuel !== "Diesel";
    h += '<div class="veh-item" data-reg="' + esc(reg) + '" onclick="selectReg(\'' + reg + '\')">'
      + '<span class="veh-reg">' + esc(reg) + '</span>'
      + '<span class="veh-badge ' + (isEV ? "pill-ev" : "pill-dz") + '">' + esc(s.fuel) + '</span></div>';
  });
  document.getElementById("veh-list").innerHTML = h;
  if (DATA.regs.length) selectReg(DATA.regs[0]);
}
document.addEventListener("DOMContentLoaded", init);
"""


def render_dashboard_html(
    full: dict[str, dict],
    *,
    version: str,
    source_path: Path,
    detail_regs: set[str] | None = None,
) -> str:
    """Render the scan result to a single self-contained interactive HTML doc.

    ``detail_regs`` is the set of registrations that have a ``detail_<REG>.html``
    drill-down page (generated this run or already on disk); each such vehicle is
    embedded with ``hasDetail: true`` so its calendar cells link to that page.
    """
    regs = sorted(full.keys())
    if not regs:
        raise ValueError("No availability data to render.")
    detail_regs = detail_regs or set()

    # Fleet-wide month axis spans BOTH bases of ALL vehicles, so toggling
    # between Events / Raw never re-flows the month blocks (only colours change).
    all_dates = [d for reg in regs for d in full[reg]["avail"]]
    all_dates += [
        d for reg in regs for d in full[reg].get("avail_raw", {})
    ]
    if not all_dates:
        raise ValueError("No availability data to render.")
    d_min, d_max = min(all_dates), max(all_dates)

    # Operator assignment + trial type, DATA-DRIVEN from each vehicle's reports'
    # ``Operator`` column (the SRF-resolved code), with plot_config.json kept only
    # as a fallback for vehicles with no operator in the data. Vehicles with no
    # operator anywhere get an empty period list and "—" trial type, handled
    # gracefully by the JS.
    op_days_by_reg = {reg: full[reg].get("operator_days", {}) for reg in regs}
    operator = build_operator_assignment(_load_plot_config(), op_days_by_reg)
    op_vehicles = operator["vehicles"]

    # Embedded data blob: per reg, the stat block + operator assignment + two
    # compact date->"tlc" maps (``avail`` = report-based Events, ``avail_raw`` =
    # on-disk Raw files).
    vehicles = {
        reg: {
            "stats": full[reg]["stats"],
            "trial_type": op_vehicles.get(reg, {}).get("trial_type", "—"),
            "periods": op_vehicles.get(reg, {}).get("periods", []),
            "hasDetail": reg in detail_regs,
            "avail": {
                d.isoformat(): _combo_key(cats)
                for d, cats in sorted(full[reg]["avail"].items())
            },
            "avail_raw": {
                d.isoformat(): _combo_key(cats)
                for d, cats in sorted(full[reg].get("avail_raw", {}).items())
            },
        }
        for reg in regs
    }
    data_obj = {
        "version": version,
        "regs": regs,
        # Fleet-wide month axis (shared by every vehicle's calendar): the global
        # min/max month across all vehicles, e.g. "2024-06" → "2026-04".
        "firstMonth": d_min.strftime("%Y-%m"),
        "lastMonth": d_max.strftime("%Y-%m"),
        # Fleet-level operator colour key + prettified display names + the
        # neutral grey used for round-robin gaps (referenced by the JS overlay).
        "operatorColors": operator["operatorColors"],
        "operatorNames": operator["operatorNames"],
        "neutralColour": operator["neutralColour"],
        "vehicles": vehicles,
    }
    # Escape "</" so a stray substring can never close the inline <script>.
    data_json = json.dumps(data_obj, ensure_ascii=False).replace("</", "<\\/")
    cat_short = {_CATEGORY_KEY[c]: CATEGORY_COLOURS[c] for c in CATEGORY_ORDER}
    cat_labels_short = {_CATEGORY_KEY[c]: CATEGORY_LABELS[c] for c in CATEGORY_ORDER}

    data_prelude = (
        "const DATA = " + data_json + ";\n"
        "const CAT = " + json.dumps(cat_short) + ";\n"
        "const CAT_LABELS = " + json.dumps(cat_labels_short) + ";\n"
        "const EMPTY = " + json.dumps(EMPTY_COLOUR) + ";\n"
    )

    # Static legend (built once; same for every vehicle).
    legend_items = [
        f'<span class="legend-item"><span class="swatch" '
        f'style="background:{CATEGORY_COLOURS[c]}"></span>{CATEGORY_LABELS[c]}</span>'
        for c in CATEGORY_ORDER
    ]
    legend_items.append(
        '<span class="legend-item"><span class="swatch" style="background:'
        f'{_combo_background(set(CATEGORY_ORDER))}"></span>Multiple</span>'
    )
    legend_items.append(
        '<span class="legend-item"><span class="swatch" style="background:'
        f'{EMPTY_COLOUR}"></span>No data</span>'
    )
    legend_html = "".join(legend_items)

    generated = date.today().isoformat()
    src_disp = html.escape(str(source_path))
    ver_disp = html.escape(version)

    head = (
        "<!DOCTYPE html>\n"
        '<html lang="en">\n<head>\n'
        '<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f"<title>JOLT data dashboard — {ver_disp}</title>\n"
        "<style>\n"
    )
    body = f"""</style>
</head>
<body>
<header class="topbar">
  <h1>JOLT data-availability dashboard</h1>
  <div class="meta">
    Version <strong>{ver_disp}</strong>
    &nbsp;&middot;&nbsp; Generated from <code>{src_disp}</code>
    &nbsp;&middot;&nbsp; {len(regs)} vehicles
    &nbsp;&middot;&nbsp; {d_min.isoformat()} &rarr; {d_max.isoformat()}
    &nbsp;&middot;&nbsp; built {generated}
  </div>
</header>
<div class="dashboard">
  <div class="panel panel-left">
    <div class="panel-head">Vehicles ({len(regs)})</div>
    <div class="panel-body" id="veh-list"></div>
  </div>
  <div class="panel panel-mid">
    <div class="panel-head">Vehicle info</div>
    <div class="panel-body" id="info"></div>
  </div>
  <div class="panel panel-right">
    <div class="panel-head cal-head">
      <span>Data availability</span>
      <div class="seg" role="group" aria-label="Availability basis">
        <button class="seg-btn active" data-mode="events"
                onclick="setMode('events')">Events</button>
        <button class="seg-btn" data-mode="raw"
                onclick="setMode('raw')">Raw data</button>
      </div>
    </div>
    <div class="panel-body cal-panel-body">
      <div class="cal-caption">
        <strong>Events</strong> = days with processed legs in the Excel report
        &nbsp;&middot;&nbsp;
        <strong>Raw data</strong> = days with raw files fetched from SRF
      </div>
      <div class="legend">{legend_html}</div>
      <div class="legend operator-legend" id="op-legend"></div>
      <div class="cal-scroll-area"><div id="calendar"></div></div>
    </div>
  </div>
</div>
<script>
"""
    tail = "\n</script>\n</body>\n</html>\n"
    return head + _CSS + body + data_prelude + _JS + tail


# ── Orchestration + CLI ──────────────────────────────────────────────────────
def generate_dashboard(
    db_root: Path,
    version: str,
    out_path: Path,
    *,
    detail_regs: set[str] | None = None,
    legs_by_reg: dict[str, list[dict]] | None = None,
) -> tuple[dict[str, dict], Path]:
    """Scan the database and write the dashboard HTML. Returns (full_scan, path).

    ``full_scan`` is ``{REG: {"avail", "avail_raw", "stats"}}`` (both bases), so
    callers can summarise the Events and Raw day counts side by side.
    ``detail_regs`` flags vehicles with a drill-down detail page (calendar links);
    ``legs_by_reg`` may be supplied to reuse already-read report legs.
    """
    full = scan_report_database_full(db_root, version, legs_by_reg=legs_by_reg)
    if not full:
        raise ValueError(f"No reports found under {db_root / version}")
    html_doc = render_dashboard_html(
        full, version=version, source_path=db_root / version,
        detail_regs=detail_regs,
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html_doc, encoding="utf-8")
    return full, out_path


def _day_counts(day_map: dict[date, set[str]]) -> dict[str, int]:
    """Per-category day counts for one ``{date: set_of_categories}`` map."""
    return {
        c: sum(1 for cats in day_map.values() if c in cats) for c in CATEGORY_ORDER
    }


def _print_summary(full: dict[str, dict]) -> None:
    """Log the per-vehicle Events-vs-Raw day-count table (T / L / C each)."""
    LOG.info("=" * 78)
    LOG.info("Per-vehicle day counts — Events (report) vs Raw (files), T/L/C:")
    LOG.info(
        "%-9s  %-22s  %-22s", "REG", "events  T    L    C", "raw     T    L    C"
    )
    LOG.info("-" * 78)
    tot_ev = {c: 0 for c in CATEGORY_ORDER}
    tot_raw = {c: 0 for c in CATEGORY_ORDER}
    for reg in sorted(full):
        ev = _day_counts(full[reg]["avail"])
        raw = _day_counts(full[reg].get("avail_raw", {}))
        for c in CATEGORY_ORDER:
            tot_ev[c] += ev[c]
            tot_raw[c] += raw[c]
        LOG.info(
            "%-9s  events %4d %4d %4d   raw  %4d %4d %4d",
            reg,
            ev[TELEMATICS], ev[LOGGER], ev[CHARGER],
            raw[TELEMATICS], raw[LOGGER], raw[CHARGER],
        )
    LOG.info("-" * 78)
    LOG.info(
        "%-9s  events %4d %4d %4d   raw  %4d %4d %4d   (%d vehicles)",
        "FLEET",
        tot_ev[TELEMATICS], tot_ev[LOGGER], tot_ev[CHARGER],
        tot_raw[TELEMATICS], tot_raw[LOGGER], tot_raw[CHARGER],
        len(full),
    )


def _project_root() -> Path:
    # src/jolt_toolkit/report_generator/data_dashboard.py -> repo root
    return Path(__file__).resolve().parents[3]


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--version",
        default=__version__,
        help=(
            "Report-database version subdirectory (default: the installed "
            f"jolt_toolkit version, currently {__version__})"
        ),
    )
    ap.add_argument(
        "--db-root",
        default=None,
        help="Report database root (default: <repo>/excel_report_database)",
    )
    ap.add_argument(
        "--out",
        default=None,
        help=(
            "Output HTML path (default: "
            "<db-root>/<version>/dashboard/data_dashboard.html). The dashboard "
            "and all detail_<REG>.html pages are written together in this folder."
        ),
    )
    ap.add_argument(
        "--details",
        default="none",
        help=(
            "Generate per-vehicle drill-down detail pages: 'none' (default), "
            "'all', or a comma-separated list of registrations (e.g. "
            "YK73WFN,LN25NKE)."
        ),
    )
    ap.add_argument(
        "--fetch-uplot",
        action="store_true",
        help="Download the vendored uPlot JS/CSS from jsDelivr, then continue.",
    )
    args = ap.parse_args(argv)

    logging.basicConfig(format="%(levelname)-5s %(message)s", level=logging.INFO)

    # One-off offline-asset fetch (the only network step in this module).
    if args.fetch_uplot:
        from jolt_toolkit.report_generator.data_dashboard_detail import fetch_uplot

        dest = fetch_uplot()
        LOG.info("uPlot assets refreshed under %s", dest)

    db_root = (
        Path(args.db_root)
        if args.db_root
        else _project_root() / "excel_report_database"
    )
    # Dashboard + every detail_<REG>.html live together in a dedicated
    # <version>/dashboard/ folder (out_dir). The raw-data SCAN still reads from
    # <version>/<REG>/raw_* (the parent version dir) — only the OUTPUT moves here.
    out_path = (
        Path(args.out)
        if args.out
        else db_root / args.version / "dashboard" / "data_dashboard.html"
    )
    out_dir = out_path.parent

    LOG.info("Database root: %s", db_root)
    LOG.info("Version:       %s", args.version)

    # Read every vehicle's report legs once and reuse them for both the dashboard
    # scan and the drill-down detail pages (segments come from these legs).
    legs_by_reg = _collect_legs_by_reg(db_root, args.version)
    if not legs_by_reg:
        raise ValueError(f"No reports found under {db_root / args.version}")

    # Resolve the --details request against the vehicles that actually have data.
    detail_request = _parse_details_arg(args.details, sorted(legs_by_reg))
    written_detail: dict[str, Path] = {}
    if detail_request:
        from jolt_toolkit.report_generator.data_dashboard_detail import (
            write_detail_pages,
        )

        LOG.info("Generating %d detail page(s): %s",
                 len(detail_request), ", ".join(detail_request))
        written_detail = write_detail_pages(
            db_root, args.version, detail_request,
            out_dir=out_dir, legs_by_reg=legs_by_reg,
        )

    # hasDetail = detail pages written this run ∪ pages already on disk, so the
    # calendar links stay live for previously-generated vehicles too.
    existing_detail = {
        p.stem[len("detail_"):]
        for p in out_dir.glob("detail_*.html")
    }
    detail_regs = set(written_detail) | existing_detail

    full, written = generate_dashboard(
        db_root, args.version, out_path,
        detail_regs=detail_regs, legs_by_reg=legs_by_reg,
    )
    _print_summary(full)
    LOG.info("=" * 78)
    if detail_regs:
        LOG.info("Detail pages live for: %s", ", ".join(sorted(detail_regs)))
    LOG.info("Dashboard written: %s", written)
    return 0


def _parse_details_arg(value: str, available: list[str]) -> list[str]:
    """Resolve the ``--details`` argument to a concrete list of registrations.

    ``none`` -> ``[]``; ``all`` -> every available reg; otherwise a comma list
    (unknown regs are dropped with a warning).
    """
    v = (value or "none").strip()
    if not v or v.lower() == "none":
        return []
    if v.lower() == "all":
        return list(available)
    avail = set(available)
    out: list[str] = []
    for reg in (r.strip() for r in v.split(",")):
        if not reg:
            continue
        if reg in avail:
            out.append(reg)
        else:
            LOG.warning("--details: %s not found in the database — skipped", reg)
    return out


if __name__ == "__main__":
    raise SystemExit(main())
