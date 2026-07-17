"""
charger_patcher.py
==================
Standalone charger data backfill tool.

Reads a generated Excel report, fetches charger events from the SRF API, and
backfills the report's missing Charger Link column.

Usage:
    from jolt_toolkit.report_generator.charger_patcher import ChargerPatcher
    patcher = ChargerPatcher()
    patcher.patch_file("excel_report_database/1.0.0/KY24LHT/jolt_report_KY24LHT_20250101_20250131.xlsx")
"""

from __future__ import annotations

import datetime
import logging
import os
from pathlib import Path

import pandas as pd
import srf_client
from openpyxl import load_workbook
from openpyxl.styles import Font
from srf_client import paging

from jolt_toolkit.report_generator.paths import get_cache_dir
from jolt_toolkit.report_generator.report_builder import (
    HEADERS,
    _build_charger_url,
)
from jolt_toolkit.report_generator.segment_algorithms import VEHICLE_CONFIG
from jolt_toolkit.report_generator.xlsx_patch_common import (
    _cell_is_empty,
    _parse_report_filename,
    _to_timestamp,
    make_srf_client,
)

logger = logging.getLogger(__name__)

# ── Excel column indices (1-based, openpyxl convention) ─────────────────
_COL_LEG_TYPE = 2
_COL_CHARGER = 4  # Charger Link
_COL_START_TIME = 6
_COL_END_TIME = 9
_COL_CHARGER_ENERGY = 33  # Energy Output from Charger (kWh)

# Cheap sanity check: the hard-coded 1-based Excel columns above must stay in
# step with report_builder.HEADERS (the xlsx has no extra leading column, so
# Excel col == HEADERS.index(name) + 1). If HEADERS is ever reordered this
# fails loudly at import instead of silently patching the wrong cell.
assert _COL_LEG_TYPE == HEADERS.index("Leg Type") + 1
assert _COL_CHARGER == HEADERS.index("Charger Link") + 1
assert _COL_START_TIME == HEADERS.index("Start Time (UTC)") + 1
assert _COL_END_TIME == HEADERS.index("End Time (UTC)") + 1
assert _COL_CHARGER_ENERGY == HEADERS.index("Energy Output from Charger (kWh)") + 1

# CSV filename for the accumulated per-vehicle charger transactions.
CHARGER_CSV_NAME = "charger_transactions.csv"

# Environment variable holding the SRF API key (also read from repo .env by the CLI).
_SRF_API_KEY_ENV = "SRF_API_KEY"


# ── Utility functions ─────────────────────────────────────────────────────
# _parse_report_filename / _cell_is_empty / _to_timestamp are shared with the
# logger patcher — see report_generator.xlsx_patch_common.


def _find_charger_matches(windows: list, t_start, t_end, tol_min: float = 4) -> list:
    """Collect ALL charger windows overlapping ``[t_start, t_end]`` (± tolerance).

    A single report leg can span several charger transactions: dual-gun chargers
    (e.g. a Nidec DC-360-360) record two interleaved meter series, and a long stop
    can contain multiple sequential sessions. Returning only the first match (the
    old behaviour) therefore dropped the energy of every extra transaction.

    Args:
        windows:  list of ``(start, end, uri, energy_kwh)`` tuples.
        t_start:  leg start time.
        t_end:    leg end time.
        tol_min:  overlap tolerance in minutes (each side).

    Returns:
        The overlapping windows as ``(ws, we, uri, energy_kwh)`` tuples, sorted by
        window start time (earliest first). Empty list if none overlap.
    """
    tol = pd.Timedelta(minutes=tol_min)
    t_s = pd.Timestamp(t_start)
    t_e = pd.Timestamp(t_end)
    if t_s.tzinfo is None:
        t_s = t_s.tz_localize("UTC")
    if t_e.tzinfo is None:
        t_e = t_e.tz_localize("UTC")
    ext_s = t_s - tol
    ext_e = t_e + tol
    matches = []
    for ws, we, uri, energy_kwh in windows:
        try:
            ws = pd.Timestamp(ws)
            we = pd.Timestamp(we)
            if ws.tzinfo is None:
                ws = ws.tz_localize("UTC")
            if we.tzinfo is None:
                we = we.tz_localize("UTC")
            if ws <= ext_e and we >= ext_s:
                matches.append((ws, we, uri, energy_kwh))
        except Exception:
            continue
    matches.sort(key=lambda m: m[0])
    return matches


# ── raw_charger CSV persistence (shared by _generator + the backfill CLI) ──────


def _charger_transaction_to_row(ct) -> dict:
    """Build one CSV row from a ``ChargerTransaction`` (schema unchanged since v2.2).

    Accesses ``ct.charger`` to record the charger metadata (label / make / model /
    max_power / dc), so callers that do not need metadata should avoid this path.
    """
    energy_kwh = None
    if ct.start_meter is not None and ct.end_meter is not None:
        energy_kwh = round(ct.end_meter - ct.start_meter, 3)
    charger = ct.charger
    return {
        "start_time": str(ct.start_time),
        "end_time": str(ct.end_time),
        "uri": ct.uri,
        "start_meter_kwh": ct.start_meter,
        "end_meter_kwh": ct.end_meter,
        "energy_delivered_kwh": energy_kwh,
        "charger_label": getattr(charger, "label", None),
        "charger_make": getattr(charger, "make", None),
        "charger_model": getattr(charger, "model", None),
        "charger_max_power_kw": getattr(charger, "max_power", None),
        "charger_dc": getattr(charger, "dc", None),
    }


def merge_save_charger_transactions(charger_objects, out_dir) -> int:
    """Merge-append charger transactions into ``raw_charger/charger_transactions.csv``.

    Each report-generation run only sees its own period's transactions, so a plain
    overwrite made the per-vehicle CSV forget everything outside the latest window.
    This helper instead builds one row per transaction, concatenates them with any
    existing CSV, de-duplicates by ``uri`` (keeping the newest occurrence), sorts by
    ``start_time`` and writes back. The schema is identical to the original dump, so
    downstream readers (dashboard raw-data base, data-collection monitor) are
    unaffected.

    Shared by ``_generator._save_charger_data`` and the backfill CLI so both persist
    identically.

    Returns the total row count after the merge (0 if nothing to write).
    """
    if not charger_objects:
        return 0
    out_dir = Path(out_dir)
    charger_dir = out_dir / "raw_charger"
    charger_dir.mkdir(exist_ok=True)

    rows = []
    for ct in charger_objects:
        try:
            rows.append(_charger_transaction_to_row(ct))
        except Exception as exc:
            logger.warning("Charger-transaction parse failed: %s", exc)
    if not rows:
        return 0

    df_new = pd.DataFrame(rows)
    csv_path = charger_dir / CHARGER_CSV_NAME
    if csv_path.exists():
        try:
            df_old = pd.read_csv(csv_path)
            # Align columns before concat so old CSVs missing a metadata column
            # still merge cleanly (new schema wins).
            df_new = pd.concat([df_old, df_new], ignore_index=True)
        except Exception as exc:
            logger.warning(
                "Existing charger CSV read failed, overwriting instead: %s", exc
            )

    # New rows are appended last → keep='last' retains the freshest copy of a uri.
    df_new = df_new.drop_duplicates(subset="uri", keep="last")
    df_new = df_new.sort_values("start_time").reset_index(drop=True)
    df_new.to_csv(csv_path, index=False)
    logger.info("Saved charger transactions: %d rows -> %s", len(df_new), csv_path.name)
    return len(df_new)


# ── Main class ────────────────────────────────────────────────────────────


class ChargerPatcher:
    """
    Charger data backfill tool.

    Reads a generated xlsx report file, fetches charger events from the SRF API,
    and writes the Charger Link URL into the report by time-window matching.

    Args:
        srf_data:    optional existing SRF client instance (share the connection to avoid recreating it)
        cache_dir:   SRF API cache directory
    """

    def __init__(self, srf_data=None, cache_dir: str | None = None):
        if cache_dir is None:
            cache_dir = str(get_cache_dir())
        if srf_data is not None:
            self._srf_data = srf_data
        else:
            api_key = os.environ.get("SRF_API_KEY")
            if not api_key:
                logger.warning("ChargerPatcher: SRF_API_KEY is not set")
            self._srf_data = make_srf_client(cache_dir, api_key=api_key)

    # ── Public interface ──────────────────────────────────────────────────

    def patch_file(
        self, xlsx_path: str | Path, *, charger_windows: list | None = None
    ) -> int:
        """
        Backfill a single xlsx report's Charger Link.

        Args:
            xlsx_path:        report file path
            charger_windows:  optional preloaded list of charger windows [(start, end, uri), ...]
                              if None, fetched automatically from the SRF API

        Returns:
            The number of rows backfilled.
        """
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            logger.error("File does not exist: %s", xlsx_path)
            return 0

        # 1. Obtain the charger windows
        if charger_windows is None:
            charger_windows = self._fetch_charger_windows(xlsx_path)
            if charger_windows is None:
                return 0

        if not charger_windows:
            logger.info(
                "ChargerPatcher: no charger events, skipping %s", xlsx_path.name
            )
            return 0

        # 2. Open the xlsx
        logger.info("ChargerPatcher: backfilling %s", xlsx_path.name)
        wb = load_workbook(str(xlsx_path))
        if "Report" not in wb.sheetnames:
            logger.error("  'Report' worksheet not found: %s", xlsx_path.name)
            wb.close()
            return 0
        ws = wb["Report"]

        # 3. Iterate rows, matching charge segments
        patched = 0
        for row_idx in range(2, ws.max_row + 1):
            if not _cell_is_empty(ws.cell(row_idx, _COL_CHARGER)):
                continue

            # Backfill the Charger Link only for charge segments. Besides the AC/DC
            # detail legs, the report builder also produces the generic types
            # 'Charge Home' / 'Charge Away' for charge legs whose AC/DC counters
            # cannot be attributed (all of EX74JXW / EX74JXY / LN25NKE / YN25RSY /
            # YN75NMA are like this, with a few scattered across every other
            # vehicle too), so the filter must also include 'Charge'.
            leg_type = ws.cell(row_idx, _COL_LEG_TYPE).value
            if not isinstance(leg_type, str):
                continue
            if not any(kw in leg_type for kw in ("AC", "DC", "Charge")):
                continue

            t_s = _to_timestamp(ws.cell(row_idx, _COL_START_TIME).value)
            t_e = _to_timestamp(ws.cell(row_idx, _COL_END_TIME).value)
            if t_s is None or t_e is None:
                continue

            matches = _find_charger_matches(charger_windows, t_s, t_e, tol_min=4)
            if not matches:
                continue

            # Sum the energy of ALL overlapping windows in this leg (multi-transaction
            # legs: dual-gun / consecutive sessions); the hyperlink points at the
            # earliest-starting window's charger URI (matches already sorted by ws ascending).
            uri = matches[0][2]
            energies = [m[3] for m in matches if m[3] is not None]
            total_energy = sum(energies) if energies else None

            url = _build_charger_url(uri, t_s, t_e)
            if url:
                cell = ws.cell(row_idx, _COL_CHARGER)
                cell.value = "Link"
                cell.hyperlink = url
                cell.font = Font(color="0000FF", underline="single")
            if total_energy is not None:
                ws.cell(row_idx, _COL_CHARGER_ENERGY).value = round(total_energy, 3)
            patched += 1

        if patched > 0:
            wb.save(str(xlsx_path))
            logger.info("  Backfilled %d Charger Link rows, saved", patched)
        else:
            logger.info("  No Charger Link to backfill")

        wb.close()
        return patched

    def patch_folder(self, folder_path: str | Path) -> dict[str, int]:
        """Backfill the Charger Link of all jolt_report_*.xlsx under a folder."""
        folder = Path(folder_path)
        if not folder.is_dir():
            logger.error("Folder does not exist: %s", folder)
            return {}

        xlsx_files = sorted(folder.glob("jolt_report_*.xlsx"))
        if not xlsx_files:
            logger.info("ChargerPatcher: no report files under %s", folder)
            return {}

        results = {}
        for fp in xlsx_files:
            results[fp.name] = self.patch_file(fp)
        return results

    # ── Internal methods ──────────────────────────────────────────────────

    def _fetch_charger_windows(self, xlsx_path: Path) -> list | None:
        """Fetch the list of charger-event windows from the SRF API (for patch_file's auto-fetch path)."""
        parsed = _parse_report_filename(xlsx_path)
        if parsed is None:
            logger.error(
                "  Cannot parse vehicle info from the file name: %s", xlsx_path.name
            )
            return None

        reg, ds_str, de_str = parsed
        cfg = VEHICLE_CONFIG.get(reg)
        if cfg is None:
            logger.error("  Vehicle %s is not registered in vehicles.json", reg)
            return None

        reg_srf = cfg["srf_reg"]
        ds = datetime.datetime.strptime(ds_str, "%Y%m%d")
        de = datetime.datetime.strptime(de_str, "%Y%m%d")

        objects = self._fetch_charger_transactions(reg_srf, ds, de)
        if objects is None:
            return []
        return self._windows_from_transactions(objects)

    def _fetch_charger_transactions(self, reg_srf: str, ds, de) -> list | None:
        """Fetch full ``ChargerTransaction`` objects for a vehicle over ``[ds, de]``.

        Returns the transaction objects (whose lazily-resolved ``.charger`` gives
        the metadata that ``merge_save_charger_transactions`` persists). ``None`` on
        an outright API failure. ``ds``/``de`` are dates (or datetimes); the query
        range is the full ``[ds 00:00, de 23:59:59]`` UTC day span.
        """
        logger.info(
            "  Fetching charger events from the SRF API: %s  %s ~ %s",
            reg_srf,
            ds.strftime("%Y%m%d"),
            de.strftime("%Y%m%d"),
        )
        params = {
            "start_time": srf_client.filter.between(
                datetime.datetime.combine(ds, datetime.time.min, datetime.timezone.utc),
                datetime.datetime.combine(de, datetime.time.max, datetime.timezone.utc),
            ),
            "sort": srf_client.sort.asc("startTime"),
        }
        objects = []
        try:
            for ct in paging.paged_items(
                self._srf_data.transactions.find_all(
                    **params, **{"vehicle.registration": reg_srf}
                )
            ):
                objects.append(ct)
        except Exception as exc:
            logger.warning("  Charger-event fetch failed: %s", exc)
            return None
        logger.info("  Fetched %d charger events", len(objects))
        return objects

    @staticmethod
    def _windows_from_transactions(objects: list) -> list:
        """Derive ``(start, end, uri, energy_kwh)`` windows from transaction objects.

        Only touches meter attributes (never ``.charger``), so it triggers no extra
        API round-trips when charger metadata is not needed.
        """
        windows = []
        for ct in objects:
            try:
                energy_kwh = None
                if ct.start_meter is not None and ct.end_meter is not None:
                    energy_kwh = ct.end_meter - ct.start_meter
                windows.append((ct.start_time, ct.end_time, ct.uri, energy_kwh))
            except AttributeError:
                pass
        return windows


# ── .env loading (dependency-free, a fallback for the CLI when the environment variable is missing) ──


def _load_srf_key_from_dotenv() -> None:
    """Populate ``SRF_API_KEY`` from the repo-root ``.env`` if it is unset.

    Dependency-free manual parse (``KEY=VALUE`` lines; ``#`` comments and blank
    lines ignored), mirroring the data-collection-monitor's run_monitor.py pattern.
    Walks up from this module until it finds a ``.env``; existing environment values
    are never overwritten.
    """
    if os.environ.get(_SRF_API_KEY_ENV):
        return
    for parent in Path(__file__).resolve().parents:
        env_path = parent / ".env"
        if not env_path.exists():
            continue
        try:
            for line in env_path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
        except Exception as exc:
            logger.warning(".env parse failed (%s): %s", env_path, exc)
        break


# ── Backfill CLI ─────────────────────────────────────────────────────────────


def _collect_report_files(target: Path) -> list:
    """Resolve a CLI target into the list of report xlsx files to patch.

    Accepts (a) a single ``jolt_report_*.xlsx`` file, (b) a vehicle directory
    holding such files directly, or (c) a version directory whose sub-directories
    are vehicle directories. ``*_finetuned*`` reports and the ``dashboard`` sub-dir
    are skipped.
    """
    if target.is_file():
        return [target]
    if not target.is_dir():
        return []

    def _reports_in(d: Path) -> list:
        return sorted(
            p for p in d.glob("jolt_report_*.xlsx") if "_finetuned" not in p.name
        )

    direct = _reports_in(target)
    if direct:
        return direct  # vehicle directory

    # version directory: iterate vehicle sub-directories
    reports = []
    for sub in sorted(target.iterdir()):
        if not sub.is_dir() or sub.name == "dashboard":
            continue
        reports.extend(_reports_in(sub))
    return reports


def main(argv: list | None = None) -> int:
    """CLI entry: backfill Charger Links into existing reports (idempotent).

    ``python -m jolt_toolkit.report_generator.charger_patcher <target> [--persist-raw]``

    For every report it fetches the SRF charger transactions for that report's own
    period (keyed by ``vehicle.registration``) and fills only the EMPTY Charger Link
    cells, so re-running patches nothing new. Diesel vehicles are skipped.
    """
    import argparse

    parser = argparse.ArgumentParser(
        prog="python -m jolt_toolkit.report_generator.charger_patcher",
        description="Backfill Charger Links into already-generated JOLT reports by "
        "fetching SRF charger transactions for each report period. Only "
        "empty Charger Link cells are filled (idempotent).",
    )
    parser.add_argument(
        "target",
        help="A single jolt_report_*.xlsx, a vehicle directory, or a version "
        "directory (e.g. excel_report_database/2.2.7).",
    )
    parser.add_argument(
        "--persist-raw",
        action="store_true",
        help="Also merge each vehicle's fetched transactions into its "
        "raw_charger/charger_transactions.csv (accumulating, deduped by uri).",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(message)s")
    _load_srf_key_from_dotenv()

    target = Path(args.target)
    reports = _collect_report_files(target)
    if not reports:
        logger.error("No jolt_report_*.xlsx found: %s", target)
        return 1

    patcher = ChargerPatcher()
    total_patched = 0
    n_reports = 0
    for xlsx_path in reports:
        parsed = _parse_report_filename(xlsx_path)
        if parsed is None:
            logger.warning("Skipping (file name unparseable): %s", xlsx_path.name)
            continue
        reg, ds_str, de_str = parsed
        cfg = VEHICLE_CONFIG.get(reg)
        if cfg is None:
            logger.warning(
                "Skipping (%s not registered in vehicles.json): %s", reg, xlsx_path.name
            )
            continue
        if str(cfg.get("fuel_type", "")).upper() == "DIESEL":
            logger.info("Skipping diesel vehicle %s: %s", reg, xlsx_path.name)
            continue

        reg_srf = cfg["srf_reg"]
        ds = datetime.datetime.strptime(ds_str, "%Y%m%d")
        de = datetime.datetime.strptime(de_str, "%Y%m%d")

        objects = patcher._fetch_charger_transactions(reg_srf, ds, de) or []
        windows = ChargerPatcher._windows_from_transactions(objects)

        patched = patcher.patch_file(xlsx_path, charger_windows=windows)
        total_patched += patched
        n_reports += 1
        logger.info("%s: +%d Charger Links", xlsx_path.name, patched)

        if args.persist_raw and objects:
            merge_save_charger_transactions(objects, xlsx_path.parent)

    logger.info(
        "Done: %d reports, %d Charger Links backfilled in total",
        n_reports,
        total_patched,
    )
    return 0


if __name__ == "__main__":
    import sys

    sys.exit(main())
